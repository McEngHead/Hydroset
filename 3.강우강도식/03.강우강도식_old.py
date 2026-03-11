import os
import sys
import json
import re
import traceback
import pandas as pd
import numpy as np
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib import rc
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from ctypes import windll, byref, sizeof, c_int
import warnings

# 경고 무시
warnings.filterwarnings("ignore")

# --- [CustomTkinter 설정] ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# 폰트 설정
FONT_TITLE = ("맑은 고딕", 22, "bold")
FONT_HEADER = ("맑은 고딕", 14, "bold")
FONT_BODY = ("맑은 고딕", 12)
FONT_BTN = ("맑은 고딕", 12, "bold")
FONT_LOG = ("consolas", 11)

# -----------------------------------------------------------
# 1. 수학적 모델 정의
# -----------------------------------------------------------
def general_eq(t, a, b, n): return a / (t**n + b)
def japanese_eq_root(t, a, b): return a / (np.sqrt(t) + b)
def talbot_eq(t, a, b): return a / (t + b)
def sherman_eq(t, a, n): return a / (t ** n)
def semilog_eq(t, a, b): return a + b * np.log10(t) # [추가됨: Plan 1]

def log_poly_val(t, coeffs):
    t = np.array(t, dtype=float)
    log_t = np.log(t)
    log_i = np.polyval(coeffs, log_t)
    return np.exp(log_i)

def calc_stats(y_true, y_pred):
    ss_res = np.sum((y_true - y_pred) ** 2)
    ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
    r2 = 1 - (ss_res / ss_tot)
    rmse = np.sqrt(np.mean((y_true - y_pred) ** 2))
    return r2, rmse

# -----------------------------------------------------------
# 2. 분석 엔진 클래스 (RainfallAnalyzer)
# -----------------------------------------------------------
class RainfallAnalyzer:
    def __init__(self, df_input):
        self.df = df_input.copy()
        
        if 'Return Period(Year)' in self.df.columns:
            self.df.set_index('Return Period(Year)', inplace=True)
        
        valid_cols = []
        for col in self.df.columns:
            try:
                float(col)
                valid_cols.append(col)
            except: pass
            
        self.df = self.df[valid_cols]
        self.df.columns = [float(c) for c in self.df.columns]
        self.df = self.df.sort_index(axis=1)
        
        self.durations_input = np.array(self.df.columns, dtype=float)
        self.pivot = 120 
        self.log_buffer = []
        self.result_data = {}
        self.selected_method_type = ""
        
        self.graph_opts = {200: True, 300: True, 500: True}

        # 강우량 -> 강우강도 변환
        for duration in self.df.columns:
            self.df[duration] = self.df[duration] * (60.0 / duration)

    def set_graph_options(self, opts):
        self.graph_opts.update(opts)

    def run_analysis(self, method_mode):
        self.log_buffer = [["Freq", "Model", "R2", "RMSE", "Note"]]
        self.result_data = {}
        
        if "General" in method_mode:
            self.selected_method_type = "General"
        # [참고] SemiLog 모드 설정 로직은 백엔드에 준비되었으나 GUI에서 호출하지 않음
        elif "SemiLog" in method_mode:
            self.selected_method_type = "SemiLog"
        else:
            self.selected_method_type = "LogPoly"

        for freq in self.df.index:
            y_obs = self.df.loc[freq].values.astype(float)
            x_obs = self.durations_input.astype(float)
            
            # --- 통계 분석 ---
            # 1. Talbot
            try:
                popt, _ = curve_fit(talbot_eq, x_obs, y_obs, p0=[1000, 10], maxfev=10000)
                r2, rmse = calc_stats(y_obs, talbot_eq(x_obs, *popt))
                self.log_buffer.append([freq, "Talbot", f"{r2:.5f}", f"{rmse:.4f}", ""])
            except: pass

            # 2. Sherman
            try:
                popt, _ = curve_fit(sherman_eq, x_obs, y_obs, p0=[1000, 0.5], maxfev=10000)
                r2, rmse = calc_stats(y_obs, sherman_eq(x_obs, *popt))
                self.log_buffer.append([freq, "Sherman", f"{r2:.5f}", f"{rmse:.4f}", ""])
            except: pass

            # 3. Japanese
            try:
                popt, _ = curve_fit(japanese_eq_root, x_obs, y_obs, p0=[1000, 10], maxfev=10000)
                r2, rmse = calc_stats(y_obs, japanese_eq_root(x_obs, *popt))
                self.log_buffer.append([freq, "Japanese", f"{r2:.5f}", f"{rmse:.4f}", ""])
            except: pass

            # 4. Semi-Log [추가됨: Plan 2 - 분석 로직만 추가]
            popt_semilog = None
            try:
                # I = a + b * log10(t)
                popt_semilog, _ = curve_fit(semilog_eq, x_obs, y_obs, p0=[100, -10], maxfev=10000)
                r2, rmse = calc_stats(y_obs, semilog_eq(x_obs, *popt_semilog))
                self.log_buffer.append([freq, "Semi-Log", f"{r2:.5f}", f"{rmse:.4f}", "Backend Only"])
            except: pass

            # 5. General (Unified)
            popt_unified = None
            try:
                popt_unified, _ = curve_fit(general_eq, x_obs, y_obs, p0=[1000, 10, 0.5], maxfev=10000)
                r2, rmse = calc_stats(y_obs, general_eq(x_obs, *popt_unified))
                self.log_buffer.append([freq, "General(Unified)", f"{r2:.5f}", f"{rmse:.4f}", ""])
            except: pass

            # 6. General (Split)
            split_res = self._calc_split_general(x_obs, y_obs)
            if split_res:
                self.log_buffer.append([freq, "General(Short)", f"{split_res['r2_s']:.5f}", f"{split_res['rmse_s']:.4f}", "0~120min"])
                self.log_buffer.append([freq, "General(Long)", f"{split_res['r2_l']:.5f}", f"{split_res['rmse_l']:.4f}", "Constraint"])

            # 7. LogPoly
            poly_coeffs = {}
            try:
                mask = (x_obs > 0) & (y_obs > 0)
                if np.any(mask):
                    log_x = np.log(x_obs[mask])
                    log_y = np.log(y_obs[mask])
                    for d in [3, 4, 5, 6]:
                        if len(log_x) > d:
                            c = np.polyfit(log_x, log_y, d)
                            y_p = np.exp(np.polyval(c, np.log(x_obs)))
                            r2, rmse = calc_stats(y_obs, y_p)
                            self.log_buffer.append([freq, f"LogPoly({d}th)", f"{r2:.5f}", f"{rmse:.4f}", ""])
                            poly_coeffs[d] = c
            except: pass

            # --- 결과 저장 ---
            if method_mode == "General_Unified":
                if popt_unified is not None:
                    self.result_data[freq] = {"Mode": "Unified", "a": popt_unified[0], "b": popt_unified[1], "n": popt_unified[2]}
            elif method_mode == "General_Split":
                if split_res:
                    self.result_data[freq] = {"Mode": "Split", "Short": split_res['popt_s'], "Long": split_res['popt_l']}
            # [추가됨: Plan 2 - 데이터 저장 로직 준비]
            elif method_mode == "SemiLog":
                if popt_semilog is not None:
                    self.result_data[freq] = {"Mode": "SemiLog", "a": popt_semilog[0], "b": popt_semilog[1]}
            elif "LogPoly" in method_mode:
                degree = int(method_mode.split("_")[1])
                if degree in poly_coeffs:
                    self.result_data[freq] = {"Mode": "LogPoly", "Order": degree, "Coeffs": poly_coeffs[degree]}

    def _calc_split_general(self, x_all, y_all):
        mask_s = x_all <= self.pivot
        x_s, y_s = x_all[mask_s], y_all[mask_s]
        mask_l = x_all >= self.pivot
        x_l, y_l = x_all[mask_l], y_all[mask_l]
        try:
            popt_s, _ = curve_fit(general_eq, x_s, y_s, p0=[1000, 10, 0.5], maxfev=20000)
            target = general_eq(self.pivot, *popt_s)
            r2_s, rmse_s = calc_stats(y_s, general_eq(x_s, *popt_s))
            
            def constrained_eq(t, b, n):
                a = target * (self.pivot**n + b)
                return a / (t**n + b)
            
            popt_l_sub, _ = curve_fit(constrained_eq, x_l, y_l, p0=[10, 0.5], maxfev=20000)
            b_l, n_l = popt_l_sub
            a_l = target * (self.pivot**n_l + b_l)
            popt_l = [a_l, b_l, n_l]
            r2_l, rmse_l = calc_stats(y_l, general_eq(x_l, *popt_l))
            return {"popt_s": popt_s, "r2_s": r2_s, "rmse_s": rmse_s,
                    "popt_l": popt_l, "r2_l": r2_l, "rmse_l": rmse_l}
        except: return None

    # --- [그래프 커스터마이징 메서드] ---
    def _get_marker_style(self, freq, idx):
        if freq == 200: return r'$X$', 'black', 'black'
        if freq == 300: return r'$◈$', 'black', 'black'
        if freq == 500: return r'$⊙$', 'black', 'black'
        shapes = ['s', '^', 'v', 'o', 'D']
        cycle_idx = idx % 10
        shape = shapes[cycle_idx % 5]
        if cycle_idx < 5: return shape, 'black', 'black'
        else: return shape, 'none', 'black'

    def _should_plot(self, freq):
        if freq in [200, 300, 500]:
            return self.graph_opts.get(freq, True)
        return True

    def save_graphs(self, save_prefix):
        if not self.result_data: return
        first_key = list(self.result_data.keys())[0]
        mode = self.result_data[first_key]["Mode"]

        if mode == "Split":
            self._save_graphs_split_mode(save_prefix)
        else:
            self._save_graphs_standard_mode(save_prefix)

    def _save_graphs_split_mode(self, prefix):
        freqs = list(self.result_data.keys())
        
        # 1. Whole Period
        plt.figure(figsize=(10, 7))
        t_whole = np.logspace(np.log10(10), np.log10(1440), 500)
        for i, freq in enumerate(freqs):
            if not self._should_plot(freq): continue
            data = self.result_data[freq]
            y_vals = []
            for t in t_whole:
                if t <= self.pivot: y_vals.append(general_eq(t, *data["Short"]))
                else: y_vals.append(general_eq(t, *data["Long"]))
            marker, fc, ec = self._get_marker_style(freq, i)
            plt.plot(t_whole, y_vals, color='black', linestyle='-', linewidth=1.0)
            plt.scatter(self.durations_input, self.df.loc[freq], 
                        marker=marker, facecolors=fc, edgecolors=ec, s=30, zorder=5, label=f"{freq}yr")
        plt.xscale('log'); plt.yscale('log')
        plt.title("IDF Curves - Whole Period (10~1440min)")
        plt.xlabel("Duration (min) [Log]")
        plt.ylabel("Intensity (mm/hr) [Log]")
        plt.grid(True, which='both', ls=':', alpha=0.5)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title="Return Period")
        plt.tight_layout()
        plt.savefig(f"{prefix}Graph_1_Whole.png"); plt.close()

        # 2. Short Term
        plt.figure(figsize=(10, 7))
        t_short_log = np.logspace(np.log10(10), np.log10(120), 200)
        mask_obs_s = self.durations_input <= 120
        x_obs_s = self.durations_input[mask_obs_s]
        for i, freq in enumerate(freqs):
            if not self._should_plot(freq): continue
            data = self.result_data[freq]
            y_vals = general_eq(t_short_log, *data["Short"])
            marker, fc, ec = self._get_marker_style(freq, i)
            plt.plot(t_short_log, y_vals, color='black', linestyle='-', linewidth=1.0)
            y_obs_s = self.df.loc[freq].values[mask_obs_s]
            plt.scatter(x_obs_s, y_obs_s, marker=marker, facecolors=fc, edgecolors=ec, s=30, zorder=5, label=f"{freq}yr")
        plt.xscale('log'); plt.yscale('log')
        plt.title("IDF Curves - Short Term (~120min)")
        plt.xlabel("Duration (min) [Log]")
        plt.ylabel("Intensity (mm/hr) [Log]")
        plt.grid(True, which='both', ls=':', alpha=0.5)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title="Return Period")
        plt.tight_layout()
        plt.savefig(f"{prefix}Graph_2_Short.png"); plt.close()

        # 3. Long Term
        plt.figure(figsize=(10, 7))
        t_long_log = np.logspace(np.log10(120), np.log10(1440), 300)
        mask_obs_l = self.durations_input >= 120
        x_obs_l = self.durations_input[mask_obs_l]
        for i, freq in enumerate(freqs):
            if not self._should_plot(freq): continue
            data = self.result_data[freq]
            y_vals = general_eq(t_long_log, *data["Long"])
            marker, fc, ec = self._get_marker_style(freq, i)
            plt.plot(t_long_log, y_vals, color='black', linestyle='-', linewidth=1.0)
            y_obs_l = self.df.loc[freq].values[mask_obs_l]
            plt.scatter(x_obs_l, y_obs_l, marker=marker, facecolors=fc, edgecolors=ec, s=30, zorder=5, label=f"{freq}yr")
        plt.xscale('log'); plt.yscale('log')
        plt.title("IDF Curves - Long Term (120min~)")
        plt.xlabel("Duration (min) [Log]")
        plt.ylabel("Intensity (mm/hr) [Log]")
        plt.grid(True, which='both', ls=':', alpha=0.5)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title="Return Period")
        plt.tight_layout()
        plt.savefig(f"{prefix}Graph_3_Long.png"); plt.close()

    def _save_graphs_standard_mode(self, prefix):
        t_whole = np.logspace(np.log10(10), np.log10(1440), 500)
        plt.figure(figsize=(10, 7))
        freqs = list(self.result_data.keys())
        
        for i, freq in enumerate(freqs):
            if not self._should_plot(freq): continue
            data = self.result_data[freq]
            mode = data["Mode"]
            marker, fc, ec = self._get_marker_style(freq, i)

            if mode == "Unified":
                y_vals = general_eq(t_whole, data['a'], data['b'], data['n'])
            elif mode == "LogPoly":
                lx = np.log(t_whole)
                y_vals = np.exp(np.polyval(data["Coeffs"], lx))
            
            plt.plot(t_whole, y_vals, color='black', linestyle='-', linewidth=1.0)
            plt.scatter(self.durations_input, self.df.loc[freq], 
                        marker=marker, facecolors=fc, edgecolors=ec, s=30, label=f"{freq}yr")

        plt.xscale('log'); plt.yscale('log')
        plt.title(f"IDF Curves - {mode}")
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title="Return Period")
        plt.grid(True, which='both', ls=':', alpha=0.5)
        plt.tight_layout()
        plt.savefig(f"{prefix}Graph_Standard.png")
        plt.close()

    def export_excel(self, filename, custom_durations=None):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Analysis_Log"
        for row in self.log_buffer:
            ws1.append(row)

        if custom_durations is None or len(custom_durations) == 0:
            custom_durations = [10] + list(range(60, 1441, 60))

        if self.selected_method_type == "General":
            self._write_excel_general(wb, custom_durations)
        else:
            self._write_excel_logpoly(wb, custom_durations)
            
        try:
            wb.save(filename)
        except PermissionError:
            raise PermissionError(f"파일이 열려 있습니다: {filename}")

    def _write_excel_general(self, wb, durations):
        ws2 = wb.create_sheet("General_Parameters")
        ws2.append(["Freq", "Type", "Short_a", "Short_b", "Short_n", "Long_a", "Long_b", "Long_n"])
        ws3 = wb.create_sheet("Probable_Rainfall")
        
        ws3.cell(1, 1, "Return Period")
        for i, d in enumerate(durations):
            ws3.cell(1, i+2, d) 

        row_idx = 2
        for freq in self.df.index:
            if freq not in self.result_data: continue
            data = self.result_data[freq]
            
            if data["Mode"] == "Unified":
                ws2.append([freq, "Unified", data['a'], data['b'], data['n'], "", "", ""])
                sa, sb, sn = f"{data['a']:.4f}", f"{data['b']:.4f}", f"{data['n']:.4f}"
                la, lb, ln = sa, sb, sn
            else: 
                s, l = data['Short'], data['Long']
                ws2.append([freq, "Split", s[0], s[1], s[2], l[0], l[1], l[2]])
                sa, sb, sn = f"{s[0]:.4f}", f"{s[1]:.4f}", f"{s[2]:.4f}"
                la, lb, ln = f"{l[0]:.4f}", f"{l[1]:.4f}", f"{l[2]:.4f}"

            ws3.cell(row_idx, 1, freq)
            for i, d in enumerate(durations):
                col_let = get_column_letter(i+2)
                ref = f"{col_let}$1"
                eq_s = f"{sa}/(POWER({ref},{sn})+{sb})"
                eq_l = f"{la}/(POWER({ref},{ln})+{lb})"
                
                if data["Mode"] == "Unified":
                    formula = f"=ROUND(({eq_s})*{ref}/60, 2)"
                else:
                    formula = f"=ROUND((IF({ref}<=120, {eq_s}, {eq_l}))*{ref}/60, 2)"
                
                ws3.cell(row_idx, i+2).value = formula
            row_idx += 1

    def _write_excel_logpoly(self, wb, durations):
        ws2 = wb.create_sheet("LogPoly_Coefficients")
        ws2.append(["Freq", "Order", "a", "b", "c", "d", "e", "f", "g"])
        ws3 = wb.create_sheet("Probable_Rainfall")
        
        ws3.cell(1, 1, "Return Period")
        for i, d in enumerate(durations):
            ws3.cell(1, i+2, d)

        row_idx = 2
        for freq in self.df.index:
            if freq not in self.result_data: continue
            data = self.result_data[freq]
            coeffs = data["Coeffs"] 
            rev_c = coeffs[::-1]
            row_c = [freq, f"{data['Order']}th"]
            keys = ['a', 'b', 'c', 'd', 'e', 'f', 'g']
            for k_i, k in enumerate(keys):
                if k_i < len(rev_c):
                    row_c.append(f"{rev_c[k_i]:.14f}")
                else:
                    row_c.append("")
            ws2.append(row_c)
            ws3.cell(row_idx, 1, freq)
            for i, d in enumerate(durations):
                intensity = log_poly_val(d, coeffs)
                depth = intensity * (d / 60.0)
                ws3.cell(row_idx, i+2).value = round(depth, 2)
            row_idx += 1

# -----------------------------------------------------------
# 3. Main GUI Application
# -----------------------------------------------------------
class RainfallIntensityApp(ctk.CTk):
    def __init__(self, project_path=None):
        super().__init__()
        
        self.project_path = project_path if project_path else os.getcwd()
        self.config_file = os.path.join(self.project_path, "project_config.json")
        self.project_name = os.path.basename(self.project_path)

        self.title(f"강우강도식 산정 - [{self.project_name}]")
        self.geometry("1200x850")
        self.change_title_bar_color()

        self.analyzer = None
        self.setup_ui()
        self.load_step2_data()

    def change_title_bar_color(self):
        try:
            hwnd = windll.user32.GetParent(self.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
        except: pass

    def setup_ui(self):
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(main_frame, text="강우강도식 유도 (General, Log-Poly)", font=FONT_TITLE).pack(pady=(0, 20))

        setting_frame = ctk.CTkFrame(main_frame)
        setting_frame.pack(fill="x", pady=(0, 20))

        # [1] 분석 방법 선택
        method_frame = ctk.CTkFrame(setting_frame, fg_color="transparent")
        method_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        ctk.CTkLabel(method_frame, text="[1] 분석 방법 선택", font=FONT_HEADER).pack(anchor="w", pady=(0, 5))
        
        self.method_var = tk.StringVar(value="General_Split")
        methods = [
            ("General형 분할 (120분 Constraint)", "General_Split"),
            ("General형 일체 (전구간 단일식)", "General_Unified"),
            ("전대수다항식 6차", "LogPoly_6"),
            ("전대수다항식 5차", "LogPoly_5"),
            ("전대수다항식 4차", "LogPoly_4")
        ]
        for txt, val in methods:
            ctk.CTkRadioButton(method_frame, text=txt, variable=self.method_var, value=val, font=FONT_BODY).pack(anchor="w", pady=2)

        # [2] 출력 설정
        right_frame = ctk.CTkFrame(setting_frame, fg_color="transparent")
        right_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        ctk.CTkLabel(right_frame, text="[2] 출력 설정", font=FONT_HEADER).pack(anchor="w", pady=(0, 5))
        
        dur_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        dur_frame.pack(anchor="w", fill="x", pady=5)
        self.dur_mode_var = tk.IntVar(value=1)
        ctk.CTkRadioButton(dur_frame, text="기본값 (10, 60~1440분)", variable=self.dur_mode_var, value=1, font=FONT_BODY).pack(anchor="w", pady=2)
        ctk.CTkRadioButton(dur_frame, text="10분 단위 증가 (~최대값)", variable=self.dur_mode_var, value=2, font=FONT_BODY).pack(anchor="w", pady=2)
        
        max_dur_frame = ctk.CTkFrame(dur_frame, fg_color="transparent")
        max_dur_frame.pack(anchor="w", padx=25)
        ctk.CTkLabel(max_dur_frame, text="▶ 최대(분):", font=FONT_BODY).pack(side="left")
        self.entry_max_dur = ctk.CTkEntry(max_dur_frame, width=60, font=FONT_BODY)
        self.entry_max_dur.insert(0, "1440")
        self.entry_max_dur.pack(side="left", padx=5)

        chk_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        chk_frame.pack(anchor="w", fill="x", pady=(10, 0))
        ctk.CTkLabel(chk_frame, text="▶ 그래프 표시 빈도:", font=FONT_BODY).pack(anchor="w", pady=(0, 2))
        
        self.var_200 = ctk.BooleanVar(value=True)
        self.var_300 = ctk.BooleanVar(value=True)
        self.var_500 = ctk.BooleanVar(value=True)
        
        chk_inner = ctk.CTkFrame(chk_frame, fg_color="transparent")
        chk_inner.pack(anchor="w", padx=10)
        ctk.CTkCheckBox(chk_inner, text="200년", variable=self.var_200, font=FONT_BODY, width=60).pack(side="left", padx=5)
        ctk.CTkCheckBox(chk_inner, text="300년", variable=self.var_300, font=FONT_BODY, width=60).pack(side="left", padx=5)
        ctk.CTkCheckBox(chk_inner, text="500년", variable=self.var_500, font=FONT_BODY, width=60).pack(side="left", padx=5)

        self.btn_run = ctk.CTkButton(main_frame, text="⚡ 강우강도식 산정 및 저장", command=self.run_analysis,
                                     font=FONT_BTN, height=50, fg_color="#2980b9", hover_color="#3498db")
        self.btn_run.pack(fill="x", pady=(0, 20))

        self.tabview = ctk.CTkTabview(main_frame)
        self.tabview.pack(fill="both", expand=True)
        
        self.tab_log = self.tabview.add(" 분석 로그 ")
        self.tab_params = self.tabview.add(" 산정된 매개변수 ")
        self.tab_graph = self.tabview.add(" IDF 곡선 그래프 ")

        self.txt_log = ctk.CTkTextbox(self.tab_log, font=FONT_LOG, activate_scrollbars=True)
        self.txt_log.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.txt_params = ctk.CTkTextbox(self.tab_params, font=FONT_LOG, activate_scrollbars=True)
        self.txt_params.pack(fill="both", expand=True, padx=10, pady=10)

        self.graph_frame = ctk.CTkFrame(self.tab_graph, fg_color="transparent")
        self.graph_frame.pack(fill="both", expand=True, padx=5, pady=5)

    def load_step2_data(self):
        self.log(f"📂 프로젝트: {self.project_name}")
        self.log("📂 2단계(확률강우량) 결과를 로드합니다...")
        
        if not os.path.exists(self.config_file):
            messagebox.showerror("오류", "프로젝트 설정 파일이 없습니다.")
            self.btn_run.configure(state="disabled")
            return

        try:
            with open(self.config_file, 'r', encoding='utf-8') as f: config = json.load(f)
            
            if 'step2_probability' not in config or config['step2_probability'].get('status') != 'completed':
                messagebox.showwarning("경고", "2단계 확률강우량 분석이 완료되지 않았습니다.")
                self.btn_run.configure(state="disabled")
                return

            step2_file = config['step2_probability'].get('output_file')
            file_path = os.path.join(self.project_path, step2_file)
            
            if not os.path.exists(file_path):
                file_path = config['step2_probability'].get('full_path', '')
                if not os.path.exists(file_path):
                    messagebox.showerror("오류", f"2단계 파일을 찾을 수 없습니다.\n{step2_file}")
                    return

            self.log(f"-> 파일 로드 성공: {os.path.basename(file_path)}")
            
            df_input = pd.read_excel(file_path, sheet_name='Probability_Rainfall')
            self.analyzer = RainfallAnalyzer(df_input)
            
            self.log("-> 데이터 파싱 완료. 분석 준비 끝.")
            
        except Exception as e:
            messagebox.showerror("오류", f"데이터 로드 실패: {e}")
            self.log(f"Error: {e}")

    def run_analysis(self):
        if not self.analyzer: return
        
        try:
            mode = self.method_var.get()
            max_dur = int(self.entry_max_dur.get())
            dur_mode = self.dur_mode_var.get()
            
            # 그래프 옵션 전달
            graph_opts = {
                200: self.var_200.get(),
                300: self.var_300.get(),
                500: self.var_500.get()
            }
            self.analyzer.set_graph_options(graph_opts)
            
            custom_durations = []
            if dur_mode == 1:
                custom_durations = [10] + list(range(60, max_dur+1, 60))
            else:
                custom_durations = list(range(10, max_dur+1, 10))

            self.log(f"\n🚀 분석 시작 (Mode: {mode})")
            
            self.analyzer.run_analysis(mode)
            
            for row in self.analyzer.log_buffer:
                self.log("\t".join(map(str, row)))
            
            self.update_params_tab()
            
            graph_prefix = os.path.join(self.project_path, f"{self.project_name}_C_")
            self.analyzer.save_graphs(graph_prefix)
            self.log("💾 그래프 파일 3종 저장 완료 (_C_Graph_...png)")
            
            self.draw_graphs()
            
            self.save_results(custom_durations)
            
            messagebox.showinfo("완료", "강우강도식 산정이 완료되었습니다.")
            self.tabview.set(" IDF 곡선 그래프 ")

        except Exception as e:
            messagebox.showerror("오류", f"분석 중 오류 발생:\n{traceback.format_exc()}")

    def update_params_tab(self):
        self.txt_params.delete("1.0", tk.END)
        self.txt_params.insert(tk.END, "[ 매개변수 산정 결과 ]\n\n")
        
        for freq, data in self.analyzer.result_data.items():
            self.txt_params.insert(tk.END, f"▶ Return Period: {freq}년\n")
            if data["Mode"] == "Unified":
                self.txt_params.insert(tk.END, f"   Type: Unified General\n")
                self.txt_params.insert(tk.END, f"   a: {data['a']:.4f}, b: {data['b']:.4f}, n: {data['n']:.4f}\n")
            elif data["Mode"] == "Split":
                s, l = data['Short'], data['Long']
                self.txt_params.insert(tk.END, f"   Type: Split General (Pivot: 120min)\n")
                self.txt_params.insert(tk.END, f"   Short(<=120): a={s[0]:.4f}, b={s[1]:.4f}, n={s[2]:.4f}\n")
                self.txt_params.insert(tk.END, f"   Long (>120) : a={l[0]:.4f}, b={l[1]:.4f}, n={l[2]:.4f}\n")
            elif data["Mode"] == "LogPoly":
                self.txt_params.insert(tk.END, f"   Type: Log-Polynomial ({data['Order']}th)\n")
                self.txt_params.insert(tk.END, f"   Coeffs: {data['Coeffs']}\n")
            self.txt_params.insert(tk.END, "-"*50 + "\n")

    def draw_graphs(self):
        for widget in self.graph_frame.winfo_children():
            widget.destroy()

        fig = plt.figure(figsize=(10, 6), dpi=100)
        ax = fig.add_subplot(111)
        ax.set_title(f"IDF Curves - {self.analyzer.selected_method_type}")
        ax.set_xlabel("Duration (min) [Log Scale]")
        ax.set_ylabel("Intensity (mm/hr) [Log Scale]")
        ax.set_xscale('log')
        ax.set_yscale('log')
        ax.grid(True, which="both", ls=":", alpha=0.5)

        freqs = list(self.analyzer.result_data.keys())
        colors = plt.cm.jet(np.linspace(0, 1, len(freqs)))
        t_plot = np.logspace(np.log10(10), np.log10(1440), 100)

        for i, freq in enumerate(freqs):
            if not self.analyzer._should_plot(freq): continue

            data = self.analyzer.result_data[freq]
            y_vals = []
            
            if data["Mode"] == "Unified":
                y_vals = general_eq(t_plot, data['a'], data['b'], data['n'])
            elif data["Mode"] == "Split":
                for t in t_plot:
                    if t <= 120: y_vals.append(general_eq(t, *data["Short"]))
                    else: y_vals.append(general_eq(t, *data["Long"]))
            elif data["Mode"] == "LogPoly":
                y_vals = np.exp(np.polyval(data["Coeffs"], np.log(t_plot)))
            
            marker, fc, ec = self.analyzer._get_marker_style(freq, i)
            ax.plot(t_plot, y_vals, color='black', linestyle='-', linewidth=1.0)
            
            obs_y = self.analyzer.df.loc[freq]
            # s=30 적용 (점 크기 축소)
            ax.scatter(obs_y.index, obs_y.values, marker=marker, facecolors=fc, edgecolors=ec, s=30, label=f"{freq}yr")

        ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title="Return Period")
        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    def save_results(self, custom_durations):
        filename = f"{self.project_name}_C_Rainfall_Intensity.xlsx"
        save_path = os.path.join(self.project_path, filename)
        
        try:
            self.analyzer.export_excel(save_path, custom_durations)
            self.log(f"💾 엑셀 결과 저장 완료: {filename}")
            
            with open(self.config_file, 'r', encoding='utf-8') as f: config = json.load(f)
            config['step3_intensity'] = {
                "status": "completed",
                "output_file": filename,
                "full_path": save_path,
                "method": self.method_var.get(),
                "timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
                
        except Exception as e:
            messagebox.showerror("저장 오류", f"파일 저장 실패: {e}")

    def log(self, message):
        self.txt_log.insert(tk.END, f"{message}\n")
        self.txt_log.see(tk.END)

if __name__ == "__main__":
    if sys.platform.startswith('win'): rc('font', family='Malgun Gothic')
    elif sys.platform.startswith('darwin'): rc('font', family='AppleGothic')
    
    project_path_arg = sys.argv[1] if len(sys.argv) > 1 else None
    app = RainfallIntensityApp(project_path=project_path_arg)
    app.mainloop()