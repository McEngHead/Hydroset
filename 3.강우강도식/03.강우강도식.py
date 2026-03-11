import os
import sys
import json
import re
import traceback
import pandas as pd
import numpy as np
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
from scipy.optimize import curve_fit, OptimizeWarning
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib import rc
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from ctypes import windll, byref, sizeof, c_int
import warnings

# [수정] NumPy 2.x 호환성 및 scipy OptimizeWarning 제어
warnings.filterwarnings("ignore", category=RuntimeWarning)
warnings.filterwarnings("ignore", category=OptimizeWarning)  # ✅ 추가

# --- [CustomTkinter 설정] ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# 폰트 설정
FONT_TITLE = ("맑은 고딕", 22, "bold")
FONT_HEADER = ("맑은 고딕", 14, "bold")
FONT_BODY = ("맑은 고딕", 12)
FONT_BTN = ("맑은 고딕", 12, "bold")
FONT_LOG = ("consolas", 11)

# -----------------------------------------------------------
# 1. 수학적 모델 정의
# -----------------------------------------------------------
def general_eq(t, a, b, n):
    t = np.asanyarray(t, dtype=np.float64)
    return a / (t**n + b)

def japanese_eq_root(t, a, b):
    t = np.asanyarray(t, dtype=np.float64)
    return a / (np.sqrt(t) + b)

def talbot_eq(t, a, b):
    t = np.asanyarray(t, dtype=np.float64)
    return a / (t + b)

def sherman_eq(t, a, n):
    t = np.asanyarray(t, dtype=np.float64)
    return a / (t ** n)

def semilog_eq(t, a, b):
    t = np.asanyarray(t, dtype=np.float64)
    return a + b * np.log10(t)

def log_poly_val(t, coeffs):
    t = np.asanyarray(t, dtype=np.float64)
    log_t = np.log(t)
    log_i = np.polyval(coeffs, log_t)
    return np.exp(log_i)

def calc_stats(y_true, y_pred):
    y_true = np.asanyarray(y_true, dtype=np.float64)
    y_pred = np.asanyarray(y_pred, dtype=np.float64)
    ss_res = np.sum((y_true - y_pred) ** 2)
    ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
    r2 = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0
    rmse = np.sqrt(np.mean((y_true - y_pred) ** 2))
    # Matplotlib 충돌 방지를 위해 파이썬 표준 float로 변환
    return float(r2), float(rmse)

# -----------------------------------------------------------
# 2. 분석 엔진 클래스 (RainfallAnalyzer)
# -----------------------------------------------------------
# -----------------------------------------------------------
# 2. 분석 엔진 클래스 (RainfallAnalyzer)
# -----------------------------------------------------------
class RainfallAnalyzer:
    def __init__(self, df_input):
        self.df = df_input.copy()
        
        rp_col = next((c for c in ['Return Period(Year)', 'Return Period'] if c in self.df.columns), None)
        if rp_col:
            rp_vals = []
            for v in self.df[rp_col]:
                try:
                    rp_vals.append(int(float(str(v).lower().replace('yr', '').replace('year', '').strip())))
                except:
                    rp_vals.append(v)
            self.df = self.df.drop(columns=[rp_col])
            self.df.index = rp_vals
        
        valid_cols = []
        for col in self.df.columns:
            try:
                float(col)
                valid_cols.append(col)
            except: pass
            
        self.df = self.df[valid_cols]
        self.df.columns = [float(c) for c in self.df.columns]
        self.df = self.df.sort_index(axis=1)
        
        self.durations_input = np.array(self.df.columns, dtype=np.float64)
        self.pivot = 120 
        self.log_buffer = []
        self.result_data = {}
        self.selected_method_type = ""
        self.graph_opts = {200: True, 300: True, 500: True}

        for duration in self.df.columns:
            # 강우량 -> 강우강도 변환 시 타입 명시
            self.df[duration] = self.df[duration].astype(np.float64) * (60.0 / duration)
        
        
    def set_graph_options(self, opts):
        self.graph_opts.update(opts)

    def run_analysis(self, method_mode):
        self.log_buffer = [["Freq", "Model", "R2", "RMSE", "Note"]]
        self.result_data = {}
        
        if "General" in method_mode:
            self.selected_method_type = "General"
        elif "SemiLog" in method_mode:
            self.selected_method_type = "SemiLog"
        else:
            self.selected_method_type = "LogPoly"

        for freq in self.df.index:
            y_obs = self.df.loc[freq].values.astype(np.float64)
            x_obs = self.durations_input
            
            # --- 다중 모델 분석 로직 ---
            try:
                popt_t, _ = curve_fit(talbot_eq, x_obs, y_obs, p0=[1000, 10], maxfev=10000)
                r2, rmse = calc_stats(y_obs, talbot_eq(x_obs, *popt_t))
                self.log_buffer.append([freq, "Talbot", f"{r2:.5f}", f"{rmse:.4f}", ""])
            except: pass

            try:
                popt_s, _ = curve_fit(sherman_eq, x_obs, y_obs, p0=[1000, 0.5], maxfev=10000)
                r2, rmse = calc_stats(y_obs, sherman_eq(x_obs, *popt_s))
                self.log_buffer.append([freq, "Sherman", f"{r2:.5f}", f"{rmse:.4f}", ""])
            except: pass

            try:
                popt_j, _ = curve_fit(japanese_eq_root, x_obs, y_obs, p0=[1000, 10], maxfev=10000)
                r2, rmse = calc_stats(y_obs, japanese_eq_root(x_obs, *popt_j))
                self.log_buffer.append([freq, "Japanese", f"{r2:.5f}", f"{rmse:.4f}", ""])
            except: pass

            # General (Unified)
            popt_unified = None
            try:
                popt_unified, _ = curve_fit(general_eq, x_obs, y_obs, p0=[1000, 10, 0.5], maxfev=10000)
                r2, rmse = calc_stats(y_obs, general_eq(x_obs, *popt_unified))
                self.log_buffer.append([freq, "General(Unified)", f"{r2:.5f}", f"{rmse:.4f}", ""])
            except: pass

            # General (Split)
            split_res = self._calc_split_general(x_obs, y_obs)
            if split_res:
                self.log_buffer.append([freq, "General(Short)", f"{split_res['r2_s']:.5f}", f"{split_res['rmse_s']:.4f}", "0~120min"])
                self.log_buffer.append([freq, "General(Long)", f"{split_res['r2_l']:.5f}", f"{split_res['rmse_l']:.4f}", "Constraint"])

            # LogPoly
            poly_coeffs = {}
            try:
                mask = (x_obs > 0) & (y_obs > 0)
                if np.any(mask):
                    log_x = np.log(x_obs[mask]).astype(np.float64)
                    log_y = np.log(y_obs[mask]).astype(np.float64)
                    for d in [3, 4, 5, 6]:
                        if len(log_x) > d:
                            c = np.polyfit(log_x, log_y, d)
                            y_p = np.exp(np.polyval(c, np.log(x_obs)))
                            r2, rmse = calc_stats(y_obs, y_p)
                            self.log_buffer.append([freq, f"LogPoly({d}th)", f"{r2:.5f}", f"{rmse:.4f}", ""])
                            poly_coeffs[d] = c
            except: pass

            # 결과 저장 (UI 선택 모드에 따라)
            if method_mode == "General_Unified":
                if popt_unified is not None:
                    self.result_data[freq] = {"Mode": "Unified", "a": popt_unified[0], "b": popt_unified[1], "n": popt_unified[2]}
            elif method_mode == "General_Split":
                if split_res:
                    self.result_data[freq] = {
                        "Mode": "Split",
                        "Short": split_res['popt_s'],
                        "Long": split_res['popt_l'],
                        "r2_s": split_res['r2_s'],
                        "rmse_s": split_res['rmse_s'],
                        "r2_l": split_res['r2_l'],
                        "rmse_l": split_res['rmse_l']
                    }
            elif "LogPoly" in method_mode:
                degree = int(method_mode.split("_")[1])
                if degree in poly_coeffs:
                    self.result_data[freq] = {"Mode": "LogPoly", "Order": degree, "Coeffs": poly_coeffs[degree]}

    def _calc_split_general(self, x_all, y_all):
        # [수정] NumPy 2.x 대응: 명시적 float64 변환으로 정밀도 확보
        x_all = np.asanyarray(x_all, dtype=np.float64)
        y_all = np.asanyarray(y_all, dtype=np.float64)

        mask_s = x_all <= self.pivot
        x_s, y_s = x_all[mask_s], y_all[mask_s]
        mask_l = x_all >= self.pivot
        x_l, y_l = x_all[mask_l], y_all[mask_l]

        try:
            # 데이터 포인트 부족 시 최적화 건너뛰기
            if len(x_s) < 3: return None

            # [수정] 초기값 개선 - 데이터 스케일에 맞춰 조정
            y_mean = float(np.mean(y_s))
            x_mean = float(np.mean(x_s))
            
            # 더 나은 초기값: a는 y 스케일, b는 x 스케일, n은 0.5~1.0 범위
            p0_improved = [y_mean * x_mean, x_mean, 0.7]
            
            # [수정] bounds 추가로 안정성 향상
            bounds = ([0, 0, 0.1], [np.inf, np.inf, 2.0])
            
            popt_s, pcov_s = curve_fit(
                general_eq, x_s, y_s, 
                p0=p0_improved,
                bounds=bounds,
                maxfev=20000
            )
            
            # [수정] NumPy 2.x 스칼라를 Python float로 변환 (안정성 확보)
            target = float(general_eq(self.pivot, *popt_s))
            
            def constrained_eq(t, b, n):
                a = target * (self.pivot**n + b)
                return a / (t**n + b)
            
            # Long 구간도 bounds 적용
            popt_l_sub, _ = curve_fit(
                constrained_eq, x_l, y_l, 
                p0=[x_mean, 0.7],
                bounds=([0, 0.1], [np.inf, 2.0]),
                maxfev=20000
            )
            b_l, n_l = popt_l_sub
            a_l = target * (self.pivot**n_l + b_l)
            
            # 통계량 계산 시에도 명시적 타입 변환 적용
            r2_s, rmse_s = calc_stats(y_s, general_eq(x_s, *popt_s))
            r2_l, rmse_l = calc_stats(y_l, general_eq(x_l, a_l, b_l, n_l))
            
            return {
                "popt_s": popt_s, "r2_s": float(r2_s), "rmse_s": float(rmse_s),
                "popt_l": [float(a_l), float(b_l), float(n_l)], 
                "r2_l": float(r2_l), "rmse_l": float(rmse_l)
            }
        except Exception as e:
            # 디버깅을 위해 에러 메시지 출력 (선택사항)
            # print(f"Split General 계산 실패: {e}")
            return None

    # --- [그래프 커스터마이징 메서드] ---
    def _get_marker_style(self, freq, idx):
        if freq == 200: return r'$X$', 'black', 'black'
        if freq == 300: return r'$◈$', 'black', 'black'
        if freq == 500: return r'$⊙$', 'black', 'black'
        shapes = ['s', '^', 'v', 'o', 'D']
        cycle_idx = idx % 10
        shape = shapes[cycle_idx % 5]
        if cycle_idx < 5: return shape, 'black', 'black'
        else: return shape, 'none', 'black'

    def _should_plot(self, freq):
        if freq in [200, 300, 500]:
            return self.graph_opts.get(freq, True)
        return True

    def save_graphs(self, save_prefix):
        if not self.result_data: return
        first_key = list(self.result_data.keys())[0]
        mode = self.result_data[first_key]["Mode"]

        if mode == "Split":
            self._save_split_graph(save_prefix)
        else:
            self._save_single_graph(save_prefix)

    def _save_split_graph(self, save_prefix):
        fig, axes = plt.subplots(2, 2, figsize=(14, 10), dpi=100)
        ax_short, ax_long, ax_full, ax_obs = axes.flatten()

        freqs = list(self.result_data.keys())
        colors = plt.cm.jet(np.linspace(0, 1, len(freqs)))
        t_short = np.linspace(10, 120, 50)
        t_long = np.linspace(120, 1440, 100)
        t_all = np.logspace(np.log10(10), np.log10(1440), 100)

        for i, freq in enumerate(freqs):
            if not self._should_plot(freq): continue
            data = self.result_data[freq]
            if data["Mode"] != "Split": continue

            a_s, b_s, n_s = data["Short"]
            a_l, b_l, n_l = data["Long"]
            y_s = general_eq(t_short, a_s, b_s, n_s)
            y_l = general_eq(t_long, a_l, b_l, n_l)
            y_all = [general_eq(t, a_s, b_s, n_s) if t <= 120 else general_eq(t, a_l, b_l, n_l) for t in t_all]
            
            marker, fc, ec = self._get_marker_style(freq, i)

            ax_short.plot(t_short, y_s, color=colors[i], linewidth=1.5, label=f"{freq}yr")
            ax_long.plot(t_long, y_l, color=colors[i], linewidth=1.5, label=f"{freq}yr")
            ax_full.plot(t_all, y_all, color=colors[i], linewidth=1.5, label=f"{freq}yr")
            
            obs_y = self.df.loc[freq]
            ax_obs.scatter(obs_y.index, obs_y.values, marker=marker, facecolors=fc, edgecolors=ec, s=30, label=f"{freq}yr")

        for ax, title in zip([ax_short, ax_long, ax_full, ax_obs],
                              ["Short (≤120min)", "Long (>120min)", "Full Range", "Observed vs Fitted"]):
            ax.set_title(title)
            ax.set_xlabel("Duration (min)" + (" [Log]" if ax == ax_full else ""))
            ax.set_ylabel("Intensity (mm/hr)" + (" [Log]" if ax == ax_full else ""))
            if ax == ax_full or ax == ax_obs:
                ax.set_xscale('log')
                ax.set_yscale('log')
            ax.grid(True, which="both", ls=":", alpha=0.5)
            ax.legend(fontsize=8, loc='best')

        plt.tight_layout()
        plt.savefig(f"{save_prefix}Graph_Split.png", dpi=150)
        plt.close()

    def _save_single_graph(self, save_prefix):
        fig = plt.figure(figsize=(10, 6), dpi=100)
        ax = fig.add_subplot(111)
        ax.set_title(f"IDF Curves - {self.selected_method_type}")
        ax.set_xlabel("Duration (min) [Log Scale]")
        ax.set_ylabel("Intensity (mm/hr) [Log Scale]")
        ax.set_xscale('log')
        ax.set_yscale('log')
        ax.grid(True, which="both", ls=":", alpha=0.5)

        freqs = list(self.result_data.keys())
        colors = plt.cm.jet(np.linspace(0, 1, len(freqs)))
        t_plot = np.logspace(np.log10(10), np.log10(1440), 100)

        for i, freq in enumerate(freqs):
            if not self._should_plot(freq): continue
            data = self.result_data[freq]
            
            if data["Mode"] == "Unified":
                y_vals = general_eq(t_plot, data['a'], data['b'], data['n'])
            elif data["Mode"] == "LogPoly":
                y_vals = np.exp(np.polyval(data["Coeffs"], np.log(t_plot)))
            
            marker, fc, ec = self._get_marker_style(freq, i)
            ax.plot(t_plot, y_vals, color=colors[i], linewidth=1.5, label=f"{freq}yr")
            
            obs_y = self.df.loc[freq]
            ax.scatter(obs_y.index, obs_y.values, marker=marker, facecolors=fc, edgecolors=ec, s=30)

        ax.legend(fontsize=9, loc='best', title="Return Period")
        plt.tight_layout()
        plt.savefig(f"{save_prefix}Graph_Single.png", dpi=150)
        plt.close()

    def export_excel(self, filepath, custom_durations):
        wb = Workbook()
        ws_params = wb.active
        ws_params.title = "Parameters"
        ws_intensity = wb.create_sheet("Intensity_Table")

        ws_params.append(["Return Period (Year)", "Type", "Parameter a", "Parameter b", "Parameter n/Order", "R²", "RMSE"])

        for freq, data in self.result_data.items():
            if data["Mode"] == "Unified":
                ws_params.append([freq, "General(Unified)", data['a'], data['b'], data['n'], "", ""])
            elif data["Mode"] == "Split":
                s, l = data['Short'], data['Long']
                ws_params.append([freq, "Short(≤120)", s[0], s[1], s[2], data['r2_s'], data['rmse_s']])
                ws_params.append(["", "Long(>120)", l[0], l[1], l[2], data['r2_l'], data['rmse_l']])
            elif data["Mode"] == "LogPoly":
                coeffs_str = ", ".join([f"{c:.6e}" for c in data["Coeffs"]])
                ws_params.append([freq, f"LogPoly({data['Order']}th)", coeffs_str, "", "", "", ""])

        header = ["Duration(min)"] + [str(f) for f in self.result_data.keys()]
        ws_intensity.append(header)

        for dur in custom_durations:
            row = [dur]
            for freq in self.result_data.keys():
                data = self.result_data[freq]
                if data["Mode"] == "Unified":
                    intensity = general_eq(dur, data['a'], data['b'], data['n'])
                elif data["Mode"] == "Split":
                    if dur <= 120: intensity = general_eq(dur, *data['Short'])
                    else: intensity = general_eq(dur, *data['Long'])
                elif data["Mode"] == "LogPoly":
                    intensity = np.exp(np.polyval(data["Coeffs"], np.log(dur)))
                row.append(round(intensity, 4))
            ws_intensity.append(row)

        wb.save(filepath)

# -----------------------------------------------------------
# 3. GUI 애플리케이션 (RainfallIntensityApp)
# -----------------------------------------------------------
class RainfallIntensityApp(ctk.CTk):
    def __init__(self, project_path=None):
        super().__init__()
        
        self.title("강우강도식 산정 (3단계)")
        self.geometry("1200x800")
        
        if project_path and os.path.isdir(project_path):
            self.project_path = project_path
            self.project_name = os.path.basename(project_path)
            self.config_file = os.path.join(project_path, "project_config.json")
        else:
            messagebox.showerror("오류", "프로젝트 경로가 유효하지 않습니다.")
            sys.exit(1)

        self.analyzer = None
        
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.build_ui()
        self.load_step2_data()

    def build_ui(self):
        # 상단 타이틀
        title_frame = ctk.CTkFrame(self, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        ctk.CTkLabel(
            title_frame, text="강우강도식 산정 (3단계)", 
            font=FONT_TITLE
        ).pack(side="left")

        # 메인 레이아웃
        main_container = ctk.CTkFrame(self, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=20, pady=10)

        # 왼쪽: 설정 패널
        left_panel = ctk.CTkFrame(main_container, width=350)
        left_panel.pack(side="left", fill="y", padx=(0, 10))
        left_panel.pack_propagate(False)

        # 1) 모델 선택
        ctk.CTkLabel(left_panel, text="📊 모델 선택", font=FONT_HEADER).pack(anchor="w", padx=10, pady=(10, 5))
        self.method_var = tk.StringVar(value="General_Split")
        
        methods = [
            ("General형 (통합)", "General_Unified"),
            ("General형 (분할, 120분 기준)", "General_Split"),
            ("Log-Polynomial (6차)", "LogPoly_6"),
            ("Log-Polynomial (5차)", "LogPoly_5"),
            ("Log-Polynomial (4차)", "LogPoly_4"),
            ("Log-Polynomial (3차)", "LogPoly_3")
        ]
        for label, val in methods:
            ctk.CTkRadioButton(
                left_panel, text=label, variable=self.method_var, value=val, font=FONT_BODY
            ).pack(anchor="w", padx=20, pady=2)

        # 2) 출력 지속시간 설정
        ctk.CTkLabel(left_panel, text="⏱️ 출력 지속시간 설정", font=FONT_HEADER).pack(anchor="w", padx=10, pady=(15, 5))
        
        dur_frame = ctk.CTkFrame(left_panel, fg_color="transparent")
        dur_frame.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkLabel(dur_frame, text="최대 지속시간(분):", font=FONT_BODY).pack(side="left", padx=5)
        self.entry_max_dur = ctk.CTkEntry(dur_frame, width=80)
        self.entry_max_dur.insert(0, "1440")
        self.entry_max_dur.pack(side="left", padx=5)

        self.dur_mode_var = tk.IntVar(value=1)
        ctk.CTkRadioButton(
            left_panel, text="10분 단위 + 1시간 간격", variable=self.dur_mode_var, value=1, font=FONT_BODY
        ).pack(anchor="w", padx=20, pady=2)
        ctk.CTkRadioButton(
            left_panel, text="10분 단위 전체", variable=self.dur_mode_var, value=2, font=FONT_BODY
        ).pack(anchor="w", padx=20, pady=2)

        # 3) 그래프 옵션
        ctk.CTkLabel(left_panel, text="📈 그래프 옵션", font=FONT_HEADER).pack(anchor="w", padx=10, pady=(15, 5))
        self.var_200 = tk.BooleanVar(value=True)
        self.var_300 = tk.BooleanVar(value=True)
        self.var_500 = tk.BooleanVar(value=True)
        
        ctk.CTkCheckBox(left_panel, text="200년 빈도 표시", variable=self.var_200, font=FONT_BODY).pack(anchor="w", padx=20, pady=2)
        ctk.CTkCheckBox(left_panel, text="300년 빈도 표시", variable=self.var_300, font=FONT_BODY).pack(anchor="w", padx=20, pady=2)
        ctk.CTkCheckBox(left_panel, text="500년 빈도 표시", variable=self.var_500, font=FONT_BODY).pack(anchor="w", padx=20, pady=2)

        # 실행 버튼
        self.btn_run = ctk.CTkButton(
            left_panel, text="🚀 분석 실행", command=self.run_analysis, 
            font=FONT_BTN, height=40
        )
        self.btn_run.pack(fill="x", padx=20, pady=20)

        # 오른쪽: 탭 뷰 (로그, 매개변수, 그래프)
        right_panel = ctk.CTkFrame(main_container)
        right_panel.pack(side="left", fill="both", expand=True)

        self.tabview = ctk.CTkTabview(right_panel)
        self.tabview.pack(fill="both", expand=True, padx=5, pady=5)

        self.tab_log = self.tabview.add(" 실행 로그 ")
        self.tab_params = self.tabview.add(" 매개변수 ")
        self.tab_graph = self.tabview.add(" IDF 곡선 그래프 ")

        # 로그 탭
        self.txt_log = tk.Text(self.tab_log, font=FONT_LOG, bg="#2b2b2b", fg="white", wrap="word")
        self.txt_log.pack(fill="both", expand=True, padx=10, pady=10)

        # 매개변수 탭
        self.txt_params = tk.Text(self.tab_params, font=FONT_LOG, bg="#2b2b2b", fg="white", wrap="word")
        self.txt_params.pack(fill="both", expand=True, padx=10, pady=10)

        self.graph_frame = ctk.CTkFrame(self.tab_graph, fg_color="transparent")
        self.graph_frame.pack(fill="both", expand=True, padx=5, pady=5)

    def load_step2_data(self):
        self.log(f"📂 프로젝트: {self.project_name}")
        self.log("📂 2단계(확률강우량) 결과를 로드합니다...")
        
        if not os.path.exists(self.config_file):
            messagebox.showerror("오류", "프로젝트 설정 파일이 없습니다.")
            self.btn_run.configure(state="disabled")
            return

        try:
            with open(self.config_file, 'r', encoding='utf-8') as f: config = json.load(f)
            
            if 'step2_probability' not in config or config['step2_probability'].get('status') != 'completed':
                messagebox.showwarning("경고", "2단계 확률강우량 분석이 완료되지 않았습니다.")
                self.btn_run.configure(state="disabled")
                return

            step2_file = config['step2_probability'].get('output_file')
            file_path = os.path.join(self.project_path, step2_file)
            
            if not os.path.exists(file_path):
                file_path = config['step2_probability'].get('full_path', '')
                if not os.path.exists(file_path):
                    messagebox.showerror("오류", f"2단계 파일을 찾을 수 없습니다.\n{step2_file}")
                    return

            self.log(f"-> 파일 로드 성공: {os.path.basename(file_path)}")
            
            df_input = pd.read_excel(file_path, sheet_name='Probability_Rainfall')
            self.analyzer = RainfallAnalyzer(df_input)
            
            self.log("-> 데이터 파싱 완료. 분석 준비 끝.")
            
        except Exception as e:
            messagebox.showerror("오류", f"데이터 로드 실패: {e}")
            self.log(f"Error: {e}")

    def run_analysis(self):
        if not self.analyzer: return
        
        try:
            mode = self.method_var.get()
            max_dur = int(self.entry_max_dur.get())
            dur_mode = self.dur_mode_var.get()
            
            # 그래프 옵션 전달
            graph_opts = {
                200: self.var_200.get(),
                300: self.var_300.get(),
                500: self.var_500.get()
            }
            self.analyzer.set_graph_options(graph_opts)
            
            custom_durations = []
            if dur_mode == 1:
                custom_durations = [10] + list(range(60, max_dur+1, 60))
            else:
                custom_durations = list(range(10, max_dur+1, 10))

            self.log(f"\n🚀 분석 시작 (Mode: {mode})")
            
            self.analyzer.run_analysis(mode)
            
            for row in self.analyzer.log_buffer:
                self.log("\t".join(map(str, row)))
            
            self.update_params_tab()
            
            graph_prefix = os.path.join(self.project_path, f"{self.project_name}_C_")
            self.analyzer.save_graphs(graph_prefix)
            self.log("💾 그래프 파일 저장 완료 (_C_Graph_...png)")
            
            self.draw_graphs()
            
            self.save_results(custom_durations)
            
            messagebox.showinfo("완료", "강우강도식 산정이 완료되었습니다.")
            self.tabview.set(" IDF 곡선 그래프 ")

        except Exception as e:
            messagebox.showerror("오류", f"분석 중 오류 발생:\n{traceback.format_exc()}")

    def update_params_tab(self):
        self.txt_params.delete("1.0", tk.END)
        self.txt_params.insert(tk.END, "[ 매개변수 산정 결과 ]\n\n")
        
        for freq, data in self.analyzer.result_data.items():
            self.txt_params.insert(tk.END, f"▶ Return Period: {freq}년\n")
            if data["Mode"] == "Unified":
                self.txt_params.insert(tk.END, f"   Type: Unified General\n")
                self.txt_params.insert(tk.END, f"   a: {data['a']:.4f}, b: {data['b']:.4f}, n: {data['n']:.4f}\n")
            elif data["Mode"] == "Split":
                s, l = data['Short'], data['Long']
                self.txt_params.insert(tk.END, f"   Type: Split General (Pivot: 120min)\n")
                self.txt_params.insert(tk.END, f"   Short(<=120): a={s[0]:.4f}, b={s[1]:.4f}, n={s[2]:.4f}\n")
                self.txt_params.insert(tk.END, f"   Long (>120) : a={l[0]:.4f}, b={l[1]:.4f}, n={l[2]:.4f}\n")
            elif data["Mode"] == "LogPoly":
                self.txt_params.insert(tk.END, f"   Type: Log-Polynomial ({data['Order']}th)\n")
                self.txt_params.insert(tk.END, f"   Coeffs: {data['Coeffs']}\n")
            self.txt_params.insert(tk.END, "-"*50 + "\n")

    def draw_graphs(self):
        for widget in self.graph_frame.winfo_children():
            widget.destroy()

        fig = plt.figure(figsize=(10, 6), dpi=100)
        ax = fig.add_subplot(111)
        ax.set_title(f"IDF Curves - {self.analyzer.selected_method_type}")
        ax.set_xlabel("Duration (min) [Log Scale]")
        ax.set_ylabel("Intensity (mm/hr) [Log Scale]")
        ax.set_xscale('log')
        ax.set_yscale('log')
        ax.grid(True, which="both", ls=":", alpha=0.5)

        freqs = list(self.analyzer.result_data.keys())
        colors = plt.cm.jet(np.linspace(0, 1, len(freqs)))
        t_plot = np.logspace(np.log10(10), np.log10(1440), 100)

        # [수정] 실제로 그려질 데이터가 있는지 확인
        has_plots = False
        
        for i, freq in enumerate(freqs):
            if not self.analyzer._should_plot(freq): continue

            data = self.analyzer.result_data[freq]
            y_vals = []
            
            if data["Mode"] == "Unified":
                y_vals = general_eq(t_plot, data['a'], data['b'], data['n'])
            elif data["Mode"] == "Split":
                for t in t_plot:
                    if t <= 120: y_vals.append(general_eq(t, *data["Short"]))
                    else: y_vals.append(general_eq(t, *data["Long"]))
            elif data["Mode"] == "LogPoly":
                y_vals = np.exp(np.polyval(data["Coeffs"], np.log(t_plot)))
            
            marker, fc, ec = self.analyzer._get_marker_style(freq, i)
            ax.plot(t_plot, y_vals, color='black', linestyle='-', linewidth=1.0)
            
            obs_y = self.analyzer.df.loc[freq]
            # s=30 적용 (점 크기 축소)
            ax.scatter(obs_y.index, obs_y.values, marker=marker, facecolors=fc, edgecolors=ec, s=30, label=f"{freq}yr")
            has_plots = True

        # [수정] 그래프가 실제로 그려진 경우에만 legend 표시
        if has_plots:
            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left', title="Return Period")
        else:
            ax.text(0.5, 0.5, 'No data to display\n(모든 재현기간이 필터링됨)', 
                   ha='center', va='center', transform=ax.transAxes, fontsize=14)
        
        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    def save_results(self, custom_durations):
        filename = f"{self.project_name}_C_Rainfall_Intensity.xlsx"
        save_path = os.path.join(self.project_path, filename)
        
        try:
            self.analyzer.export_excel(save_path, custom_durations)
            self.log(f"💾 엑셀 결과 저장 완료: {filename}")
            
            with open(self.config_file, 'r', encoding='utf-8') as f: config = json.load(f)
            config['step3_intensity'] = {
                "status": "completed",
                "output_file": filename,
                "full_path": save_path,
                "method": self.method_var.get(),
                "timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
                
        except Exception as e:
            messagebox.showerror("저장 오류", f"파일 저장 실패: {e}")

    def on_closing(self):
        # Matplotlib figure 정리
        plt.close('all')

        # CustomTkinter 내부 after() 콜백 에러 방지를 위해 즉시 종료
        self.destroy()
        sys.exit(0)
    
    def log(self, message):
        self.txt_log.insert(tk.END, f"{message}\n")
        self.txt_log.see(tk.END)
        


if __name__ == "__main__":
    if sys.platform.startswith('win'): rc('font', family='Malgun Gothic')
    elif sys.platform.startswith('darwin'): rc('font', family='AppleGothic')
    
    project_path_arg = sys.argv[1] if len(sys.argv) > 1 else None
    app = RainfallIntensityApp(project_path=project_path_arg)
    app.mainloop()