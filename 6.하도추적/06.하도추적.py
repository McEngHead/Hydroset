"""
06.하도추적.py — Muskingum 하도홍수추적 (Muskingum Flood Routing)
Hydro Analysis System  Module 6

Muskingum 방법 (KWRA 수문학 CH08 기준):
  S = K[xI + (1-x)O]
  O2 = C1*I2 + C2*I1 + C3*O1
  C1 = (-Kx + 0.5Δt) / (K - Kx + 0.5Δt)
  C2 = (Kx + 0.5Δt)  / (K - Kx + 0.5Δt)
  C3 = (K - Kx - 0.5Δt) / (K - Kx + 0.5Δt)
  안정 조건: 2Kx ≤ Δt ≤ 2K(1-x)
"""

import os, sys, json, traceback, warnings, copy, math
import threading
import urllib.request, urllib.error
import numpy as np
from datetime import datetime
from ctypes import windll, byref, sizeof, c_int

import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog

warnings.filterwarnings("ignore", category=RuntimeWarning)

# ── 지연 import (scipy, matplotlib, openpyxl) ────────────────────────────
# 실행/엑셀 저장 시점에 최초 1회만 import → 앱 시작 ~1초 단축
_lazy = {}

def _ensure_scipy():
    if 'PchipInterpolator' not in _lazy:
        from scipy.interpolate import PchipInterpolator
        _lazy['PchipInterpolator'] = PchipInterpolator
    return _lazy['PchipInterpolator']

def _ensure_mpl():
    if 'plt' not in _lazy:
        import matplotlib
        matplotlib.use('TkAgg')
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        _lazy['plt'] = plt
        _lazy['FigureCanvasTkAgg'] = FigureCanvasTkAgg
    return _lazy['plt'], _lazy['FigureCanvasTkAgg']

def _ensure_openpyxl():
    if 'Workbook' not in _lazy:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        _lazy['Workbook'] = Workbook
        _lazy['Font'] = Font
        _lazy['PatternFill'] = PatternFill
        _lazy['Alignment'] = Alignment
        _lazy['get_column_letter'] = get_column_letter
    return _lazy

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

OLLAMA_BASE = 'http://localhost:11434'

# ── Google Gemini API 설정 ──────────────────────────────────────────────────
_gemini_key_file = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                '..', 'GEMINI_API_KEY.json')
try:
    with open(_gemini_key_file, 'r', encoding='utf-8') as _f:
        GEMINI_API_KEY = json.load(_f)['GEMINI_API_KEY']
except Exception:
    GEMINI_API_KEY = ''
GEMINI_MODELS = [
    'gemini-2.5-flash',
    'gemini-2.5-pro',
]
# ────────────────────────────────────────────────────────────────────────────

FONT_TITLE  = ("맑은 고딕", 18, "bold")
FONT_HEADER = ("맑은 고딕", 12, "bold")
FONT_BODY   = ("맑은 고딕", 11)
FONT_BTN    = ("맑은 고딕", 11, "bold")
FONT_SMALL  = ("맑은 고딕", 10)
FONT_LOG    = ("Consolas", 10)

HUFF_PRESETS = {
    "1분위": [0.000, 0.130, 0.262, 0.415, 0.546, 0.649, 0.726, 0.795, 0.861, 0.931, 1.000],
    "2분위": [0.000, 0.009, 0.053, 0.101, 0.184, 0.331, 0.538, 0.756, 0.916, 0.975, 1.000],
    "3분위": [0.000, 0.008, 0.041, 0.086, 0.154, 0.263, 0.437, 0.636, 0.833, 0.953, 1.000],
    "4분위": [0.000, 0.007, 0.030, 0.062, 0.110, 0.186, 0.310, 0.492, 0.700, 0.892, 1.000],
}

# 노드 타입별 스타일
NODE_STYLES = {
    'SUBBASIN':  {'label': '소유역',  'fill': '#1d5c33', 'outline': '#27ae60', 'shape': 'round_rect'},
    'RESERVOIR': {'label': '저수지',  'fill': '#0d3b5e', 'outline': '#2e86de', 'shape': 'hexagon'},
    'JUNCTION':  {'label': '합류점',  'fill': '#1c3d5a', 'outline': '#5dade2', 'shape': 'circle'},
    'OUTLET':    {'label': '출구',    'fill': '#4a1c6e', 'outline': '#8e44ad', 'shape': 'rect'},
}
# Backward-compat style entry for any old REACH nodes that might still be in memory
_REACH_STYLE = {'label': '하도추적', 'fill': '#1a2d4a', 'outline': '#2980b9', 'shape': 'parallelogram'}

# =============================================================================
# Muskingum 추적 엔진
# =============================================================================

class MuskingumEngine:
    @staticmethod
    def compute_coefficients(K, X, dt_hr):
        denom = K - K * X + 0.5 * dt_hr
        if denom <= 1e-12:
            return 0.0, 1.0, 0.0, False
        C1 = (-K * X + 0.5 * dt_hr) / denom
        C2 = (K * X  + 0.5 * dt_hr) / denom
        C3 = (K - K * X - 0.5 * dt_hr) / denom
        stable = (C1 >= -1e-9) and (C3 >= -1e-9)
        return C1, C2, C3, stable

    @staticmethod
    def auto_nstps(K, X, dt_hr):
        if K <= 1e-12: return 1
        upper = 2.0 * K * (1.0 - X)
        if dt_hr <= upper + 1e-9: return 1
        return max(1, int(np.ceil(dt_hr / upper)))

    def route(self, inflow, K, X, dt_hr, NSTPS=0, Q0=None):
        n = len(inflow)
        if n == 0:
            return np.array([]), True, [], 0.0, 1.0, 0.0, 1
        if NSTPS <= 0:
            NSTPS = self.auto_nstps(K, X, dt_hr)
        dt_sub = dt_hr / NSTPS
        C1, C2, C3, stable = self.compute_coefficients(K, X, dt_sub)
        warns = []
        if not stable:
            warns.append(f"불안정: dt_sub={dt_sub:.3f}hr, 2K(1-x)={2*K*(1-X):.3f}hr")
        O_prev = inflow[0] if Q0 is None else float(Q0)
        outflow = np.zeros(n)
        for i in range(n):
            I_curr = float(inflow[i])
            I_prev = float(inflow[i-1]) if i > 0 else float(inflow[0])
            if NSTPS == 1:
                O_prev = C1 * I_curr + C2 * I_prev + C3 * O_prev
            else:
                for j in range(NSTPS):
                    t0 = j / NSTPS; t1 = (j+1) / NSTPS
                    I0s = I_prev + t0 * (I_curr - I_prev)
                    I1s = I_prev + t1 * (I_curr - I_prev)
                    O_prev = C1 * I1s + C2 * I0s + C3 * O_prev
            outflow[i] = max(0.0, O_prev)
        return outflow, stable, warns, C1, C2, C3, NSTPS


class ReservoirEngine:
    """Modified Puls (Storage-Indication Method) 저수지추적."""

    @staticmethod
    def _build_si(A_avg, Cd, L, Hc, dt):
        """Build (SI, Q) arrays vs H (200 points, H from 0 to 3*Hc+5)."""
        Hmax = max(Hc * 3.0 + 5.0, 10.0)
        H  = np.linspace(0.0, Hmax, 300)
        S  = A_avg * H
        Q  = np.where(H > Hc, Cd * L * (H - Hc) ** 1.5, 0.0)
        SI = 2.0 * S / dt + Q
        return SI, Q, H

    @staticmethod
    def route(inflow, A_avg, Cd, L, Hc, S0, dt_hr, NQ=None):
        """
        Modified Puls routing.
        inflow  : array (m³/s)
        A_avg   : average surface area (m²)
        Cd, L   : spillway discharge coef, length (m)
        Hc      : spillway crest height above bottom (m)
        S0      : initial storage (m³)
        dt_hr   : time step (hours)
        Returns : outflow array (m³/s)
        """
        if NQ is None:
            NQ = len(inflow)
        dt = dt_hr * 3600.0
        SI_tbl, Q_tbl, H_tbl = ReservoirEngine._build_si(A_avg, Cd, L, Hc, dt)

        H0 = S0 / max(A_avg, 1.0)
        O0 = float(np.interp(H0, H_tbl, Q_tbl))
        # Initial SI value: 2*S0/dt + O0
        si_prev = 2.0 * S0 / dt + O0
        O_prev  = O0

        outflow = np.zeros(NQ)
        for i in range(NQ):
            I1  = float(inflow[i - 1]) if i > 0 else 0.0
            I2  = float(inflow[i])     if i < len(inflow) else 0.0
            # 2S1/dt - O1 = SI_prev - 2*O_prev
            rhs = max(0.0, I1 + I2 + si_prev - 2.0 * O_prev)
            O2  = float(np.interp(rhs, SI_tbl, Q_tbl))
            si_prev = rhs   # = 2S2/dt + O2
            O_prev  = O2
            outflow[i] = O2
        return outflow


# =============================================================================
# Clark 단위도 + 유효우량 엔진
# =============================================================================

class ClarkEngine:
    def effective_rainfall(self, total_precip, tr_min, dt_min, cn, huff_pc):
        n_step = int(tr_min / dt_min) + 1
        t_huff = np.linspace(0.0, 1.0, len(huff_pc))
        t_norm = np.linspace(0.0, 1.0, n_step)
        pc_interp = np.clip(_ensure_scipy()(t_huff, np.array(huff_pc, dtype=float))(t_norm), 0.0, 1.0)
        cum_rain = pc_interp * total_precip
        S = (25400.0 / cn) - 254.0
        Ia = 0.2 * S
        cum_excess = np.zeros(n_step)
        for i, P in enumerate(cum_rain):
            if P > Ia:
                cum_excess[i] = (P - Ia) ** 2 / (P - Ia + S)
        return np.maximum(0.0, np.diff(cum_excess, prepend=0.0))

    def clark_uh(self, A, Tc, R, dt_hr):
        uh_len = int((Tc + R * 10.0) / dt_hr) + 5
        t_vals = np.arange(uh_len) * dt_hr
        ai = np.zeros(uh_len)
        for i, t in enumerate(t_vals):
            T = t / Tc if Tc > 1e-12 else 1.0
            if   T >= 1.0: ai[i] = 1.0
            elif T <  0.5: ai[i] = 1.414 * T ** 1.5
            else:          ai[i] = 1.0 - 1.414 * (1.0 - T) ** 1.5
        vol_const = A * 1000.0
        I_flow = np.zeros(uh_len)
        for i in range(1, uh_len):
            dAI = max(0.0, ai[i] - ai[i-1])
            I_flow[i] = dAI * vol_const / (dt_hr * 3600.0)
        c_ca = dt_hr / (R + 0.5 * dt_hr)
        c_cb = 1.0 - c_ca
        O_inst = np.zeros(uh_len)
        for i in range(1, uh_len):
            O_inst[i] = c_ca * I_flow[i] + c_cb * O_inst[i-1]
        uh = np.zeros(uh_len)
        for i in range(1, uh_len):
            uh[i] = 0.5 * (O_inst[i] + O_inst[i-1])
        calc_vol = np.sum(uh) * dt_hr * 3600.0
        target_vol = A * 10000.0
        if calc_vol > 1e-6:
            uh *= target_vol / calc_vol
        return uh

    def compute_runoff(self, A, Tc, R, total_precip, tr_min, dt_min, cn, huff_pc, NQ):
        dt_hr  = dt_min / 60.0
        excess = self.effective_rainfall(total_precip, tr_min, dt_min, cn, huff_pc)
        uh     = self.clark_uh(A, Tc, R, dt_hr)
        full_conv = np.convolve(excess, uh) / 10.0
        result = np.zeros(NQ)
        n = min(len(full_conv), NQ)
        result[:n] = full_conv[:n]
        return np.maximum(0.0, result)


# =============================================================================
# 수문망 처리기
# =============================================================================

class HydroNetworkProcessor:
    def __init__(self):
        self.clark = ClarkEngine()
        self.musk  = MuskingumEngine()
        self.results  = {}
        self.warnings = []
        self.summary  = []

    def run(self, operations, dt_min, NQ, tr_min, huff_pc, baseflow=0.0):
        self.results  = {}
        self.warnings = []
        self.summary  = []
        stack    = []
        cum_area = 0.0
        dt_hr    = dt_min / 60.0

        for op in operations:
            t    = op['type']
            name = op['name']

            if t == 'BASIN':
                A  = float(op.get('A',  1.0))
                PB = float(op.get('PB', 100.0))
                CN = float(op.get('CN', 80.0))
                Tc = float(op.get('Tc', 1.0))
                R  = float(op.get('R',  1.5))
                node_huff_pc = op.get('huff_pc') or huff_pc
                flow = self.clark.compute_runoff(A, Tc, R, PB, tr_min, dt_min, CN, node_huff_pc, NQ)
                stack.append(flow.copy())
                cum_area += A
                pidx = int(np.argmax(flow))
                self.results[name] = {'flow': flow, 'type': 'BASIN', 'A': A,
                    'peak_q': float(flow[pidx]), 'peak_hr': pidx * dt_hr, 'cum_area': cum_area}
                self.summary.append({'op': '수문곡선', 'station': name,
                    'peak_q': float(flow[pidx]), 'peak_hr': pidx * dt_hr, 'cum_area': cum_area})

            elif t == 'ROUTE':
                if not stack:
                    self.warnings.append(f"[{name}] 추적 오류: 스택 비어있음"); continue
                K     = float(op.get('K',    1.0))
                X     = float(op.get('X',    0.20))
                NSTPS = int(  op.get('NSTPS', 0))
                inflow  = stack.pop()
                outflow, stable, rw, C1, C2, C3, ns = self.musk.route(inflow, K, X, dt_hr, NSTPS)
                stack.append(outflow.copy())
                for w in rw: self.warnings.append(f"[{name}] {w}")
                pidx = int(np.argmax(outflow))
                self.results[name] = {'flow': outflow, 'type': 'ROUTE',
                    'K': K, 'X': X, 'NSTPS': ns, 'stable': stable,
                    'C1': C1, 'C2': C2, 'C3': C3,
                    'peak_q': float(outflow[pidx]), 'peak_hr': pidx * dt_hr, 'cum_area': cum_area}
                self.summary.append({'op': f'Muskingum(K={K:.2f},X={X:.2f})', 'station': name,
                    'peak_q': float(outflow[pidx]), 'peak_hr': pidx * dt_hr, 'cum_area': cum_area})

            elif t == 'COMBINE':
                N = int(op.get('N', 2))
                if len(stack) < N:
                    self.warnings.append(f"[{name}] 합류 오류: 스택 {len(stack)}개, 요청 {N}개")
                    N = len(stack)
                if N <= 0:
                    self.warnings.append(f"[{name}] 합류 오류: 합산 대상 없음"); continue
                combined = np.zeros(NQ)
                for _ in range(N):
                    h = stack.pop()
                    ln = min(len(h), NQ)
                    combined[:ln] += h[:ln]
                stack.append(combined.copy())
                pidx = int(np.argmax(combined))
                self.results[name] = {'flow': combined, 'type': 'COMBINE', 'N': N,
                    'peak_q': float(combined[pidx]), 'peak_hr': pidx * dt_hr, 'cum_area': cum_area}
                self.summary.append({'op': f'{N}개 합류', 'station': name,
                    'peak_q': float(combined[pidx]), 'peak_hr': pidx * dt_hr, 'cum_area': cum_area})

            elif t == 'RESERVOIR':
                if not stack:
                    self.warnings.append(f"[{name}] 저수지추적 오류: 스택 비어있음"); continue
                A_avg = float(op.get('A_avg', 10000.0))
                Cd    = float(op.get('Cd',    0.42))
                L     = float(op.get('L',     10.0))
                Hc    = float(op.get('Hc',    3.0))
                S0    = float(op.get('S0',    0.0))
                inflow  = stack.pop()
                outflow = ReservoirEngine.route(inflow, A_avg, Cd, L, Hc, S0, dt_hr, NQ)
                stack.append(outflow.copy())
                pidx = int(np.argmax(outflow))
                self.results[name] = {'flow': outflow, 'type': 'RESERVOIR',
                    'peak_q': float(outflow[pidx]), 'peak_hr': pidx * dt_hr, 'cum_area': cum_area}
                self.summary.append({'op': f'저수지추적(Cd={Cd:.2f},L={L:.1f}m)', 'station': name,
                    'peak_q': float(outflow[pidx]), 'peak_hr': pidx * dt_hr, 'cum_area': cum_area})

        if stack and baseflow > 0.0:
            stack[-1] += baseflow
            last_name = self.summary[-1]['station'] if self.summary else None
            if last_name and last_name in self.results:
                self.results[last_name]['flow'] = stack[-1].copy()
                pidx = int(np.argmax(stack[-1]))
                self.results[last_name]['peak_q'] = float(stack[-1][pidx])
                self.results[last_name]['peak_hr'] = pidx * dt_hr
                self.summary[-1]['peak_q'] = self.results[last_name]['peak_q']
                self.summary[-1]['peak_hr'] = self.results[last_name]['peak_hr']

        return self.results


# =============================================================================
# 네트워크 데이터 모델
# =============================================================================

class NetworkNode:
    _counter = 0

    def __init__(self, type_, name, x, y, params=None):
        NetworkNode._counter += 1
        self.id     = NetworkNode._counter
        self.type   = type_
        self.name   = name
        self.x      = float(x)
        self.y      = float(y)
        self.params = params if params is not None else self._default_params()

    def _default_params(self):
        if self.type == 'SUBBASIN':
            return {'A': 10.0, 'PB': 200.0, 'CN': 80.0, 'Tc': 1.0, 'R': 1.5}
        elif self.type == 'RESERVOIR':
            return {'A_avg': 10000.0, 'Cd': 0.42, 'L': 10.0, 'Hc': 3.0, 'S0': 0.0}
        elif self.type == 'REACH':   # backward compat
            return {'K': 1.0, 'X': 0.20, 'NSTPS': 0}
        return {}

    # Half-sizes for hit-testing
    _HW = {'SUBBASIN': 58, 'RESERVOIR': 52, 'REACH': 62, 'JUNCTION': 34, 'OUTLET': 30}
    _HH = {'SUBBASIN': 26, 'RESERVOIR': 26, 'REACH': 20, 'JUNCTION': 34, 'OUTLET': 26}

    def hit_test(self, px, py):
        hw = self._HW.get(self.type, 40)
        hh = self._HH.get(self.type, 20)
        if self.type == 'JUNCTION':
            return (px - self.x)**2 + (py - self.y)**2 <= hw**2
        return abs(px - self.x) <= hw and abs(py - self.y) <= hh

    def port(self, direction):
        """Returns port position for direction 'E','W','N','S'."""
        hw = self._HW.get(self.type, 40)
        hh = self._HH.get(self.type, 20)
        return {
            'E': (self.x + hw, self.y),
            'W': (self.x - hw, self.y),
            'S': (self.x, self.y + hh),
            'N': (self.x, self.y - hh),
        }[direction]


class NetworkEdge:
    _counter = 0

    def __init__(self, src_id, dst_id, src_dir=None, dst_dir=None, reach_params=None, label=None):
        NetworkEdge._counter += 1
        self.id           = NetworkEdge._counter
        self.src          = src_id
        self.dst          = dst_id
        self.src_dir      = src_dir   # 'E','W','N','S' or None=auto
        self.dst_dir      = dst_dir
        self.reach_params = reach_params  # dict {'K','X','NSTPS'} or None
        self.label        = label         # display name for reach edge


# =============================================================================
# 네트워크 캔버스
# =============================================================================

PORT_R = 4  # port circle radius
HIT_R  = 12 # port hit radius


class NetworkCanvas(tk.Canvas):

    GRID        = 40
    _SCROLL_PAD = 1200   # 콘텐츠 경계 바깥 여유 공간 (canvas px)

    def __init__(self, master, on_select=None, on_edge_select=None, on_log=None, **kwargs):
        super().__init__(master, bg='#12121e', highlightthickness=0, **kwargs)
        self.nodes   = {}   # id -> NetworkNode
        self.edges   = {}   # id -> NetworkEdge
        self._sel_node = None
        self._sel_edge = None
        self._mode     = 'select'
        self._drag_start = None       # (ex, ey, nx, ny)
        self._conn_src      = None       # node id being connected from
        self._conn_src_dir  = None       # direction of connection start
        self._conn_reach    = False      # True → next created edge gets reach_params
        self._mouse_xy      = (0, 0)
        self._on_select      = on_select       # callback(node or None)
        self._on_edge_select = on_edge_select  # callback(edge or None)
        self._on_log         = on_log          # callback(str)
        self._zoom       = 1.0
        self._world_bbox = (0, 0, 3000, 1200)
        self._pan_last   = None
        self._snap_on    = True
        self._sel_nodes    = set()   # ids of multi-selected nodes
        self._rubber_start = None    # world (wx,wy) where rubber-band drag started
        self._multi_origin = {}      # {node_id:(ox,oy)} for multi-drag
        self._undo_stack   = []      # list of (nodes_copy, edges_copy)
        self._redo_stack   = []
        self._reach_drag   = None    # (edge, src_ox, src_oy, dst_ox, dst_oy) — reach label drag
        self._edge_reconnect = None  # {'edge', 'end':'src'|'dst', 'fixed_port':(x,y)}
        self._drag_redraw_id = None  # after() ID for drag-throttle

        self.bind('<Button-1>',        self._click)
        self.bind('<B1-Motion>',       self._drag)
        self.bind('<ButtonRelease-1>', self._release)
        self.bind('<Motion>',          self._motion)
        self.bind('<Delete>',          self._delete)
        self.bind('<BackSpace>',       self._delete)
        self.bind('<Escape>',          self._escape)
        self.bind('<Double-Button-1>', self._dbl_click)
        self.bind('<Configure>',       lambda e: self.redraw())
        self.bind('<MouseWheel>',      self._on_zoom)
        self.bind('<Button-3>',        self._right_click)
        self.bind('<Button-2>',        self._on_pan_press)
        self.bind('<B2-Motion>',       self._on_pan_drag)
        self.bind('<ButtonRelease-2>', self._on_pan_release)
        self.bind('<Double-Button-2>', self._zoom_extents)
        self.bind('<Left>',            self._on_arrow)
        self.bind('<Right>',           self._on_arrow)
        self.bind('<Up>',              self._on_arrow)
        self.bind('<Down>',            self._on_arrow)
        self.bind('<Control-z>',       self._undo)
        self.bind('<Control-y>',       self._redo)
        self.bind('<Control-Alt-z>',   self._redo)

    # ── coordinate helpers ───────────────────────────────────────────────────

    def _cx(self, ex): return self.canvasx(ex)
    def _cy(self, ey): return self.canvasy(ey)

    # ── mode ─────────────────────────────────────────────────────────────────

    def set_mode(self, mode):
        self._mode = mode
        self.configure(cursor='crosshair' if mode != 'select' else 'arrow')

    # ── drawing ──────────────────────────────────────────────────────────────

    def redraw(self):
        self.delete('all')
        self._draw_grid()
        self._draw_edges()
        self._draw_nodes()
        if self._mode == 'connect' and self._conn_src:
            self._draw_rubber_band()
        if self._rubber_start:
            self._draw_sel_box()
        if self._edge_reconnect:
            self._draw_reconnect_preview()

    def _redraw_throttled(self):
        """드래그 중 redraw 스로틀: 16ms(≈60fps) 상한으로 실행 횟수 제한."""
        if self._drag_redraw_id is not None:
            return
        self._drag_redraw_id = self.after(16, self._redraw_throttled_fire)

    def _redraw_throttled_fire(self):
        self._drag_redraw_id = None
        self.redraw()

    def _draw_grid(self):
        w = max(self.winfo_width(),  self.winfo_reqwidth(),  100)
        h = max(self.winfo_height(), self.winfo_reqheight(), 100)
        try:
            x0 = int(self.canvasx(0)); x1 = int(self.canvasx(w))
            y0 = int(self.canvasy(0)); y1 = int(self.canvasy(h))
        except Exception:
            x0, x1, y0, y1 = 0, w, 0, h
        g = max(10, int(self.GRID * self._zoom))
        for x in range((x0 // g) * g, x1 + g, g):
            self.create_line(x, y0, x, y1, fill='#1e1e30', width=1)
        for y in range((y0 // g) * g, y1 + g, g):
            self.create_line(x0, y, x1, y, fill='#1e1e30', width=1)

    def _draw_edges(self):
        z  = self._zoom
        CR = max(6, int(8 * z))
        for edge in self.edges.values():
            src = self.nodes.get(edge.src)
            dst = self.nodes.get(edge.dst)
            if not src or not dst: continue
            sd, dd = self._port_dirs(src, dst, edge.src_dir, edge.dst_dir)
            wx1, wy1 = src.port(sd)
            wx2, wy2 = dst.port(dd)
            x1, y1 = wx1*z, wy1*z
            x2, y2 = wx2*z, wy2*z
            sel = self._sel_edge and self._sel_edge.id == edge.id
            col = '#f39c12' if sel else '#3498db'
            lw  = 2 if sel else 1
            self._draw_ortho_edge(x1, y1, sd, x2, y2, dd, col, lw, f'edge_{edge.id}', CR)
            if edge.reach_params:
                self._draw_reach_label(wx1, wy1, wx2, wy2, edge, z, f'edge_{edge.id}')
            if sel:
                # 끝점 재연결 핸들 (src=녹색, dst=빨간색)
                hr = max(5, int(6 * z))
                self.create_oval(x1-hr, y1-hr, x1+hr, y1+hr,
                                 fill='#27ae60', outline='white', width=1, tags=f'edge_{edge.id}')
                self.create_oval(x2-hr, y2-hr, x2+hr, y2+hr,
                                 fill='#e74c3c', outline='white', width=1, tags=f'edge_{edge.id}')

    def _draw_reach_label(self, wx1, wy1, wx2, wy2, edge, z, tag):
        """Draw parallelogram label at edge midpoint for reach edges."""
        mx = (wx1 + wx2) / 2 * z
        my = (wy1 + wy2) / 2 * z
        hw = 28 * z
        hh = 11 * z
        off = 7 * z
        pts = [mx-hw+off, my-hh, mx+hw+off, my-hh, mx+hw-off, my+hh, mx-hw-off, my+hh]
        sel = self._sel_edge and self._sel_edge.id == edge.id
        fill_c = '#1a2d4a'
        out_c  = '#f39c12' if sel else '#2980b9'
        lw2 = max(1, int((2 if sel else 1.5) * z))
        self.create_polygon(*pts, fill=fill_c, outline=out_c, width=lw2, tags=tag)
        fs = min(max(6, int(8 * z)), 13)
        lbl = edge.label or 'RC'
        self.create_text(mx, my, text=lbl, fill='white',
                         font=('맑은 고딕', fs, 'bold'), tags=tag)

    def _draw_nodes(self):
        for node in self.nodes.values():
            self._draw_node(node)

    def _draw_sel_box(self):
        """Draw rubber-band selection rectangle (tagged for fast coord-only update)."""
        z = self._zoom
        x0w, y0w = self._rubber_start
        x1w, y1w = self._mouse_xy
        self.create_rectangle(x0w*z, y0w*z, x1w*z, y1w*z,
                              outline='#5dade2', width=1, dash=(4,3), tags='_sel_box')

    def _draw_reconnect_preview(self):
        """Draw dashed line from fixed end to mouse during edge reconnect drag."""
        z  = self._zoom
        fx, fy = self._edge_reconnect['fixed_port']
        mx, my = self._mouse_xy
        self.create_line(fx*z, fy*z, mx*z, my*z,
                         fill='#f39c12', width=2, dash=(6, 3))

    def _draw_node(self, node):
        style = NODE_STYLES.get(node.type) or (
            _REACH_STYLE if node.type == 'REACH' else {})
        fill    = style.get('fill',    '#333333')
        outline = style.get('outline', '#ffffff')
        sel = (self._sel_node and self._sel_node.id == node.id) or (node.id in self._sel_nodes)
        if sel:
            outline = '#f39c12'
        lw = max(1, int((2 if sel else 1) * self._zoom))

        z  = self._zoom
        x, y = node.x * z, node.y * z
        hw = node._HW.get(node.type, 40) * z
        hh = node._HH.get(node.type, 20) * z
        tag = f'node_{node.id}'

        if node.type == 'SUBBASIN':
            self._round_rect(x-hw, y-hh, x+hw, y+hh, 10*z,
                             fill=fill, outline=outline, width=lw, tags=tag)
        elif node.type == 'RESERVOIR':
            pts = []
            for i in range(6):
                ang = math.pi / 2 + i * math.pi / 3
                pts.extend([x + hw * math.cos(ang), y + hh * math.sin(ang)])
            self.create_polygon(*pts, fill=fill, outline=outline, width=lw, tags=tag)
        elif node.type == 'REACH':   # backward compat (old loaded nodes)
            off = 10 * z
            pts = [x-hw+off, y-hh, x+hw+off, y-hh, x+hw-off, y+hh, x-hw-off, y+hh]
            self.create_polygon(*pts, fill=fill, outline=outline, width=lw, tags=tag)
        elif node.type == 'JUNCTION':
            self.create_oval(x-hw, y-hh, x+hw, y+hh,
                             fill=fill, outline=outline, width=lw, tags=tag)
        elif node.type == 'OUTLET':
            self.create_rectangle(x-hw, y-hh, x+hw, y+hh,
                                  fill=fill, outline=outline, width=lw, tags=tag)

        lbl = style.get('label', node.type)
        fs_name = min(max(6, int(9*z)), 18)
        fs_type = min(max(5, int(8*z)), 16)
        self.create_text(x, y - 7*z, text=node.name, fill='white',
                         font=('맑은 고딕', fs_name, 'bold'), tags=tag)
        self.create_text(x, y + 8*z, text=f'[{lbl}]', fill='#888888',
                         font=('맑은 고딕', fs_type), tags=tag)

        # Four-direction port circles — select 모드에서 비선택 노드는 생략 (성능)
        show_ports = sel or self._mode != 'select'
        if show_ports:
            has_out      = node.type != 'OUTLET'
            port_fill    = '#e67e22' if has_out else '#27ae60'
            port_outline = '#f39c12' if has_out else '#2ecc71'
            pr = max(2, int(PORT_R * z))
            for d in ('E', 'W', 'N', 'S'):
                wx, wy = node.port(d)
                px, py = wx*z, wy*z
                self.create_oval(px-pr, py-pr, px+pr, py+pr,
                                 fill=port_fill, outline=port_outline, width=1,
                                 tags=f'port_{d}_{node.id}')

    def _round_rect(self, x1, y1, x2, y2, r, **kw):
        pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r,
               x2,y2-r, x2,y2, x2-r,y2, x1+r,y2,
               x1,y2, x1,y2-r, x1,y1+r, x1,y1, x1+r,y1]
        return self.create_polygon(*pts, smooth=True, **kw)

    def _draw_rubber_band(self):
        src = self.nodes.get(self._conn_src)
        if not src: return
        d = self._conn_src_dir or 'E'
        z = self._zoom
        wx1, wy1 = src.port(d)
        x1, y1 = wx1*z, wy1*z
        wx2, wy2 = self._mouse_xy   # world coords
        x2, y2 = wx2*z, wy2*z
        DX = {'E': 1, 'W': -1, 'N': 0, 'S': 0}
        DY = {'E': 0, 'W':  0, 'N':-1, 'S': 1}
        off = max(abs(x2-x1)*0.4, abs(y2-y1)*0.4, 30)
        self.create_line(x1, y1, x1+DX[d]*off, y1+DY[d]*off, x2, y2,
                         smooth=True, fill='#5dade2', width=1, dash=(6,3))

    @staticmethod
    def _port_dirs(src, dst, src_hint=None, dst_hint=None):
        """위치 기반 자동 포트 방향 선택 (힌트 무시, 항상 최적 방향)."""
        dx = dst.x - src.x
        dy = dst.y - src.y
        if abs(dx) >= abs(dy):
            return ('E', 'W') if dx >= 0 else ('W', 'E')
        else:
            return ('S', 'N') if dy >= 0 else ('N', 'S')

    @staticmethod
    def _bezier_cps(x1, y1, d1, x2, y2, d2):
        """Compute bezier control points based on port exit/entry directions."""
        dist = max(abs(x2-x1), abs(y2-y1), 1)
        off  = max(dist * 0.45, 40)
        DX = {'E': 1, 'W': -1, 'N': 0, 'S': 0}
        DY = {'E': 0, 'W':  0, 'N':-1, 'S': 1}
        return ((x1 + DX[d1]*off, y1 + DY[d1]*off),
                (x2 + DX[d2]*off, y2 + DY[d2]*off))

    # ── orthogonal edge routing ───────────────────────────────────────────────
    _ARC_P = {                              # corner arc (start_angle, extent)
        # tkinter: positive extent = CW on screen (0=E, 90=S, 180=W, 270=N)
        ('E','S'):(180,-90), ('E','N'):(180, 90),
        ('W','S'):(0,   90), ('W','N'):(0,  -90),
        ('S','E'):(270, 90), ('S','W'):(270,-90),
        ('N','E'):(90, -90), ('N','W'):(90,  90),
    }

    @staticmethod
    def _ortho_wpts(x1, y1, dep, x2, y2, arr):
        """Return orthogonal waypoints: L-shape, Z-shape, or U-shape."""
        OPP = {'E':'W','W':'E','N':'S','S':'N'}
        H   = {'E','W'}
        if dep == arr:                    # same travel direction → Z or straight
            if dep in H:
                if abs(y1-y2) < 1: return [(x1,y1),(x2,y2)]
                mid = (x1+x2)/2
                return [(x1,y1),(mid,y1),(mid,y2),(x2,y2)]
            else:
                if abs(x1-x2) < 1: return [(x1,y1),(x2,y2)]
                mid = (y1+y2)/2
                return [(x1,y1),(x1,mid),(x2,mid),(x2,y2)]
        if dep == OPP[arr]:               # opposite directions → U-shape
            EXT = 80
            if dep in H:
                mx = (max(x1,x2)+EXT) if dep=='E' else (min(x1,x2)-EXT)
                return [(x1,y1),(mx,y1),(mx,y2),(x2,y2)]
            else:
                my = (max(y1,y2)+EXT) if dep=='S' else (min(y1,y2)-EXT)
                return [(x1,y1),(x1,my),(x2,my),(x2,y2)]
        # perpendicular → L-shape (1 corner)
        if dep in H:
            return [(x1,y1),(x2,y1),(x2,y2)]
        return [(x1,y1),(x1,y2),(x2,y2)]

    def _stroke_ortho(self, pts, CR, col, lw, tag):
        """Draw straight orthogonal segments through waypoints (no arc corners)."""
        n = len(pts)
        if n < 2: return
        flat = [v for p in pts for v in p]
        self.create_line(*flat, fill=col, width=lw,
                         arrow=tk.LAST, arrowshape=(7, 9, 3), tags=tag)

    def _draw_ortho_edge(self, x1, y1, d1, x2, y2, d2, col, lw, tag, CR):
        OPP = {'E':'W','W':'E','N':'S','S':'N'}
        arr = OPP[d2]
        pts = self._ortho_wpts(x1, y1, d1, x2, y2, arr)
        self._stroke_ortho(pts, CR, col, lw, tag)

    # ── events ───────────────────────────────────────────────────────────────

    def _motion(self, event):
        z = self._zoom
        self._mouse_xy = (self._cx(event.x) / z, self._cy(event.y) / z)
        if self._mode == 'connect':
            self._redraw_throttled()

    def _click(self, event):
        self.focus_set()
        self.after(0, self.focus_set)   # 콜백 내 위젯 생성/삭제 후 포커스 복원
        z = self._zoom
        x, y = self._cx(event.x) / z, self._cy(event.y) / z

        # --- PLACE MODE ---
        if self._mode.startswith('place:'):
            ntype = self._mode.split(':')[1]
            self._place_node(ntype, x, y)
            self.set_mode('select')
            if self._on_select: self._on_select(self._sel_node)
            return

        # --- CONNECT MODE ---
        if self._mode == 'connect':
            src_node = self.nodes.get(self._conn_src)
            for node in reversed(list(self.nodes.values())):
                if node.id == self._conn_src: continue
                if node.type == 'SUBBASIN': continue  # no inputs
                # Check 4 port circles
                for d in ('E', 'W', 'N', 'S'):
                    px, py = node.port(d)
                    if abs(x-px) <= HIT_R and abs(y-py) <= HIT_R:
                        sd, dd = self._port_dirs(src_node, node, self._conn_src_dir, d)
                        self._create_edge(self._conn_src, node.id, sd, d)
                        self._conn_src = None; self._conn_src_dir = None
                        self.set_mode('select'); self.redraw(); return
                # Or click node body
                if node.hit_test(x, y):
                    sd, dd = self._port_dirs(src_node, node, self._conn_src_dir, None) \
                             if src_node else ('E', 'W')
                    self._create_edge(self._conn_src, node.id, sd, dd)
                    self._conn_src = None; self._conn_src_dir = None
                    self.set_mode('select'); self.redraw(); return
            self._conn_src = None; self._conn_src_dir = None
            self.set_mode('select'); self.redraw(); return

        # --- SELECT MODE ---
        # 1. Check edges first (port 영역 제외) — 포트보다 먼저 검사해야 선택 가능
        OPP = {'E':'W','W':'E','N':'S','S':'N'}
        for edge in self.edges.values():
            src = self.nodes.get(edge.src)
            dst = self.nodes.get(edge.dst)
            if src and dst:
                sd, dd = self._port_dirs(src, dst, edge.src_dir, edge.dst_dir)
                p1 = src.port(sd); p2 = dst.port(dd)
                # 포트 근처 클릭은 포트 감지에 양보
                if abs(x-p1[0]) <= HIT_R and abs(y-p1[1]) <= HIT_R: continue
                if abs(x-p2[0]) <= HIT_R and abs(y-p2[1]) <= HIT_R: continue
                pts = self._ortho_wpts(p1[0], p1[1], sd, p2[0], p2[1], OPP[dd])
                hit = any(self._pt_seg_dist(x, y, pts[i][0], pts[i][1],
                                            pts[i+1][0], pts[i+1][1]) <= 8
                          for i in range(len(pts) - 1))
                if hit:
                    self._sel_edge = edge
                    self._sel_node = None
                    self._sel_nodes.clear()
                    if self._on_select: self._on_select(None)
                    if self._on_edge_select: self._on_edge_select(edge)
                    self.redraw()
                    return

        # 2. Check output ports (start connection)
        for node in reversed(list(self.nodes.values())):
            if node.type == 'OUTLET': continue  # no outputs
            for d in ('E', 'W', 'N', 'S'):
                px, py = node.port(d)
                if abs(x-px) <= HIT_R and abs(y-py) <= HIT_R:
                    self._conn_src = node.id
                    self._conn_src_dir = d
                    self.set_mode('connect')
                    self.redraw()
                    return

        # 3. Check selected edge endpoint handles (reconnect drag)
        if self._sel_edge:
            edge = self._sel_edge
            src  = self.nodes.get(edge.src)
            dst  = self.nodes.get(edge.dst)
            if src and dst:
                sd, dd = self._port_dirs(src, dst, edge.src_dir, edge.dst_dir)
                p1 = src.port(sd); p2 = dst.port(dd)
                HR = 10   # hit radius in world coords
                for end, px, py in (('src', p1[0], p1[1]), ('dst', p2[0], p2[1])):
                    if abs(x - px) <= HR and abs(y - py) <= HR:
                        self._push_undo()
                        fixed = p2 if end == 'src' else p1
                        self._edge_reconnect = {
                            'edge': edge, 'end': end, 'fixed_port': fixed
                        }
                        self._drag_start = (x, y, 0, 0)
                        self.redraw()
                        return

        # 4. Check reach edge labels (parallelogram) — drag moves src+dst together
        for edge in self.edges.values():
            if not edge.reach_params: continue
            src = self.nodes.get(edge.src)
            dst = self.nodes.get(edge.dst)
            if src and dst:
                sd, dd = self._port_dirs(src, dst, edge.src_dir, edge.dst_dir)
                p1 = src.port(sd); p2 = dst.port(dd)
                mx = (p1[0] + p2[0]) / 2
                my = (p1[1] + p2[1]) / 2
                if abs(x - mx) <= 35 and abs(y - my) <= 13:
                    self._push_undo()
                    self._sel_edge = edge
                    self._sel_node = None
                    self._sel_nodes.clear()
                    self._drag_start  = (x, y, 0, 0)
                    self._reach_drag  = (edge, src.x, src.y, dst.x, dst.y)
                    if self._on_select: self._on_select(None)
                    if self._on_edge_select: self._on_edge_select(edge)
                    self.redraw()
                    return

        # 5. Check node bodies
        for node in reversed(list(self.nodes.values())):
            if node.hit_test(x, y):
                if self._sel_nodes and node.id in self._sel_nodes:
                    # Start multi-drag — keep entire selection
                    self._push_undo()
                    self._drag_start = (x, y, node.x, node.y)
                    self._multi_origin = {nid: (self.nodes[nid].x, self.nodes[nid].y)
                                          for nid in self._sel_nodes if nid in self.nodes}
                else:
                    self._push_undo()
                    self._sel_nodes.clear()
                    self._multi_origin = {}
                    self._sel_node = node
                    self._sel_edge = None
                    self._drag_start = (x, y, node.x, node.y)
                    if self._on_select: self._on_select(node)
                self.redraw()
                return

        # 4. Nothing hit — start rubber-band selection
        self._sel_node = None
        self._sel_edge = None
        self._sel_nodes.clear()
        self._rubber_start = (x, y)
        if self._on_select: self._on_select(None)
        self.redraw()

    def _drag(self, event):
        z = self._zoom
        x, y = self._cx(event.x) / z, self._cy(event.y) / z
        if self._rubber_start:
            self._mouse_xy = (x, y)
            # 빠른 경로: 선택 사각형 좌표만 갱신 (전체 redraw 없음)
            x0w, y0w = self._rubber_start
            if self.find_withtag('_sel_box'):
                self.coords('_sel_box', x0w*z, y0w*z, x*z, y*z)
            else:
                self.redraw()
            return
        if self._edge_reconnect and self._drag_start:
            self._mouse_xy = (x, y)
            self._redraw_throttled()
            return
        if self._reach_drag and self._drag_start:
            dx = x - self._drag_start[0]
            dy = y - self._drag_start[1]
            edge, sox, soy, dox, doy = self._reach_drag
            src = self.nodes.get(edge.src)
            dst = self.nodes.get(edge.dst)
            if src:
                nx, ny = sox + dx, soy + dy
                if self._snap_on: nx, ny = self._snap(nx), self._snap(ny)
                src.x, src.y = nx, ny
            if dst:
                nx, ny = dox + dx, doy + dy
                if self._snap_on: nx, ny = self._snap(nx), self._snap(ny)
                dst.x, dst.y = nx, ny
            self._redraw_throttled()
            return
        if self._drag_start:
            dx = x - self._drag_start[0]
            dy = y - self._drag_start[1]
            if self._multi_origin:
                for nid, (ox, oy) in self._multi_origin.items():
                    node = self.nodes.get(nid)
                    if node:
                        nx = ox + dx
                        ny = oy + dy
                        if self._snap_on:
                            nx, ny = self._snap(nx), self._snap(ny)
                        node.x, node.y = nx, ny
            elif self._sel_node:
                nx = self._drag_start[2] + dx
                ny = self._drag_start[3] + dy
                if self._snap_on:
                    nx, ny = self._snap(nx), self._snap(ny)
                self._sel_node.x = nx
                self._sel_node.y = ny
            self._redraw_throttled()

    def _release(self, event):
        if self._edge_reconnect:
            z = self._zoom
            x = self._cx(event.x) / z
            y = self._cy(event.y) / z
            info = self._edge_reconnect
            self._edge_reconnect = None
            self._drag_start = None
            # 가장 가까운 포트 탐색 (스냅 거리 20)
            best_node, best_dir, best_dist = None, None, 20
            edge = info['edge']
            for node in self.nodes.values():
                # src 이동 시 dst 노드 제외, dst 이동 시 src 노드 제외
                if info['end'] == 'src' and node.id == edge.dst: continue
                if info['end'] == 'dst' and node.id == edge.src: continue
                if node.type == 'SUBBASIN' and info['end'] == 'dst': continue
                for d in ('E', 'W', 'N', 'S'):
                    px, py = node.port(d)
                    dist = math.hypot(x - px, y - py)
                    if dist < best_dist:
                        best_dist, best_node, best_dir = dist, node, d
            if best_node:
                if info['end'] == 'src':
                    edge.src     = best_node.id
                    edge.src_dir = best_dir
                else:
                    edge.dst     = best_node.id
                    edge.dst_dir = best_dir
            self.redraw()
            return
        if self._rubber_start:
            z = self._zoom
            x = self._cx(event.x) / z
            y = self._cy(event.y) / z
            x0, y0 = self._rubber_start
            xmin, xmax = min(x0, x), max(x0, x)
            ymin, ymax = min(y0, y), max(y0, y)
            self._rubber_start = None
            self.delete('_sel_box')   # 선택 사각형 즉시 제거
            if xmax - xmin > 4 or ymax - ymin > 4:
                hit = {n.id for n in self.nodes.values()
                       if xmin <= n.x <= xmax and ymin <= n.y <= ymax}
                if len(hit) == 1:
                    nid = next(iter(hit))
                    self._sel_node = self.nodes.get(nid)
                    self._sel_nodes.clear()
                    if self._on_select: self._on_select(self._sel_node)
                elif hit:
                    self._sel_nodes = hit
                    self._sel_node  = None
                    if self._on_select: self._on_select(None)
            self.after(0, self.redraw)
        # 드래그 종료 시 대기 중인 스로틀 취소 후 즉시 최종 redraw
        if self._drag_redraw_id is not None:
            self.after_cancel(self._drag_redraw_id)
            self._drag_redraw_id = None
        self._drag_start     = None
        self._multi_origin   = {}
        self._reach_drag     = None
        self._edge_reconnect = None

    def _delete(self, event):
        if self._sel_edge:
            self._push_undo()
            edge_label = self._sel_edge.label or str(self._sel_edge.id)
            eid = self._sel_edge.id
            if eid in self.edges: del self.edges[eid]
            self._sel_edge = None
            if self._on_log: self._on_log(f'연결선 삭제: {edge_label}')
            self.redraw()
        elif self._sel_nodes:
            self._push_undo()
            cnt = len(self._sel_nodes)
            for nid in list(self._sel_nodes):
                for eid in [e for e, ed in self.edges.items() if ed.src == nid or ed.dst == nid]:
                    del self.edges[eid]
                if nid in self.nodes:
                    del self.nodes[nid]
            self._sel_nodes.clear()
            if self._on_select: self._on_select(None)
            if self._on_log: self._on_log(f'다중 삭제: {cnt}개 요소')
            self.redraw()
        elif self._sel_node:
            del_name = self._sel_node.name
            del_type = NODE_STYLES.get(self._sel_node.type, {}).get('label', self._sel_node.type)
            self._push_undo()
            nid = self._sel_node.id
            for eid in [eid for eid, e in self.edges.items() if e.src == nid or e.dst == nid]:
                del self.edges[eid]
            del self.nodes[nid]
            self._sel_node = None
            if self._on_log: self._on_log(f'{del_type} 삭제: {del_name}')
            if self._on_select: self._on_select(None)
            self.redraw()

    def _right_click(self, event):
        z = self._zoom
        x, y = self._cx(event.x) / z, self._cy(event.y) / z
        OPP = {'E':'W','W':'E','N':'S','S':'N'}
        hit_edge = None
        for edge in self.edges.values():
            src = self.nodes.get(edge.src)
            dst = self.nodes.get(edge.dst)
            if src and dst:
                sd, dd = self._port_dirs(src, dst, edge.src_dir, edge.dst_dir)
                p1 = src.port(sd); p2 = dst.port(dd)
                pts = self._ortho_wpts(p1[0], p1[1], sd, p2[0], p2[1], OPP[dd])
                if any(self._pt_seg_dist(x, y, pts[i][0], pts[i][1],
                                         pts[i+1][0], pts[i+1][1]) <= 8
                       for i in range(len(pts) - 1)):
                    hit_edge = edge
                    break
        if hit_edge is None:
            return
        # 선택 상태로 표시
        self._sel_edge = hit_edge
        self._sel_node = None
        self._sel_nodes.clear()
        self.redraw()
        # 컨텍스트 메뉴
        menu = tk.Menu(self, tearoff=0, bg='#1e1e2e', fg='white',
                       activebackground='#c0392b', activeforeground='white',
                       font=('맑은 고딕', 10))
        lbl = hit_edge.label or hit_edge.id[:8]
        menu.add_command(label=f'연결선 삭제: {lbl}',
                         command=lambda e=hit_edge: self._delete_edge(e))
        menu.tk_popup(event.x_root, event.y_root)

    def _delete_edge(self, edge):
        self._push_undo()
        eid = edge.id
        lbl = edge.label or eid[:8]
        if eid in self.edges:
            del self.edges[eid]
        if self._sel_edge and self._sel_edge.id == eid:
            self._sel_edge = None
        if self._on_edge_select: self._on_edge_select(None)
        if self._on_log: self._on_log(f'연결선 삭제: {lbl}')
        self.redraw()

    def _escape(self, event):
        self._conn_src = None
        self._conn_src_dir = None
        self.set_mode('select')
        self.redraw()

    # ── undo / redo ──────────────────────────────────────────────────────────

    def _push_undo(self):
        state = (copy.deepcopy(self.nodes), copy.deepcopy(self.edges))
        self._undo_stack.append(state)
        if len(self._undo_stack) > 50:
            self._undo_stack.pop(0)
        self._redo_stack.clear()

    def _undo(self, event=None):
        if not self._undo_stack: return 'break'
        cur = (copy.deepcopy(self.nodes), copy.deepcopy(self.edges))
        self._redo_stack.append(cur)
        self.nodes, self.edges = self._undo_stack.pop()
        self._sel_node = None
        self._sel_edge = None
        self._sel_nodes.clear()
        if self._on_select: self._on_select(None)
        self.redraw()
        return 'break'

    def _redo(self, event=None):
        if not self._redo_stack: return
        cur = (copy.deepcopy(self.nodes), copy.deepcopy(self.edges))
        self._undo_stack.append(cur)
        self.nodes, self.edges = self._redo_stack.pop()
        self._sel_node = None
        self._sel_edge = None
        self._sel_nodes.clear()
        if self._on_select: self._on_select(None)
        self.redraw()

    # ── zoom / pan ───────────────────────────────────────────────────────────

    def _update_scrollregion(self):
        PAD = self._SCROLL_PAD
        W = self._world_bbox[2] * self._zoom
        H = self._world_bbox[3] * self._zoom
        self.configure(scrollregion=(-PAD, -PAD, int(W) + PAD, int(H) + PAD))

    def _on_zoom(self, event):
        factor   = 1.1 if event.delta > 0 else 1/1.1
        new_zoom = max(0.15, min(4.0, self._zoom * factor))
        if abs(new_zoom - self._zoom) < 1e-9: return
        cx_m = self.canvasx(event.x)
        cy_m = self.canvasy(event.y)
        wx_m = cx_m / self._zoom
        wy_m = cy_m / self._zoom
        self._zoom = new_zoom
        self._update_scrollregion()
        PAD = self._SCROLL_PAD
        W  = self._world_bbox[2] * self._zoom
        H  = self._world_bbox[3] * self._zoom
        tW = W + 2 * PAD
        tH = H + 2 * PAD
        new_cx = wx_m * self._zoom
        new_cy = wy_m * self._zoom
        # 클램프 없음 — 음수 캔버스 영역으로 자유롭게 이동
        if tW > 0: self.xview_moveto((new_cx - event.x + PAD) / tW)
        if tH > 0: self.yview_moveto((new_cy - event.y + PAD) / tH)
        self.redraw()

    def _on_pan_press(self, event):
        self.configure(cursor='fleur')
        self.scan_mark(event.x, event.y)

    def _on_pan_drag(self, event):
        self.scan_dragto(event.x, event.y, gain=1)
        self.redraw()

    def _on_pan_release(self, event):
        self.configure(cursor='crosshair' if self._mode != 'select' else 'arrow')
        self._pan_last = None

    def _zoom_extents(self, event=None):
        """휠버튼 더블클릭: 모든 노드가 보이도록 줌·스크롤 조정 (CAD Zoom Extents)."""
        if not self.nodes: return
        xs = [n.x for n in self.nodes.values()]
        ys = [n.y for n in self.nodes.values()]
        PAD_W = 4 * self.GRID          # 노드 외곽 여유 (world units)
        wx0 = min(xs) - PAD_W;  wx1 = max(xs) + PAD_W
        wy0 = min(ys) - PAD_W;  wy1 = max(ys) + PAD_W
        ww = wx1 - wx0;  wh = wy1 - wy0
        if ww < 1 or wh < 1: return
        cw = self.winfo_width();  ch = self.winfo_height()
        if cw < 10 or ch < 10: return
        self._zoom = max(0.15, min(4.0, min(cw / ww, ch / wh)))
        self._update_scrollregion()
        PAD = self._SCROLL_PAD
        W = self._world_bbox[2] * self._zoom + 2 * PAD
        H = self._world_bbox[3] * self._zoom + 2 * PAD
        cx = (wx0 + ww / 2) * self._zoom
        cy = (wy0 + wh / 2) * self._zoom
        if W > 0: self.xview_moveto((cx - cw / 2 + PAD) / W)
        if H > 0: self.yview_moveto((cy - ch / 2 + PAD) / H)
        self.redraw()

    def _on_arrow(self, event):
        step = 30.0
        PAD = self._SCROLL_PAD
        tW = self._world_bbox[2] * self._zoom + 2 * PAD
        tH = self._world_bbox[3] * self._zoom + 2 * PAD
        if   event.keysym == 'Left':  self.xview_moveto(self.xview()[0] - step / tW)
        elif event.keysym == 'Right': self.xview_moveto(self.xview()[0] + step / tW)
        elif event.keysym == 'Up':    self.yview_moveto(self.yview()[0] - step / tH)
        elif event.keysym == 'Down':  self.yview_moveto(self.yview()[0] + step / tH)

    def _dbl_click(self, event):
        x, y = self._cx(event.x) / self._zoom, self._cy(event.y) / self._zoom
        for node in reversed(list(self.nodes.values())):
            if node.hit_test(x, y):
                self._edit_node_dialog(node)
                return

    # ── helpers ──────────────────────────────────────────────────────────────

    def _place_node(self, ntype, x, y):
        self._push_undo()
        if self._snap_on:
            x, y = self._snap(x), self._snap(y)
        count = sum(1 for n in self.nodes.values() if n.type == ntype) + 1
        prefix = {'SUBBASIN': 'SB', 'RESERVOIR': 'RS', 'JUNCTION': 'JN', 'OUTLET': 'OUT'}
        name = f"{prefix.get(ntype,'ND')}{count:02d}"
        node = NetworkNode(ntype, name, x, y)
        self.nodes[node.id] = node
        self._sel_node = node
        self._sel_edge = None
        self.redraw()
        if self._on_log:
            self._on_log(f'{NODE_STYLES.get(ntype, {}).get("label", ntype)} 생성: {name}')

    def _snap(self, v):
        """Snap world coordinate to nearest grid point."""
        return round(v / self.GRID) * self.GRID

    @staticmethod
    def _pt_seg_dist(px, py, ax, ay, bx, by):
        """Distance from point (px,py) to segment (ax,ay)-(bx,by)."""
        dx, dy = bx - ax, by - ay
        if dx == 0 and dy == 0:
            return math.hypot(px - ax, py - ay)
        t = max(0.0, min(1.0, ((px - ax) * dx + (py - ay) * dy) / (dx * dx + dy * dy)))
        return math.hypot(px - (ax + t * dx), py - (ay + t * dy))

    def _create_edge(self, src_id, dst_id, src_dir=None, dst_dir=None):
        if any(e.src == src_id and e.dst == dst_id for e in self.edges.values()):
            self._conn_reach = False
            return
        if src_id == dst_id:
            self._conn_reach = False
            return
        self._push_undo()
        rp = {'K': 1.0, 'X': 0.20, 'NSTPS': 0} if self._conn_reach else None
        cnt = sum(1 for e in self.edges.values() if e.reach_params) + 1
        lbl = f'RC{cnt:02d}' if rp else None
        edge = NetworkEdge(src_id, dst_id, src_dir, dst_dir, reach_params=rp, label=lbl)
        self.edges[edge.id] = edge
        self._conn_reach = False
        if self._on_log:
            sn = self.nodes.get(src_id)
            dn = self.nodes.get(dst_id)
            sname = sn.name if sn else str(src_id)
            dname = dn.name if dn else str(dst_id)
            reach_tag = ' [하도추적]' if rp else ''
            self._on_log(f'연결: {sname} → {dname}{reach_tag}')
        self.redraw()

    def _edit_node_dialog(self, node):
        dlg = NodeEditDialog(self.winfo_toplevel(), node)
        self.winfo_toplevel().wait_window(dlg)
        self.redraw()
        if self._on_select: self._on_select(node)

    def add_node(self, node):
        self.nodes[node.id] = node

    def add_edge(self, edge):
        self.edges[edge.id] = edge

    def clear(self):
        self.nodes.clear()
        self.edges.clear()
        self._sel_node = None
        self._sel_edge = None
        self._conn_src = None
        self.redraw()

    # ── graph → operations (DFS) ─────────────────────────────────────────────

    def build_operations(self):
        outlets = [n for n in self.nodes.values() if n.type == 'OUTLET']
        if not outlets:
            return [], "출구(OUTLET) 노드가 없습니다."

        outlet = outlets[0]
        preds = {n.id: [] for n in self.nodes.values()}
        for e in self.edges.values():
            preds[e.dst].append(e.src)

        ops = []
        visited = set()
        errors  = []

        # Build edge lookup: (src_id, dst_id) -> edge
        edge_map = {(e.src, e.dst): e for e in self.edges.values()}

        def emit_reach(uid, dst_id):
            """If edge uid→dst_id has reach_params, emit ROUTE op."""
            e = edge_map.get((uid, dst_id))
            if e and e.reach_params:
                rp = e.reach_params
                ops.append({'type': 'ROUTE', 'name': e.label or f'RC_{uid}_{dst_id}',
                             'K': rp.get('K', 1.0), 'X': rp.get('X', 0.20),
                             'NSTPS': rp.get('NSTPS', 0)})

        def dfs(nid):
            if nid in visited:
                errors.append(f"순환 연결 감지: {self.nodes[nid].name}")
                return
            visited.add(nid)
            node = self.nodes.get(nid)
            if not node: return
            up_ids = preds.get(nid, [])

            if node.type == 'SUBBASIN':
                ops.append({'type': 'BASIN', 'name': node.name, **node.params})

            elif node.type == 'RESERVOIR':
                for uid in up_ids:
                    dfs(uid)
                    emit_reach(uid, nid)
                ops.append({'type': 'RESERVOIR', 'name': node.name, **node.params})

            elif node.type == 'REACH':   # backward compat for old loaded nodes
                for uid in up_ids:
                    dfs(uid)
                ops.append({'type': 'ROUTE', 'name': node.name, **node.params})

            elif node.type == 'JUNCTION':
                for uid in up_ids:
                    dfs(uid)
                    emit_reach(uid, nid)
                N = len(up_ids)
                if N >= 2:
                    ops.append({'type': 'COMBINE', 'name': node.name, 'N': N})
                elif N == 1:
                    pass
                else:
                    errors.append(f"합류점 '{node.name}' 에 입력 없음")

            elif node.type == 'OUTLET':
                for uid in up_ids:
                    dfs(uid)
                    emit_reach(uid, nid)
                N = len(up_ids)
                if N >= 2:
                    errors.append(f"출구 '{node.name}'에 {N}개 직접 연결 — 합류점(JUNCTION) 노드를 사용하세요.")

        dfs(outlet.id)

        if errors:
            return ops, "\n".join(errors)
        return ops, None

    # ── flat ops → graph (for loading from config) ────────────────────────────

    def load_operations(self, operations):
        """Convert flat operations list → visual graph with auto-layout."""
        self.clear()
        # stack of node names (strings)
        stack  = []
        # pending reach: if the top of stack has a ROUTE op applied,
        # store it as pending_reach[(src_name)] = (reach_name, reach_params)
        pending_reach = {}   # src_name -> (reach_name, reach_params)
        name_to_node  = {}

        for op in operations:
            t  = op['type']
            nm = op['name']
            params = {k: v for k, v in op.items() if k not in ('type', 'name')}

            if t == 'BASIN':
                node = NetworkNode('SUBBASIN', nm, 0, 0, params)
                self.nodes[node.id] = node
                name_to_node[nm] = node
                stack.append(nm)

            elif t == 'ROUTE':
                # Mark the top-of-stack node as having a pending reach
                if stack:
                    src = stack[-1]
                    rp = {'K': params.get('K', 1.0), 'X': params.get('X', 0.20),
                          'NSTPS': params.get('NSTPS', 0)}
                    pending_reach[src] = (nm, rp)
                    # keep the same src name on the stack (the reach is on the edge, not a node)

            elif t == 'RESERVOIR':
                node = NetworkNode('RESERVOIR', nm, 0, 0, params)
                self.nodes[node.id] = node
                name_to_node[nm] = node
                if stack:
                    prev = stack.pop()
                    pr = pending_reach.pop(prev, None)
                    if pr:
                        reach_nm, rp = pr
                        edge = NetworkEdge(name_to_node[prev].id, node.id,
                                           reach_params=rp, label=reach_nm)
                    else:
                        edge = NetworkEdge(name_to_node[prev].id, node.id)
                    self.edges[edge.id] = edge
                stack.append(nm)

            elif t == 'COMBINE':
                N = int(op.get('N', 2))
                node = NetworkNode('JUNCTION', nm, 0, 0, {})
                self.nodes[node.id] = node
                name_to_node[nm] = node
                pop_n = min(N, len(stack))
                for _ in range(pop_n):
                    prev = stack.pop()
                    pr = pending_reach.pop(prev, None)
                    if pr:
                        reach_nm, rp = pr
                        if prev in name_to_node:
                            edge = NetworkEdge(name_to_node[prev].id, node.id,
                                               reach_params=rp, label=reach_nm)
                            self.edges[edge.id] = edge
                    else:
                        if prev in name_to_node:
                            edge = NetworkEdge(name_to_node[prev].id, node.id)
                            self.edges[edge.id] = edge
                stack.append(nm)

        # Add OUTLET — 스택에 1개만 남은 경우에만 OUT 자동 생성
        # 2개 이상 남은 경우 HC N + OUT 카드 누락 → 원본 그대로 배열 (경고는 호출부에서 표시)
        if len(stack) == 1:
            out_node = NetworkNode('OUTLET', 'OUT', 0, 0, {})
            self.nodes[out_node.id] = out_node
            nm = stack[0]
            if nm in name_to_node:
                pr = pending_reach.pop(nm, None)
                if pr:
                    reach_nm, rp = pr
                    edge = NetworkEdge(name_to_node[nm].id, out_node.id,
                                       reach_params=rp, label=reach_nm)
                else:
                    edge = NetworkEdge(name_to_node[nm].id, out_node.id)
                self.edges[edge.id] = edge

        self._auto_layout()
        self.redraw()

    def load_canvas_state(self, data):
        """캔버스 전체 상태(노드 위치 포함) 복원."""
        self.clear()
        # 노드 복원
        max_nid = 0
        for nd in data.get('nodes', []):
            n = NetworkNode(nd['type'], nd['name'], nd['x'], nd['y'],
                            params=nd.get('params'))
            n.id = nd['id']
            self.nodes[n.id] = n
            if n.id > max_nid: max_nid = n.id
        NetworkNode._counter = max_nid

        # 엣지 복원
        max_eid = 0
        for ed in data.get('edges', []):
            e = NetworkEdge(ed['src'], ed['dst'],
                            src_dir=ed.get('src_dir'),
                            dst_dir=ed.get('dst_dir'),
                            reach_params=ed.get('reach_params'),
                            label=ed.get('label'))
            e.id = ed['id']
            self.edges[e.id] = e
            if e.id > max_eid: max_eid = e.id
        NetworkEdge._counter = max_eid

        # world_bbox를 노드 위치 기반으로 재계산
        if self.nodes:
            G = self.GRID
            max_x = max(n.x for n in self.nodes.values()) + 6 * G
            max_y = max(n.y for n in self.nodes.values()) + 6 * G
            self._world_bbox = (0, 0, int(max(max_x, 2000)), int(max(max_y, 900)))
        self._update_scrollregion()
        self.redraw()

    def _auto_layout(self):
        if not self.nodes: return

        # ── 인접 리스트 ──────────────────────────────────────────────────────
        preds = {n.id: [] for n in self.nodes.values()}
        succs = {n.id: [] for n in self.nodes.values()}
        for e in self.edges.values():
            if e.src in preds and e.dst in preds:
                preds[e.dst].append(e.src)
                succs[e.src].append(e.dst)

        G      = self.GRID        # 40
        STEP_X = 5 * G            # 200  수평 간격
        STEP_Y = 5 * G            # 200  수직 밴드 간격
        MX     = 8 * G            # 320  좌측 여백
        MAIN_Y = 10 * G           # 400  본류 Y

        BACKBONE = {'JUNCTION', 'RESERVOIR', 'OUTLET'}

        # ── 헬퍼 ─────────────────────────────────────────────────────────────
        def has_reach(pred_id, junc_id):
            return any(e.src == pred_id and e.dst == junc_id and e.reach_params
                       for e in self.edges.values())

        def upstream_size(nid):
            seen, q = set(), [nid]
            while q:
                n = q.pop()
                if n in seen: continue
                seen.add(n); q.extend(preds[n])
            return len(seen)

        # ── OUTLET 판별 ────────────────────────────────────────────────────
        # 명시적 OUTLET 없으면 backbone successor 없는 terminal 중 가장 큰 upstream
        outlets = [n.id for n in self.nodes.values() if n.type == 'OUTLET']
        if outlets:
            outlet_id = outlets[0]
        else:
            terminal_bb = [n.id for n in self.nodes.values()
                           if n.type in BACKBONE
                           and not any(self.nodes[s].type in BACKBONE
                                       for s in succs[n.id] if s in self.nodes)]
            if not terminal_bb:
                self._layout_bfs(preds, succs); return
            outlet_id = max(terminal_bb, key=upstream_size)

        # ── Step 2: Backbone depth BFS (backbone 노드만) ─────────────────────
        depth = {outlet_id: 0}
        queue = [outlet_id]
        while queue:
            nid = queue.pop(0)
            for c in preds[nid]:
                if self.nodes[c].type in BACKBONE:
                    nd = depth[nid] + 1
                    if nd > depth.get(c, -1):
                        depth[c] = nd; queue.append(c)
        for n in self.nodes.values():
            if n.id not in depth:
                depth[n.id] = 0

        # ── Step 3: Main chain (backbone only) ───────────────────────────────
        main_chain = [outlet_id]
        cur = outlet_id
        while True:
            cands = [p for p in preds[cur] if self.nodes[p].type in BACKBONE]
            if not cands: break
            nxt = max(cands, key=lambda n: (
                depth.get(n, 0),
                1 if has_reach(n, cur) else 0,
                upstream_size(n)
            ))
            main_chain.append(nxt); cur = nxt
        main_chain.reverse()   # 상류 → 하류
        main_set = set(main_chain)

        # ── Step 4: Sub-junction 탐색 ────────────────────────────────────────
        # backbone이지만 main_chain에 없는 노드 중 main_chain 노드에 직접 연결되는 것
        sub_junctions = {}   # sj_id → mc_id (연결되는 본류 합류점)
        for mc_id in main_chain:
            for p in preds[mc_id]:
                if p not in main_set and self.nodes[p].type in {'JUNCTION', 'RESERVOIR'}:
                    sub_junctions[p] = mc_id

        # ── Step 5: 본류 노드 X 위치 ─────────────────────────────────────────
        max_depth = max(depth[n] for n in main_chain) if main_chain else 0
        positions = {}
        taken = set()
        for nid in main_chain:
            col = max_depth - depth[nid]
            x = MX + col * STEP_X
            positions[nid] = (x, MAIN_Y)
            taken.add((x, MAIN_Y))

        # ── Step 6: 서브 합류점 배치 (본류 노드 바로 위, 동일 X) ─────────────
        SJ_Y = MAIN_Y - STEP_Y   # 200
        for sj_id, mc_id in sub_junctions.items():
            x = positions[mc_id][0]
            # 같은 X에 서브 합류점이 이미 있으면 한 칸 왼쪽으로 이동
            y = SJ_Y
            while (x, y) in taken:
                x -= STEP_X
            positions[sj_id] = (x, y)
            taken.add((x, y))

        # ── Step 6.5: 서브 합류점 reach-소유역이 점유할 X 위치 사전 계산 ─────
        # Step 7에서 위쪽 배치 시 해당 위치와 충돌하지 않도록 forced_below_mcs 구성
        forced_below_mcs = set()
        for sj_id, mc_id in sub_junctions.items():
            # 서브 합류점 자체가 (mc_x, SJ_Y) 를 점유 → 해당 mc_id 아래 강제
            forced_below_mcs.add(mc_id)
            sj_x = positions[sj_id][0]
            # reach-연결 소유역이 (sj_x - STEP_X, SJ_Y) 쪽에 배치될 예정
            for sb_id in preds[sj_id]:
                if self.nodes[sb_id].type == 'SUBBASIN' and has_reach(sb_id, sj_id):
                    x = sj_x - STEP_X
                    while (x, SJ_Y) in taken:
                        x -= STEP_X
                    # 이 X를 가진 본류 합류점도 아래 강제
                    for other_mc in main_chain:
                        if positions[other_mc][0] == x:
                            forced_below_mcs.add(other_mc)

        # ── Step 7: 본류 소유역 배치 ─────────────────────────────────────────
        above_cnt = {}   # mc_id → 위 방향 배치 횟수
        below_cnt = {}   # mc_id → 아래 방향 배치 횟수

        for mc_id in main_chain:
            mc_x = positions[mc_id][0]
            force_below = mc_id in forced_below_mcs
            direct_sb = [p for p in preds[mc_id]
                         if p not in main_set
                         and p not in sub_junctions
                         and self.nodes[p].type == 'SUBBASIN']
            # reach 연결 먼저, 이후 name 순
            direct_sb.sort(key=lambda n: (0 if has_reach(n, mc_id) else 1,
                                          self.nodes[n].name))
            for i, sb_id in enumerate(direct_sb):
                if not force_below and i % 2 == 0:   # 짝수 → 위 (강제 아래 아닐 때만)
                    row = above_cnt.get(mc_id, 0)
                    y = MAIN_Y - (row + 1) * STEP_Y
                    above_cnt[mc_id] = row + 1
                    while (mc_x, y) in taken:
                        y -= STEP_Y
                else:                                 # 아래
                    row = below_cnt.get(mc_id, 0)
                    y = MAIN_Y + (row + 1) * STEP_Y
                    below_cnt[mc_id] = row + 1
                    while (mc_x, y) in taken:
                        y += STEP_Y
                positions[sb_id] = (mc_x, y)
                taken.add((mc_x, y))

        # ── Step 8: 서브 합류점 소유역 배치 ──────────────────────────────────
        for sj_id in sub_junctions:
            sj_x, sj_y = positions[sj_id]
            sb_list = [p for p in preds[sj_id]
                       if self.nodes[p].type == 'SUBBASIN']
            sb_list.sort(key=lambda n: (0 if has_reach(n, sj_id) else 1,
                                        self.nodes[n].name))
            for sb_id in sb_list:
                if has_reach(sb_id, sj_id):
                    # reach 연결 → 서브 합류점 왼쪽, 동일 Y (수평 직선)
                    x = sj_x - STEP_X
                    while (x, sj_y) in taken:
                        x -= STEP_X
                    positions[sb_id] = (x, sj_y)
                else:
                    # 서브 합류점 위쪽 (수직 직선)
                    y = sj_y - STEP_Y
                    while (sj_x, y) in taken:
                        y -= STEP_Y
                    positions[sb_id] = (sj_x, y)
                taken.add(positions[sb_id])

        # ── Step 9: 고아 노드 처리 (클러스터 인식) ───────────────────────────
        orphan_ids = [n.id for n in self.nodes.values() if n.id not in positions]
        if orphan_ids:
            orphan_set = set(orphan_ids)
            o_preds = {n: [p for p in preds[n] if p in orphan_set] for n in orphan_set}
            o_succs = {n: [s for s in succs[n] if s in orphan_set] for n in orphan_set}
            cluster_roots = [n for n in orphan_set if not o_succs[n]] or list(orphan_set)
            fb_x = MX + (max_depth + 2) * STEP_X
            fb_y = MAIN_Y
            placed = set()
            for root in sorted(cluster_roots, key=lambda n: self.nodes[n].name):
                while (fb_x, fb_y) in taken:
                    fb_y += STEP_Y
                positions[root] = (fb_x, fb_y)
                taken.add((fb_x, fb_y))
                placed.add(root)
                above_c = 0
                for p in sorted(o_preds.get(root, []),
                                key=lambda n: (0 if has_reach(n, root) else 1,
                                               self.nodes[n].name)):
                    if p in placed:
                        continue
                    if has_reach(p, root):
                        x = fb_x - STEP_X
                        while (x, fb_y) in taken:
                            x -= STEP_X
                        positions[p] = (x, fb_y)
                    else:
                        above_c += 1
                        y = fb_y - above_c * STEP_Y
                        while (fb_x, y) in taken:
                            y -= STEP_Y
                        positions[p] = (fb_x, y)
                    taken.add(positions[p])
                    placed.add(p)
                fb_y += STEP_Y
            for nid in orphan_ids:
                if nid not in placed:
                    while (fb_x, fb_y) in taken:
                        fb_y += STEP_Y
                    positions[nid] = (fb_x, fb_y)
                    taken.add((fb_x, fb_y))
                    fb_y += STEP_Y

        # ── Step 10: 위치 적용 + y < G 보정 + world_bbox 갱신 ────────────────
        for nid, (x, y) in positions.items():
            if nid in self.nodes:
                self.nodes[nid].x = float(x)
                self.nodes[nid].y = float(y)

        min_y = min(n.y for n in self.nodes.values())
        if min_y < G:
            shift = round((G - min_y) / G) * G
            for n in self.nodes.values():
                n.y += shift

        max_x = max(n.x for n in self.nodes.values()) + 3 * STEP_X
        max_y = max(n.y for n in self.nodes.values()) + 3 * STEP_Y
        self._world_bbox = (0, 0, int(max(max_x, 2000)), int(max(max_y, 900)))
        self._update_scrollregion()

    def _layout_bfs(self, preds, succs):
        """Fallback BFS level layout for networks without a clear outlet."""
        levels  = {n.id: 0 for n in self.nodes.values()}
        sources = [n.id for n in self.nodes.values() if not preds[n.id]]
        queue   = list(sources)
        while queue:
            nid = queue.pop(0)
            for s in succs[nid]:
                lv = levels[nid] + 1
                if lv > levels[s]:
                    levels[s] = lv
                    queue.append(s)
        by_level = {}
        for nid, lv in levels.items():
            by_level.setdefault(lv, []).append(nid)
        G = self.GRID
        STEP_X = 5 * G; STEP_Y = 3 * G; MARGIN_X = 3 * G; MARGIN_Y = 8 * G
        for lv, nids in sorted(by_level.items()):
            x = MARGIN_X + lv * STEP_X
            total_h = (len(nids) - 1) * STEP_Y
            start_y = MARGIN_Y + max(0, (500 - total_h) // 2)
            for i, nid in enumerate(nids):
                self.nodes[nid].x = float(x)
                self.nodes[nid].y = float(start_y + i * STEP_Y)
        self._world_bbox = (0, 0, 2000, 1200)
        self._update_scrollregion()


# =============================================================================
# 노드 편집 다이얼로그
# =============================================================================

class NodeEditDialog(ctk.CTkToplevel):
    def __init__(self, parent, node):
        super().__init__(parent)
        self.node = node
        self.title(f"노드 편집 — {node.name}")
        self.geometry("340x400")
        self.resizable(False, False)
        self.grab_set()
        self.focus()
        self._set_dark()
        self._entries = {}
        self._build()

    def _set_dark(self):
        try:
            hwnd = windll.user32.GetParent(self.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
        except Exception: pass

    def _build(self):
        node = self.node
        style = NODE_STYLES.get(node.type, {})
        frame = ctk.CTkScrollableFrame(self)
        frame.pack(fill='both', expand=True, padx=12, pady=8)
        frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(frame, text=f"[ {style.get('label', node.type)} ]",
                     font=FONT_HEADER,
                     text_color=style.get('outline', 'white')).grid(
            row=0, column=0, columnspan=2, sticky='w', pady=(0,8))

        def row(r, label, key, default):
            ctk.CTkLabel(frame, text=label, font=FONT_SMALL, anchor='w').grid(
                row=r, column=0, sticky='w', padx=4, pady=3)
            ent = ctk.CTkEntry(frame, font=FONT_SMALL, justify='right')
            ent.insert(0, str(default))
            ent.grid(row=r, column=1, sticky='ew', padx=4, pady=3)
            self._entries[key] = ent

        r = 1
        row(r, '노드 이름', 'name', node.name); r += 1

        if node.type == 'SUBBASIN':
            row(r, '유역면적 A (km²)',   'A',  node.params.get('A',  10.0));  r += 1
            row(r, '총강우량 PB (mm)',   'PB', node.params.get('PB', 200.0)); r += 1
            row(r, '유출곡선지수 CN',    'CN', node.params.get('CN', 80.0));  r += 1
            row(r, '도달시간 Tc (hr)',   'Tc', node.params.get('Tc', 1.0));   r += 1
            row(r, '저류상수 R (hr)',    'R',  node.params.get('R',  1.5));   r += 1
        elif node.type == 'REACH':
            row(r, '저류계수 K (hr)',        'K',    node.params.get('K',    1.0));  r += 1
            row(r, '가중계수 X (0~0.5)',     'X',    node.params.get('X',    0.20)); r += 1
            row(r, '세분할 NSTPS (0=자동)',  'NSTPS', node.params.get('NSTPS', 0));  r += 1
            ctk.CTkLabel(frame, text='※ 안정: 2Kx ≤ Δt ≤ 2K(1-x)\n   NSTPS=0 → 자동',
                         font=('맑은 고딕', 9), text_color='gray',
                         wraplength=208, justify='left').grid(
                row=r, column=0, columnspan=2, pady=(0,6)); r += 1
        elif node.type == 'JUNCTION':
            ctk.CTkLabel(frame, text='입력 수 N은 연결에서 자동 결정됩니다.',
                         font=FONT_SMALL, text_color='gray').grid(
                row=r, column=0, columnspan=2, pady=6)

        btn = ctk.CTkFrame(self, fg_color='transparent')
        btn.pack(fill='x', padx=12, pady=8)
        ctk.CTkButton(btn, text='확인', command=self._ok, font=FONT_BTN,
                      fg_color='#27ae60', hover_color='#2ecc71', width=120).pack(side='left', padx=4)
        ctk.CTkButton(btn, text='취소', command=self.destroy, font=FONT_BTN,
                      fg_color='#c0392b', hover_color='#e74c3c', width=120).pack(side='right', padx=4)

    def _ok(self):
        try:
            if 'name' in self._entries:
                nm = self._entries['name'].get().strip()
                if nm: self.node.name = nm
            for key, ent in self._entries.items():
                if key == 'name': continue
                v = ent.get().strip()
                if key == 'NSTPS':
                    self.node.params[key] = int(float(v))
                else:
                    self.node.params[key] = float(v)
            self.destroy()
        except ValueError as e:
            messagebox.showerror('입력 오류', str(e), parent=self)


# =============================================================================
# 팔레트 패널 (좌측)
# =============================================================================

class PalettePanel(ctk.CTkFrame):
    def __init__(self, master, on_type_select, **kwargs):
        super().__init__(master, width=280, **kwargs)
        self.pack_propagate(False)
        self._callback = on_type_select
        self._build()

    def _build(self):
        for ntype in ('SUBBASIN',):
            self._item(ntype, NODE_STYLES[ntype])
        self._item_reach_edge()
        for ntype in ('RESERVOIR', 'JUNCTION', 'OUTLET'):
            self._item(ntype, NODE_STYLES[ntype])

        # ── 활동 로그 ─────────────────────────────────────────────────────────
        ctk.CTkFrame(self, height=1, fg_color='#333344').pack(fill='x', padx=8, pady=6)
        ctk.CTkLabel(self, text='활동 로그', font=('맑은 고딕', 9),
                     text_color='#5dade2').pack(anchor='w', padx=10)
        log_outer = tk.Frame(self, bg='#0d0d1a')
        log_outer.pack(fill='both', expand=True, padx=6, pady=(2, 6))
        ysb = tk.Scrollbar(log_outer, orient='vertical')
        ysb.pack(side='right', fill='y')
        self._log_text = tk.Text(log_outer, wrap='word', bg='#0d0d1a', fg='#999999',
                                 font=('맑은 고딕', 8), relief='flat', width=1,
                                 yscrollcommand=ysb.set, state='disabled')
        self._log_text.pack(fill='both', expand=True, padx=2, pady=2)
        ysb.configure(command=self._log_text.yview)

    def log(self, msg):
        from datetime import datetime
        self._log_text.configure(state='normal')
        self._log_text.insert('end', f'[{datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self._log_text.see('end')
        self._log_text.configure(state='disabled')


    def _item(self, ntype, style):
        frame = ctk.CTkFrame(self, corner_radius=8, fg_color='#1a1a2e',
                             border_width=1, border_color='#333344')
        frame.pack(fill='x', padx=8, pady=4)

        mini = tk.Canvas(frame, width=148, height=48, bg='#1a1a2e', highlightthickness=0)
        mini.pack(pady=4)
        self._draw_mini(mini, ntype, style)

        for w in (frame, mini):
            w.bind('<Button-1>', lambda e, t=ntype: self._callback(t))
            w.configure(cursor='hand2')

    def _item_reach_edge(self):
        frame = ctk.CTkFrame(self, corner_radius=8, fg_color='#1a1a2e',
                             border_width=1, border_color='#333344')
        frame.pack(fill='x', padx=8, pady=4)

        mini = tk.Canvas(frame, width=148, height=48, bg='#1a1a2e', highlightthickness=0)
        mini.pack(pady=4)
        cy = 24
        mini.create_line(4, cy, 44, cy, fill='#2980b9', width=1)
        off = 6
        pts = [44+off, cy-12, 104+off, cy-12, 104-off, cy+12, 44-off, cy+12]
        mini.create_polygon(*pts, fill='#1a2d4a', outline='#2980b9', width=1)
        mini.create_line(104, cy, 140, cy, fill='#2980b9', width=1,
                         arrow=tk.LAST, arrowshape=(7, 9, 3))
        mini.create_text(74, cy, text='하도추적', fill='white',
                         font=('맑은 고딕', 9, 'bold'))

        hint = ctk.CTkLabel(frame, text='연결선 선택 후 클릭',
                            font=('맑은 고딕', 8), text_color='#7f8c8d')
        hint.pack(pady=(0, 4))

        for w in (frame, mini, hint):
            w.bind('<Button-1>', lambda e: self._callback('REACH_EDGE'))
            w.configure(cursor='hand2')

    def _draw_mini(self, c, ntype, style):
        cx, cy = 74, 24
        fill = style['fill']; out = style['outline']
        if ntype == 'SUBBASIN':
            self._round_rect_mini(c, cx-52, cy-17, cx+52, cy+17, 8, fill=fill, outline=out, width=1)
        elif ntype == 'RESERVOIR':
            pts = []
            for i in range(6):
                ang = math.pi / 2 + i * math.pi / 3
                pts.extend([cx + 26 * math.cos(ang), cy + 20 * math.sin(ang)])
            c.create_polygon(*pts, fill=fill, outline=out, width=1)
        elif ntype == 'JUNCTION':
            c.create_oval(cx-22, cy-22, cx+22, cy+22, fill=fill, outline=out, width=1)
        elif ntype == 'OUTLET':
            c.create_rectangle(cx-24, cy-17, cx+24, cy+17, fill=fill, outline=out, width=1)
        c.create_text(cx, cy, text=style['label'], fill='white',
                      font=('맑은 고딕', 9, 'bold'))

    def _round_rect_mini(self, c, x1, y1, x2, y2, r, **kw):
        pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r,
               x2,y2-r, x2,y2, x2-r,y2, x1+r,y2,
               x1,y2, x1,y2-r, x1,y1+r, x1,y1, x1+r,y1]
        c.create_polygon(*pts, smooth=True, **kw)


# =============================================================================
# 속성 패널 (우측)
# =============================================================================

class PropertiesPanel(tk.Frame):
    def __init__(self, master, redraw_cb=None, canvas_ref=None, on_log=None, **kwargs):
        kwargs.pop('corner_radius', None)   # tk.Frame 미지원 인자 제거
        super().__init__(master, bg='#1e1e2e', **kwargs)
        self._node       = None
        self._edge       = None
        self._entries    = {}
        self._svars      = {}
        self._redraw     = redraw_cb
        self._canvas_ref = canvas_ref
        self._on_log     = on_log

        # ── 헤더 (파일명 + 구분선 + 제목) ──
        hdr = tk.Frame(self, bg='#1a1a2e')
        hdr.pack(fill='x', side='top')
        self._filename_lbl = ctk.CTkLabel(
            hdr, text='(저장 파일 없음)',
            font=FONT_SMALL, text_color='gray',
            wraplength=210, justify='center')
        self._filename_lbl.pack(fill='x', padx=6, pady=(6, 2))
        tk.Frame(hdr, height=1, bg='#3a3a5e').pack(fill='x', padx=6, pady=2)
        ctk.CTkLabel(hdr, text='노드속성', font=FONT_HEADER,
                     text_color='#5dade2').pack(pady=(4, 6))

        # ── 분리창 (드래그 가능 구분선) ──
        # tk.PanedWindow 의 pane 자식은 반드시 tk.Frame 이어야 함
        self._pane = tk.PanedWindow(self, orient='vertical',
                                    bg='#2a2a4a', sashwidth=6, sashpad=1,
                                    sashrelief='raised', handlesize=0,
                                    showhandle=False)
        self._pane.pack(fill='both', expand=True)

        # 상단: 노드속성 스크롤 — tk.Frame 래퍼 안에 CTkScrollableFrame
        scroll_wrapper = tk.Frame(self._pane, bg='#1e1e2e')
        self._pane.add(scroll_wrapper, minsize=60, stretch='always')
        self._scroll = ctk.CTkScrollableFrame(scroll_wrapper, corner_radius=0)
        self._scroll.pack(fill='both', expand=True)

        # 하단: HEC-1 .dat 미리보기 — tk.Frame 래퍼
        dat_wrapper = tk.Frame(self._pane, bg='#0d1117')
        self._pane.add(dat_wrapper, minsize=60, stretch='always')

        dat_hdr = tk.Frame(dat_wrapper, bg='#1a1a2e')
        dat_hdr.pack(fill='x')
        self._dat_title_lbl = ctk.CTkLabel(dat_hdr, text='(파일 없음) .dat 미리보기',
                                           font=FONT_SMALL, text_color='#5dade2')
        self._dat_title_lbl.pack(side='left', padx=8, pady=3)
        self._dat_text = ctk.CTkTextbox(
            dat_wrapper, font=FONT_LOG,
            fg_color='#0d1117', text_color='#7ec8e3',
            corner_radius=0, state='disabled', wrap='none')
        self._dat_text.pack(fill='both', expand=True)

        # 초기 sash 위치: 전체 높이의 55%
        self.after(300, self._set_initial_sash)
        self._show_empty()

    def _set_initial_sash(self):
        h = self._pane.winfo_height()
        if h > 20:
            self._pane.sash_place(0, 0, max(60, int(h * 0.55)))

    def set_filename(self, path):
        name = os.path.basename(path) if path else '(저장 파일 없음)'
        self._filename_lbl.configure(text=name)

    def set_dat_title(self, title: str):
        self._dat_title_lbl.configure(text=title)

    def update_dat_preview(self, content: str):
        self._dat_text.configure(state='normal')
        self._dat_text.delete('1.0', 'end')
        self._dat_text.insert('1.0', content)
        self._dat_text.configure(state='disabled')

    def _show_empty(self):
        for w in self._scroll.winfo_children(): w.destroy()
        ctk.CTkLabel(self._scroll, text='노드를 선택하면\n속성이 여기에 표시됩니다.',
                     font=FONT_SMALL, text_color='gray',
                     justify='center').pack(pady=30)

    def show_node(self, node):
        self._node = node
        self._entries.clear()
        self._svars.clear()
        for w in self._scroll.winfo_children(): w.destroy()

        if node is None:
            self._show_empty()
            return

        style = NODE_STYLES.get(node.type, {})
        col   = style.get('outline', 'white')

        badge = ctk.CTkFrame(self._scroll, fg_color=style.get('fill', '#333'), corner_radius=6)
        badge.pack(fill='x', padx=10, pady=(8, 4))
        ctk.CTkLabel(badge, text=f'● {style.get("label", node.type)}',
                     font=FONT_HEADER, text_color=col).pack(anchor='w', padx=8, pady=6)

        self._add_field('이름', 'name', node.name, is_str=True)

        if node.type == 'SUBBASIN':
            self._section('─── 유역 매개변수 ───')
            self._add_field('유역면적 A (km²)',   'A',  node.params.get('A',  10.0))
            self._add_field('총강우량 PB (mm)',   'PB', node.params.get('PB', 200.0))
            self._add_field('유출곡선지수 CN',    'CN', node.params.get('CN', 80.0))
            self._add_field('도달시간 Tc (hr)',   'Tc', node.params.get('Tc', 1.0))
            self._add_field('저류상수 R (hr)',    'R',  node.params.get('R',  1.5))
            self._section('─── Huff 강우시간분포 ───')
            self._add_huff_field(node)

        elif node.type == 'RESERVOIR':
            self._section('─── 저수지 매개변수 ───')
            self._add_field('평균 수면적 A_avg (m²)', 'A_avg', node.params.get('A_avg', 10000.0))
            self._add_field('여수로 방류계수 Cd',      'Cd',    node.params.get('Cd',    0.42))
            self._add_field('여수로 길이 L (m)',       'L',     node.params.get('L',     10.0))
            self._add_field('여수로 마루고 Hc (m)',    'Hc',    node.params.get('Hc',    3.0))
            self._add_field('초기 저류량 S0 (m³)',     'S0',    node.params.get('S0',    0.0))
            ctk.CTkLabel(self._scroll,
                         text='※ Q = Cd×L×(H-Hc)^1.5 (광정위어)\n   S = A_avg × H',
                         font=FONT_SMALL, text_color='gray',
                         wraplength=210, justify='left').pack(anchor='w', padx=12, pady=2)

        elif node.type == 'REACH':
            self._section('─── Muskingum 매개변수 ───')
            self._add_field('저류계수 K (hr)', 'K', node.params.get('K', 1.0))
            self._add_slider('가중계수 X',     'X', node.params.get('X', 0.20), 0.0, 0.5)
            self._add_field('NSTPS (0=자동)',  'NSTPS', node.params.get('NSTPS', 0))
            ctk.CTkLabel(self._scroll, text='※ 안정: 2Kx ≤ Δt ≤ 2K(1-x)',
                         font=FONT_SMALL, text_color='gray',
                         wraplength=200).pack(anchor='w', padx=12, pady=2)

        elif node.type in ('JUNCTION', 'OUTLET'):
            n_in = 0
            if self._canvas_ref:
                n_in = sum(1 for e in self._canvas_ref.edges.values()
                           if e.dst == node.id)
            label_txt = ('합류점' if node.type == 'JUNCTION' else '출구') + ' 노드'
            ctk.CTkLabel(self._scroll, text=label_txt,
                         font=FONT_SMALL, text_color='gray').pack(pady=(8, 2))
            n_color = '#27ae60' if n_in >= 2 else ('#e74c3c' if n_in == 0 else '#f39c12')
            ctk.CTkLabel(self._scroll, text=f'현재 N = {n_in}  (연결 수 자동 계산)',
                         font=FONT_HEADER, text_color=n_color).pack(pady=(2, 8))

        ctk.CTkButton(self._scroll, text='적용', command=self._apply,
                      font=FONT_BTN, fg_color='#27ae60', hover_color='#2ecc71',
                      height=34).pack(fill='x', padx=10, pady=(14, 4))

    def _section(self, text):
        ctk.CTkLabel(self._scroll, text=text, font=FONT_SMALL,
                     text_color='gray').pack(fill='x', padx=12, pady=(8, 2))

    def _add_field(self, label, key, default, is_str=False):
        ctk.CTkLabel(self._scroll, text=label, font=FONT_SMALL,
                     anchor='w').pack(fill='x', padx=12, pady=(4, 0))
        ent = ctk.CTkEntry(self._scroll, font=FONT_SMALL,
                           justify='left' if is_str else 'right')
        ent.insert(0, str(default))
        ent.pack(fill='x', padx=10, pady=(0, 2))
        self._entries[key] = ent

    def _add_slider(self, label, key, default, from_, to):
        lbl = ctk.CTkLabel(self._scroll, text=f'{label}: {default:.2f}',
                           font=FONT_SMALL, anchor='w')
        lbl.pack(fill='x', padx=12, pady=(4, 0))
        var = ctk.DoubleVar(value=default)
        def on_change(v, _lbl=lbl, _label=label):
            _lbl.configure(text=f'{_label}: {float(v):.2f}')
        slider = ctk.CTkSlider(self._scroll, from_=from_, to=to, variable=var,
                               number_of_steps=50, command=on_change)
        slider.pack(fill='x', padx=10, pady=(0, 2))
        self._svars[key] = var

    def _add_huff_field(self, node):
        """SUBBASIN 전용: Huff 분위 콤보 + 커스텀 PC 입력 텍스트박스."""
        preset_key = node.params.get('huff_preset', '3분위')
        pc_raw = node.params.get('huff_pc', HUFF_PRESETS.get(preset_key, HUFF_PRESETS['3분위']))

        # 분위 콤보
        combo_row = tk.Frame(self._scroll, bg='#1e1e2e')
        combo_row.pack(fill='x', padx=10, pady=(4, 0))
        ctk.CTkLabel(combo_row, text='분위', font=FONT_SMALL,
                     width=50, anchor='w').pack(side='left')
        preset_var = ctk.StringVar(value=preset_key)
        self._svars['huff_preset'] = preset_var

        def _on_preset(choice):
            if choice in HUFF_PRESETS:
                vals = HUFF_PRESETS[choice]
                pc_box.configure(state='normal')
                pc_box.delete('1.0', 'end')
                pc_box.insert('1.0', '  '.join(f'{v:.3f}' for v in vals))
                pc_box.configure(state='normal')

        combo = ctk.CTkComboBox(combo_row, values=list(HUFF_PRESETS.keys()),
                                variable=preset_var, width=120,
                                font=FONT_SMALL, command=_on_preset)
        combo.pack(side='right')

        # PC 값 텍스트박스 (빈칸 또는 컴마 구분)
        ctk.CTkLabel(self._scroll, text='누적 백분율 (빈칸/컴마 구분)',
                     font=FONT_SMALL, text_color='gray',
                     anchor='w').pack(fill='x', padx=12, pady=(4, 0))
        pc_box = ctk.CTkTextbox(self._scroll, font=FONT_LOG,
                                height=52, corner_radius=4, wrap='word')
        pc_box.insert('1.0', '  '.join(f'{v:.3f}' for v in pc_raw))
        pc_box.pack(fill='x', padx=10, pady=(0, 4))
        self._entries['__huff_pc_box__'] = pc_box

    @staticmethod
    def _parse_huff_pc(text: str) -> list:
        """빈칸·탭·컴마 구분자로 PC 값 파싱 (HEC-1_FINAL.py 동일 로직)."""
        import re
        parts = re.split(r'[,\s\t]+', text.strip())
        result = []
        for p in parts:
            p = p.strip()
            if p:
                try:
                    result.append(float(p))
                except ValueError:
                    pass
        return result

    def _apply(self):
        if not self._node: return
        try:
            nm = self._entries.get('name')
            if nm:
                v = nm.get().strip()
                if v: self._node.name = v
            for key, ent in self._entries.items():
                if key in ('name', '__huff_pc_box__', '__label__'): continue
                v = ent.get().strip()
                self._node.params[key] = int(float(v)) if key == 'NSTPS' else float(v)
            for key, var in self._svars.items():
                if key == 'huff_preset':
                    self._node.params['huff_preset'] = var.get()
                else:
                    self._node.params[key] = round(float(var.get()), 4)
            # Huff PC 파싱 저장
            pc_box = self._entries.get('__huff_pc_box__')
            if pc_box and self._node.type == 'SUBBASIN':
                pc_list = self._parse_huff_pc(pc_box.get('1.0', 'end'))
                if pc_list:
                    self._node.params['huff_pc'] = pc_list
            if self._redraw: self._redraw()
            if self._on_log:
                lbl = NODE_STYLES.get(self._node.type, {}).get('label', self._node.type)
                self._on_log(f'{lbl} 속성 변경: {self._node.name}')
        except ValueError as e:
            messagebox.showerror('입력 오류', str(e))

    def show_edge(self, edge):
        self._node = None
        self._edge = edge if hasattr(edge, 'reach_params') else None
        self._entries.clear()
        self._svars.clear()
        for w in self._scroll.winfo_children(): w.destroy()

        if edge is None:
            self._show_empty()
            return

        if not getattr(edge, 'reach_params', None):
            # 일반 연결선 패널 (reach_params 없음)
            badge = ctk.CTkFrame(self._scroll, fg_color='#2a2a3e', corner_radius=6)
            badge.pack(fill='x', padx=10, pady=(8, 4))
            ctk.CTkLabel(badge, text='─ 연결선 ─',
                         font=FONT_HEADER, text_color='#3498db').pack(anchor='w', padx=8, pady=6)

            # 연결 정보 표시
            src_name, dst_name = '?', '?'
            if self._canvas_ref:
                sn = self._canvas_ref.nodes.get(edge.src)
                dn = self._canvas_ref.nodes.get(edge.dst)
                src_name = sn.name if sn else str(edge.src)
                dst_name = dn.name if dn else str(edge.dst)
            ctk.CTkLabel(self._scroll,
                         text=f'{src_name}  →  {dst_name}',
                         font=FONT_BODY, text_color='#aaaacc',
                         wraplength=220).pack(anchor='w', padx=12, pady=(6, 2))

            ctk.CTkLabel(self._scroll,
                         text='※ 하도추적 적용 시 팔레트에서\n   [하도추적] 버튼을 클릭하세요.',
                         font=FONT_SMALL, text_color='gray',
                         wraplength=220, justify='left').pack(anchor='w', padx=12, pady=(4, 8))

            ctk.CTkButton(self._scroll, text='연결선 삭제',
                          command=self._delete_plain_edge,
                          font=FONT_BTN, fg_color='#7f3b3b', hover_color='#c0392b',
                          height=34).pack(fill='x', padx=10, pady=(6, 4))
            return

        badge = ctk.CTkFrame(self._scroll, fg_color='#1a2d4a', corner_radius=6)
        badge.pack(fill='x', padx=10, pady=(8, 4))
        ctk.CTkLabel(badge, text='▶─[하도추적]─▶',
                     font=FONT_HEADER, text_color='#2980b9').pack(anchor='w', padx=8, pady=6)

        ctk.CTkLabel(self._scroll, text='구간 이름', font=FONT_SMALL,
                     anchor='w').pack(fill='x', padx=12, pady=(4, 0))
        ent_lbl = ctk.CTkEntry(self._scroll, font=FONT_SMALL, justify='left')
        ent_lbl.insert(0, edge.label or '')
        ent_lbl.pack(fill='x', padx=10, pady=(0, 2))
        self._entries['__label__'] = ent_lbl

        rp = edge.reach_params
        self._section('─── Muskingum 매개변수 ───')
        self._add_field('저류계수 K (hr)', 'K', rp.get('K', 1.0))
        self._add_slider('가중계수 X',     'X', rp.get('X', 0.20), 0.0, 0.5)
        self._add_field('NSTPS (0=자동)',  'NSTPS', rp.get('NSTPS', 0))
        ctk.CTkLabel(self._scroll, text='※ 안정: 2Kx ≤ Δt ≤ 2K(1-x)',
                     font=FONT_SMALL, text_color='gray',
                     wraplength=200).pack(anchor='w', padx=12, pady=2)

        ctk.CTkButton(self._scroll, text='적용', command=self._apply_edge,
                      font=FONT_BTN, fg_color='#27ae60', hover_color='#2ecc71',
                      height=34).pack(fill='x', padx=10, pady=(14, 4))

    def _apply_edge(self):
        edge = getattr(self, '_edge', None)
        if not edge or not edge.reach_params: return
        try:
            lbl_ent = self._entries.get('__label__')
            if lbl_ent:
                v = lbl_ent.get().strip()
                if v: edge.label = v
            for key, ent in self._entries.items():
                if key == '__label__': continue
                v = ent.get().strip()
                edge.reach_params[key] = int(float(v)) if key == 'NSTPS' else float(v)
            for key, var in self._svars.items():
                edge.reach_params[key] = round(float(var.get()), 4)
            if self._redraw: self._redraw()
            if self._on_log:
                self._on_log(f'하도추적 속성 변경: {edge.label or str(edge.id)[:8]}')
        except ValueError as e:
            messagebox.showerror('입력 오류', str(e))

    def _delete_plain_edge(self):
        """일반 연결선 삭제 (PropertiesPanel 버튼 → canvas에 위임)."""
        if not self._canvas_ref: return
        canvas = self._canvas_ref
        # _sel_edge가 설정된 경우 canvas._delete 직접 호출
        if canvas._sel_edge and not getattr(canvas._sel_edge, 'reach_params', None):
            canvas._push_undo()
            eid = canvas._sel_edge.id
            edge_label = str(eid)
            if eid in canvas.edges:
                del canvas.edges[eid]
            canvas._sel_edge = None
            if canvas._on_log: canvas._on_log(f'연결선 삭제: {edge_label}')
            canvas.redraw()
            self._show_empty()


# =============================================================================
# 수문망 편집기 창
# =============================================================================

class NetworkEditorWindow(ctk.CTkToplevel):

    def __init__(self, parent, on_apply,
                 dt_min=60, tr_min=1440, NQ=300, baseflow=0.0, huff_pc=None):
        super().__init__(parent)
        self.title('하천망 편집기 — 하도추적')
        self.geometry('1380x720')
        self.minsize(900, 500)
        self.after(50, lambda: self.state('zoomed'))
        self._on_apply   = on_apply
        self._dt_min     = dt_min
        self._tr_min     = tr_min
        self._NQ         = NQ
        self._baseflow   = baseflow
        self._huff_pc    = huff_pc or [0.0, 0.008, 0.041, 0.086, 0.154,
                                       0.263, 0.437, 0.636, 0.833, 0.953, 1.0]
        self._current_path   = None   # 현재 저장/로드된 .json 경로
        self._selected_node  = None   # N 실시간 표시용
        self._set_dark()
        # 창 수준 단축키 (캔버스 포커스 없이도 동작)
        self.bind('<Control-z>',     lambda e: self._canvas._undo())
        self.bind('<Control-Alt-z>', lambda e: self._canvas._redo())
        self.bind('<Delete>',        self._window_delete)
        self.bind('<BackSpace>',     self._window_delete)

        self.grid_columnconfigure(0, weight=0, minsize=280) # 팔레트
        self.grid_columnconfigure(1, weight=1)            # 캔버스 (확장)
        self.grid_columnconfigure(2, weight=0, minsize=280)  # 속성창 (고정 280px)
        self.grid_rowconfigure(0, weight=1)

        # Left palette (고정 열)
        self._palette = PalettePanel(self, on_type_select=self._palette_clicked,
                                     corner_radius=0)
        self._palette.grid(row=0, column=0, sticky='nsew', rowspan=2)

        # ── 캔버스 영역 (column=1) ──
        left_frame = tk.Frame(self, bg='#12121e')
        left_frame.grid(row=0, column=1, sticky='nsew', rowspan=2)
        left_frame.grid_rowconfigure(0, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)

        # 캔버스 + 스크롤바
        xsb = tk.Scrollbar(left_frame, orient='horizontal')
        ysb = tk.Scrollbar(left_frame, orient='vertical')
        xsb.grid(row=1, column=0, sticky='ew')
        ysb.grid(row=0, column=1, sticky='ns')

        self._canvas = NetworkCanvas(left_frame,
                                     on_select=self._node_selected,
                                     on_edge_select=self._edge_selected,
                                     on_log=self._log,
                                     xscrollcommand=xsb.set,
                                     yscrollcommand=ysb.set)
        self._canvas.grid(row=0, column=0, sticky='nsew')
        self._canvas.configure(scrollregion=(0, 0, 3000, 1200))
        xsb.configure(command=self._canvas.xview)
        ysb.configure(command=self._canvas.yview)

        # 툴바 (캔버스 하단)
        toolbar = ctk.CTkFrame(left_frame, height=42, corner_radius=0, fg_color='#1a1a2e')
        toolbar.grid(row=2, column=0, columnspan=2, sticky='ew')

        self._mode_lbl = ctk.CTkLabel(toolbar, text='모드: 선택',
                                      font=FONT_SMALL, text_color='#5dade2', width=100)
        self._mode_lbl.pack(side='left', padx=12)

        ctk.CTkLabel(toolbar,
                     text='포트→연결  |  드래그=이동  |  Del=삭제  |  Esc=취소',
                     font=('맑은 고딕', 9), text_color='gray').pack(side='left', padx=6)

        for txt, cmd, col, w in [
            ('예제 로드',         self._load_example,         '#5d6d7e', 90),
            ('초기화',            self._clear,                '#7f3b3b', 70),
            ('PNG로 저장',        self._save_png,             '#1a5276', 90),
            ('닫기',              self._close,                '#4a4a4a', 60),
            ('다른이름으로 저장', self._save_network,         '#1a5276', 120),
            ('저장하기',          self._save_network_current, '#1a5276', 80),
            ('불러오기',          self._load_network,         '#1a5276', 80),
            ('적용',              self._apply_network,        '#1a6b3a', 60),
            ('배열최적화',        self._optimize_layout,      '#2e4057', 90),
            ('업데이트',          self._apply,                '#1a6b3a', 80),
        ]:
            ctk.CTkButton(toolbar, text=txt, command=cmd,
                          font=FONT_SMALL, height=30, width=w,
                          fg_color=col).pack(side='right', padx=4, pady=6)

        # ── 속성창 (column=2, 고정 280px) ──
        self._props = PropertiesPanel(self,
                                      redraw_cb=self._make_redraw_cb(),
                                      canvas_ref=self._canvas,
                                      on_log=self._log,
                                      corner_radius=0)
        self._props.grid(row=0, column=2, sticky='nsew', rowspan=2)

        # 편집기 초기 로드: 캔버스가 비어있으면 Sample_Redraw.json 자동 표시
        _ex = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Sample_Redraw.json')
        if not self._canvas.nodes and os.path.exists(_ex):
            try:
                with open(_ex, 'r', encoding='utf-8') as _f:
                    _d = json.load(_f)
                self._canvas.load_canvas_state(_d)
                self._current_path = _ex
                self._props.set_filename(_ex)
                self._refresh_dat_preview()
            except Exception:
                pass

    def _set_dark(self):
        try:
            hwnd = windll.user32.GetParent(self.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
        except Exception: pass

    def _window_delete(self, event):
        """창 수준 Delete/BackSpace — Entry·Text에 포커스 없을 때만 캔버스에 위임."""
        focused = self.focus_get()
        if focused and focused.winfo_class() in ('Entry', 'Text'):
            return   # 위젯이 직접 처리하도록 위임
        self._canvas._delete(event)

    def _make_redraw_cb(self):
        """redraw 후 선택된 JUNCTION/OUTLET N값 갱신."""
        def cb():
            self._canvas.redraw()
            if self._selected_node and self._selected_node.id in self._canvas.nodes:
                self._props.show_node(self._selected_node)
        return cb

    def _palette_clicked(self, ntype):
        if ntype == 'REACH_EDGE':
            # 연결선이 선택된 상태면 즉시 하도추적으로 변환
            sel_edge = self._canvas._sel_edge
            if sel_edge is not None and sel_edge.id in self._canvas.edges:
                edge = self._canvas.edges[sel_edge.id]
                if edge.reach_params is None:
                    self._canvas._push_undo()
                    edge.reach_params = {'K': 1.0, 'X': 0.20, 'NSTPS': 0}
                    if not edge.label:
                        dst_node = self._canvas.nodes.get(edge.dst)
                        edge.label = f'{dst_node.name}R' if dst_node else f'RC{sum(1 for e in self._canvas.edges.values() if e.reach_params):02d}'
                    self._props.show_edge(edge)
                    self._canvas.redraw()
                    self._mode_lbl.configure(text='하도추적 변환 완료 ✓')
                    self._log(f'하도추적 변환: {edge.label}')
                    self.after(2000, lambda: self._mode_lbl.configure(text='모드: 선택'))
                else:
                    self._mode_lbl.configure(text='이미 하도추적 연결선입니다')
                    self.after(2000, lambda: self._mode_lbl.configure(text='모드: 선택'))
                return
            # 연결선 미선택 시 기존 방식 (포트 클릭 연결 모드)
            self._canvas._conn_reach = True
            self._canvas.set_mode('select')
            self._mode_lbl.configure(text='모드: 하도추적 연결 (포트 클릭 → 포트 클릭)')
            self._canvas.focus_set()
            return
        self._canvas.set_mode(f'place:{ntype}')
        label_map = {
            'SUBBASIN':  '소유역 배치 중 (캔버스 클릭)',
            'RESERVOIR': '저수지 배치 중 (캔버스 클릭)',
            'JUNCTION':  '합류점 배치 중 (캔버스 클릭)',
            'OUTLET':    '출구 배치 중 (캔버스 클릭)',
        }
        self._mode_lbl.configure(text=f'모드: {label_map.get(ntype, ntype)}')
        self._canvas.focus_set()


    def _log(self, msg):
        if hasattr(self._palette, 'log'):
            self._palette.log(msg)

    def _node_selected(self, node):
        self._selected_node = node
        self._props.show_node(node)
        self._mode_lbl.configure(text='모드: 선택')
        self._canvas.set_mode('select')

    def _edge_selected(self, edge):
        self._props.show_edge(edge)
        self._mode_lbl.configure(text='모드: 선택')
        self._canvas.set_mode('select')

    def _load_example(self):
        if self._canvas.nodes:
            if not messagebox.askyesno('확인', '기존 네트워크를 지우고 예제를 로드하시겠습니까?',
                                       parent=self):
                return
        example_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Sample_Redraw.json')
        if not os.path.exists(example_path):
            messagebox.showerror('오류', f'예제 파일을 찾을 수 없습니다:\n{example_path}', parent=self)
            return
        try:
            with open(example_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, dict) and data.get('format') == 'hydroset_canvas_v1':
                self._canvas.load_canvas_state(data)
            else:
                ops = data if isinstance(data, list) else []
                self._canvas.load_operations(ops)
            self._current_path = example_path
            self._props.set_filename(example_path)
            self._props.show_node(None)
            self._refresh_dat_preview()
        except Exception as e:
            messagebox.showerror('예제 로드 오류', str(e), parent=self)

    def _clear(self):
        if messagebox.askyesno('초기화', '네트워크를 모두 삭제 및 초기화하시겠습니까?', parent=self):
            self._canvas.clear()
            self._props.show_node(None)

    @staticmethod
    def _parse_dat_ops(path):
        """HEC-1 .dat 파일 → operations 리스트 파싱."""
        ops = []
        cur_name = None
        pending_route = None   # {'name','K','X','NSTPS'}
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            lines = f.readlines()
        for raw in lines:
            if len(raw) < 2: continue
            card = raw[:2].upper()
            body = raw[2:].strip()
            if card == 'KK':
                cur_name = body[:8].strip()
            elif card == 'BA':
                # BASIN — collect params across subsequent cards
                ops.append({'type': 'BASIN', 'name': cur_name or '',
                            'A': float(body.split()[0]) if body.split() else 0.0,
                            'PB': 0.0, 'CN': 0.0, 'Tc': 0.0, 'R': 0.0})
            elif card == 'PB':
                if ops and ops[-1]['type'] == 'BASIN':
                    ops[-1]['PB'] = float(body.split()[0]) if body.split() else 0.0
            elif card == 'LS':
                if ops and ops[-1]['type'] == 'BASIN':
                    parts = body.split()
                    if parts:
                        ops[-1]['CN'] = float(parts[0])
            elif card == 'UC':
                if ops and ops[-1]['type'] == 'BASIN':
                    parts = body.split()
                    if len(parts) >= 2:
                        ops[-1]['Tc'] = float(parts[0])
                        ops[-1]['R']  = float(parts[1])
            elif card == 'RM':
                parts = body.split()
                # RM  nstps  K  X  [nstps_auto]
                K = float(parts[1]) if len(parts) > 1 else 1.0
                X = float(parts[2]) if len(parts) > 2 else 0.2
                pending_route = {'type': 'ROUTE', 'name': cur_name or '',
                                 'K': K, 'X': X, 'NSTPS': 0}
                ops.append(pending_route)
                pending_route = None
            elif card == 'HC':
                N = int(body.split()[0]) if body.split() else 2
                ops.append({'type': 'COMBINE', 'name': cur_name or '', 'N': N})
        return ops

    def _load_network(self):
        if self._canvas.nodes:
            if not messagebox.askyesno('확인', '기존 네트워크를 지우고 불러오시겠습니까?', parent=self):
                return
        paths = filedialog.askopenfilenames(
            parent=self, title='네트워크 불러오기',
            filetypes=[('지원 파일', '*.json *.dat'),
                       ('JSON 파일', '*.json'),
                       ('HEC-1 DAT', '*.dat'),
                       ('모든 파일', '*.*')])
        if not paths: return

        json_path = next((p for p in paths if p.lower().endswith('.json')), None)
        dat_path  = next((p for p in paths if p.lower().endswith('.dat')),  None)

        try:
            if json_path:
                # JSON 있으면 JSON로 캔버스 상태 복원 (DAT는 미리보기만)
                with open(json_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                if isinstance(data, dict) and data.get('format') == 'hydroset_canvas_v1':
                    self._canvas.load_canvas_state(data)
                else:
                    ops = data if isinstance(data, list) else []
                    self._canvas.load_operations(ops)
                self._current_path = json_path
                self._props.set_filename(json_path)
                if dat_path:
                    # DAT도 있으면 미리보기에 표시
                    dat_name = os.path.basename(dat_path)
                    self._props.set_dat_title(f'{dat_name} 미리보기')
                    with open(dat_path, 'r', encoding='utf-8', errors='replace') as f:
                        self._props.update_dat_preview(f.read())
                else:
                    self._refresh_dat_preview()
            elif dat_path:
                # DAT만 선택 → 파싱 후 auto-layout
                ops = self._parse_dat_ops(dat_path)
                # 최종 스택 확인: 미합산 수문곡선 경고
                _sim = []
                for _op in ops:
                    if _op['type'] == 'BASIN':
                        _sim.append(_op['name'])
                    elif _op['type'] == 'RESERVOIR':
                        if _sim: _sim.pop()
                        _sim.append(_op['name'])
                    elif _op['type'] == 'COMBINE':
                        _n = min(int(_op.get('N', 2)), len(_sim))
                        for _ in range(_n): _sim.pop()
                        _sim.append(_op['name'])
                if len(_sim) > 1:
                    messagebox.showwarning(
                        'DAT 파싱 경고',
                        f'DAT 파일 종료 시 {len(_sim)}개 수문곡선이 미합산 상태입니다:\n'
                        f'  {", ".join(_sim)}\n\n'
                        f'최종 HC {len(_sim)} + 출구 카드가 누락된 것으로 보입니다.',
                        parent=self)
                self._canvas.load_operations(ops)
                self._canvas._auto_layout()
                self._canvas.redraw()
                self._canvas._zoom_extents()
                self._current_path = None
                self._props.set_filename(dat_path)
                dat_name = os.path.basename(dat_path)
                self._props.set_dat_title(f'{dat_name} 미리보기')
                with open(dat_path, 'r', encoding='utf-8', errors='replace') as f:
                    self._props.update_dat_preview(f.read())
            fname = os.path.basename(json_path or dat_path or '')
            self._log(f'불러오기: {fname}')
            self._props.show_node(None)
        except Exception as e:
            messagebox.showerror('불러오기 오류', str(e), parent=self)

    def _save_network(self):
        if not self._canvas.nodes:
            messagebox.showwarning('알림', '저장할 네트워크가 없습니다.', parent=self)
            return
        path = filedialog.asksaveasfilename(
            parent=self, title='네트워크 저장',
            defaultextension='.json',
            filetypes=[('JSON 파일', '*.json'), ('모든 파일', '*.*')])
        if not path: return
        self._do_save(path)

    def _save_network_current(self):
        if not self._canvas.nodes:
            messagebox.showwarning('알림', '저장할 네트워크가 없습니다.', parent=self)
            return
        if not self._current_path:
            self._save_network()
            return
        self._do_save(self._current_path)

    def _do_save(self, path):
        """JSON + DAT 두 파일 동시 저장."""
        state = {
            'format': 'hydroset_canvas_v1',
            'nodes': [
                {'id': n.id, 'type': n.type, 'name': n.name,
                 'x': n.x, 'y': n.y, 'params': n.params}
                for n in self._canvas.nodes.values()
            ],
            'edges': [
                {'id': e.id, 'src': e.src, 'dst': e.dst,
                 'src_dir': e.src_dir, 'dst_dir': e.dst_dir,
                 'reach_params': e.reach_params, 'label': e.label}
                for e in self._canvas.edges.values()
            ],
        }
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        # .dat 동시 저장
        ops, _ = self._canvas.build_operations()
        if ops:
            dat_content = self._build_dat_content(ops)
            dat_path = os.path.splitext(path)[0] + '.dat'
            with open(dat_path, 'w', encoding='utf-8') as f:
                f.write(dat_content)
            self._props.update_dat_preview(dat_content)
        self._current_path = path
        self._props.set_filename(path)
        messagebox.showinfo('저장 완료', f'저장되었습니다:\n{path}', parent=self)

    def _save_png(self):
        path = filedialog.asksaveasfilename(
            parent=self, title='PNG로 저장',
            defaultextension='.png',
            filetypes=[('PNG 파일', '*.png'), ('모든 파일', '*.*')])
        if not path: return
        try:
            from PIL import ImageGrab
            self.update_idletasks()
            c = self._canvas
            x = c.winfo_rootx(); y = c.winfo_rooty()
            w = c.winfo_width(); h = c.winfo_height()
            img = ImageGrab.grab(bbox=(x, y, x + w, y + h))
            img.save(path)
            messagebox.showinfo('저장 완료', f'PNG 저장:\n{path}', parent=self)
        except ImportError:
            messagebox.showerror('오류', 'Pillow(PIL) 라이브러리가 필요합니다.\npip install Pillow', parent=self)

    def _refresh_dat_preview(self):
        """현재 캔버스 상태로 .dat 미리보기를 즉시 갱신한다."""
        ops, _ = self._canvas.build_operations()
        if ops:
            self._props.update_dat_preview(self._build_dat_content(ops))
            if self._current_path:
                self._props.set_dat_title(f'{os.path.basename(self._current_path)} .dat 미리보기')
            else:
                self._props.set_dat_title('편집 중 .dat 미리보기')

    def _apply(self):
        ops, err = self._canvas.build_operations()
        if err:
            messagebox.showerror('오류', err, parent=self)
            return
        if not ops:
            messagebox.showwarning('알림', '네트워크가 비어 있습니다.', parent=self)
            return
        self._on_apply(ops)
        dat_content = self._build_dat_content(ops)
        self._props.update_dat_preview(dat_content)
        self._write_dat_file(ops)

    def _close(self):
        if self._canvas.nodes:
            ans = messagebox.askyesnocancel('닫기', '네트워크를 저장하시겠습니까?', parent=self)
            if ans is None:
                return
            if ans:
                self._save_network_current()
        self.destroy()

    def _apply_network(self):
        """요소 관계 반영 → JSON + .dat 저장 (무확인) + 미리보기 갱신."""
        if not self._canvas.nodes:
            messagebox.showwarning('알림', '저장할 네트워크가 없습니다.', parent=self)
            return
        if not self._current_path:
            self._save_network()   # 경로 지정 필요 시 파일 대화상자 호출
            return
        # JSON 저장
        state = {
            'format': 'hydroset_canvas_v1',
            'nodes': [
                {'id': n.id, 'type': n.type, 'name': n.name,
                 'x': n.x, 'y': n.y, 'params': n.params}
                for n in self._canvas.nodes.values()
            ],
            'edges': [
                {'id': e.id, 'src': e.src, 'dst': e.dst,
                 'src_dir': e.src_dir, 'dst_dir': e.dst_dir,
                 'reach_params': e.reach_params, 'label': e.label}
                for e in self._canvas.edges.values()
            ],
        }
        with open(self._current_path, 'w', encoding='utf-8') as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        # .dat 저장 + 미리보기 갱신
        ops, err = self._canvas.build_operations()
        if err:
            messagebox.showerror('오류', err, parent=self)
            return
        if ops:
            self._on_apply(ops)
            dat_content = self._build_dat_content(ops)
            dat_path = os.path.splitext(self._current_path)[0] + '.dat'
            with open(dat_path, 'w', encoding='utf-8') as f:
                f.write(dat_content)
            self._props.update_dat_preview(dat_content)
        self._props.set_filename(self._current_path)
        self._log(f'저장/적용: {os.path.basename(self._current_path)}')
        self._mode_lbl.configure(text='적용 완료 ✓')
        self.after(2000, lambda: self._mode_lbl.configure(text='모드: 선택'))

    def _optimize_layout(self):
        """HIERARCHY.txt 순서에 맞게 하천망 자동 배열."""
        if not self._canvas.nodes:
            messagebox.showwarning('알림', '배열할 네트워크가 없습니다.', parent=self)
            return
        self._canvas._auto_layout()
        self._canvas.redraw()
        # 배열 후 전체 보기 (zoom-extents)
        self._canvas._zoom_extents()
        self._log('배열최적화 완료')
        self._mode_lbl.configure(text='배열최적화 완료 ✓')
        self.after(2000, lambda: self._mode_lbl.configure(text='모드: 선택'))

    @staticmethod
    def _strip_lead0(s):
        """HEC-1 스타일: ' 0.' → '  .' (소수점 앞 불필요한 0 제거)"""
        return s.replace(' 0.', '  .')

    def _build_dat_content(self, ops):
        """ops 리스트 → HEC-1 .dat 형식 문자열 (8자 필드 기준)."""
        outlet_name = ''
        for op in reversed(ops):
            if op['type'] in ('COMBINE', 'BASIN'):
                outlet_name = op['name']
                break
        dt_min   = self._dt_min
        tr_min   = self._tr_min
        NQ       = self._NQ
        baseflow = self._baseflow
        huff_pc  = self._huff_pc
        sl       = self._strip_lead0

        n_step = int(tr_min / dt_min) + 1
        t_huff = np.linspace(0, 1, len(huff_pc))
        t_norm = np.linspace(0, 1, n_step)
        pc_vals = np.clip(_ensure_scipy()(t_huff, np.array(huff_pc, dtype=float))(t_norm),
                          0.0, 1.0)

        lns = [
            'ID Hydroset',
            '*DIAGRAM',
            'IM',
            f'IO{0:8d}{1:8d}',
            f'IT{int(dt_min):8d}{"01JAN00":>8s}{"0000":>8s}{int(NQ):8d}',
        ]
        if outlet_name:
            lns.append(f'VS  {outlet_name}')
        lns.append('VV' + sl(f'{baseflow:8.2f}'))
        lns.append('*')

        for op in ops:
            t = op['type']; name = op['name']
            if t == 'BASIN':
                lns.append(f'KK{name}')
                lns.append(f'IN{int(dt_min):8d}{"01JAN00":>8s}{"0000":>8s}')
                lns.append('BA' + sl(f'{op["A"]:8.1f}'))
                lns.append('PB' + sl(f'{op["PB"]:8.1f}'))
                fmt = [sl(f'{v:8.3f}') for v in pc_vals]
                for i in range(0, len(fmt), 10):
                    lns.append('PC' + ''.join(fmt[i:i+10]))
                lns.append('LS' + '        ' + sl(f'{op["CN"]:8.1f}'))
                lns.append('UC' + sl(f'{op["Tc"]:8.2f}') + sl(f'{op["R"]:8.2f}'))
                lns.append('*')
            elif t == 'RESERVOIR':
                A_avg = float(op.get('A_avg', 10000.0))
                Cd    = float(op.get('Cd',    0.42))
                L     = float(op.get('L',     10.0))
                Hc    = float(op.get('Hc',    3.0))
                S0    = float(op.get('S0',    0.0))
                lns.append(f'KK{name}')
                lns.append(f'* RESERVOIR  A_avg={A_avg:.1f}m2  Cd={Cd:.3f}  L={L:.2f}m  Hc={Hc:.2f}m  S0={S0:.1f}m3')
                lns.append(f'* Q=Cd*L*(H-Hc)^1.5 (광정위어)  S=A_avg*H')
                # S-A-E 표 (Modified Puls): H=Hc~Hc+5m 구간 10점
                lns.append('SA' + ''.join(f'{A_avg/1e6:8.3f}' for _ in range(10)))
                h_vals = [Hc + i * 0.5 for i in range(10)]
                q_vals = [Cd * L * max(h - Hc, 0) ** 1.5 for h in h_vals]
                s_vals = [S0 + A_avg * (h - Hc) for h in h_vals]
                lns.append('SE' + ''.join(f'{s/1e6:8.4f}' for s in s_vals))
                lns.append('SQ' + ''.join(f'{q:8.3f}' for q in q_vals))
                lns.append('*')
            elif t == 'ROUTE':
                nstps = max(1, int(op.get('NSTPS', 0) or 1))
                lns.append(f'KK{name}')
                lns.append('RM' + f'{nstps:8d}' + sl(f'{op["K"]:8.2f}') + sl(f'{op["X"]:8.2f}'))
                lns.append('*')
            elif t == 'COMBINE':
                lns.append(f'KK{name}')
                lns.append(f'HC{int(op["N"]):8d}')
                lns.append('*')
        lns.append('ZZ')
        return '\n'.join(lns)

    def _write_dat_file(self, ops):
        """현재 _current_path 기반으로 .dat 파일 저장 (경로 없으면 스킵)."""
        if not self._current_path: return
        dat_path = os.path.splitext(self._current_path)[0] + '.dat'
        content = self._build_dat_content(ops)
        with open(dat_path, 'w', encoding='utf-8') as f:
            f.write(content)

    def load_operations(self, ops):
        """Load existing ops into visual editor."""
        self._canvas.load_operations(ops)

    def open_json_file(self, path):
        """JSON 파일을 편집기에 직접 로드한다."""
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, dict) and data.get('format') == 'hydroset_canvas_v1':
                self._canvas.load_canvas_state(data)
            else:
                ops = data if isinstance(data, list) else []
                self._canvas.load_operations(ops)
            self._current_path = path
            self._props.set_filename(path)
            self._refresh_dat_preview()
        except Exception as e:
            self._log(f'파일 로드 오류: {e}')

    def get_node_count(self):
        return len(self._canvas.nodes), len(self._canvas.edges)


# =============================================================================
# 예제 조작 (100-1440-SW00.dat 기반)
# =============================================================================

EXAMPLE_OPERATIONS = [
    {'type':'BASIN',   'name':'MG24',   'A':79.9,  'PB':279.3, 'CN':89.8, 'Tc':1.29, 'R':1.83},
    {'type':'ROUTE',   'name':'MG23R',  'K':0.63,  'X':0.20,   'NSTPS':0},
    {'type':'BASIN',   'name':'XMG23',  'A':27.5,  'PB':287.3, 'CN':92.9, 'Tc':0.90, 'R':1.05},
    {'type':'COMBINE', 'name':'MG23',   'N':2},
    {'type':'ROUTE',   'name':'MG18R',  'K':0.34,  'X':0.20,   'NSTPS':0},
    {'type':'BASIN',   'name':'XMG18',  'A':33.5,  'PB':305.2, 'CN':91.7, 'Tc':1.10, 'R':1.27},
    {'type':'COMBINE', 'name':'MG18',   'N':2},
    {'type':'BASIN',   'name':'GSC04',  'A':68.6,  'PB':307.7, 'CN':80.8, 'Tc':1.54, 'R':2.54},
    {'type':'ROUTE',   'name':'GSC00R', 'K':0.65,  'X':0.20,   'NSTPS':0},
    {'type':'BASIN',   'name':'XGSC00', 'A':77.1,  'PB':310.9, 'CN':86.9, 'Tc':1.76, 'R':2.89},
    {'type':'COMBINE', 'name':'GSC00',  'N':2},
    {'type':'BASIN',   'name':'XMG17',  'A':0.1,   'PB':309.5, 'CN':10.0, 'Tc':0.10, 'R':0.18},
    {'type':'COMBINE', 'name':'MG17',   'N':3},
    {'type':'ROUTE',   'name':'MG15R',  'K':1.53,  'X':0.20,   'NSTPS':0},
    {'type':'BASIN',   'name':'XMG15',  'A':68.5,  'PB':302.4, 'CN':85.9, 'Tc':1.87, 'R':3.06},
    {'type':'COMBINE', 'name':'MG15',   'N':2},
    {'type':'BASIN',   'name':'SYC00',  'A':152.7, 'PB':287.7, 'CN':87.8, 'Tc':2.15, 'R':3.55},
    {'type':'BASIN',   'name':'XMG14',  'A':0.1,   'PB':309.5, 'CN':10.0, 'Tc':0.10, 'R':0.18},
    {'type':'COMBINE', 'name':'MG14',   'N':3},
    {'type':'ROUTE',   'name':'MG13R',  'K':0.48,  'X':0.20,   'NSTPS':0},
    {'type':'BASIN',   'name':'XMG13',  'A':13.9,  'PB':255.2, 'CN':91.6, 'Tc':0.59, 'R':1.01},
    {'type':'COMBINE', 'name':'MG13',   'N':2},
    {'type':'BASIN',   'name':'JJC10',  'A':31.7,  'PB':310.4, 'CN':85.0, 'Tc':0.99, 'R':1.64},
    {'type':'BASIN',   'name':'SW01',   'A':26.0,  'PB':289.8, 'CN':88.9, 'Tc':0.71, 'R':1.02},
    {'type':'ROUTE',   'name':'SW00R',  'K':0.13,  'X':0.20,   'NSTPS':0},
    {'type':'BASIN',   'name':'XSW00',  'A':2.9,   'PB':314.2, 'CN':87.7, 'Tc':0.19, 'R':0.27},
    {'type':'COMBINE', 'name':'SW00',   'N':2},
]


# =============================================================================
# 메인 GUI
# =============================================================================

class FloodRoutingApp(ctk.CTk):

    def __init__(self, project_path='', input_file=''):
        super().__init__()
        self.project_path = project_path.strip() or os.getcwd()
        self.project_name = os.path.basename(self.project_path)
        self.config_file  = os.path.join(self.project_path, 'project_config.json')

        self.processor  = HydroNetworkProcessor()
        self.operations = []
        self._editor    = None   # NetworkEditorWindow reference
        self._mpl_canvas = None
        self._net_json_name = None  # 편집기에서 적용된 JSON 파일명 (basename)
        self._net_json_path = None  # 편집기에서 적용된 JSON 전체 경로

        # matplotlib rcParams는 그래프 생성 시 지연 설정

        self.title(f'하도홍수추적 (6단계) — [{self.project_name}]')
        self.geometry('1280x800')
        self.minsize(900, 600)
        self._set_dark()
        self.tk.eval('proc bgerror {msg} {}')
        self.protocol('WM_DELETE_WINDOW', self._on_close)

        self._huff_var  = ctk.StringVar(value='3분위')
        self._pc_values = list(HUFF_PRESETS['3분위'])

        self._build_ui()
        self._load_config()
        pass  # init end

    def _set_dark(self):
        try:
            hwnd = windll.user32.GetParent(self.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
            try: windll.uxtheme.SetPreferredAppMode(2)
            except AttributeError: pass
        except Exception: pass

    # ── UI 구성 ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=0, minsize=280)  # 좌측 패널 고정 폭
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=11)  # 스크롤 영역
        self.grid_rowconfigure(1, weight=1)   # 실행 로그 영역
        self._build_left()
        self._build_right()

    def _build_left(self):
        scroll = ctk.CTkScrollableFrame(self, width=280, corner_radius=0)
        scroll.grid(row=0, column=0, sticky='nsew')
        scroll.grid_columnconfigure(0, weight=1)
        self._left_scroll = scroll

        def section(text):
            ctk.CTkLabel(scroll, text=text, font=FONT_HEADER, anchor='w',
                         text_color='#5dade2').pack(fill='x', padx=12, pady=(12, 3))

        def sep():
            ctk.CTkFrame(scroll, height=1, fg_color='#444444').pack(fill='x', padx=8, pady=5)

        def field(label, key, default):
            row = ctk.CTkFrame(scroll, fg_color='transparent')
            row.pack(fill='x', padx=8, pady=2)
            ctk.CTkLabel(row, text=label, font=FONT_SMALL, width=120, anchor='w').pack(side='left')
            ent = ctk.CTkEntry(row, font=FONT_SMALL, width=65, justify='right')
            ent.insert(0, default)
            ent.pack(side='right')
            self._entries[key] = ent

        self._entries = {}

        section('[ 계산 설정 ]')
        field('계산 시간간격 Δt (분)', 'DT_MIN', '60')
        field('강우 지속기간 TR (분)', 'TR_MIN', '1440')
        field('계산 스텝 수 NQ',      'NQ',     '300')
        sep()

        section('[ 하천망 ]')
        self._net_lbl = ctk.CTkLabel(scroll, text='파일로드 없음',
                                     font=FONT_SMALL, text_color='red', anchor='w')
        self._net_lbl.pack(fill='x', padx=12, pady=4)

        _net_row = tk.Frame(scroll, bg='#2b2b2b')
        _net_row.pack(fill='x', padx=8, pady=3)
        _net_row.grid_columnconfigure(0, weight=3)
        _net_row.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(_net_row, text='하천망편집기',
                      command=self._open_editor,
                      font=FONT_BTN, height=38,
                      fg_color='#2c3e50', hover_color='#3d5166',
                      ).grid(row=0, column=0, sticky='ew', padx=(0, 2))
        ctk.CTkButton(_net_row, text='하천망 검토',
                      command=self._open_review,
                      font=FONT_BTN, height=38,
                      fg_color='#6c3483', hover_color='#7d3c98',
                      ).grid(row=0, column=1, sticky='ew', padx=(2, 0))
        sep()

        section('[ 실행 ]')
        for label, cmd, color in [
            ('분석 실행  ▶', self._run,        '#27ae60'),
            ('결과 저장',    self._save_excel, '#2980b9'),
        ]:
            ctk.CTkButton(scroll, text=label, command=cmd,
                          font=FONT_BTN, height=38,
                          fg_color=color).pack(fill='x', padx=8, pady=3)
        sep()

        section('[ 최종 출구 결과 ]')
        self._result_labels = {}
        for key, label, color in [
            ('outlet',   '최종 유출점',      '#f0f0f0'),
            ('peak_q',   '첨두유량 (m³/s)',  '#e74c3c'),
            ('peak_hr',  '첨두 발생 (hr)',   '#f39c12'),
            ('cum_area', '총 유역면적 (km²)','#2ecc71'),
        ]:
            row = ctk.CTkFrame(scroll, fg_color='transparent')
            row.pack(fill='x', padx=8, pady=2)
            ctk.CTkLabel(row, text=label+':', font=FONT_SMALL,
                         width=150, anchor='w').pack(side='left')
            lbl = ctk.CTkLabel(row, text='─', font=FONT_HEADER,
                               text_color=color, anchor='w')
            lbl.pack(side='left', padx=4)
            self._result_labels[key] = lbl

        # ── 실행 로그 (스크롤 밖 row=1, 바닥까지 확장) ────────────────────────
        log_frame = ctk.CTkFrame(self, corner_radius=0, fg_color='#1a1a1a')
        log_frame.grid(row=1, column=0, sticky='nsew')
        log_frame.grid_columnconfigure(0, weight=1)
        log_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(log_frame, text='[ 실행 로그 ]', font=FONT_HEADER, anchor='w',
                     text_color='#5dade2', fg_color='transparent').grid(
                     row=0, column=0, sticky='w', padx=12, pady=(6, 2))
        self._txt_log = tk.Text(log_frame, width=1, height=6, font=FONT_LOG,
                                bg='#1a1a1a', fg='#cccccc',
                                insertbackground='white', wrap='word', relief='flat')
        self._txt_log.grid(row=1, column=0, sticky='nsew', padx=8, pady=(0, 8))

    def _build_right(self):
        right = ctk.CTkFrame(self, corner_radius=0, fg_color='transparent')
        right.grid(row=0, column=1, rowspan=2, sticky='nsew', padx=6, pady=6)
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(0, weight=3)
        right.grid_rowconfigure(1, weight=1)

        self._graph_frame = ctk.CTkFrame(right, corner_radius=6)
        self._graph_frame.grid(row=0, column=0, sticky='nsew', pady=(0, 4))
        ctk.CTkLabel(self._graph_frame,
                     text='분석 실행 후 수문곡선이 여기에 표시됩니다.',
                     font=FONT_BODY, text_color='gray').pack(expand=True)

        self._tbl_frame = ctk.CTkScrollableFrame(right, height=160, corner_radius=6)
        self._tbl_frame.grid(row=1, column=0, sticky='nsew')
        ctk.CTkLabel(self._tbl_frame,
                     text='분석 실행 후 RUNOFF SUMMARY가 표시됩니다.',
                     font=FONT_SMALL, text_color='gray').pack()

    # ── 편집기 ────────────────────────────────────────────────────────────────

    def _open_editor(self):
        if self._editor and self._editor.winfo_exists():
            self._editor.lift()
            self._editor.focus_force()
            return
        try:
            dt_min = float(self._entries['DT_MIN'].get())
            tr_min = float(self._entries['TR_MIN'].get())
            NQ     = int(float(self._entries['NQ'].get()))
        except Exception:
            dt_min, tr_min, NQ = 60.0, 1440.0, 300
        self._editor = NetworkEditorWindow(
            self, on_apply=self._on_network_applied,
            dt_min=dt_min, tr_min=tr_min, NQ=NQ,
            baseflow=0.0, huff_pc=list(self._pc_values))
        # 저장된 JSON 경로 or 예제 파일로 편집기 오픈
        json_path = self._net_json_path
        if not json_path:
            json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     'Sample_Redraw.json')
        if os.path.exists(json_path):
            self._editor.open_json_file(json_path)
        self._editor.after(100, lambda: (
            self._editor.lift(),
            self._editor.focus_force()
        ) if self._editor and self._editor.winfo_exists() else None)

    # ── 하천망 검토 (Ollama) ──────────────────────────────────────────────────

    def _list_ollama_models(self):
        """Ollama에 설치된 모델 목록 반환."""
        try:
            req = urllib.request.Request(
                f'{OLLAMA_BASE}/api/tags',
                headers={'Content-Type': 'application/json'})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read())
                return [m['name'] for m in data.get('models', [])]
        except Exception:
            return []

    def _open_review(self):
        """모델 선택 (Gemini API + Ollama 로컬) 후 하천망 표준지침 검토 실행."""
        if not self.operations:
            messagebox.showwarning('알림', '하천망 데이터가 없습니다.\n편집기에서 네트워크를 먼저 로드하세요.')
            return

        # ── 모델 목록 구성: 자동분석 + Gemini API + Ollama 로컬 ──
        auto_models = ['[자동분석] 표준지침 4.3.3']
        gemini_models = [f'[Gemini] {m}' for m in GEMINI_MODELS] if GEMINI_API_KEY else []
        ollama_models = [f'[Ollama] {m}' for m in self._list_ollama_models()]
        models = auto_models + gemini_models + ollama_models

        if not models:
            messagebox.showwarning(
                '모델 없음',
                'Gemini API 키가 없고, Ollama 서버에도 연결할 수 없습니다.\n\n'
                '• Gemini: GEMINI_API_KEY.json 파일을 확인하세요.\n'
                '• Ollama: ollama pull qwen3:8b 로 모델을 설치하세요.'
            )
            return

        # 기본값: 자동분석 (models[0])
        default_model = models[0]

        # 모델 선택 다이얼로그
        dlg = ctk.CTkToplevel(self)
        dlg.title('모델 선택 — 하천망 검토')
        dlg.geometry('460x230')
        dlg.resizable(False, False)
        dlg.grab_set()
        try:
            hwnd = windll.user32.GetParent(dlg.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
        except Exception:
            pass

        ctk.CTkLabel(dlg, text='LLM 모델 선택', font=FONT_HEADER).pack(pady=(16, 4))
        model_var = ctk.StringVar(value=default_model)
        ctk.CTkComboBox(dlg, values=models, variable=model_var, width=380).pack(pady=6)

        think_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(dlg, text='추론 과정 출력',
                        variable=think_var, font=FONT_SMALL,
                        checkbox_width=16, checkbox_height=16).pack(pady=(8, 2))

        selected = [None]

        def _ok():
            chosen = model_var.get()
            selected[0] = (chosen, think_var.get())
            try:
                cfg = {}
                try:
                    with open(self.config_file, encoding='utf-8') as f:
                        cfg = json.load(f)
                except Exception:
                    pass
                cfg['review_model'] = chosen
                with open(self.config_file, 'w', encoding='utf-8') as f:
                    json.dump(cfg, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            dlg.destroy()

        def _cancel():
            dlg.destroy()

        btn_row = ctk.CTkFrame(dlg, fg_color='transparent')
        btn_row.pack(pady=10)
        ctk.CTkButton(btn_row, text='검토 시작', command=_ok, width=120,
                      fg_color='#6c3483', hover_color='#7d3c98').pack(side='left', padx=6)
        ctk.CTkButton(btn_row, text='취소', command=_cancel, width=80,
                      fg_color='#555', hover_color='#666').pack(side='left', padx=6)

        dlg.wait_window()
        if selected[0] is None:
            return

        chosen_model, think_mode = selected[0]
        if chosen_model.startswith('[자동분석]'):
            self._run_guideline_checker()
        elif chosen_model.startswith('[Gemini] '):
            gemini_id = chosen_model.replace('[Gemini] ', '')
            self._run_gemini_review(gemini_id, think_mode)
        elif chosen_model.startswith('[Ollama] '):
            ollama_id = chosen_model.replace('[Ollama] ', '')
            self._run_ollama_review(ollama_id, think_mode)
        else:
            self._run_ollama_review(chosen_model, think_mode)

    def _run_gemini_review(self, model_id, think_mode=True):
        """Google Gemini API 스트리밍 요청 → 송신/추론/수신 전체 구분 출력."""
        win = ctk.CTkToplevel(self)
        win.title(f'하천망 검토 — {model_id}')
        win.geometry('980x800')
        try:
            hwnd = windll.user32.GetParent(win.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
        except Exception:
            pass

        top_bar = ctk.CTkFrame(win, fg_color='transparent')
        top_bar.pack(fill='x', padx=10, pady=(8, 2))
        status_lbl = ctk.CTkLabel(top_bar, text=f'  [Gemini {model_id}]  대기 중...',
                                   font=FONT_SMALL, text_color='#f39c12', anchor='w')
        status_lbl.pack(side='left', fill='x', expand=True)

        stop_event  = threading.Event()
        resp_holder = [None]

        def _copy():
            win.clipboard_clear()
            win.clipboard_append(txt.get('1.0', 'end'))

        def _stop():
            stop_event.set()
            if resp_holder[0] is not None:
                try: resp_holder[0].close()
                except Exception: pass
            status_lbl.configure(text=f'  [Gemini {model_id}]  중단됨', text_color='#e74c3c')
            stop_btn.configure(state='disabled', fg_color='#333')

        stop_btn = ctk.CTkButton(top_bar, text='■ 중단', command=_stop, width=70, height=26,
                                 font=FONT_SMALL, fg_color='#922b21', hover_color='#a93226')
        stop_btn.pack(side='right', padx=(0, 4))
        ctk.CTkButton(top_bar, text='전체 복사', command=_copy, width=80, height=26,
                      font=FONT_SMALL, fg_color='#444', hover_color='#555').pack(side='right')

        txt_frame = ctk.CTkFrame(win, fg_color='#0d0d0d')
        txt_frame.pack(fill='both', expand=True, padx=8, pady=(2, 8))
        txt = tk.Text(txt_frame, wrap='word', bg='#0d0d0d', fg='#e8e8e8',
                      font=('맑은 고딕', 10), relief='flat', padx=14, pady=10,
                      insertbackground='white', selectbackground='#444')
        sb = ctk.CTkScrollbar(txt_frame, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        txt.pack(fill='both', expand=True)

        # 색상 태그 정의
        txt.tag_config('hdr_send',   foreground='#3498db', font=('맑은 고딕', 10, 'bold'))
        txt.tag_config('prompt',     foreground='#85c1e9', font=('맑은 고딕',  9))
        txt.tag_config('hdr_think',  foreground='#f39c12', font=('맑은 고딕', 10, 'bold'))
        txt.tag_config('thinking',   foreground='#e59866', font=('맑은 고딕',  9, 'italic'))
        txt.tag_config('hdr_recv',   foreground='#2ecc71', font=('맑은 고딕', 10, 'bold'))
        txt.tag_config('response',   foreground='#e8e8e8', font=('맑은 고딕', 10))
        txt.tag_config('diagram',    foreground='#a8d8a8', font=('Consolas',   10))
        txt.tag_config('sep',        foreground='#444444')

        _DIAG_CHARS = set('╔╗╚╝║═╠╣╦╩╬─│┌┐└┘├┤┬┴┼▶◀▲▼⚠✓')
        _DIAG_STARTS = ('[', '(', '──', '  [', '  (', '  ╔', '  ╚', '  ║')

        def _is_diagram_line(line):
            if any(c in line for c in _DIAG_CHARS):
                return True
            stripped = line.strip()
            return any(stripped.startswith(s.strip()) for s in _DIAG_STARTS
                       if s.strip() in ('[', '(', '──'))

        def _ins(text, tag=''):
            if tag == 'response' and '\n' in text:
                lines = text.split('\n')
                for i, line in enumerate(lines):
                    seg = line + ('\n' if i < len(lines) - 1 else '')
                    t = 'diagram' if _is_diagram_line(line) else 'response'
                    txt.insert('end', seg, t)
            elif tag:
                txt.insert('end', text, tag)
            else:
                txt.insert('end', text)
            txt.see('end')

        def _sep(char='─', n=80):
            _ins(char * n + '\n', 'sep')

        base_prompt = self._build_review_prompt()

        if think_mode:
            think_instruction = (
                "Before responding, follow this format strictly:\n"
                "1. Open a <추론> tag and write your full reasoning process (any language is fine).\n"
                "2. Close with </추론>.\n"
                "3. Then write the final report **in Korean**.\n\n"
                "Example format:\n"
                "<추론>\n"
                "Total area is X km², which exceeds 250 km². Therefore ...\n"
                "</추론>\n\n"
                "## 1. 적합성 판정표\n...\n\n"
            )
            prompt = think_instruction + base_prompt
        else:
            prompt = base_prompt

        # ── 송신 프롬프트 출력 ──
        _ins('▶ 송신 프롬프트\n', 'hdr_send')
        _sep()
        _ins(prompt + '\n', 'prompt')
        _sep()
        _ins('\n')

        # 추론 태그 파싱용 (Ollama와 동일)
        THINK_TAG_PAIRS = [
            ('<추론>',    '</추론>'),
            ('<think>',   '</think>'),
            ('<thought>', '</thought>'),
            ('<thinking>','</thinking>'),
        ]
        MAX_OPEN_LEN = max(len(o) for o, _ in THINK_TAG_PAIRS)

        def _find_open_tag(text):
            best = None
            for o, c in THINK_TAG_PAIRS:
                i = text.find(o)
                if i != -1:
                    if best is None or i < best[0]:
                        best = (i, o, c)
            return best

        def _stream():
            try:
                win.after(0, lambda: status_lbl.configure(
                    text=f'  [Gemini {model_id}]  전송 중...', text_color='#f39c12'))

                # Gemini REST API — streamGenerateContent (SSE)
                api_url = (
                    f"https://generativelanguage.googleapis.com/v1beta/models/"
                    f"{model_id}:streamGenerateContent?alt=sse&key={GEMINI_API_KEY}"
                )
                body = json.dumps({
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {
                        "temperature": 0.7,
                        "maxOutputTokens": 16384,
                    }
                }).encode('utf-8')
                req = urllib.request.Request(
                    api_url, data=body,
                    headers={'Content-Type': 'application/json'})

                think_started = False
                resp_started  = False
                in_think      = False
                close_tag     = ''
                buf           = ''
                total_chars   = 0

                with urllib.request.urlopen(req, timeout=300) as resp:
                    resp_holder[0] = resp
                    for raw_line in resp:
                        if stop_event.is_set():
                            break
                        raw_line = raw_line.decode('utf-8', errors='replace').strip()
                        if not raw_line:
                            continue
                        # SSE 형식: "data: {...}"
                        if not raw_line.startswith('data: '):
                            continue
                        json_str = raw_line[6:]
                        try:
                            obj = json.loads(json_str)
                        except Exception:
                            continue

                        # Gemini 응답 구조: candidates[0].content.parts[0].text
                        candidates = obj.get('candidates', [])
                        if not candidates:
                            continue
                        parts = candidates[0].get('content', {}).get('parts', [])
                        chunk = ''
                        for part in parts:
                            chunk += part.get('text', '')
                        if not chunk:
                            continue

                        total_chars += len(chunk)

                        # ── 추론 태그 파싱 (Ollama 로직 재사용) ──
                        buf += chunk
                        while buf:
                            if in_think:
                                idx = buf.find(close_tag)
                                if idx == -1:
                                    safe = len(buf) - len(close_tag) + 1
                                    if safe <= 0:
                                        break
                                    emit = buf[:safe]
                                    buf  = buf[safe:]
                                    win.after(0, lambda c=emit: _ins(c, 'thinking'))
                                else:
                                    emit = buf[:idx]
                                    buf  = buf[idx + len(close_tag):]
                                    if emit:
                                        win.after(0, lambda c=emit: _ins(c, 'thinking'))
                                    in_think  = False
                                    close_tag = ''
                                    if not resp_started:
                                        resp_started = True
                                        win.after(0, lambda: (
                                            _ins('\n'),
                                            _ins('◆ 수신 응답\n', 'hdr_recv'),
                                            _sep()))
                            else:
                                found = _find_open_tag(buf)
                                if found is None:
                                    safe = len(buf) - MAX_OPEN_LEN + 1
                                    if safe <= 0:
                                        break
                                    emit = buf[:safe]
                                    buf  = buf[safe:]
                                    if not resp_started:
                                        resp_started = True
                                        win.after(0, lambda: (
                                            _ins('◆ 수신 응답\n', 'hdr_recv'),
                                            _sep()))
                                    win.after(0, lambda c=emit: _ins(c, 'response'))
                                else:
                                    idx, open_t, close_t = found
                                    pre = buf[:idx]
                                    buf = buf[idx + len(open_t):]
                                    if pre.strip():
                                        if not resp_started:
                                            resp_started = True
                                            win.after(0, lambda: (
                                                _ins('◆ 수신 응답\n', 'hdr_recv'),
                                                _sep()))
                                        win.after(0, lambda c=pre: _ins(c, 'response'))
                                    in_think  = True
                                    close_tag = close_t
                                    if not think_started:
                                        think_started = True
                                        label = f'◆ 추론 과정  ({open_t} 태그)\n'
                                        win.after(0, lambda lb=label: (
                                            _ins(lb, 'hdr_think'),
                                            _sep('·')))

                # ── 버퍼 잔여 flush ──
                if buf.strip():
                    tag = 'thinking' if in_think else 'response'
                    if not resp_started and tag == 'response':
                        resp_started = True
                        win.after(0, lambda: (
                            _ins('◆ 수신 응답\n', 'hdr_recv'),
                            _sep()))
                    win.after(0, lambda c=buf, t=tag: _ins(c, t))

                if not stop_event.is_set():
                    stats = f"총 {total_chars:,}자 수신"
                    win.after(0, lambda s=stats: (
                        _ins('\n'),
                        _sep(),
                        _ins(f'◆ 완료  {s}\n', 'hdr_recv'),
                        status_lbl.configure(
                            text=f'  [Gemini {model_id}]  완료', text_color='#27ae60'),
                        stop_btn.configure(state='disabled', fg_color='#333')))

            except urllib.error.HTTPError as e:
                try:
                    body = e.read().decode('utf-8', errors='replace')
                    detail = body[:500]
                except Exception:
                    detail = str(e)
                win.after(0, lambda d=detail: (
                    _ins(f'\n[HTTP {e.code}] {d}\n', 'hdr_think'),
                    status_lbl.configure(
                        text=f'  HTTP {e.code}', text_color='#e74c3c'),
                    stop_btn.configure(state='disabled', fg_color='#333')))

            except Exception as e:
                win.after(0, lambda: (
                    _ins(f'\n[오류] {e}\n', 'hdr_think'),
                    status_lbl.configure(
                        text=f'  오류: {e}', text_color='#e74c3c'),
                    stop_btn.configure(state='disabled', fg_color='#333')))

        win.after(0, lambda: status_lbl.configure(
            text=f'  [Gemini {model_id}]  송신 완료 — 응답 대기 중...', text_color='#f39c12'))
        threading.Thread(target=_stream, daemon=True).start()

    def _run_ollama_review(self, model, think_mode=True):
        """Ollama 스트리밍 요청 → 송신/추론/수신 전체 구분 출력."""
        win = ctk.CTkToplevel(self)
        win.title(f'하천망 검토 — {model}')
        win.geometry('980x800')
        try:
            hwnd = windll.user32.GetParent(win.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
        except Exception:
            pass

        top_bar = ctk.CTkFrame(win, fg_color='transparent')
        top_bar.pack(fill='x', padx=10, pady=(8, 2))
        status_lbl = ctk.CTkLabel(top_bar, text=f'  [{model}]  대기 중...',
                                   font=FONT_SMALL, text_color='#f39c12', anchor='w')
        status_lbl.pack(side='left', fill='x', expand=True)

        stop_event  = threading.Event()   # 중단 신호
        resp_holder = [None]              # HTTP 연결 참조 (강제 close용)

        def _copy():
            win.clipboard_clear()
            win.clipboard_append(txt.get('1.0', 'end'))

        def _stop():
            stop_event.set()
            if resp_holder[0] is not None:
                try: resp_holder[0].close()
                except Exception: pass
            status_lbl.configure(text=f'  [{model}]  중단됨', text_color='#e74c3c')
            stop_btn.configure(state='disabled', fg_color='#333')

        stop_btn = ctk.CTkButton(top_bar, text='■ 중단', command=_stop, width=70, height=26,
                                 font=FONT_SMALL, fg_color='#922b21', hover_color='#a93226')
        stop_btn.pack(side='right', padx=(0, 4))
        ctk.CTkButton(top_bar, text='전체 복사', command=_copy, width=80, height=26,
                      font=FONT_SMALL, fg_color='#444', hover_color='#555').pack(side='right')

        txt_frame = ctk.CTkFrame(win, fg_color='#0d0d0d')
        txt_frame.pack(fill='both', expand=True, padx=8, pady=(2, 8))
        txt = tk.Text(txt_frame, wrap='word', bg='#0d0d0d', fg='#e8e8e8',
                      font=('맑은 고딕', 10), relief='flat', padx=14, pady=10,
                      insertbackground='white', selectbackground='#444')
        sb = ctk.CTkScrollbar(txt_frame, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        txt.pack(fill='both', expand=True)

        # 색상 태그 정의
        txt.tag_config('hdr_send',   foreground='#3498db', font=('맑은 고딕', 10, 'bold'))
        txt.tag_config('prompt',     foreground='#85c1e9', font=('맑은 고딕',  9))
        txt.tag_config('hdr_think',  foreground='#f39c12', font=('맑은 고딕', 10, 'bold'))
        txt.tag_config('thinking',   foreground='#e59866', font=('맑은 고딕',  9, 'italic'))
        txt.tag_config('hdr_recv',   foreground='#2ecc71', font=('맑은 고딕', 10, 'bold'))
        txt.tag_config('response',   foreground='#e8e8e8', font=('맑은 고딕', 10))
        txt.tag_config('diagram',    foreground='#a8d8a8', font=('Consolas',   10))
        txt.tag_config('hdr_raw',    foreground='#888888', font=('맑은 고딕',  9, 'bold'))
        txt.tag_config('raw',        foreground='#666666', font=('Consolas',   8))
        txt.tag_config('sep',        foreground='#444444')

        # 다이어그램 라인 감지 문자
        _DIAG_CHARS = set('╔╗╚╝║═╠╣╦╩╬─│┌┐└┘├┤┬┴┼▶◀▲▼⚠✓')
        _DIAG_STARTS = ('[', '(', '──', '  [', '  (', '  ╔', '  ╚', '  ║')

        def _is_diagram_line(line):
            if any(c in line for c in _DIAG_CHARS):
                return True
            stripped = line.strip()
            return any(stripped.startswith(s.strip()) for s in _DIAG_STARTS
                       if s.strip() in ('[', '(', '──'))

        def _ins(text, tag=''):
            """태그가 'response'이면 라인별로 다이어그램 감지 후 자동 전환."""
            if tag == 'response' and '\n' in text:
                lines = text.split('\n')
                for i, line in enumerate(lines):
                    seg = line + ('\n' if i < len(lines) - 1 else '')
                    t = 'diagram' if _is_diagram_line(line) else 'response'
                    txt.insert('end', seg, t)
            elif tag:
                txt.insert('end', text, tag)
            else:
                txt.insert('end', text)
            txt.see('end')

        def _sep(char='─', n=80):
            _ins(char * n + '\n', 'sep')

        base_prompt = self._build_review_prompt()

        # think_mode: 프롬프트 앞에 추론 출력 지시 삽입
        if think_mode:
            think_instruction = (
                "Before responding, follow this format strictly:\n"
                "1. Open a <추론> tag and write your full reasoning process (any language is fine).\n"
                "2. Close with </추론>.\n"
                "3. Then write the final report **in Korean**.\n\n"
                "Example format:\n"
                "<추론>\n"
                "Total area is X km², which exceeds 250 km². Therefore ...\n"
                "</추론>\n\n"
                "## 1. 적합성 판정표\n...\n\n"
            )
            prompt = think_instruction + base_prompt
        else:
            prompt = base_prompt

        # ── 송신 프롬프트 출력 ──────────────────────────────────────────────
        _ins('▶ 송신 프롬프트\n', 'hdr_send')
        _sep()
        _ins(prompt + '\n', 'prompt')
        _sep()
        _ins('\n')

        # 모델이 사용할 수 있는 추론 태그 쌍 (우선순위 순)
        THINK_TAG_PAIRS = [
            ('<추론>',    '</추론>'),
            ('<think>',   '</think>'),
            ('<thought>', '</thought>'),
            ('<thinking>','</thinking>'),
        ]
        # 버퍼 보류 길이: 가장 긴 오픈태그 길이
        MAX_OPEN_LEN = max(len(o) for o, _ in THINK_TAG_PAIRS)

        def _find_open_tag(text):
            """text 안에서 가장 먼저 나오는 오픈태그 반환. (idx, open, close) or None"""
            best = None
            for o, c in THINK_TAG_PAIRS:
                i = text.find(o)
                if i != -1:
                    if best is None or i < best[0]:
                        best = (i, o, c)
            return best

        def _stream():
            try:
                win.after(0, lambda: status_lbl.configure(
                    text=f'  [{model}]  전송 중...', text_color='#f39c12'))

                def _make_req(use_ctx=True):
                    body = {'model': model, 'prompt': prompt, 'stream': True,
                            'keep_alive': 0}
                    if use_ctx:
                        body['options'] = {'num_ctx': 16384}
                    return urllib.request.Request(
                        f'{OLLAMA_BASE}/api/generate',
                        data=json.dumps(body).encode('utf-8'),
                        headers={'Content-Type': 'application/json'})

                req = _make_req(use_ctx=True)

                think_started = False
                resp_started  = False
                in_think      = False
                close_tag     = ''   # 현재 열린 추론태그에 대응하는 닫힘태그
                buf           = ''

                with urllib.request.urlopen(req, timeout=300) as resp:
                    resp_holder[0] = resp
                    for line in resp:
                        if stop_event.is_set():
                            break
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            obj = json.loads(line)
                        except Exception:
                            win.after(0, lambda l=line: _ins(f'[파싱오류] {l}\n', 'raw'))
                            continue

                        # ── 네이티브 thinking 필드 (DeepSeek-R1 등) ────────
                        native_think = obj.get('thinking', '')
                        if native_think:
                            if not think_started:
                                think_started = True
                                win.after(0, lambda: (
                                    _ins('◆ 추론 과정  (thinking field)\n', 'hdr_think'),
                                    _sep('·')))
                            win.after(0, lambda c=native_think: _ins(c, 'thinking'))

                        # ── 응답 청크: 다중 추론 태그 파싱 ───────────────
                        chunk = obj.get('response', '')
                        if chunk:
                            buf += chunk
                            while buf:
                                if in_think:
                                    idx = buf.find(close_tag)
                                    if idx == -1:
                                        safe = len(buf) - len(close_tag) + 1
                                        if safe <= 0:
                                            break
                                        emit = buf[:safe]
                                        buf  = buf[safe:]
                                        win.after(0, lambda c=emit: _ins(c, 'thinking'))
                                    else:
                                        emit = buf[:idx]
                                        buf  = buf[idx + len(close_tag):]
                                        if emit:
                                            win.after(0, lambda c=emit: _ins(c, 'thinking'))
                                        in_think  = False
                                        close_tag = ''
                                        if not resp_started:
                                            resp_started = True
                                            win.after(0, lambda: (
                                                _ins('\n'),
                                                _ins('◆ 수신 응답\n', 'hdr_recv'),
                                                _sep()))
                                else:
                                    found = _find_open_tag(buf)
                                    if found is None:
                                        safe = len(buf) - MAX_OPEN_LEN + 1
                                        if safe <= 0:
                                            break
                                        emit = buf[:safe]
                                        buf  = buf[safe:]
                                        if not resp_started:
                                            resp_started = True
                                            win.after(0, lambda: (
                                                _ins('◆ 수신 응답\n', 'hdr_recv'),
                                                _sep()))
                                        win.after(0, lambda c=emit: _ins(c, 'response'))
                                    else:
                                        idx, open_t, close_t = found
                                        pre = buf[:idx]
                                        buf = buf[idx + len(open_t):]
                                        if pre.strip():
                                            if not resp_started:
                                                resp_started = True
                                                win.after(0, lambda: (
                                                    _ins('◆ 수신 응답\n', 'hdr_recv'),
                                                    _sep()))
                                            win.after(0, lambda c=pre: _ins(c, 'response'))
                                        in_think  = True
                                        close_tag = close_t
                                        if not think_started:
                                            think_started = True
                                            label = f'◆ 추론 과정  ({open_t} 태그)\n'
                                            win.after(0, lambda lb=label: (
                                                _ins(lb, 'hdr_think'),
                                                _sep('·')))

                        # ── 완료 ──────────────────────────────────────────
                        if obj.get('done', False):
                            if buf.strip():
                                tag = 'thinking' if in_think else 'response'
                                win.after(0, lambda c=buf, t=tag: _ins(c, t))
                            stats = (
                                f"prompt_tokens={obj.get('prompt_eval_count','?')}  "
                                f"response_tokens={obj.get('eval_count','?')}  "
                                f"total_duration={obj.get('total_duration',0)//1_000_000}ms"
                            )
                            win.after(0, lambda s=stats: (
                                _ins('\n'),
                                _sep(),
                                _ins(f'◆ 완료  {s}\n', 'hdr_recv'),
                                status_lbl.configure(
                                    text=f'  [{model}]  완료', text_color='#27ae60'),
                                stop_btn.configure(state='disabled', fg_color='#333')))

            except urllib.error.HTTPError as e:
                # Ollama 에러 본문 파싱
                try:
                    body = e.read().decode('utf-8', errors='replace')
                    obj  = json.loads(body)
                    detail = obj.get('error', body)
                except Exception:
                    detail = str(e)

                # 모델 로드 실패: 재시도해도 소용없음
                _load_fail_keywords = ('failed to load', 'resource limitation', 'cannot load')
                is_load_fail = any(kw in detail.lower() for kw in _load_fail_keywords)

                if e.code == 500 and is_load_fail:
                    win.after(0, lambda d=detail: (
                        _ins(f'\n[HTTP 500 — 모델 로드 실패] {d}\n'
                             f'→ 재시도 불가. 다른 모델을 선택하세요.\n', 'hdr_think'),
                        status_lbl.configure(
                            text=f'  [{model}]  로드 실패 — 다른 모델 사용',
                            text_color='#e74c3c'),
                        stop_btn.configure(state='disabled', fg_color='#333')))
                elif e.code == 500:
                    win.after(0, lambda d=detail: _ins(
                        f'\n[HTTP 500] {d}\n→ num_ctx 옵션 없이 재시도...\n', 'hdr_think'))
                    try:
                        req2 = _make_req(use_ctx=False)
                        think_started = resp_started = in_think = False
                        close_tag = buf = ''
                        with urllib.request.urlopen(req2, timeout=300) as resp2:
                            resp_holder[0] = resp2
                            for line in resp2:
                                if stop_event.is_set():
                                    break
                                line = line.strip()
                                if not line:
                                    continue
                                try:
                                    obj2 = json.loads(line)
                                    chunk = obj2.get('response', '')
                                    if chunk:
                                        if not resp_started:
                                            resp_started = True
                                            win.after(0, lambda: (
                                                _ins('◆ 수신 응답 (재시도)\n', 'hdr_recv'),
                                                _sep()))
                                        win.after(0, lambda c=chunk: _ins(c, 'response'))
                                    if obj2.get('done', False):
                                        stats2 = (
                                            f"prompt_tokens={obj2.get('prompt_eval_count','?')}  "
                                            f"response_tokens={obj2.get('eval_count','?')}  "
                                            f"total_duration={obj2.get('total_duration',0)//1_000_000}ms"
                                        )
                                        win.after(0, lambda s=stats2: (
                                            _ins('\n'), _sep(),
                                            _ins(f'◆ 완료 (재시도)  {s}\n', 'hdr_recv'),
                                            status_lbl.configure(
                                                text=f'  [{model}]  완료', text_color='#27ae60'),
                                            stop_btn.configure(state='disabled', fg_color='#333')))
                                except Exception:
                                    pass
                    except Exception as e2:
                        win.after(0, lambda err=e2: (
                            _ins(f'\n[재시도 오류] {err}\n', 'hdr_think'),
                            status_lbl.configure(text=f'  재시도 오류: {err}', text_color='#e74c3c')))
                else:
                    win.after(0, lambda d=detail: (
                        _ins(f'\n[HTTP {e.code}] {d}\n', 'hdr_think'),
                        status_lbl.configure(text=f'  HTTP {e.code}: {d[:60]}', text_color='#e74c3c')))

            except Exception as e:
                win.after(0, lambda: (
                    _ins(f'\n[오류] {e}\n', 'hdr_think'),
                    status_lbl.configure(
                        text=f'  오류: {e}', text_color='#e74c3c')))

        win.after(0, lambda: status_lbl.configure(
            text=f'  [{model}]  송신 완료 — 응답 대기 중...', text_color='#f39c12'))
        threading.Thread(target=_stream, daemon=True).start()

    # ── 표준지침 4.3.3 자동분석 (규칙 기반) ─────────────────────────────────
    def _run_guideline_checker(self):
        """네트워크 JSON에서 DAG를 구성, 표준지침 4.3.3조 적합성을 코드로 판정."""
        # ── 1. 네트워크 JSON 로드 ──
        net_path = getattr(self, '_net_json_path', None)
        if not net_path or not os.path.isfile(net_path):
            messagebox.showwarning('알림', '네트워크 JSON 파일이 없습니다.')
            return
        with open(net_path, encoding='utf-8') as f:
            net = json.load(f)

        nodes_list = net.get('nodes', [])
        edges_list = net.get('edges', [])

        # ── 2. DAG 자료구조 구축 ──
        nodes = {}  # id -> dict
        for n in nodes_list:
            nodes[n['id']] = n
        children = {}   # id -> [(child_id, edge)]
        parents  = {}   # id -> [(parent_id, edge)]
        for nid in nodes:
            children[nid] = []
            parents[nid]  = []
        for e in edges_list:
            children[e['src']].append((e['dst'], e))
            parents[e['dst']].append((e['src'], e))

        # ── 3. 소유역/합류점/출구 분류 ──
        subbasins = {nid: n for nid, n in nodes.items() if n['type'] == 'SUBBASIN'}
        junctions = {nid: n for nid, n in nodes.items() if n['type'] == 'JUNCTION'}
        outlets   = {nid: n for nid, n in nodes.items() if n['type'] == 'OUTLET'}
        reservoirs= {nid: n for nid, n in nodes.items() if n['type'] == 'RESERVOIR'}

        # ── 4. 누적면적 계산 (상류→하류 BFS) ──
        cum_area = {}
        def calc_cum(nid):
            if nid in cum_area:
                return cum_area[nid]
            n = nodes[nid]
            if n['type'] == 'SUBBASIN':
                a = n.get('params', {}).get('A', 0)
                cum_area[nid] = a
                return a
            total = 0
            for pid, _ in parents[nid]:
                total += calc_cum(pid)
            cum_area[nid] = total
            return total

        outlet_id = list(outlets.keys())[0] if outlets else None
        if outlet_id is None:
            messagebox.showwarning('알림', '출구(OUTLET) 노드가 없습니다.')
            return
        total_area = calc_cum(outlet_id)

        # ── 5. 하도추적 엣지 목록 ──
        reach_edges = [e for e in edges_list if e.get('reach_params')]
        reach_labels = [e.get('label', f"E{e['id']}") for e in reach_edges]

        # ── 6. 순차 추적 패턴 감지 ──
        # 본류 = 출구에서 가장 먼(깊은) 상류까지의 경로
        # 깊이(depth) 계산: 각 노드에서 최상류까지의 홉 수
        depth = {}
        def calc_depth(nid):
            if nid in depth:
                return depth[nid]
            plist = parents.get(nid, [])
            if not plist:
                depth[nid] = 0
                return 0
            d = max(calc_depth(pid) for pid, _ in plist) + 1
            depth[nid] = d
            return d
        calc_depth(outlet_id)

        main_path_edges = []  # 본류 경로의 reach 엣지
        def trace_main_stem(nid):
            """출구에서 상류로 본류 추적 (최장경로=depth 최대)"""
            cur = nid
            while True:
                plist = parents.get(cur, [])
                if not plist:
                    break
                # 깊이 최대인 상류 선택 (동점 시 면적 최대)
                best_pid, best_edge = max(
                    plist, key=lambda pe: (depth.get(pe[0], 0), cum_area.get(pe[0], 0)))
                if best_edge.get('reach_params'):
                    main_path_edges.append(best_edge)
                cur = best_pid
        trace_main_stem(outlet_id)

        sequential_count = len(main_path_edges)
        has_sequential = sequential_count >= 2

        # ── 7. 적합성 판정 ──
        checks = []

        # 7-1. 총면적 250km² 초과 여부
        area_over = total_area > 250
        checks.append({
            'item': '유역면적 250 km² 기준',
            'result': '해당' if area_over else '미해당',
            'detail': f'총면적 {total_area:.1f} km² → {"250 km² 초과, 중규모 유역 분할 필요" if area_over else "250 km² 이하, 분할 불필요"}'
        })

        # 7-2. 개별 소유역 면적 체크
        large_basins = [(n['name'], n['params'].get('A', 0)) for n in subbasins.values()
                        if n.get('params', {}).get('A', 0) > 250]
        checks.append({
            'item': '개별 소유역 면적 상한',
            'result': '부적합' if large_basins else '적합',
            'detail': (f'면적 초과 소유역: {", ".join(f"{nm}({a:.1f}km²)" for nm,a in large_basins)}'
                       if large_basins else '모든 소유역 250 km² 이하')
        })

        # 7-3. 하도추적 방식 (순차 추적 여부)
        if area_over and has_sequential:
            route_verdict = '부적합'
            route_detail = (f'본류 경로에 {sequential_count}개 하도추적 구간이 순차 배치됨 → '
                            '표준지침 위반 (구간별 순차 추적 금지)')
        elif area_over and not has_sequential:
            route_verdict = '적합'
            route_detail = '순차 추적 패턴 미발견'
        else:
            route_verdict = '해당없음'
            route_detail = '250 km² 이하 유역 — 순차 추적 허용'
        checks.append({
            'item': '하도추적 방식 (순차추적 금지)',
            'result': route_verdict,
            'detail': route_detail
        })

        # 7-4. 자체홍수량 산정 방식
        # 현재 구성에서 소유역 개별 합산 여부 체크
        if area_over:
            # 개별 BASIN → COMBINE 패턴이 있으면 문제
            basin_count = len(subbasins)
            checks.append({
                'item': '자체홍수량 산정 (소유역 개별합산 금지)',
                'result': '부적합' if basin_count > 3 else '주의',
                'detail': (f'{basin_count}개 소유역이 개별 UH로 산정 후 합산됨 → '
                           '중규모 유역 단위 통합 UH 적용 필요')
            })
        else:
            checks.append({
                'item': '자체홍수량 산정',
                'result': '해당없음',
                'detail': '250 km² 이하 유역 — 현행 방식 허용'
            })

        # 7-5. 댐/저수지 처리
        has_dam = len(reservoirs) > 0
        checks.append({
            'item': '댐/저수지 처리',
            'result': '적합' if has_dam else '해당없음',
            'detail': (f'저수지 {len(reservoirs)}개 존재 → 저수지 홍수추적 적용'
                       if has_dam else '댐/저수지 없음')
        })

        # ── 8. 중규모 유역 자동 분할 ──
        medium_watersheds = []
        if area_over:
            # 본류 경로의 합류점 (JUNCTION)에서 누적면적 ≤ 250 km² 기준으로 컷
            # 전략: 출구에서 역추적, 누적면적이 250을 넘는 첫 지점에서 분할
            # → 상류에서 하류로 순서 정렬

            # 본류 합류점 순서 (상류→하류) — depth 기준 본류 추적
            main_junctions = []
            cur = outlet_id
            while True:
                plist = parents.get(cur, [])
                if not plist:
                    break
                best_pid, _ = max(
                    plist, key=lambda pe: (depth.get(pe[0], 0), cum_area.get(pe[0], 0)))
                if nodes[cur]['type'] in ('JUNCTION', 'OUTLET'):
                    main_junctions.append(cur)
                cur = best_pid
            main_junctions.reverse()  # 상류→하류

            # 각 합류점에 유입되는 소유역 목록 (DFS)
            def collect_subbasins(nid, stop_set):
                """nid 상류의 소유역 수집, stop_set에 있는 노드는 진입 안함"""
                result = []
                stack = [nid]
                visited = set()
                while stack:
                    c = stack.pop()
                    if c in visited:
                        continue
                    visited.add(c)
                    n = nodes.get(c)
                    if not n:
                        continue
                    if n['type'] == 'SUBBASIN':
                        result.append(n)
                        continue
                    for pid, _ in parents.get(c, []):
                        if pid not in stop_set:
                            stack.append(pid)
                return result

            # 분할 로직: 상류→하류 순회하며 누적면적 250 초과 시 직전 합류점에서 컷
            cut_points = set()
            mw_idx = 0
            base_area = 0  # 마지막 분할점의 누적면적

            last_valid = None  # 250 이내인 마지막 합류점
            for jid in main_junctions:
                ca = cum_area.get(jid, 0)
                incremental = ca - base_area
                if incremental > 250:
                    if last_valid is not None:
                        cut_points.add(last_valid)
                        base_area = cum_area.get(last_valid, 0)
                        last_valid = None
                        # 현재 jid 재평가
                        if ca - base_area <= 250:
                            last_valid = jid
                    else:
                        # 첫 합류점부터 250 초과 — 여기서 강제 분할
                        cut_points.add(jid)
                        base_area = ca
                else:
                    last_valid = jid

            # 분할점이 없으면 중간 합류점에서 가장 균등하게 분할
            if not cut_points and total_area > 250:
                best_j = None
                best_diff = float('inf')
                for jid in main_junctions:
                    if nodes[jid]['type'] != 'JUNCTION':
                        continue
                    ca = cum_area.get(jid, 0)
                    rest = total_area - ca
                    if ca <= 250 and rest <= 250:
                        diff = abs(ca - rest)
                        if diff < best_diff:
                            best_diff = diff
                            best_j = jid
                if best_j is None:
                    # 250 이하 조건 완화 — 가장 균등하게
                    for jid in main_junctions:
                        if nodes[jid]['type'] != 'JUNCTION':
                            continue
                        ca = cum_area.get(jid, 0)
                        rest = total_area - ca
                        diff = abs(ca - rest)
                        if diff < best_diff:
                            best_diff = diff
                            best_j = jid
                if best_j:
                    cut_points.add(best_j)

            # 분할점 순서 정렬 (상류→하류)
            cut_list = sorted(cut_points, key=lambda jid: cum_area.get(jid, 0))

            # 각 중규모 유역 구성
            boundaries = [None] + cut_list + [outlet_id]
            for i in range(len(boundaries) - 1):
                upstream_boundary = boundaries[i]
                downstream_boundary = boundaries[i + 1]

                stop = set()
                if upstream_boundary is not None:
                    stop.add(upstream_boundary)
                    # 상류 경계 노드의 상류도 제외
                    def add_upstream(nid):
                        for pid, _ in parents.get(nid, []):
                            stop.add(pid)
                            add_upstream(pid)
                    add_upstream(upstream_boundary)

                subs = collect_subbasins(downstream_boundary, stop)
                sub_area = sum(s.get('params', {}).get('A', 0) for s in subs)

                # 경계 하도추적 구간: src가 stop에 속하고 dst가 stop에 속하지 않는 reach
                boundary_reaches = []
                # 내부 노드: stop에 없고 downstream_boundary까지 도달 가능한 노드
                interior = {s['id'] for s in subs}
                if upstream_boundary is not None:
                    for e in reach_edges:
                        # src가 상류 유역(stop), dst가 이 유역 내부 → 경계 reach
                        if e['src'] in stop and e['dst'] not in stop:
                            boundary_reaches.append(e)
                        # src와 dst 모두 이 유역 내부이고 본류 경로상 → 내부 reach이지만
                        # 경계 추적 시 통합 대상
                        elif e['src'] not in stop and e['dst'] not in stop:
                            src_ca = cum_area.get(e['src'], 0)
                            dst_ca = cum_area.get(e['dst'], 0)
                            up_ca = cum_area.get(upstream_boundary, 0)
                            dn_ca = cum_area.get(downstream_boundary, 0)
                            if src_ca >= up_ca and dst_ca <= dn_ca:
                                boundary_reaches.append(e)

                mw_idx += 1
                roman = ['Ⅰ', 'Ⅱ', 'Ⅲ', 'Ⅳ', 'Ⅴ', 'Ⅵ'][min(mw_idx - 1, 5)]
                mw_name = f'A-{roman}'

                # 내부 하도추적 구간 (유역 내부 노드 간의 reach)
                # 유역 내부 노드 = stop에 속하지 않고 downstream_boundary까지 도달 가능
                interior_nodes = set()
                def collect_interior(nid):
                    if nid in interior_nodes or nid in stop:
                        return
                    interior_nodes.add(nid)
                    for pid, _ in parents.get(nid, []):
                        if pid not in stop:
                            collect_interior(pid)
                collect_interior(downstream_boundary)

                internal_reaches = []
                for e in reach_edges:
                    if e['src'] in interior_nodes and e['dst'] in interior_nodes:
                        if e not in boundary_reaches:
                            internal_reaches.append(e)

                medium_watersheds.append({
                    'name': mw_name,
                    'subbasins': sorted(subs, key=lambda s: s['name']),
                    'area': sub_area,
                    'cum_area': cum_area.get(downstream_boundary, 0),
                    'boundary_node': nodes.get(downstream_boundary, {}).get('name', ''),
                    'upstream_node': nodes.get(upstream_boundary, {}).get('name', '') if upstream_boundary else '(최상류)',
                    'internal_reaches': internal_reaches,
                    'boundary_reaches': boundary_reaches,
                })

        # ── 9. 보고서 생성 ──
        report = []
        report.append('=' * 80)
        report.append('■ 표준지침 4.3.3조 적합성 자동분석 보고서')
        report.append('=' * 80)
        report.append('')

        # 9-1. 기본 정보
        report.append('【기본 정보】')
        report.append(f'  총 유역면적: {total_area:.1f} km²')
        report.append(f'  소유역 수: {len(subbasins)}개')
        report.append(f'  하도추적 구간: {len(reach_edges)}개 ({", ".join(reach_labels)})')
        report.append(f'  합류 지점: {len(junctions)}개')
        report.append(f'  저수지: {len(reservoirs)}개')
        report.append('')

        # 소유역 목록
        report.append('【소유역 목록】')
        for n in sorted(subbasins.values(), key=lambda s: s['name']):
            p = n.get('params', {})
            report.append(f"  {n['name']:10s}  A={p.get('A',0):7.1f} km²  "
                          f"Tc={p.get('Tc',0):5.2f}hr  R={p.get('R',0):5.2f}hr  "
                          f"CN={p.get('CN',0):5.1f}")
        report.append('')

        # 9-2. 적합성 판정표
        report.append('─' * 80)
        report.append('■ 1. 적합성 판정표')
        report.append('─' * 80)
        report.append(f'  {"검토 항목":<35s} {"판정":<10s} {"근거"}')
        report.append('  ' + '─' * 76)
        for c in checks:
            report.append(f"  {c['item']:<35s} {c['result']:<10s} {c['detail']}")
        report.append('')

        # 전체 판정
        has_fail = any(c['result'] == '부적합' for c in checks)
        overall = '부적합 — 개선 필요' if has_fail else '적합'
        report.append(f'  ▶ 종합 판정: {overall}')
        report.append('')

        # 9-3. 문제점 상세 분석
        report.append('─' * 80)
        report.append('■ 2. 문제점 상세 분석')
        report.append('─' * 80)
        if not has_fail:
            report.append('  특이사항 없음 — 현행 구성이 표준지침에 적합합니다.')
        else:
            if area_over and has_sequential:
                report.append('')
                report.append('  [문제 1] 순차적 하도추적에 의한 첨두홍수량 과다 산정')
                report.append('')
                report.append(f'  현재 구성: 본류 경로에 {sequential_count}개 하도추적 구간이 순차 배치')
                report.append(f'  관련 구간: {", ".join(e.get("label", "?") for e in main_path_edges)}')
                report.append('')
                report.append('  ◆ 과다산정 메커니즘:')
                report.append('    1차: 작은 소유역 → 짧은 Tc → 뾰족한 UH(종거 큼) → 높은 첨두')
                report.append('    2차: 하도 저류효과 미미 → 감쇠 없이 하류 누적')
                report.append('    결과: 하류로 갈수록 첨두홍수량이 비현실적으로 증가')
                report.append('')
                report.append('  ◆ 표준지침 해결 방향:')
                report.append('    → 중규모 유역(≤250 km²)으로 묶어 통합 UH 적용')
                report.append('    → 중규모 유역 내부 하도추적 제거')
                report.append('    → 중규모 유역 간 경계는 한꺼번에 하도추적')
                report.append('')

            if area_over:
                report.append('')
                report.append('  [문제 2] 소유역 개별 UH 합산 방식')
                report.append('')
                report.append(f'  현재: {len(subbasins)}개 소유역 각각 개별 UH 산정 → 합산')
                report.append('  문제: 작은 소유역의 뾰족한 UH가 첨두를 과대 추정')
                report.append('  개선: 중규모 유역 단위 통합 면적에 단일 UH 적용 필요')

        report.append('')

        # 9-4. 텍스트 다이어그램
        report.append('─' * 80)
        report.append('■ 3. 텍스트 다이어그램')
        report.append('─' * 80)
        report.append('')

        # 다이어그램 A: 현재 구조
        report.append('  [다이어그램 A] 현재 구조 — 문제점 표시')
        report.append('')

        # 본류 경로 역순으로 다이어그램 생성
        def build_current_diagram():
            lines = []
            cur = outlet_id
            visited_diag = set()
            level = 0

            def draw_node(nid, indent):
                if nid in visited_diag:
                    return
                visited_diag.add(nid)
                n = nodes.get(nid)
                if not n:
                    return

                plist = parents.get(nid, [])
                prefix = '    ' * indent

                if n['type'] == 'SUBBASIN':
                    a = n.get('params', {}).get('A', 0)
                    lines.append(f'{prefix}[{n["name"]} {a:.1f}km²]')
                elif n['type'] == 'OUTLET':
                    lines.append(f'{prefix}({n["name"]}) ◀── 출구')
                elif n['type'] == 'JUNCTION':
                    lines.append(f'{prefix}({n["name"]}) ◀── 합류')
                elif n['type'] == 'RESERVOIR':
                    lines.append(f'{prefix}⬡{n["name"]}⬡ ◀── 저수지')

                for pid, edge in sorted(plist, key=lambda pe: -cum_area.get(pe[0], 0)):
                    rp = edge.get('reach_params')
                    if rp:
                        k = rp.get('K', 0)
                        label = edge.get('label', '?')
                        warn = ' ⚠ 순차추적' if (area_over and edge in main_path_edges) else ''
                        lines.append(f'{prefix}  ├── {label} K={k:.2f}hr ──┤{warn}')
                    else:
                        lines.append(f'{prefix}  ├──────────┤')
                    draw_node(pid, indent + 1)

            draw_node(outlet_id, 1)
            return lines

        diag_lines = build_current_diagram()
        for dl in diag_lines:
            report.append(dl)
        report.append('')

        # 다이어그램 B: 개선 후 구조
        if medium_watersheds:
            report.append('  [다이어그램 B] 개선 후 구조 (표준지침 적용)')
            report.append('')
            for mw in medium_watersheds:
                sub_names = [s['name'] for s in mw['subbasins']]
                sub_areas = [s.get('params', {}).get('A', 0) for s in mw['subbasins']]
                report.append(f'    ╔══ {mw["name"]} ({mw["area"]:.1f} km²) ═══{"═" * 40}╗')
                report.append(f'    ║  상류경계: {mw["upstream_node"]}')
                report.append(f'    ║  하류경계: {mw["boundary_node"]} (누적 {mw["cum_area"]:.1f} km²)')
                report.append(f'    ║  소유역: {", ".join(sub_names)}')
                report.append(f'    ║  면적: {" + ".join(f"{a:.1f}" for a in sub_areas)} = {mw["area"]:.1f} km²')
                if mw['internal_reaches']:
                    r_labels = [e.get('label', '?') for e in mw['internal_reaches']]
                    report.append(f'    ║  내부 하도추적 (제거 대상): {", ".join(r_labels)}')
                else:
                    report.append(f'    ║  내부 하도추적: 없음 ✓')
                if mw['boundary_reaches']:
                    r_labels = [e.get('label', '?') for e in mw['boundary_reaches']]
                    k_sum = sum(e.get('reach_params', {}).get('K', 0) for e in mw['boundary_reaches'])
                    report.append(f'    ║  경계 하도추적 (통합 적용): {", ".join(r_labels)} → 통합 K={k_sum:.2f}hr')
                report.append(f'    ╚{"═" * 60}╝')
                report.append('')
        report.append('')

        # 9-5. 개선 방안
        report.append('─' * 80)
        report.append('■ 4. 표준지침 기반 개선 방안')
        report.append('─' * 80)

        if not medium_watersheds:
            report.append('  총면적 250 km² 이하 — 현행 방식 유지 가능')
        else:
            report.append(f'  전체 유역({total_area:.1f} km²)을 {len(medium_watersheds)}개 중규모 유역으로 분할:')
            report.append('')

            for mw in medium_watersheds:
                sub_names = [s['name'] for s in mw['subbasins']]
                report.append(f'  ◆ {mw["name"]} ({mw["area"]:.1f} km², 경계: {mw["upstream_node"]} → {mw["boundary_node"]})')
                report.append(f'    소유역: {", ".join(sub_names)}')

                if mw['upstream_node'] == '(최상류)':
                    report.append(f'    방법: 유역추적만 (하도추적 제외)')
                    report.append(f'    → {mw["area"]:.1f} km² 전체에 통합 UH 적용')
                    if mw['internal_reaches']:
                        r_labels = [e.get('label', '?') for e in mw['internal_reaches']]
                        report.append(f'    → 내부 하도추적 구간 제거: {", ".join(r_labels)}')
                else:
                    report.append(f'    방법: 직상류 출구 홍수량 → 경계 하도추적 + 자체홍수량 합성')
                    if mw['boundary_reaches']:
                        r_labels = [e.get('label', '?') for e in mw['boundary_reaches']]
                        k_sum = sum(e.get('reach_params', {}).get('K', 0) for e in mw['boundary_reaches'])
                        report.append(f'    → 경계 하도추적: {" + ".join(r_labels)} = 통합 K={k_sum:.2f}hr (한꺼번에 추적)')
                    report.append(f'    → 자체홍수량: {mw["area"]:.1f} km²에 통합 UH 적용 (개별 합산 금지)')
                report.append('')

        # 9-6. 개선 전후 효과
        report.append('─' * 80)
        report.append('■ 5. 개선 전후 예상 효과')
        report.append('─' * 80)
        if has_fail:
            report.append('')
            report.append('  ◆ 첨두홍수량 변화:')
            report.append('    - 현행: 소유역 개별 UH(짧은 Tc) + 순차 추적 → 첨두 과다 산정')
            report.append('    - 개선: 중규모 유역 통합 UH(긴 Tc) + 통합 추적 → 첨두 저감')
            report.append(f'    - 예상: 통합 UH 적용 시 Tc 증가로 종거 감소 → 첨두 10~30% 저감 가능')
            report.append('')
            report.append('  ◆ 실무 적용 시 주의사항:')
            report.append('    - 중규모 유역 면적이 250 km²에 근접할수록 효과 증가')
            report.append('    - 통합 K값 산정 시 하도 연장·경사 등을 재검토해야 함')
            report.append('    - 댐이 있는 경우 댐 지점 기준 저수지 추적 선행 필요')
        else:
            report.append('  현행 구성이 표준지침에 적합하므로 추가 개선 불필요.')

        report.append('')
        report.append('=' * 80)
        report.append('※ 본 보고서는 표준지침 4.3.3조 규칙 기반 자동분석 결과입니다.')
        report.append('=' * 80)

        full_report = '\n'.join(report)

        # ── 10. 결과 출력 창 ──
        win = ctk.CTkToplevel(self)
        win.title('하천망 검토 — 표준지침 4.3.3 자동분석')
        win.geometry('980x800')
        try:
            hwnd = windll.user32.GetParent(win.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
        except Exception:
            pass

        top_bar = ctk.CTkFrame(win, fg_color='transparent')
        top_bar.pack(fill='x', padx=10, pady=(8, 2))
        status_lbl = ctk.CTkLabel(top_bar, text='  [자동분석]  표준지침 4.3.3 — 분석 완료',
                                   font=FONT_SMALL, text_color='#2ecc71', anchor='w')
        status_lbl.pack(side='left', fill='x', expand=True)

        def _copy():
            win.clipboard_clear()
            win.clipboard_append(txt.get('1.0', 'end'))
        ctk.CTkButton(top_bar, text='전체 복사', command=_copy, width=80, height=26,
                      font=FONT_SMALL, fg_color='#444', hover_color='#555').pack(side='right')

        txt_frame = ctk.CTkFrame(win, fg_color='#0d0d0d')
        txt_frame.pack(fill='both', expand=True, padx=8, pady=(2, 8))
        txt = tk.Text(txt_frame, wrap='word', bg='#0d0d0d', fg='#e8e8e8',
                      font=('맑은 고딕', 10), relief='flat', padx=14, pady=10,
                      insertbackground='white', selectbackground='#444')
        sb = ctk.CTkScrollbar(txt_frame, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        txt.pack(fill='both', expand=True)

        # 색상 태그
        txt.tag_config('title',    foreground='#f1c40f', font=('맑은 고딕', 11, 'bold'))
        txt.tag_config('header',   foreground='#3498db', font=('맑은 고딕', 10, 'bold'))
        txt.tag_config('ok',       foreground='#2ecc71')
        txt.tag_config('fail',     foreground='#e74c3c')
        txt.tag_config('warn',     foreground='#f39c12')
        txt.tag_config('diagram',  foreground='#a8d8a8', font=('Consolas', 10))
        txt.tag_config('sep',      foreground='#444444')
        txt.tag_config('body',     foreground='#e8e8e8', font=('맑은 고딕', 10))

        # 보고서 삽입 (줄별 태그 적용)
        for line in report:
            if line.startswith('═') or line.startswith('=' * 10):
                txt.insert('end', line + '\n', 'sep')
            elif line.startswith('─') or '─' * 10 in line:
                txt.insert('end', line + '\n', 'sep')
            elif line.startswith('■'):
                txt.insert('end', line + '\n', 'title')
            elif '부적합' in line:
                txt.insert('end', line + '\n', 'fail')
            elif '적합' in line and '부적합' not in line:
                txt.insert('end', line + '\n', 'ok')
            elif '해당없음' in line or '미해당' in line:
                txt.insert('end', line + '\n', 'warn')
            elif line.strip().startswith('╔') or line.strip().startswith('╚') or line.strip().startswith('║'):
                txt.insert('end', line + '\n', 'diagram')
            elif '⚠' in line:
                txt.insert('end', line + '\n', 'fail')
            elif '✓' in line:
                txt.insert('end', line + '\n', 'ok')
            elif line.strip().startswith('◆') or line.strip().startswith('▶'):
                txt.insert('end', line + '\n', 'header')
            elif line.strip().startswith('[') or line.strip().startswith('(') or line.strip().startswith('├'):
                txt.insert('end', line + '\n', 'diagram')
            else:
                txt.insert('end', line + '\n', 'body')

        txt.configure(state='disabled')

    def _build_review_prompt(self):
        """하천망 구조를 파싱하여 표준지침 검토용 프롬프트 생성."""
        ops = self.operations
        basins   = [op for op in ops if op['type'] == 'BASIN']
        routes   = [op for op in ops if op['type'] == 'ROUTE']
        combines = [op for op in ops if op['type'] == 'COMBINE']
        total_area = sum(op.get('A', 0) for op in basins)

        basin_lines = [
            f"  - {op['name']}: A={op.get('A',0):.1f} km², "
            f"Tc={op.get('Tc',0):.2f}hr, R={op.get('R',0):.2f}hr, CN={op.get('CN',0):.1f}"
            for op in basins
        ]
        route_lines = [
            f"  - {op['name']}: K={op.get('K',0):.2f}hr, X={op.get('X',0):.3f}, NSTPS={op.get('NSTPS',0)}"
            for op in routes
        ]
        combine_lines = [
            f"  - {op['name']}: {op.get('N',2)}개 수문곡선 합류"
            for op in combines
        ]
        order_lines = []
        for i, op in enumerate(ops, 1):
            t = op['type']
            if t == 'BASIN':
                order_lines.append(f"  {i:2d}. [소유역]   {op['name']} (A={op.get('A',0):.1f} km²)")
            elif t == 'ROUTE':
                order_lines.append(f"  {i:2d}. [하도추적] {op['name']} (K={op.get('K',0):.2f}hr, X={op.get('X',0):.3f})")
            elif t == 'COMBINE':
                order_lines.append(f"  {i:2d}. [합류]     {op['name']} ({op.get('N',2)}개 합산)")

        nl = '\n'
        return f"""당신은 수문학 전문가입니다. 아래 제시한 "홍수량 산정 표준지침(환경부, 2019)" 4.3.3조의 규정을 완전히 숙지한 후, 검토 대상 하천망이 이 규정에 적합한지 판정하고 문제점 및 개선방안을 한국어 보고서로 작성하세요.

================================================================================
■ 홍수량 산정 표준지침(환경부, 2019) 제4.3.3조 전문
================================================================================

【가. 단위도 적용을 위한 소유역 분할 기준】

(1) 대규모 유역의 경우 단위도의 기본 가정을 벗어나는 강우-유출관계가 예상되므로,
    전체 유역을 적절한 개수의 유역과 하도구간으로 나누어 단위도의 적용과 하도추적을
    병행하여 수문곡선을 합성해 나가는 것이 타당한 방법이다.
    유역면적이 약 250 km² 이하에서 단위도 사용을 추천하며, 250 km² 이상에서는
    정확도가 감소한다.

(2) 표준지침에서는 유역면적이 250 km² 이상인 유역에서의 유출계산은 전체 유역을
    적절한 개수의 소유역과 하도구간으로 분할하여 소유역에 단위도를 적용하고
    하도구간에 대해 홍수추적을 축차적으로 행하여 홍수 수문곡선을 합성한다.

(3) 유역면적 250 km² 이하인 유역에서도 홍수량 저감효과가 있는 구조물이 위치하는
    경우와 하도저류효과를 고려할 수 있는 하천유역에서는 소유역을 분할해서
    홍수량을 산정할 수 있다.

【나. 소유역을 분할한 경우의 홍수량 산정 절차】

(1) 대상 유역면적이 250 km² 이상인 경우는 소유역을 분할해서 홍수량을 산정한다.
    「설계홍수량 산정 요령(국토해양부, 2012)」에서는 홍수량 산정지점(P-Ⅰ-1, P-Ⅰ-2,
    ……, P-Ⅲ-2 등)을 기준으로 유역을 여러 개의 소유역으로 분할하여 상류에서 하류
    방향으로 유역추적과 하도추적을 반복하면서 홍수량을 산정하는 방법을 제시하였다.

(2) 소유역을 분할하여 하도추적을 실시할 경우, 분할된 소유역의 홍수도달시간이 짧으면
    단위도의 종거가 커지기 때문에 첨두홍수량이 커질 수 있다. 이때 하도의 저류효과가
    크지 않을 경우 첨두홍수량의 감소가 미미하므로 하류로 내려가면서 이 영향이 누적되어
    홍수량이 너무 커지는 문제가 발생할 수 있다.
    이러한 문제점을 해소하기 위해 다음과 같은 하도추적 방법의 적용을 추천한다:

  1) 분석 대상유역의 홍수량 산정지점(P-Ⅰ-1, P-Ⅰ-2, ……, P-Ⅲ-2 등) 기준으로
     여러 개의 소유역(A-Ⅰ-1, A-Ⅰ-2, ……, A-Ⅲ-2 등)으로 분할하기 이전에,
     전체 대상 유역을 상류로부터 하류 방향으로 유역면적이 약 250 km² 이하인
     몇 개의 중규모 유역(A-Ⅰ, A-Ⅱ, A-Ⅲ 등)으로 먼저 분할하고,
     최상류 중규모 유역의 홍수량을 유역추적 방법으로 산정한 후,
     이를 직하류 구간에 대한 하도추적 대상 홍수량으로 채택한다.
     ※ 최상류 중규모 유역 내부에서는 하도추적 없이 소유역 홍수량을 합산한다.

  2) 하류 방향으로의 하도추적은 홍수량 산정지점 구간별로 순차적으로 시행하는 것이
     아니라, 분할된 직상류의 중규모 유역 홍수량을 홍수량 산정지점별 해당 하도구간에
     대하여 한꺼번에 실시한다.
     ※ 예: P-Ⅱ-3 지점 → P-Ⅰ-3의 홍수량을 L-Ⅱ-1 + L-Ⅱ-2 + L-Ⅱ-3 구간을
            합한 하나의 구간에 대해 한꺼번에 하도추적. (P-Ⅱ-1 → P-Ⅱ-2 → P-Ⅱ-3
            순서의 구간별 순차 추적 금지)

  3) 하도추적 후 합성되는 자체 소유역의 홍수량은 홍수량 산정지점별로,
     직상류 중규모 유역 출구지점에서 해당 홍수량 산정지점까지의 면적에 대하여
     한꺼번에 산정한다.
     ※ 예: P-Ⅱ-3 지점의 자체 소유역 홍수량 → A-Ⅱ-1 + A-Ⅱ-2 + A-Ⅱ-3을
            하나의 유역으로 묶어 한꺼번에 단위도 적용.
            (A-Ⅱ-1, A-Ⅱ-2, A-Ⅱ-3 각각 개별 산정 후 합산 금지)

  4) 만약 대상 하천수계 내에 홍수조절용량을 가진 댐이 위치하고 있을 경우에는,
     댐 지점을 기준으로 하여 저수지 홍수추적을 실시하고 댐 직하류부터
     하도추적을 계속해야 한다.

(3) 하도추적 포함방법에 의한 홍수량 산정절차 (예시: 총면적 약 600 km²):

  [1단계] 전체 유역을 250 km² 이하로 분할:
    - 중규모 유역Ⅰ: 240 km² (최상류)
    - 중규모 유역Ⅱ: 230 km²
    - 중규모 유역Ⅲ: 130 km² (최하류)

  [2단계] 중규모 유역Ⅰ 홍수량 산정:
    - P-Ⅰ-1, P-Ⅰ-2, P-Ⅰ-3 지점 모두 하도추적 제외(유역추적만) 방법으로 산정
    - 즉, 유역Ⅰ 내 소유역들의 홍수량을 하도추적 없이 합산하여 P-Ⅰ-3 출구 홍수량 확정

  [3단계] 중규모 유역Ⅱ 각 산정지점 홍수량:
    - P-Ⅱ-1: P-Ⅰ-3 홍수량 → L-Ⅱ-1 구간 하도추적 + A-Ⅱ-1 자체홍수량 합성
    - P-Ⅱ-2: P-Ⅰ-3 홍수량 → (L-Ⅱ-1+L-Ⅱ-2) 합한 하나의 구간 하도추적
              + (A-Ⅱ-1+A-Ⅱ-2) 합한 하나의 유역 자체홍수량 합성
    - P-Ⅱ-3: P-Ⅰ-3 홍수량 → (L-Ⅱ-1+L-Ⅱ-2+L-Ⅱ-3) 합한 하나의 구간 하도추적
              + (A-Ⅱ-1+A-Ⅱ-2+A-Ⅱ-3) 합한 하나의 유역 자체홍수량 합성

  [4단계] 중규모 유역Ⅲ 각 산정지점 홍수량:
    - P-Ⅲ-1: P-Ⅱ-3 홍수량 → L-Ⅲ-1 구간 하도추적 + A-Ⅲ-1 자체홍수량 합성
    - P-Ⅲ-2: P-Ⅱ-3 홍수량 → (L-Ⅲ-1+L-Ⅲ-2) 합한 하나의 구간 하도추적
              + (A-Ⅲ-1+A-Ⅲ-2) 합한 하나의 유역 자체홍수량 합성

================================================================================
■ 검토 대상 하천망
================================================================================

총 유역면적: {total_area:.1f} km²
소유역 수: {len(basins)}개 / 하도추적 구간: {len(routes)}개 / 합류 지점: {len(combines)}개

【소유역 목록】
{nl.join(basin_lines) if basin_lines else '  (없음)'}

【하도추적 구간】
{nl.join(route_lines) if route_lines else '  (없음)'}

【합류 지점】
{nl.join(combine_lines) if combine_lines else '  (없음)'}

【수문 분석 실행 순서 (DFS 기준)】
{nl.join(order_lines)}

================================================================================
■ 보고서 작성 요청
================================================================================

위 표준지침 전문을 기준으로 다음 5개 항목으로 보고서를 작성하세요:

1. 적합성 판정표
   - 항목별(소유역 면적 상한 / 중규모 유역 분할 / 하도추적 방식 / 자체 소유역 합산 / 댐 처리)
   - 각 항목: 적합 / 불합리 / 해당없음 판정 + 근거

2. 문제점 상세 분석
   - 불합리 항목별로 현재 구성의 구체적 문제 설명
   - 첨두홍수량 과다 산정이 우려되는 지점 및 이유

3. 텍스트 다이어그램 (반드시 포함)

   아래 두 가지 다이어그램을 ASCII/텍스트 기호로 작성하세요.
   기호 규칙:
     [소유역명 면적km²]   : 소유역 박스
     (합류점명)           : 합류 지점
     ──K=X.Xhr──▶        : 하도추적 엣지 (K값 표시)
     ──────▶              : 하도추적 없는 직접 연결
     ╔══중규모유역명══╗   : 중규모 유역 그룹 경계
     ⚠                   : 문제 지점 표시
     ✓                   : 표준지침 적합 지점

   [다이어그램 A] 현재 구조 — 문제점 표시
   - 전체 하천망을 상류→하류 방향으로 표현
   - ⚠ 기호로 순차 추적 문제 지점 명시
   - 각 소유역 면적 표시

   [다이어그램 B] 개선 후 구조 (표준지침 적용)
   - 중규모 유역 A-Ⅰ, A-Ⅱ 등을 ╔══╗ 박스로 구분
   - 각 중규모 유역 내부: 소유역 합산 방식 표시
   - 중규모 유역 간 경계: 한꺼번에 하도추적하는 구간 명시
   - ✓ 기호로 표준지침 적합 처리 표시

4. 표준지침 기반 개선 방안 (문자 설명)
   - 중규모 유역 분할 기준 및 각 유역 구성 소유역 목록
   - 각 중규모 유역별 하도추적 재구성 방법
   - 자체 소유역 묶음 방법

5. 개선 전후 예상 효과
   - 첨두홍수량 변화 방향 및 크기 예상
   - 실무 적용 시 주의사항

실무 수문 엔지니어가 즉시 활용할 수 있도록 구체적인 소유역명과 수치를 포함하여 작성하세요.
다이어그램은 등폭 텍스트(monospace)로 작성하며, 한눈에 전체 구조를 파악할 수 있도록 충분한 공간을 사용하세요."""

    def _on_network_applied(self, ops):
        self.operations = ops
        if self._editor and self._editor.winfo_exists():
            p = getattr(self._editor, '_current_path', None)
            if p:
                self._net_json_path = p
                self._net_json_name = os.path.basename(p)
        self._update_net_label()
        n = len(ops)
        bc = sum(1 for o in ops if o['type'] == 'BASIN')
        rc = sum(1 for o in ops if o['type'] == 'ROUTE')
        self._log(f'네트워크 적용: {n}개 조작 (소유역 {bc}개, 추적 {rc}구간)')

    def _update_net_label(self):
        if self._net_json_name:
            self._net_lbl.configure(
                text=self._net_json_name,
                text_color='#5dade2')
        else:
            self._net_lbl.configure(
                text='파일로드 없음',
                text_color='red')

    # ── 예제 로드 ─────────────────────────────────────────────────────────────

    def _load_example(self):
        if self.operations:
            if not messagebox.askyesno('확인', '기존 조작을 지우고 예제를 로드하시겠습니까?'):
                return
        self._set_entry('DT_MIN', '60')
        self._set_entry('TR_MIN', '1440')
        self._set_entry('NQ',     '300')
        self._huff_var.set('3분위')
        self._pc_values = list(HUFF_PRESETS['3분위'])
        self.operations = copy.deepcopy(EXAMPLE_OPERATIONS)
        self._update_net_label()
        self._log('예제 네트워크 로드 완료 (100-1440-SW00.dat, 521.9 km²)')
        # Sync to editor if open
        if self._editor and self._editor.winfo_exists():
            self._editor.load_operations(self.operations)

    # ── 분석 실행 ─────────────────────────────────────────────────────────────

    def _run(self):
        # If editor is open, get latest ops from it
        if self._editor and self._editor.winfo_exists():
            ops, err = self._editor._canvas.build_operations()
            if err:
                messagebox.showerror('네트워크 오류', err)
                return
            self.operations = ops
            self._update_net_label()

        if not self.operations:
            messagebox.showwarning('알림', '하천망이 없습니다. 편집기 또는 예제 로드를 사용하세요.')
            return

        try:
            dt_min = float(self._entries['DT_MIN'].get())
            tr_min = float(self._entries['TR_MIN'].get())
            NQ     = int(float(self._entries['NQ'].get()))
        except ValueError as e:
            messagebox.showerror('입력 오류', str(e))
            return

        self._log(f'분석 시작: Δt={dt_min}min TR={tr_min}min NQ={NQ}')

        try:
            results = self.processor.run(
                self.operations, dt_min, NQ, tr_min, self._pc_values, 0.0)
        except Exception:
            self._log(f'오류: {traceback.format_exc()}')
            messagebox.showerror('계산 오류', traceback.format_exc()[:300])
            return

        for w in self.processor.warnings:
            self._log(f'⚠ {w}')

        if self.processor.summary:
            last = self.processor.summary[-1]
            self._result_labels['outlet'].configure(text=last['station'])
            self._result_labels['peak_q'].configure(text=f"{last['peak_q']:.2f} m³/s")
            self._result_labels['peak_hr'].configure(text=f"{last['peak_hr']:.2f} hr")
            self._result_labels['cum_area'].configure(text=f"{last['cum_area']:.2f} km²")
            self._log(f"완료 | 출구={last['station']} | 첨두={last['peak_q']:.2f} m³/s @ {last['peak_hr']:.2f}hr")

        self._plot_results(results, dt_min, NQ)
        self._update_summary_table()

    def _plot_results(self, results, dt_min, NQ):
        for w in self._graph_frame.winfo_children(): w.destroy()
        dt_hr    = dt_min / 60.0
        time_hr  = np.arange(NQ) * dt_hr

        plt, FigureCanvasTkAgg = _ensure_mpl()
        plt.rcParams['font.family']        = 'Malgun Gothic'
        plt.rcParams['axes.unicode_minus'] = False
        fig, axes = plt.subplots(1, 2, figsize=(12, 4.5))
        fig.patch.set_facecolor('#1e1e1e')
        for ax in axes:
            ax.set_facecolor('#1e1e1e')
            ax.tick_params(colors='white')
            ax.xaxis.label.set_color('white')
            ax.yaxis.label.set_color('white')
            ax.title.set_color('white')
            for sp in ax.spines.values(): sp.set_edgecolor('#555555')

        colors_plot = plt.cm.tab10(np.linspace(0, 1, min(len(results), 10)))
        shown = 0
        for (name, rdata), col in zip(results.items(), colors_plot):
            if rdata['type'] in ('BASIN', 'COMBINE') and shown < 8:
                flow = rdata['flow']
                n    = min(len(flow), NQ)
                axes[0].plot(time_hr[:n], flow[:n], color=col, label=name, linewidth=1.2)
                shown += 1
        axes[0].set_title('지점별 수문곡선', fontproperties='Malgun Gothic')
        axes[0].set_xlabel('시간 (hr)', fontproperties='Malgun Gothic')
        axes[0].set_ylabel('유량 (m³/s)', fontproperties='Malgun Gothic')
        if shown > 0:
            axes[0].legend(fontsize=7, facecolor='#2a2a2a', labelcolor='white')
        axes[0].grid(True, color='#333333', alpha=0.7)

        if self.processor.summary:
            last_name = self.processor.summary[-1]['station']
            if last_name in results:
                flow = results[last_name]['flow']
                n    = min(len(flow), NQ)
                axes[1].fill_between(time_hr[:n], flow[:n], alpha=0.3, color='#3498db')
                axes[1].plot(time_hr[:n], flow[:n], color='#3498db', linewidth=2,
                             label=f'{last_name} 최종 유출')
                pidx = int(np.argmax(flow[:n]))
                axes[1].axvline(time_hr[pidx], color='#e74c3c', linestyle='--', alpha=0.7)
                axes[1].scatter([time_hr[pidx]], [flow[pidx]], color='#e74c3c', zorder=5, s=60)
                axes[1].annotate(
                    f"첨두 {flow[pidx]:.1f} m³/s\n@ {time_hr[pidx]:.1f}hr",
                    xy=(time_hr[pidx], flow[pidx]),
                    xytext=(10, -30), textcoords='offset points',
                    color='#e74c3c', fontsize=8,
                    arrowprops=dict(arrowstyle='->', color='#e74c3c'))
        axes[1].set_title('최종 출구 수문곡선', fontproperties='Malgun Gothic')
        axes[1].set_xlabel('시간 (hr)', fontproperties='Malgun Gothic')
        axes[1].set_ylabel('유량 (m³/s)', fontproperties='Malgun Gothic')
        axes[1].legend(fontsize=8, facecolor='#2a2a2a', labelcolor='white')
        axes[1].grid(True, color='#333333', alpha=0.7)
        fig.tight_layout(pad=2.0)

        if self._mpl_canvas:
            try: self._mpl_canvas.get_tk_widget().destroy()
            except Exception: pass
        self._mpl_canvas = FigureCanvasTkAgg(fig, master=self._graph_frame)
        self._mpl_canvas.draw()
        self._mpl_canvas.get_tk_widget().pack(fill='both', expand=True)

    def _update_summary_table(self):
        for w in self._tbl_frame.winfo_children(): w.destroy()
        headers = ['#', '조작', '지점명', '첨두유량(m³/s)', '첨두시간(hr)', '누가면적(km²)']
        col_w   = [30, 180, 90, 130, 110, 120]
        hrow    = ctk.CTkFrame(self._tbl_frame, fg_color='#2c3e50', corner_radius=0)
        hrow.pack(fill='x', padx=2)
        for h, w in zip(headers, col_w):
            ctk.CTkLabel(hrow, text=h, font=('맑은 고딕', 10, 'bold'),
                         width=w, anchor='center', text_color='white').pack(side='left', padx=1)
        for i, row in enumerate(self.processor.summary):
            bg = '#1e1e1e' if i % 2 == 0 else '#242424'
            fr = ctk.CTkFrame(self._tbl_frame, fg_color=bg, corner_radius=0)
            fr.pack(fill='x', padx=2)
            vals = [str(i+1), row['op'], row['station'],
                    f"{row['peak_q']:.2f}", f"{row['peak_hr']:.2f}", f"{row['cum_area']:.2f}"]
            for v, w in zip(vals, col_w):
                ctk.CTkLabel(fr, text=v, font=FONT_SMALL, width=w,
                             anchor='center', text_color='#cccccc').pack(side='left', padx=1)

    # ── Excel 저장 ────────────────────────────────────────────────────────────

    def _save_excel(self):
        if not self.processor.results:
            messagebox.showwarning('알림', '먼저 분석을 실행하세요.')
            return
        ts    = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'{self.project_name}_F_Flood_Routing_{ts}.xlsx'
        out   = os.path.join(self.project_path, fname)
        try:
            self._write_excel(out)
            ops, _ = self._canvas.build_operations()
            out_path = os.path.splitext(out)[0] + '.out'
            self._write_out_file(out_path, ops)
            self._log(f'Excel 저장: {fname}')
            self._save_config(fname, out)
            messagebox.showinfo('저장 완료', f'저장되었습니다:\n{out}')
        except Exception:
            self._log(f'Excel 저장 오류: {traceback.format_exc()}')
            messagebox.showerror('저장 오류', traceback.format_exc()[:300])

    def _write_out_file(self, path, ops):
        """HEC-1 스타일 .out 텍스트 파일 저장."""
        now = datetime.now()
        lines = []
        lines.append('HEC-1 STYLE ANALYSIS OUTPUT (Hydroset)')
        lines.append(f'RUN DATE  {now.strftime("%d-%b-%y").upper()}  TIME {now.strftime("%H:%M:%S")}')
        lines.append('')

        # ── 입력 에코 ──────────────────────────────────────────────────────────
        lines.append('HEC-1 INPUT ECHO')
        header = '  LINE    ID' + ''.join(f'.......{i}' for i in range(1, 9))
        lines.append(header)
        dat_content = self._build_dat_content(ops)
        for idx, ln in enumerate(dat_content.splitlines(), 1):
            lines.append(f'{idx:6d}    {ln}')
        lines.append('')

        # ── Runoff Summary ─────────────────────────────────────────────────────
        lines.append('RUNOFF SUMMARY')
        lines.append('-' * 70)
        lines.append(f'{"#":>4}  {"OP":<16}  {"STATION":<12}  {"PEAK_Q(m³/s)":>14}  {"PEAK_HR":>8}  {"CUM_AREA":>10}')
        lines.append('-' * 70)
        for i, row in enumerate(self.processor.summary, 1):
            lines.append(
                f'{i:4d}  {row["op"]:<16}  {row["station"]:<12}  '
                f'{row["peak_q"]:14.3f}  {row["peak_hr"]:8.2f}  {row["cum_area"]:10.2f}'
            )
        lines.append('')

        # ── 지점별 수문곡선 ────────────────────────────────────────────────────
        dt_hr = self._dt_min / 60.0
        for row in self.processor.summary:
            station = row['station']
            res = self.processor.results.get(station)
            if res is None:
                continue
            q_arr = res['flow']
            lines.append(f'STATION: {station}  ({row["op"]})')
            lines.append(f'{"TIME(hr)":>10}  {"FLOW(m³/s)":>12}')
            for j, q in enumerate(q_arr):
                lines.append(f'{j * dt_hr:10.2f}  {q:12.3f}')
            lines.append('')

        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

    def _write_excel(self, path):
        oxl = _ensure_openpyxl()
        Workbook        = oxl['Workbook']
        Font            = oxl['Font']
        PatternFill     = oxl['PatternFill']
        Alignment       = oxl['Alignment']
        get_column_letter = oxl['get_column_letter']

        wb    = Workbook()
        hfnt  = Font(bold=True, color='FFFFFF')
        hfill = PatternFill('solid', fgColor='2C3E50')
        rfill = PatternFill('solid', fgColor='1A252F')

        def hcell(ws, r, c, v):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = hfnt; cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')

        ws1 = wb.active; ws1.title = 'Runoff Summary'
        for c, h in enumerate(['순번','조작','지점명','첨두유량(m³/s)',
                                '첨두시간(hr)','누가면적(km²)'], 1):
            hcell(ws1, 1, c, h)
            ws1.column_dimensions[get_column_letter(c)].width = 18

        for i, row in enumerate(self.processor.summary, 2):
            ws1.cell(row=i, column=1, value=i-1)
            ws1.cell(row=i, column=2, value=row['op'])
            ws1.cell(row=i, column=3, value=row['station'])
            ws1.cell(row=i, column=4, value=round(row['peak_q'],  3))
            ws1.cell(row=i, column=5, value=round(row['peak_hr'], 3))
            ws1.cell(row=i, column=6, value=round(row['cum_area'],3))
            if i % 2 == 0:
                for c in range(1, 7): ws1.cell(row=i, column=c).fill = rfill

        route_ops = [(nm, d) for nm, d in self.processor.results.items() if d['type'] == 'ROUTE']
        if route_ops:
            cs = 8
            for c, h in enumerate(['추적지점','K(hr)','X','NSTPS','C1','C2','C3','안정'], cs):
                hcell(ws1, 1, c, h)
                ws1.column_dimensions[get_column_letter(c)].width = 12
            for i, (nm, d) in enumerate(route_ops, 2):
                for j, v in enumerate([nm, d['K'], d['X'], d['NSTPS'],
                                        round(d['C1'],5), round(d['C2'],5), round(d['C3'],5),
                                        'OK' if d['stable'] else '경고']):
                    ws1.cell(row=i, column=cs+j, value=v)

        dt_min = float(self._entries['DT_MIN'].get())
        NQ     = int(float(self._entries['NQ'].get()))
        dt_hr  = dt_min / 60.0

        for name, rdata in self.processor.results.items():
            ws = wb.create_sheet(name[:30])
            hcell(ws, 1, 1, '시간(hr)'); hcell(ws, 1, 2, '유량(m³/s)')
            ws.column_dimensions['A'].width = 12; ws.column_dimensions['B'].width = 14
            flow = rdata['flow']
            n    = min(len(flow), NQ)
            for i in range(n):
                ws.cell(row=i+2, column=1, value=round(i * dt_hr, 4))
                ws.cell(row=i+2, column=2, value=round(float(flow[i]), 4))
            ws.cell(row=1, column=4, value='항목').font = hfnt
            ws.cell(row=1, column=4).fill = hfill
            ws.cell(row=1, column=5, value='값').font  = hfnt
            ws.cell(row=1, column=5).fill = hfill
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 14
            params = [('유형', rdata['type']),
                      ('첨두유량(m³/s)', round(rdata['peak_q'], 3)),
                      ('첨두시간(hr)',   round(rdata['peak_hr'],3))]
            if rdata['type'] == 'BASIN':
                params += [('A(km²)', rdata['A'])]
            elif rdata['type'] == 'ROUTE':
                params += [('K(hr)', rdata['K']), ('X', rdata['X']),
                           ('NSTPS', rdata['NSTPS']),
                           ('C1', round(rdata['C1'],5)), ('C2', round(rdata['C2'],5)),
                           ('C3', round(rdata['C3'],5)),
                           ('안정', 'OK' if rdata['stable'] else '경고')]
            for pi, (lbl, val) in enumerate(params, 2):
                ws.cell(row=pi, column=4, value=lbl)
                ws.cell(row=pi, column=5, value=val)

        wb.save(path)

    # ── Config I/O ────────────────────────────────────────────────────────────

    def _load_config(self):
        try:
            with open(self.config_file, encoding='utf-8') as f:
                cfg = json.load(f)
            s6 = cfg.get('step6', {})
            for key in ('DT_MIN', 'TR_MIN', 'NQ'):
                if key in s6: self._set_entry(key, str(s6[key]))
            if 'huff' in s6 and s6['huff'] in HUFF_PRESETS:
                self._huff_var.set(s6['huff'])
                self._pc_values = list(HUFF_PRESETS[s6['huff']])
            if 'operations' in s6 and isinstance(s6['operations'], list):
                self.operations = s6['operations']
                self._update_net_label()
        except Exception:
            pass

    def _save_config(self, fname='', out_path=''):
        try:
            cfg = {}
            try:
                with open(self.config_file, encoding='utf-8') as f:
                    cfg = json.load(f)
            except Exception: pass

            s6 = {}
            for key in ('DT_MIN', 'TR_MIN', 'NQ'):
                try:
                    val = self._entries[key].get()
                    s6[key] = int(float(val)) if key == 'NQ' else float(val)
                except Exception: pass
            s6['huff']       = self._huff_var.get()
            s6['operations'] = copy.deepcopy(self.operations)

            if self.processor.summary:
                last = self.processor.summary[-1]
                s6['outlet']   = last['station']
                s6['peak_q']   = round(last['peak_q'],  3)
                s6['peak_hr']  = round(last['peak_hr'], 3)
                s6['cum_area'] = round(last['cum_area'],3)

            if fname:
                cfg['step6_flood_routing'] = {
                    'status':      'completed',
                    'output_file': fname,
                    'full_path':   out_path,
                    'outlet':      s6.get('outlet', ''),
                    'peak_q':      s6.get('peak_q',  0),
                    'peak_hr':     s6.get('peak_hr', 0),
                    'cum_area':    s6.get('cum_area',0),
                    'timestamp':   datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                }
            cfg['step6'] = s6
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception: pass

    # ── 헬퍼 ──────────────────────────────────────────────────────────────────

    def _set_entry(self, key, val):
        w = self._entries.get(key)
        if w is None: return
        if isinstance(w, ctk.CTkComboBox):
            w.set(val)
        else:
            w.delete(0, 'end'); w.insert(0, val)

    def _on_huff_change(self, choice):
        if choice in HUFF_PRESETS:
            self._pc_values = list(HUFF_PRESETS[choice])

    def _log(self, msg):
        self._txt_log.insert('end', f'[{datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self._txt_log.see('end')

    def _on_close(self):
        self._save_config()
        self.destroy()


# =============================================================================
# 진입점
# =============================================================================

if __name__ == '__main__':
    project_path = sys.argv[1] if len(sys.argv) > 1 else os.getcwd()
    input_file   = sys.argv[2] if len(sys.argv) > 2 else ''
    app = FloodRoutingApp(project_path, input_file)
    app.mainloop()
