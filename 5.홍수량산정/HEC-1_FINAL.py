import sys
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

# [필수 라이브러리 체크]
try:
    import numpy as np
    import pandas as pd
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import openpyxl
except ImportError as e:
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("라이브러리 오류", f"필수 라이브러리 설치 필요:\n{e}\n\npip install numpy pandas matplotlib openpyxl")
    sys.exit(1)

# ============================================================================
# HEC1 Engine (계산 로직)
# ============================================================================

class HEC1Engine:
    def __init__(self):
        self.params = {}
        self.results = {}
        self.clark_details = {}

    def set_params(self, params):
        self.params = params

    def calculate(self, shift_step=0):
        """
        HEC-1 Clark 단위도법 계산
        - Huff Input(0~1 누가비)을 TR_MIN/NMIN+1 지점에 선형보간 후 사용
        - conv_matrix를 결과에 포함 (엑셀 Convolution Detail 검증용)
        """
        nmin = self.params['NMIN']
        nq = self.params['NQ']
        total_p = self.params['TOTAL_PRECIP']
        tr_min = self.params['TR_MIN']

        # --------------------------------------------------------------
        # 0. Huff Input → Interpolated PC (선형 보간)
        # --------------------------------------------------------------
        pc_huff = np.array(self.params['PC'], dtype=float)  # 0~1, Nhuff개

        n_step = int(tr_min / nmin) + 1        # 예: 610/10+1 = 62
        t_pc = np.linspace(0.0, 1.0, n_step)   # 0~1
        t_huff = np.linspace(0.0, 1.0, len(pc_huff))

        pc_interp = np.interp(t_pc, t_huff, pc_huff)  # 0~1

        # 누가강우 시계열 (길이 nq+1)
        if n_step < nq + 1:
            pad_len = nq + 1 - n_step
            pc_series = np.concatenate(
                [pc_interp, np.full(pad_len, pc_interp[-1])]
            )
        else:
            pc_series = pc_interp[:nq + 1]

        cumulative_rainfall = pc_series * total_p  # 길이 nq+1

        # --------------------------------------------------------------
        # 1. 강우 및 유효우량 (SCS-CN)
        # --------------------------------------------------------------
        cn = self.params['CN']
        S = (25400.0 / cn) - 254.0
        Ia = 0.2 * S

        cumulative_excess = np.zeros(nq + 1)
        for i, P in enumerate(cumulative_rainfall):
            if P > Ia:
                cumulative_excess[i] = ((P - Ia) ** 2) / (P - Ia + S)
            else:
                cumulative_excess[i] = 0.0

        # 증가분 (길이 nq)
        incremental_rainfall = np.diff(cumulative_rainfall)
        incremental_excess = np.diff(cumulative_excess)

        # --------------------------------------------------------------
        # 2. Clark 단위도
        # --------------------------------------------------------------
        tc, r, area = self.params['TC'], self.params['R'], self.params['AREA_KM2']
        dt_hr = nmin / 60.0

        uh_duration = int((tc + r * 10) / dt_hr) + 5
        t_vals = np.arange(0, uh_duration * dt_hr, dt_hr)

        ai_vals = np.zeros(len(t_vals))
        for i, t in enumerate(t_vals):
            T = t / tc if tc > 0 else 1.0
            if T >= 1.0:
                ai_vals[i] = 1.0
            elif T < 0.5:
                ai_vals[i] = 1.414 * (T ** 1.5)
            else:
                ai_vals[i] = 1.0 - 1.414 * ((1.0 - T) ** 1.5)

        vol_const = area * 1000.0
        I_flow = np.zeros(len(t_vals))
        for i in range(1, len(t_vals)):
            dAI = ai_vals[i] - ai_vals[i - 1]
            if dAI < 0:
                dAI = 0
            I_flow[i] = dAI * vol_const / (dt_hr * 3600.0)

        c_coef = dt_hr / (r + 0.5 * dt_hr)
        O_inst = np.zeros(len(t_vals))
        for i in range(1, len(t_vals)):
            O_inst[i] = c_coef * I_flow[i] + (1 - c_coef) * O_inst[i - 1]

        uh_ordinates = np.zeros(len(t_vals))
        for i in range(1, len(t_vals)):
            uh_ordinates[i] = 0.5 * (O_inst[i] + O_inst[i - 1])

        calc_vol = np.sum(uh_ordinates) * dt_hr * 3600
        target_vol = area * 1000.0
        if calc_vol > 0:
            uh_ordinates *= (target_vol / calc_vol)

        self.clark_details = {
            'Time_hr': t_vals,
            'AI': ai_vals,
            'I': I_flow,
            'O': O_inst,
            'UH': uh_ordinates,
            'CA': c_coef,
            'CB': 1 - c_coef
        }

        # --------------------------------------------------------------
        # 3. Convolution (conv_matrix 기준)
        # --------------------------------------------------------------
        uh_vec = uh_ordinates[1:]  # 첫 값 제외
        conv_matrix = np.zeros((nq, len(incremental_excess)))

        for i, excess in enumerate(incremental_excess):
            if excess > 0:
                runoff_comp = excess * uh_vec
                start_idx = i
                end_idx = min(start_idx + len(runoff_comp), nq)
                if end_idx > start_idx:
                    conv_matrix[start_idx:end_idx, i] = runoff_comp[:end_idx - start_idx]

        raw_runoff = np.sum(conv_matrix, axis=1)

        # Shift Logic (원본과 동일)
        if shift_step == -1:
            final_hydrograph = np.append(raw_runoff[1:], 0)
        elif shift_step == 1:
            final_hydrograph = np.insert(raw_runoff, 0, 0.0)[:nq]
        else:
            final_hydrograph = np.insert(raw_runoff, 0, 0.0)[:nq]

        self.results = {
            'time_min': np.arange(0, nq * nmin, nmin),
            'rain_inc': incremental_rainfall,
            'excess_inc': incremental_excess,
            'flow': final_hydrograph,
            'peak_flow': np.max(final_hydrograph),
            'peak_time_min': np.argmax(final_hydrograph) * nmin,
            'conv_matrix': conv_matrix,
            'pc_interp': pc_series
        }

    def write_hec1_out(self, out_path):
        """HEC-1 OUT 파일 생성"""
        flows = self.results['flow']
        params = self.params
        start_date = datetime(2005, 7, 11, 18, 0)
        nmin = params['NMIN']
        
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(f"1TABLE 1    STATION      {params['STATION_ID']}\n")
            f.write("                           FLOW\n\n")
            f.write(" PER  DAY MON  HRMN\n\n")
            
            for i, flow in enumerate(flows):
                curr_time = start_date + timedelta(minutes=i * nmin)
                per_str = f"{i+1:4d}"
                day_str = f"{curr_time.day:2d}"
                mon_str = curr_time.strftime("%b").upper()
                time_str = curr_time.strftime("%H%M")
                
                if abs(flow) < 0.005: flow_str = "         .00"
                else:
                    flow_str = f"{flow:12.2f}"
                    if 0 < abs(flow) < 1: flow_str = flow_str.replace("0.", " .")
                
                f.write(f" {per_str}   {day_str} {mon_str}  {time_str}    {flow_str}\n")
            
            f.write("\n")
            f.write(f"              MAX         {np.max(flows):.2f}\n")

    def write_hec1_dat(self, dat_path):
        """HEC-1 DAT 파일 생성 - 원본 형식 준수"""
        params = self.params
        start_date = datetime(2005, 7, 11, 18, 0)
        d_str = start_date.strftime("%d%b%y").upper()
        t_str = start_date.strftime("%H%M")
        
        # 헤더 부분 (고정)
        header = f"""ID Project(Flood)
*DIAGRAM
IM
IO     5       1
IT    {params['NMIN']:2d} {d_str}    {t_str}     {params['NQ']}
VS{params['STATION_ID']}
VV  2.11
* 
"""
        
        # 변수 부분 - BA, PB
        ba_val = params['AREA_KM2']
        ba_str = f"{ba_val:.1f}".lstrip('0') if ba_val < 1 else f"{ba_val:.1f}"
        
        pb_val = params['TOTAL_PRECIP']
        pb_str = f"{pb_val:.1f}".lstrip('0') if pb_val < 1 else f"{pb_val:.1f}"
        
        variables = f"""KK{params['STATION_ID']}
IN    {params['NMIN']:2d} {d_str}    {t_str}
BA{ba_str:>6s}
PB{pb_str:>6s}
"""
        
        # PC 카드
        # pc_lines = []
        # pc_list = params['PC']
        # for i in range(0, len(pc_list), 10):
            # chunk = pc_list[i:i+10]
            # pc_line = "PC"
            # for idx, val in enumerate(chunk):
                # val_str = f"{val:.3f}".lstrip('0') if val < 1 else f"{val:.3f}"
                # if idx == 0:
                    # pc_line += f"{val_str:>6s}"
                # else:
                    # pc_line += f"{val_str:>8s}"
            # pc_lines.append(pc_line)
            
        # PC 카드 (보간된 PC 사용)
        # PC 카드 (보간된 PC 사용, 마지막 1.000 한 번만)
        pc_lines = []

        # 계산 시 만든 보간 PC: 0.0~1.0, 길이 = NQ+1
        pc_interp = self.results.get('pc_interp')
        if pc_interp is None:
            pc_list = list(params['PC'])
        else:
            pc_list = list(pc_interp)

        # 마지막 값이 1.0이면 중복 방지를 위해 하나만 유지
        if len(pc_list) >= 2 and abs(pc_list[-1] - 1.0) < 1e-6 and abs(pc_list[-2] - 1.0) < 1e-6:
            # 뒤에서부터 1.0이 아닌 첫 값 이후는 잘라냄
            i = len(pc_list) - 1
            while i > 0 and abs(pc_list[i-1] - 1.0) < 1e-6:
                i -= 1
            pc_list = pc_list[:i+1]

        for i in range(0, len(pc_list), 10):
            chunk = pc_list[i:i+10]
            pc_line = "PC"
            for idx, val in enumerate(chunk):
                val_str = f"{val:.3f}".lstrip('0') if val < 1 else f"{val:.3f}"
                if idx == 0:
                    pc_line += f"{val_str:>6s}"
                else:
                    pc_line += f"{val_str:>8s}"
            pc_lines.append(pc_line)

        
        # LS, UC
        cn_val = params['CN']
        cn_str = f"{cn_val:.1f}".lstrip('0') if cn_val < 1 else f"{cn_val:.1f}"
        
        tc_val = params['TC']
        r_val = params['R']
        tc_str = f"{tc_val:.2f}".lstrip('0') if tc_val < 1 else f"{tc_val:.2f}"
        r_str = f"{r_val:.2f}".lstrip('0') if r_val < 1 else f"{r_val:.2f}"
        
        tail = f"""LS{' '*10}{cn_str}
UC{' '*3}{tc_str}{' '*5}{r_str}
*
ZZ
"""
        
        # 파일 쓰기 (Windows 줄바꿈)
        with open(dat_path, 'w', encoding='utf-8', newline='') as f:
            f.write(header.replace('\n', '\r\n'))
            f.write(variables.replace('\n', '\r\n'))
            for pc_line in pc_lines:
                f.write(pc_line + '\r\n')
            f.write(tail.replace('\n', '\r\n'))

    def save_excel_with_formulas(self, excel_path):
        """
        수식 기반 엑셀 파일 생성 - 수정본
        - Raw Data: 파라미터 H2~I11, S/Ia는 I10/I11에 수식으로 입력
        - Clark UH Derivation: 파라미터 I2~J7, A~G에 UH 표
        - 음수 Inc Excess 문제 해결
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            red_font = Font(color="FF0000")
            black_font = Font(color="000000")

            res = self.results
            details = self.clark_details
            params = self.params

            nmin = params['NMIN']
            nq = params['NQ']

            # ===== Sheet 1: Raw Data =====
            ws_raw = wb.active
            ws_raw.title = "Raw Data"
            
            for col_idx in range(1, 13):  # 1=A, 7=G
                col_letter = get_column_letter(col_idx)
                ws_raw.column_dimensions[col_letter].width = 14
                
            for row_idx in range(1, ws_raw.max_row + 1):
                ws_raw.row_dimensions[row_idx].height = 20

            headers = ["Time (min)", "Inc Rain", "Inc Excess", "Cum Rain", "Cum Excess"]
            for col, h in enumerate(headers, 1):
                cell = ws_raw.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            # 파라미터 헤더 (H1, I1)
            param_header_h = ws_raw.cell(row=1, column=8, value="PARAMETER")
            param_header_h.font = header_font
            param_header_h.fill = header_fill
            
            param_header_i = ws_raw.cell(row=1, column=9, value="VALUES")
            param_header_i.font = header_font
            param_header_i.fill = header_fill

            # 사용자 입력 파라미터 (빨간색) vs 계산값 (검은색)
            param_list = [
                ("Area (km2)", params['AREA_KM2'], True),
                ("Total Precip (mm)", params['TOTAL_PRECIP'], True),
                ("CN", params['CN'], True),
                ("Tc (hr)", params['TC'], True),
                ("R (hr)", params['R'], True),
                ("dt (min)", params['NMIN'], True),
                ("NQ", params['NQ'], True),
                ("TR (min)", params['TR_MIN'], True),
                ("S (mm)", "=(25400/I4)-254", False),  # 계산값
                ("Ia (mm)", "=0.2*I10", False),        # 계산값
            ]

            for i, (label, value, is_user_input) in enumerate(param_list, 2):
                label_cell = ws_raw.cell(row=i, column=8, value=label)
                
                value_cell = ws_raw.cell(row=i, column=9, value=value)
                if is_user_input:
                    value_cell.font = red_font
                else:
                    value_cell.font = black_font

            # Huff Input 헤더 (K1, L1)
            huff_header_k = ws_raw.cell(row=1, column=11, value="Huff Input")
            huff_header_k.font = header_font
            huff_header_k.fill = header_fill
            
            huff_header_l = ws_raw.cell(row=1, column=12, value="Interpolated PC")
            huff_header_l.font = header_font
            huff_header_l.fill = header_fill

            # Huff Input 데이터 (K2 이후) - 사용자 입력값 빨간색
            pc_list = params['PC']
            for i, pc_val in enumerate(pc_list):
                cell = ws_raw.cell(row=i+2, column=11, value=pc_val)
                cell.font = red_font

            # Interpolated PC 데이터 (L2 이후) - 계산값 검은색
            pc_interp = res.get('pc_interp')
            if pc_interp is not None:
                for i, val in enumerate(pc_interp):
                    row = i + 2
                    ws_raw.cell(row=row, column=1, value=i * nmin)
                    cell = ws_raw.cell(row=row, column=12, value=float(val))
                    cell.font = black_font
            else:
                last_pc = pc_list[-1] if pc_list else 1.0
                for i in range(nq+1):
                    row = i + 2
                    v = pc_list[i] if i < len(pc_list) else last_pc
                    ws_raw.cell(row=row, column=1, value=i * nmin)
                    cell = ws_raw.cell(row=row, column=12, value=v)
                    cell.font = black_font

            # Cum Rain / Inc Rain / Cum Excess / Inc Excess
            # 파라미터 위치: H2~I11
            # I3: Total Precip
            # I10: S = (25400/I3)-254
            # I11: Ia = 0.2*I10
            for i in range(nq):
                r = i + 2
                
                # Cum Rain = Interpolated PC * Total Precip (I3)
                cell_d = ws_raw.cell(row=r, column=4, value=f"=L{r}*$I$3")
                cell_d.font = black_font
                
                # Inc Rain
                cell_b = ws_raw.cell(
                    row=r, column=2, 
                    value=f"=D{r}-D{r-1}" if i > 0 else f"=D{r}"
                )
                cell_b.font = black_font
                
                # Cum Excess = IF(D{r}>Ia, (D{r}-Ia)^2/(D{r}-Ia+S), 0)
                # = IF(D{r}>$I$11, POWER(D{r}-$I$11,2)/(D{r}-$I$11+$I$10), 0)
                cell_e = ws_raw.cell(
                    row=r, column=5,
                    value=(
                        f"=IF(D{r}>$I$11,"
                        f" POWER(D{r}-$I$11,2)/(D{r}-$I$11+$I$10), 0)"
                    ),
                )
                cell_e.font = black_font
                
                # Inc Excess = E{r} - E{r-1}
                cell_c = ws_raw.cell(
                    row=r, column=3, 
                    value=f"=E{r}-E{r-1}" if i > 0 else "=E2"
                )
                cell_c.font = black_font

            # ===== Sheet 2: Clark UH Derivation =====
            ws_clark = wb.create_sheet("Clark UH Derivation")
            
            denom = int(60 / nmin)
            fmt = f'? ?/{denom}'     # 예: '? ?/6'
            cell = ws_clark.cell(row=7, column=10)
            cell.number_format = fmt
            
            for col_idx in range(1, 13):  # 1=A, 7=G
                col_letter = get_column_letter(col_idx)
                ws_clark.column_dimensions[col_letter].width = 14
            for row_idx in range(1, ws_raw.max_row + 1):
                ws_clark.row_dimensions[row_idx].height = 20
            
            # UH 헤더 (A1~G1)
            uh_headers = ["Time (hr)", "T/Tc", "AI", "dAI", "I", "O", "UH"]
            for col, h in enumerate(uh_headers, 1):
                cell = ws_clark.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            # 파라미터 헤더 (I1, J1)
            param_header_i = ws_clark.cell(row=1, column=9, value="PARAMETER")
            param_header_i.font = header_font
            param_header_i.fill = header_fill
            
            param_header_j = ws_clark.cell(row=1, column=10, value="VALUES")
            param_header_j.font = header_font
            param_header_j.fill = header_fill

            # 파라미터 데이터 (I2~J7)
            # 사용자 입력값: Tc, R, Area → 빨간색
            # 계산값: CA, CB, dt_hr → 검은색
            dt_hr_val = params['NMIN'] / 60.0
            clark_params = [
                ("CA", "=$J$7/($J$5+0.5*$J$7)", False),  # J6
                ("CB", "=1-$J$2", False),                # J7
                ("Tc", params['TC'], True),          # J3
                ("R", params['R'], True),            # J4
                ("Area", params['AREA_KM2'], True),  # J5
                ("dt_hr", dt_hr_val, False),          # J2
            ]
            
            for i, (label, value, is_user_input) in enumerate(clark_params, 2):
                ws_clark.cell(row=i, column=9, value=label)
                value_cell = ws_clark.cell(row=i, column=10, value=value)
                value_cell.font = red_font if is_user_input else black_font

            # UH 데이터 (행 2부터 시작)
            for i, _ in enumerate(details['Time_hr']):
                r = i + 2
                
                # Time (hr) = (r-2)*dt_hr = (r-2)*J7
                ws_clark.cell(row=r, column=1, value=f"=({r}-2)*$J$7")
                ws_clark.cell(row=r, column=1).font = black_font
                
                # T/Tc = A{r} / Tc = A{r} / J4
                ws_clark.cell(row=r, column=2, value=f"=A{r}/$J$4")
                ws_clark.cell(row=r, column=2).font = black_font
                
                # AI (Clark AI formula)
                ws_clark.cell(
                    row=r, column=3,
                    value=(
                        f"=IF(B{r}>=1, 1,"
                        f" IF(B{r}<0.5, 1.414*POWER(B{r},1.5),"
                        f" 1-1.414*POWER(1-B{r},1.5)))"
                    ),
                )
                ws_clark.cell(row=r, column=3).font = black_font
                
                if i == 0:
                    # 첫 행: dAI, I, O, UH = 0
                    for c in range(4, 8):
                        ws_clark.cell(row=r, column=c, value=0)
                        ws_clark.cell(row=r, column=c).font = black_font
                else:
                    # dAI = C{r} - C{r-1}
                    ws_clark.cell(row=r, column=4, value=f"=C{r}-C{r-1}")
                    ws_clark.cell(row=r, column=4).font = black_font
                    
                    # I = MAX(dAI, 0) * Area * 1000 / (dt_hr * 3600)
                    # = MAX(D{r}, 0) * J6 * 1000 / (J8 * 3600)
                    ws_clark.cell(
                        row=r, column=5,
                        value=f"=MAX(D{r},0)*$J$6*1000/($J$7*3600)",
                    )
                    ws_clark.cell(row=r, column=5).font = black_font
                    
                    # O = CA * I{r} + CB * O{r-1}
                    # = J2 * E{r} + J3 * F{r-1}
                    ws_clark.cell(
                        row=r, column=6,
                        value=f"=$J$2*E{r}+$J$3*F{r-1}",
                    )
                    ws_clark.cell(row=r, column=6).font = black_font
                    
                    # UH = 0.5 * (O{r} + O{r-1})
                    ws_clark.cell(
                        row=r, column=7,
                        value=f"=0.5*(F{r}+F{r-1})",
                    )
                    ws_clark.cell(row=r, column=7).font = black_font
                    
                    
            col_idx = 1  # C열
            for row_idx in range(2, ws_clark.max_row + 1):  # 헤더(1행) 제외
                cell = ws_clark.cell(row=row_idx, column=col_idx)
                cell.number_format = fmt
                
                
            # ===== Sheet 3: Convolution Detail =====
            ws_conv = wb.create_sheet("Convolution Detail")

            ws_conv.cell(row=1, column=1, value="Time").font = header_font
            ws_conv.cell(row=1, column=2, value="Total Runoff (CMS)").font = header_font
            ws_conv.cell(row=1, column=1).fill = header_fill
            ws_conv.cell(row=1, column=2).fill = header_fill

            valid_idx = list(range(len(res['excess_inc'])))

            for idx, v_idx in enumerate(valid_idx):
                col = 3 + idx
                ws_conv.cell(row=1, column=col, value=f"P_{v_idx+1}").font = header_font

            def col_index_to_letter(idx: int) -> str:
                s = ""
                while idx > 0:
                    idx, rem = divmod(idx - 1, 26)
                    s = chr(65 + rem) + s
                return s

            for r_idx in range(nq):
                r_num = r_idx + 2
                ws_conv.cell(row=r_num, column=1, value=f"='Raw Data'!A{r_num}")

                for c_idx, v_idx in enumerate(valid_idx):
                    uh_row = r_idx - v_idx
                    if 0 <= uh_row < len(details['UH']) - 1:
                        col = 3 + c_idx
                        ws_conv.cell(
                            row=r_num,
                            column=col,
                            value=(
                                f"='Raw Data'!$C${v_idx+2}"
                                f"*'Clark UH Derivation'!$G${2+uh_row}"
                            ),
                        )

                if len(valid_idx) > 0:
                    last_col_letter = col_index_to_letter(3 + len(valid_idx) - 1)
                    ws_conv.cell(
                        row=r_num,
                        column=2,
                        value=f"=SUM(C{r_num}:{last_col_letter}{r_num})",
                    )
                else:
                    ws_conv.cell(row=r_num, column=2, value=0)

            # ===== Sheet 4: Summary =====
            ws_sum = wb.create_sheet("Summary")
            
            for col_idx in range(1, 4):  # 1=A, 7=G
                col_letter = get_column_letter(col_idx)
                ws_sum.column_dimensions[col_letter].width = 14
            for row_idx in range(1, ws_raw.max_row + 1):
                ws_sum.row_dimensions[row_idx].height = 20
                
            sum_headers = ["Time (min)", "Inc Rain", "Inc Excess", "Runoff (CMS)"]
            for col, h in enumerate(sum_headers, 1):
                cell = ws_sum.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            for i in range(nq):
                r = i + 2
                ws_sum.cell(row=r, column=1, value=f"='Raw Data'!A{r}")
                ws_sum.cell(row=r, column=2, value=f"='Raw Data'!B{r}")
                ws_sum.cell(row=r, column=3, value=f"='Raw Data'!C{r}")
                ws_sum.cell(row=r, column=4, value=f"='Convolution Detail'!B{r}")

            wb.save(excel_path)

        except Exception as e:
            raise Exception(f"엑셀 저장 중 오류: {str(e)}")

# ============================================================================
# GUI Application
# ============================================================================

class HEC1GUIApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Clark 합성단위도 홍수량 산정")
        self.root.geometry("1100x720")
        
        self.engine = HEC1Engine()
        self.fig = None
        self.canvas_widget = None

        plt.rcParams['font.family'] = 'Malgun Gothic'
        plt.rcParams['axes.unicode_minus'] = False
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Main Container
        self.main_container = tk.Frame(root, bg="white")
        self.main_container.pack(fill=tk.BOTH, expand=True)

        # === LEFT PANEL ===
        left_panel = tk.Frame(self.main_container, width=350, bg="#f0f0f0", padx=10, pady=10)
        left_panel.pack(side=tk.LEFT, fill=tk.Y)
        left_panel.pack_propagate(False)

        # Title
        tk.Label(left_panel, text="[ 입력 매개변수 ]", font=("맑은 고딕", 12, "bold"), bg="#f0f0f0").pack(pady=(0, 10))

        # Input Fields
        self.entries = {}
        params = [
            ("지점명 (Station ID)", "080800"),
            ("재현기간(년)", "50"),
            ("강우지속기간(분)", "100"),
            ("계산시간간격(IN, 분)", "10"),
            ("계산 횟수 (NQ)", "300"),
            ("유역 면적 (km2)", "5.4"),
            ("총 강우량 (mm)", "110.4"),
            ("CN (SCS)", "84.3"),
            ("도달시간 Tc (hr)", "0.40"),
            ("저류상수 R (hr)", "0.57")
        ]
        
        for label, default in params:
            row_frame = tk.Frame(left_panel, bg="#f0f0f0")
            row_frame.pack(fill=tk.X, pady=2)
            tk.Label(row_frame, text=label, width=22, anchor='w', bg="#f0f0f0", font=("Arial", 9)).pack(side=tk.LEFT)
            ent = tk.Entry(row_frame, font=("Arial", 9), width=12, justify='right')
            ent.insert(0, default)
            ent.pack(side=tk.RIGHT)
            self.entries[label] = ent

        # PC Data Input
        tk.Label(left_panel, text="Huff 강우시간분포 누적 백분율", justify='left', anchor='w', bg="#f0f0f0", font=("Arial", 9)).pack(fill=tk.X, pady=(15,2))
        self.txt_pc = tk.Text(left_panel, height=6, width=30, font=("Consolas", 8))
        self.txt_pc.pack(fill=tk.X)
        default_pc = ".000    .009    .053    .101    .184    .331    .538    .756    .916    .975   1.000"
        self.txt_pc.insert(1.0, default_pc)

        # Buttons
        self.btn_run = tk.Button(left_panel, text="분석 실행 (Run)", command=self.run_analysis, 
                              bg="#007bff", fg="white", font=("맑은 고딕", 11, "bold"), height=2)
        self.btn_run.pack(fill=tk.X, pady=(15, 5))

        self.btn_save = tk.Button(left_panel, text="결과 파일 저장", command=self.save_all_files, 
                                state=tk.DISABLED, bg="#28a745", fg="white", font=("맑은 고딕", 10))
        self.btn_save.pack(fill=tk.X, pady=2)

        # [결과 요약]
        tk.Label(left_panel, text="[ 3. 분석 결과 ]", font=("맑은 고딕", 11, "bold"), bg="#f0f0f0").pack(pady=(15, 2))
        
        self.txt_result = tk.Text(left_panel, height=4, width=30, bg="white", relief="solid", bd=1, font=("맑은 고딕", 10))
        self.txt_result.pack(fill=tk.X)
        self.txt_result.insert(1.0, "대기 중...")
        self.txt_result.config(state=tk.DISABLED)

        # === RIGHT PANEL (Graph) ===
        self.right_panel = tk.Frame(self.main_container, bg="white", bd=1, relief="solid")
        self.right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

    def on_closing(self):
        """종료 처리"""
        try:
            if self.fig is not None:
                plt.close(self.fig)
            plt.close('all')
            if self.canvas_widget is not None:
                self.canvas_widget.destroy()
            self.root.quit()
            self.root.destroy()
        except:
            pass
        finally:
            sys.exit(0)

    def get_inputs(self):
        try:
            p = {k: self.entries[k].get() for k in self.entries}
            params = {
                'STATION_ID': p["지점명 (Station ID)"],
                'RETURN_PERIOD': int(p["재현기간(년)"]),
                'TR_MIN': int(p["강우지속기간(분)"]),
                'NMIN': int(p["계산시간간격(IN, 분)"]),
                'NQ': int(p["계산 횟수 (NQ)"]), 
                'AREA_KM2': float(p["유역 면적 (km2)"]),
                'TOTAL_PRECIP': float(p["총 강우량 (mm)"]), 
                'CN': float(p["CN (SCS)"]),
                'TC': float(p["도달시간 Tc (hr)"]), 
                'R': float(p["저류상수 R (hr)"])
            }
            
            # PC 값 파싱
            pc_text = self.txt_pc.get("1.0", tk.END)
            pc_values = re.split(r'[,\s\t]+', pc_text.strip())
            pc_list = []
            for val in pc_values:
                val = val.strip()
                if val:
                    try:
                        pc_list.append(float(val))
                    except ValueError:
                        pass
            
            if len(pc_list) == 0:
                messagebox.showwarning("입력 오류", "PC 값이 없습니다.")
                return None
            
            params['PC'] = pc_list
            return params
        except ValueError as e:
            messagebox.showerror("입력 오류", str(e))
            return None

    def run_analysis(self):
        params = self.get_inputs()
        if not params: return
        
        try:
            self.engine.set_params(params)
            self.engine.calculate(shift_step=0)
            res = self.engine.results
            
            summary = (
                f"첨두 유량 : {res['peak_flow']:.2f} cms\n"
                f"첨두발생시간: {res['peak_time_min']:.1f} min\n"
                f"유효강우량 : {np.sum(res['excess_inc']):.1f} mm"
            )
            
            self.txt_result.config(state=tk.NORMAL)
            self.txt_result.delete(1.0, tk.END)
            self.txt_result.insert(1.0, summary)
            self.txt_result.config(state=tk.DISABLED)

            self.plot_results()
            self.btn_save.config(state=tk.NORMAL)
            
        except Exception as e:
            messagebox.showerror("실행 오류", str(e))

    def plot_results(self):
        for widget in self.right_panel.winfo_children():
            widget.destroy()
        
        if self.fig is not None:
            plt.close(self.fig)
        
        res = self.engine.results
        
        # ---- 여기 추가: 0.00 이상 나오는 구간까지만 사용 ----
        flows = res['flow']
        times = res['time_min']
        rains = res['rain_inc']
        excess = res['excess_inc']

        end_idx = len(flows)
        for i in range(len(flows) - 1, -1, -1):
            if round(flows[i], 2) > 0.0:   # 소수 둘째 자리까지 0.00 초과
                end_idx = i + 1
                break

        plot_times = times[:end_idx]
        plot_flows = flows[:end_idx]
        plot_rains = rains[:end_idx]
        plot_excess = excess[:end_idx]   # ✅ 잘라낸 유효우량

        # ---------------------------------------------------
        
        self.fig, ax1 = plt.subplots(figsize=(8, 6), dpi=100)
        ax2 = ax1.twinx()

        ax2.bar(res['time_min'], res['rain_inc'], width=self.engine.params['NMIN'], 
                color='#aaccff', alpha=0.6, label='강우량 (Rainfall)', align='edge')
                
        ax2.bar(plot_times, plot_excess,
        width=self.engine.params['NMIN'],
        color='#0055ff', alpha=0.9,
        label='유효우량 (Excess)', align='edge')
        
        ax2.set_ylabel('강우량/유효우량 (mm)', color='blue')
        ax2.set_ylim(0, max(res['rain_inc'])*3 if max(res['rain_inc']) > 0 else 10) 
        ax2.invert_yaxis()
        ax1.plot(res['time_min'], res['flow'], color='#e74c3c', linewidth=2, label='유출량 (Runoff)')

        # ★ 여기 추가: X축을 잘린 구간까지만 고정
        ax1.set_xlim(plot_times[0], plot_times[-1])
        ax1.set_xlabel('시간 (분)', fontweight='bold')
        ax1.set_ylabel('유량 (cms)', color='#e74c3c', fontweight='bold')
        ax1.set_ylim(bottom=0)
        ax1.grid(True, linestyle='--', alpha=0.5)
        
        plt.title(f"수문곡선 (Hydrograph) - {self.engine.params['STATION_ID']}")
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc='center right')
        
        canvas = FigureCanvasTkAgg(self.fig, master=self.right_panel)
        canvas.draw()
        self.canvas_widget = canvas.get_tk_widget()
        self.canvas_widget.pack(fill=tk.BOTH, expand=True)

    def save_all_files(self):
        """DAT, OUT, XLSX 동시 저장 + 그래프 팝업"""
        folder = filedialog.askdirectory(title="결과 파일 저장 폴더 선택")
        if not folder:
            return
        
        try:
            # 파일명 구성: 재현기간(3자리)-강우지속기간(4자리)-지점명(6자리)_날짜_시각
            return_period = self.engine.params['RETURN_PERIOD']
            tr_min = self.engine.params['TR_MIN']
            station_id = self.engine.params['STATION_ID']
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 재현기간: 3자리 0 패딩
            rp_str = f"{return_period:03d}"
            # 강우지속기간: 4자리 0 패딩
            tr_str = f"{tr_min:04d}"
            # 지점명: 6자리 (부족하면 뒤에 0 패딩, 넘으면 앞 6자리만)
            station_str = f"{station_id:<6s}"[:6].replace(' ', '0')
            
            base_filename = f"{rp_str}-{tr_str}-{station_str}_{timestamp}"
            
            # 파일 경로 생성
            dat_path = os.path.join(folder, f"{base_filename}.DAT")
            out_path = os.path.join(folder, f"{base_filename}.OUT")
            xlsx_path = os.path.join(folder, f"{base_filename}.xlsx")
            png_path = os.path.join(folder, f"{base_filename}.png")
            
            # 파일 저장
            self.engine.write_hec1_dat(dat_path)
            self.engine.write_hec1_out(out_path)
            self.engine.save_excel_with_formulas(xlsx_path)
            
            # 그래프 저장
            if self.fig:
                self.fig.savefig(png_path, dpi=150, bbox_inches='tight')
            
            # 그래프 팝업
            if self.fig:
                popup = tk.Toplevel(self.root)
                popup.title("수문곡선")
                popup.geometry("900x700")
                
                canvas = FigureCanvasTkAgg(self.fig, master=popup)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            
            messagebox.showinfo("저장 완료", 
                f"모든 파일이 저장되었습니다!\n\n"
                f"📁 위치: {folder}\n"
                f"• {os.path.basename(dat_path)}\n"
                f"• {os.path.basename(out_path)}\n"
                f"• {os.path.basename(xlsx_path)}\n"
                f"• {os.path.basename(png_path)}")
            
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = HEC1GUIApp(root)
        root.mainloop()
    except KeyboardInterrupt:
        print("\n프로그램이 사용자에 의해 중단되었습니다.")
        sys.exit(0)
    except Exception as e:
        print(f"오류 발생: {e}")
        traceback.print_exc()
    finally:
        plt.close('all')
        sys.exit(0)