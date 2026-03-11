"""
05.홍수량산정.py  —  Clark / SCS / Nakayasu 합성단위도 홍수량 비교 산정
Hydro Analysis System  Module 5
"""

import os
import sys
import json
import re
import traceback
import warnings
import numpy as np
from datetime import datetime
from scipy.interpolate import interp1d, PchipInterpolator, CubicSpline
from ctypes import windll, byref, sizeof, c_int

import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog

import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=RuntimeWarning)

# ─── CustomTkinter 전역 설정 ──────────────────────────────────────────────────
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

FONT_TITLE  = ("맑은 고딕", 20, "bold")
FONT_HEADER = ("맑은 고딕", 13, "bold")
FONT_BODY   = ("맑은 고딕", 11)
FONT_BTN    = ("맑은 고딕", 11, "bold")
FONT_SMALL  = ("맑은 고딕", 10)
FONT_LOG    = ("Consolas", 10)

# ─── DB / 배치계산 상수 ────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(__file__), 'rainfall_db.sqlite')
# RETURN_PERIOD / TR_MIN ComboBox 초기값 — Step 3 파일 로드 후 동적으로 갱신됨
COMBOBOX_OPTIONS = {
    'RETURN_PERIOD': ["전체"],   # _load_prob_rainfall_options() 에서 동적 업데이트
    'TR_MIN':        ["전체"],   # _load_prob_rainfall_options() 에서 동적 업데이트
    'NQ':            ["100", "200", "300"],
}

# ─── Huff 분위 기본값 ──────────────────────────────────────────────────────────
HUFF_PRESETS = {
    "1분위": [0.000, 0.130, 0.262, 0.415, 0.546, 0.649, 0.726, 0.795, 0.861, 0.931, 1.000],
    "2분위": [0.000, 0.009, 0.053, 0.101, 0.184, 0.331, 0.538, 0.756, 0.916, 0.975, 1.000],
    "3분위": [0.000, 0.008, 0.041, 0.086, 0.154, 0.263, 0.437, 0.636, 0.833, 0.953, 1.000],
    "4분위": [0.000, 0.007, 0.030, 0.062, 0.110, 0.186, 0.310, 0.492, 0.700, 0.892, 1.000],
    "직접 입력": None,
}

METHOD_COLORS = {
    "Clark":    "#e74c3c",
    "SCS":      "#2ecc71",
    "Nakayasu": "#3498db",
}


# ============================================================================
# 계산 엔진
# ============================================================================

class RainfallRunoffEngine:
    """강우-유출 해석 엔진 (Clark / SCS / Nakayasu)"""

    def __init__(self):
        self.rainfall_data: dict   = {}
        self.uh_data:       dict   = {}
        self.runoff_data:   dict   = {}

    # ── 유효우량 (SCS-CN + Huff) ─────────────────────────────────────────────

    def calculate_effective_rainfall(self, total_precip, tr_min, dt_min, cn, huff_pc,
                                      interp_method='pchip', pc_rounding=True):
        pc_huff = np.array(huff_pc, dtype=float)
        n_step  = int(tr_min / dt_min) + 1
        t_norm  = np.linspace(0.0, 1.0, n_step)
        t_huff  = np.linspace(0.0, 1.0, len(pc_huff))

        if interp_method == 'pchip':
            pc_interp = PchipInterpolator(t_huff, pc_huff)(t_norm)
            pc_interp = np.clip(pc_interp, 0.0, 1.0)
        elif interp_method == 'cubic':
            pc_interp = CubicSpline(t_huff, pc_huff)(t_norm)
            pc_interp = np.clip(pc_interp, 0.0, 1.0)
        else:  # 'linear'
            pc_interp = np.interp(t_norm, t_huff, pc_huff)

        if pc_rounding:
            pc_interp = np.floor(pc_interp * 1000.0 + 0.5) / 1000.0

        cum_rain = pc_interp * total_precip

        S  = (25400.0 / cn) - 254.0
        Ia = 0.2 * S

        cum_excess = np.zeros(n_step)
        for i, P in enumerate(cum_rain):
            if P > Ia:
                cum_excess[i] = (P - Ia) ** 2 / (P - Ia + S)

        rain_inc   = np.diff(cum_rain,   prepend=0.0)
        excess_inc = np.diff(cum_excess, prepend=0.0)
        time_min   = np.arange(n_step) * dt_min

        self.rainfall_data = {
            'time_min':   time_min,
            'rain_inc':   rain_inc,
            'excess_inc': excess_inc,
            'cum_rain':   cum_rain,
            'cum_excess': cum_excess,
            'pc_interp':     pc_interp,
            'huff_pc':       huff_pc,
            'tr_min':        tr_min,
            'dt_min':        dt_min,
            'S':             S,
            'Ia':            Ia,
            'interp_method': interp_method,
            'pc_rounding':   pc_rounding,
        }
        return time_min, rain_inc, excess_inc

    # ── Clark 단위도 (HEC-1 방식) ─────────────────────────────────────────────
    # 규약: target_vol = A*10000 → 10mm 기준 정규화 → 합성곱 시 /10 사용

    def get_clark_uh(self, A, Tc, R, dt_hr):
        uh_len = int((Tc + R * 10.0) / dt_hr) + 5
        t_vals = np.arange(uh_len) * dt_hr

        ai = np.zeros(uh_len)
        for i, t in enumerate(t_vals):
            T = t / Tc if Tc > 0 else 1.0
            if T >= 1.0:
                ai[i] = 1.0
            elif T < 0.5:
                ai[i] = 1.414 * T ** 1.5
            else:
                ai[i] = 1.0 - 1.414 * (1.0 - T) ** 1.5

        vol_const = A * 1000.0
        I_flow    = np.zeros(uh_len)
        for i in range(1, uh_len):
            dAI = max(0.0, ai[i] - ai[i - 1])
            I_flow[i] = dAI * vol_const / (dt_hr * 3600.0)

        c_ca   = dt_hr / (R + 0.5 * dt_hr)
        c_cb   = 1.0 - c_ca
        O_inst = np.zeros(uh_len)
        for i in range(1, uh_len):
            O_inst[i] = c_ca * I_flow[i] + c_cb * O_inst[i - 1]

        uh = np.zeros(uh_len)
        for i in range(1, uh_len):
            uh[i] = 0.5 * (O_inst[i] + O_inst[i - 1])

        # SCS/Nakayasu 와 동일한 10mm 기준으로 정규화
        calc_vol   = np.sum(uh) * dt_hr * 3600.0
        target_vol = A * 10000.0
        if calc_vol > 0:
            uh *= target_vol / calc_vol

        self.uh_data['clark'] = {
            'time_hr': t_vals, 'uh': uh,
            'AI': ai, 'I': I_flow, 'O': O_inst,
            'Tc': Tc, 'R': R, 'CA': c_ca, 'CB': c_cb,
        }
        return uh

    # ── SCS 합성단위도 ────────────────────────────────────────────────────────

    def get_scs_uh(self, A, Tc, tr_hr, dt_hr):
        tp = 0.6 * Tc
        Tp = tr_hr / 2.0 + tp
        Qp = (2.08 * A) / Tp if Tp > 0 else 0.0

        scs_ratios = np.array([
            [0.0,0.0],[0.1,0.03],[0.2,0.10],[0.3,0.19],[0.4,0.31],
            [0.5,0.47],[0.6,0.66],[0.7,0.82],[0.8,0.93],[0.9,0.99],
            [1.0,1.00],[1.1,0.99],[1.2,0.93],[1.3,0.86],[1.4,0.78],
            [1.5,0.68],[1.6,0.56],[1.7,0.46],[1.8,0.39],[1.9,0.33],
            [2.0,0.28],[2.2,0.207],[2.4,0.147],[2.6,0.107],[2.8,0.077],
            [3.0,0.055],[3.2,0.04],[3.4,0.029],[3.6,0.021],[3.8,0.015],
            [4.0,0.011],[4.5,0.005],[5.0,0.0],
        ])
        t_real = scs_ratios[:, 0] * Tp
        q_real = scs_ratios[:, 1] * Qp
        f      = interp1d(t_real, q_real, kind='linear', bounds_error=False, fill_value=0.0)
        time_axis = np.arange(0, t_real[-1] + dt_hr, dt_hr)
        uh        = f(time_axis)

        # 10mm 기준 정규화
        vol_m3   = np.sum(uh) * dt_hr * 3600.0
        depth_m  = vol_m3 / (A * 1e6) if A > 0 else 1.0
        if depth_m > 0:
            uh *= 0.01 / depth_m

        self.uh_data['scs'] = {
            'time_hr': time_axis, 'uh': uh,
            'Tc': Tc, 'tp': tp, 'Tp': Tp, 'Qp': Qp,
            'scs_ratios': scs_ratios,
        }
        return uh

    # ── Nakayasu 합성단위도 ───────────────────────────────────────────────────

    def get_nakayasu_uh(self, A, L, tr_hr, dt_hr, t03_method='alpha', alpha=1.5):
        tg = 0.4 + 0.058 * L if L > 15 else 0.21 * (L ** 0.7)
        Tp = tg + 0.8 * tr_hr
        T03 = alpha * tg if t03_method.lower() == 'alpha' else 0.47 * (A * L) ** 0.25
        Qp  = (A * 10.0) / (3.6 * (0.3 * Tp + T03)) if (0.3 * Tp + T03) > 0 else 0.0

        max_time  = Tp + 5.0 * T03
        time_axis = np.arange(0, max_time + dt_hr, dt_hr)
        ratio     = np.where(T03 > 0, (time_axis - Tp) / T03, 0.0)

        uh = np.where(
            time_axis <= Tp,
            Qp * (time_axis / Tp) ** 2.4 if Tp > 0 else 0.0,
            np.where(
                ratio <= 1.0,
                Qp * (0.3 ** ratio),
                np.where(
                    ratio <= 2.5,
                    Qp * (0.3 ** ((time_axis - Tp + 0.5 * T03) / (1.5 * T03))),
                    Qp * (0.3 ** ((time_axis - Tp + 1.5 * T03) / (2.0 * T03))),
                ),
            ),
        )
        uh = np.where(np.isnan(uh) | np.isinf(uh), 0.0, uh)

        # 10mm 기준 정규화
        vol_m3  = np.sum(uh) * dt_hr * 3600.0
        depth_m = vol_m3 / (A * 1e6) if A > 0 else 1.0
        if depth_m > 0:
            uh *= 0.01 / depth_m

        self.uh_data['nakayasu'] = {
            'time_hr': time_axis, 'uh': uh,
            'tg': tg, 'Tp': Tp, 'T03': T03, 'Qp': Qp,
            't03_method': t03_method, 'alpha': alpha,
        }
        return uh

    # ── Convolution (모든 방법 공통: /10) ────────────────────────────────────

    def convolve_runoff(self, method: str, method_params: dict):
        if not self.rainfall_data:
            raise ValueError("먼저 calculate_effective_rainfall()를 호출하세요.")

        excess_inc = self.rainfall_data['excess_inc']
        dt_min     = self.rainfall_data['dt_min']
        tr_min     = self.rainfall_data['tr_min']
        dt_hr      = dt_min / 60.0
        tr_hr      = tr_min / 60.0
        A          = method_params['A']

        if method == 'Clark':
            uh = self.get_clark_uh(A, method_params['Tc'], method_params['R'], dt_hr)
        elif method == 'SCS':
            uh = self.get_scs_uh(A, method_params['Tc'], tr_hr, dt_hr)
        elif method == 'Nakayasu':
            uh = self.get_nakayasu_uh(
                A, method_params['L'], tr_hr, dt_hr,
                method_params.get('t03_method', 'alpha'),
                method_params.get('alpha', 1.5),
            )
        else:
            raise ValueError(f"지원하지 않는 방법: {method}")

        full_conv = np.convolve(excess_inc, uh) / 10.0

        # trailing 0 제거
        threshold = 0.001
        end_idx   = len(full_conv)
        for i in range(len(full_conv) - 1, -1, -1):
            if full_conv[i] > threshold:
                end_idx = min(i + 10, len(full_conv))
                break

        runoff       = full_conv[:end_idx]
        time_min_out = np.arange(len(runoff)) * dt_min

        self.runoff_data[method] = {
            'time_min': time_min_out,
            'runoff':   runoff,
        }
        return time_min_out, runoff

    # ── 물수지 검토 ──────────────────────────────────────────────────────────
    # 이론: sum(q_full) × dt_sec / A_m2 × 1000 = sum(excess_inc) [mm]
    # 실제 차이 원인: ① pc_rounding에 의한 누가값 반올림 오차
    #                ② convolve_runoff trailing-zero 절삭

    def check_water_balance(self, area_km2: float, dt_min: float) -> dict:
        """유효우량 총량 vs 유출수문곡선 체적 물수지 검토.

        Returns
        -------
        dict:
            Pe_mm   : 총 유효우량 [mm]
            rain_mm : 총 강우량 [mm]
            methods : { 'Clark'|'SCS'|'Nakayasu' :
                         { V_mm, diff_mm, diff_pct } }
        """
        dt_sec = dt_min * 60.0
        A_m2   = area_km2 * 1e6

        rain_mm = float(np.sum(self.rainfall_data['rain_inc']))
        Pe_mm   = float(np.sum(self.rainfall_data['excess_inc']))

        methods = {}
        for method, rdata in self.runoff_data.items():
            V_m3     = float(np.sum(rdata['runoff'])) * dt_sec   # [m³]
            V_mm     = V_m3 / A_m2 * 1000.0                      # [mm]
            diff_mm  = Pe_mm - V_mm
            diff_pct = diff_mm / Pe_mm * 100.0 if Pe_mm > 0 else 0.0
            methods[method] = {
                'V_mm':     V_mm,
                'diff_mm':  diff_mm,
                'diff_pct': diff_pct,
            }

        return {'rain_mm': rain_mm, 'Pe_mm': Pe_mm, 'methods': methods}

    # ── Excel 저장 ────────────────────────────────────────────────────────────

    def save_excel(self, xlsx_path: str, values: dict):
        wb  = Workbook()
        hfnt  = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="4F81BD")
        rfnt  = Font(color="CC0000", bold=True)
        bfnt  = Font(color="000000")

        def _hrow(ws, headers, row=1, start_col=1):
            for ci, h in enumerate(headers, start_col):
                c = ws.cell(row=row, column=ci, value=h)
                c.font, c.fill = hfnt, hfill

        def _set_col_widths(ws, width=14, cols=8):
            for c in range(1, cols + 1):
                ws.column_dimensions[get_column_letter(c)].width = width

        # ── Sheet 1: Raw Data ────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Raw Data"
        _set_col_widths(ws1, 14, 8)
        _hrow(ws1, ["Time (min)", "Cum Rain (mm)", "Inc Rain (mm)",
                    "Cum Excess (mm)", "Inc Excess (mm)", "", "PARAM", "VALUE"])

        rd = self.rainfall_data
        for i in range(len(rd['time_min'])):
            r = i + 2
            ws1.cell(row=r, column=1, value=round(float(rd['time_min'][i]), 2))
            ws1.cell(row=r, column=2, value=round(float(rd['cum_rain'][i]),   4))
            ws1.cell(row=r, column=3, value=round(float(rd['rain_inc'][i]),   4))
            ws1.cell(row=r, column=4, value=round(float(rd['cum_excess'][i]), 4))
            ws1.cell(row=r, column=5, value=round(float(rd['excess_inc'][i]), 4))

        params_list = [
            ("Area (km²)",       values['AREA_KM2']),
            ("Total Precip (mm)",values['TOTAL_PRECIP']),
            ("CN",               values['CN']),
            ("dt (min)",         values['DT_MIN']),
            ("TR (min)",         values['TR_MIN']),
            ("S (mm)",           round(rd['S'],  3)),
            ("Ia (mm)",          round(rd['Ia'], 3)),
            ("Total Excess (mm)",round(float(np.sum(rd['excess_inc'])), 3)),
            ("Huff 보간 방법",    rd.get('interp_method', '-')),
            ("소수점 처리",       "3째자리 반올림" if rd.get('pc_rounding') else "반올림 없음"),
        ]
        for pi, (lbl, val) in enumerate(params_list, 2):
            ws1.cell(row=pi, column=7, value=lbl)
            ws1.cell(row=pi, column=8, value=val).font = rfnt

        # Huff Input / Interpolated PC (columns J, K)
        ws1.cell(row=1, column=10, value="Huff Input").font  = hfnt
        ws1.cell(row=1, column=10).fill = hfill
        ws1.cell(row=1, column=11, value="Interp PC").font   = hfnt
        ws1.cell(row=1, column=11).fill = hfill
        for hi, v in enumerate(rd['huff_pc']):
            ws1.cell(row=hi + 2, column=10, value=round(v, 4)).font = rfnt
        for pi, v in enumerate(rd['pc_interp']):
            ws1.cell(row=pi + 2, column=11, value=round(float(v), 4))

        # ── Sheet 2: Clark UH ────────────────────────────────────────────────
        ws2 = wb.create_sheet("Clark UH")
        _set_col_widths(ws2, 14, 10)
        _hrow(ws2, ["Time (hr)", "AI", "dAI", "I (cms)", "O (cms)", "UH (cms/10mm)"])
        _hrow(ws2, ["PARAM", "VALUE"], row=1, start_col=8)

        if 'clark' in self.uh_data:
            cd = self.uh_data['clark']
            for i, t in enumerate(cd['time_hr']):
                r = i + 2
                ws2.cell(row=r, column=1, value=round(float(t),       4))
                ws2.cell(row=r, column=2, value=round(float(cd['AI'][i]), 5))
                dAI = float(cd['AI'][i] - cd['AI'][i-1]) if i > 0 else 0.0
                ws2.cell(row=r, column=3, value=round(max(0.0, dAI),  5))
                ws2.cell(row=r, column=4, value=round(float(cd['I'][i]),  4))
                ws2.cell(row=r, column=5, value=round(float(cd['O'][i]),  4))
                ws2.cell(row=r, column=6, value=round(float(cd['uh'][i]), 5))

            clark_params = [
                ("Tc (hr)",   cd['Tc']),
                ("R (hr)",    cd['R']),
                ("CA",        round(cd['CA'], 5)),
                ("CB",        round(cd['CB'], 5)),
                ("dt (hr)",   round(values['DT_MIN']/60.0, 5)),
                ("Area (km²)",values['AREA_KM2']),
            ]
            for pi, (lbl, val) in enumerate(clark_params, 2):
                ws2.cell(row=pi, column=8, value=lbl)
                ws2.cell(row=pi, column=9, value=val).font = rfnt

        # ── Sheet 3: SCS UH ──────────────────────────────────────────────────
        ws3 = wb.create_sheet("SCS UH")
        _set_col_widths(ws3, 14, 10)
        _hrow(ws3, ["Time (hr)", "t/Tp", "q/Qp", "UH (cms/10mm)"])
        _hrow(ws3, ["PARAM", "VALUE"], row=1, start_col=6)

        if 'scs' in self.uh_data:
            sd = self.uh_data['scs']
            Tp_s = sd['Tp'] if sd['Tp'] > 0 else 1.0
            Qp_s = sd['Qp'] if sd['Qp'] > 0 else 1.0
            for i, t in enumerate(sd['time_hr']):
                r = i + 2
                ws3.cell(row=r, column=1, value=round(float(t),             4))
                ws3.cell(row=r, column=2, value=round(float(t / Tp_s),      4))
                ws3.cell(row=r, column=3, value=round(float(sd['uh'][i]/Qp_s), 5))
                ws3.cell(row=r, column=4, value=round(float(sd['uh'][i]),   5))

            scs_params = [
                ("Area (km²)", values['AREA_KM2']),
                ("Tc (hr)",    sd['Tc']),
                ("tp (hr)",    round(sd['tp'], 4)),
                ("Tp (hr)",    round(sd['Tp'], 4)),
                ("Qp (cms)",   round(sd['Qp'], 4)),
                ("TR (min)",   values['TR_MIN']),
            ]
            for pi, (lbl, val) in enumerate(scs_params, 2):
                ws3.cell(row=pi, column=6, value=lbl)
                ws3.cell(row=pi, column=7, value=val).font = rfnt

        # ── Sheet 4: Nakayasu UH ─────────────────────────────────────────────
        ws4 = wb.create_sheet("Nakayasu UH")
        _set_col_widths(ws4, 14, 10)
        _hrow(ws4, ["Time (hr)", "t/Tp", "q/Qp", "UH (cms/10mm)"])
        _hrow(ws4, ["PARAM", "VALUE"], row=1, start_col=6)

        if 'nakayasu' in self.uh_data:
            nd = self.uh_data['nakayasu']
            Tp_n = nd['Tp']  if nd['Tp']  > 0 else 1.0
            Qp_n = nd['Qp']  if nd['Qp']  > 0 else 1.0
            for i, t in enumerate(nd['time_hr']):
                r = i + 2
                ws4.cell(row=r, column=1, value=round(float(t),             4))
                ws4.cell(row=r, column=2, value=round(float(t / Tp_n),      4))
                ws4.cell(row=r, column=3, value=round(float(nd['uh'][i]/Qp_n), 5))
                ws4.cell(row=r, column=4, value=round(float(nd['uh'][i]),   5))

            naka_params = [
                ("Area (km²)",   values['AREA_KM2']),
                ("L (km)",       values['L_KM']),
                ("tg (hr)",      round(nd['tg'],  4)),
                ("Tp (hr)",      round(nd['Tp'],  4)),
                ("T0.3 (hr)",    round(nd['T03'], 4)),
                ("Qp (cms)",     round(nd['Qp'],  4)),
                ("T0.3 방법",    nd['t03_method']),
                ("alpha",        nd['alpha']),
            ]
            for pi, (lbl, val) in enumerate(naka_params, 2):
                ws4.cell(row=pi, column=6, value=lbl)
                ws4.cell(row=pi, column=7, value=val).font = rfnt

        # ── Sheet 5: 홍수량 비교 ──────────────────────────────────────────────
        ws5 = wb.create_sheet("홍수량 비교")
        _set_col_widths(ws5, 14, 8)
        _hrow(ws5, ["Time (min)", "Inc Rain (mm)", "Inc Excess (mm)",
                    "Clark (cms)", "SCS (cms)", "Nakayasu (cms)"])

        max_len = max(
            (len(self.runoff_data.get(m, {}).get('runoff', [])) for m in ['Clark','SCS','Nakayasu']),
            default=0
        )
        for i in range(max_len):
            r = i + 2
            t_val  = i * values['DT_MIN']
            rain_v = float(rd['rain_inc'][i])   if i < len(rd['rain_inc'])   else 0.0
            exc_v  = float(rd['excess_inc'][i]) if i < len(rd['excess_inc']) else 0.0
            ws5.cell(row=r, column=1, value=round(t_val,  2))
            ws5.cell(row=r, column=2, value=round(rain_v, 4))
            ws5.cell(row=r, column=3, value=round(exc_v,  4))

            for ci, method in enumerate(['Clark', 'SCS', 'Nakayasu'], 4):
                rdata = self.runoff_data.get(method, {}).get('runoff', [])
                q = float(rdata[i]) if i < len(rdata) else 0.0
                ws5.cell(row=r, column=ci, value=round(q, 4))

        # Peak 요약 행
        r_sum = max_len + 3
        ws5.cell(row=r_sum, column=1, value="첨두유량 (cms)").font = Font(bold=True)
        for ci, method in enumerate(['Clark', 'SCS', 'Nakayasu'], 4):
            rdata = self.runoff_data.get(method, {}).get('runoff', [])
            peak  = float(np.max(rdata)) if len(rdata) > 0 else 0.0
            c = ws5.cell(row=r_sum, column=ci, value=round(peak, 3))
            c.font = Font(bold=True, color="CC0000")

        wb.save(xlsx_path)


# ============================================================================
# GUI 애플리케이션
# ============================================================================

class FloodDischargeApp(ctk.CTk):

    # 입력 필드 정의: (표시명, config key, 타입, 기본값)
    _BASE_FIELDS = [
        ("관측소 ID",           "STATION_ID",   "str",   "080000"),
        ("재현기간 (년)",         "RETURN_PERIOD","int",   "전체"),
        ("강우지속기간 TR (분)",  "TR_MIN",       "int",   "전체"),
        ("단기지속기간 DT(분, 기타지속기간 DT : 10분)", "DT_MIN", "int", "1"),
        ("계산횟수 NQ",          "NQ",           "int",   "300"),
        ("유역면적 A (km²)",     "AREA_KM2",     "float", "5.4"),
        ("유출곡선지수 CN",       "CN",           "float", "84.3"),
        ("도달시간 Tc (hr)",     "TC_HR",        "float", "0.40"),
        ("저류상수 R (hr)",       "R_HR",         "float", "0.57"),
    ]
    _CLARK_FIELDS = []   # TC_HR, R_HR 이 _BASE_FIELDS 로 이동
    _SCS_FIELDS = [
        ("도달시간 Tc (hr)",     "SCS_TC_HR",    "float", "0.40"),
    ]
    _NAKA_FIELDS = [
        ("유로연장 L (km)",      "L_KM",         "float", "2.50"),
        ("alpha 값",             "ALPHA",         "float", "1.5"),
    ]

    def __init__(self, project_path: str = "", input_file: str = ""):
        super().__init__()

        self.project_path = project_path.strip() or os.getcwd()
        self.project_name = os.path.basename(self.project_path)
        self.config_file  = os.path.join(self.project_path, "project_config.json")
        self.log_file     = os.path.join(self.project_path, f"{self.project_name}_log.txt")

        self.engine  = RainfallRunoffEngine()
        self.results = {}
        self.fig     = None

        plt.rcParams['font.family']        = 'Malgun Gothic'
        plt.rcParams['axes.unicode_minus'] = False

        self.title(f"홍수량 산정 (5단계)  ─  [{self.project_name}]")
        self.geometry("1300x900")
        self.minsize(1000, 650)

        self._set_dark_title_bar()
        self.protocol("WM_DELETE_WINDOW", self._on_closing)

        self.entries: dict[str, ctk.CTkEntry]      = {}
        self._huff_var    = ctk.StringVar(value="3분위")
        self._interp_var  = ctk.StringVar(value="PCHIP (기본)")
        self._round_var   = ctk.StringVar(value="반올림 없음")
        self._wb_check_var = ctk.BooleanVar(value=False)
        # Step 3 파일 로드 후 채워지는 가용 재현기간·지속기간 목록
        self._available_periods:   list = []
        self._available_durations: list = []
        # SCS / Nakayasu 매개변수 (자식창에서 편집, 값만 보관)
        self._scs_naka_cfg: dict = {
            'SCS_TC_HR': 0.40, 'L_KM': 2.50,
            'ALPHA': 1.5, 'T03_METHOD': 'alpha',
            'WB_CHECK': False,
        }
        self._pc_values: list = list(HUFF_PRESETS["3분위"])

        self._build_ui()
        self._load_config()

    # ── 다크 타이틀바 ────────────────────────────────────────────────────────

    def _set_dark_title_bar(self):
        try:
            hwnd = windll.user32.GetParent(self.winfo_id())
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, byref(c_int(1)), sizeof(c_int))
            windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(c_int(1)), sizeof(c_int))
            try:
                windll.uxtheme.SetPreferredAppMode(2)
            except AttributeError:
                pass
        except Exception:
            pass

    # ── UI 구성 ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self._build_left()
        self._build_right()

    def _build_left(self):
        scroll = ctk.CTkScrollableFrame(self, width=330, corner_radius=0)
        scroll.grid(row=0, column=0, sticky="nsew")
        scroll.grid_columnconfigure(0, weight=1)

        def section(text):
            ctk.CTkLabel(scroll, text=text, font=FONT_HEADER, anchor="w",
                         text_color="#5dade2").pack(fill="x", padx=12, pady=(12, 3))

        def add_field(label, key, default):
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", padx=8, pady=2)
            ctk.CTkLabel(row, text=label, font=FONT_SMALL,
                         width=180, anchor="w").pack(side="left")
            if key in COMBOBOX_OPTIONS:
                w = ctk.CTkComboBox(row, values=COMBOBOX_OPTIONS[key],
                                    font=FONT_SMALL, width=88,
                                    justify="right", state="normal")
                w.set(default)
            else:
                w = ctk.CTkEntry(row, font=FONT_SMALL, width=88, justify="right")
                w.insert(0, default)
            w.pack(side="right")
            self.entries[key] = w

        def sep():
            ctk.CTkFrame(scroll, height=1, fg_color="#444444").pack(
                fill="x", padx=8, pady=6)

        def add_dt_field(label, default_dt):
            """DT_MIN Entry + SHORT_DT_THRESHOLD ComboBox"""
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", padx=8, pady=2)
            ctk.CTkLabel(row, text="단기DT(분)/기타:10분",
                         font=FONT_SMALL, anchor="w").pack(side="left")
            ent = ctk.CTkEntry(row, font=FONT_SMALL, width=44, justify="right")
            ent.insert(0, "1")
            ent.pack(side="left", padx=(6, 4))
            self.entries['DT_MIN'] = ent
            combo = ctk.CTkComboBox(
                row, values=["300"],
                font=FONT_SMALL, width=88, justify="right", state="readonly")
            combo.set("300")
            combo.pack(side="right")
            self.entries['SHORT_DT_THRESHOLD'] = combo

        # 기본 매개변수 (유출곡선지수 CN 아래 도달시간·저류상수 포함)
        for lbl, key, _, dflt in self._BASE_FIELDS:
            if key == 'DT_MIN':
                add_dt_field(lbl, dflt)
            else:
                add_field(lbl, key, dflt)
            # AREA_KM2 바로 아래에 확률강우량 자동 표시 (읽기 전용)
            if key == 'AREA_KM2':
                _row = ctk.CTkFrame(scroll, fg_color="transparent")
                _row.pack(fill="x", padx=8, pady=2)
                ctk.CTkLabel(_row, text="확률강우량 (mm)",
                             font=FONT_SMALL, width=180, anchor="w").pack(side="left")
                self._precip_lbl = ctk.CTkLabel(
                    _row, text="─", font=FONT_SMALL,
                    text_color="#f0a500", anchor="e", width=88)
                self._precip_lbl.pack(side="right")

        # 재현기간·지속기간 ComboBox에 전체/개별 감지 콜백 바인딩
        self.entries['RETURN_PERIOD'].configure(command=self._on_rp_tr_change)
        self.entries['TR_MIN'].configure(command=self._on_rp_tr_change)
        self._prob_rain_table: dict = {}   # {(rp, tr_min): mm}

        # 초기 상태 적용
        self._on_rp_tr_change()

        sep()

        # 버튼 4개 — 동일 간격
        ctk.CTkButton(
            scroll, text="[ Huff 강우시간분포 ]",
            command=self._open_huff_db_dialog,
            font=FONT_BTN, height=36,
            fg_color="#3a3a3a", hover_color="#505050",
        ).pack(fill="x", padx=8, pady=2)

        ctk.CTkButton(
            scroll, text="[ SCS, Nakayasu 합성단위도 ]",
            command=self._open_scs_naka_dialog,
            font=FONT_BTN, height=36,
            fg_color="#3a3a3a", hover_color="#505050",
        ).pack(fill="x", padx=8, pady=2)

        self._btn_run = ctk.CTkButton(
            scroll, text="분석 실행  ▶", command=self._run_analysis,
            font=FONT_BTN, height=42,
            fg_color="#3a3a3a", hover_color="#505050",
        )
        self._btn_run.pack(fill="x", padx=8, pady=2)

        sep()

        # 결과 요약
        section("[ 분석 결과 ]")
        self._result_labels: dict[str, ctk.CTkLabel] = {}
        result_rows = [
            ("Clark  첨두유량",    "clark",       METHOD_COLORS["Clark"]),
            ("SCS   첨두유량",     "scs",         METHOD_COLORS["SCS"]),
            ("Nakayasu 첨두",      "nakayasu",    METHOD_COLORS["Nakayasu"]),
            ("Clark  유출량 (m³)", "vol_clark",   METHOD_COLORS["Clark"]),
            ("SCS   유출량 (m³)",  "vol_scs",     METHOD_COLORS["SCS"]),
            ("Nakayasu 유출(m³)",  "vol_nakayasu",METHOD_COLORS["Nakayasu"]),
        ]
        for lbl_text, key, color in result_rows:
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", padx=8, pady=2)
            ctk.CTkLabel(row, text=lbl_text + ":", font=FONT_SMALL,
                         width=140, anchor="w").pack(side="left")
            lbl = ctk.CTkLabel(row, text="─", font=FONT_HEADER, text_color=color, anchor="w")
            lbl.pack(side="left", padx=4)
            self._result_labels[key] = lbl

        # 로그
        sep()
        ctk.CTkLabel(scroll, text="[ 실행 로그 ]",
                     font=FONT_HEADER, anchor="w", text_color="#5dade2").pack(
            fill="x", padx=12, pady=(8, 2))
        self._txt_log = tk.Text(
            scroll, height=6, font=FONT_LOG,
            bg="#1a1a1a", fg="#cccccc",
            insertbackground="white", wrap="word", relief="flat",
        )
        self._txt_log.pack(fill="x", padx=8, pady=(0, 12))

    def _build_right(self):
        right = ctk.CTkFrame(self, corner_radius=0, fg_color="transparent")
        right.grid(row=0, column=1, sticky="nsew", padx=8, pady=8)
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(0, weight=1)
        self._graph_frame = right

        ctk.CTkLabel(
            right,
            text="분석 실행 후 수문곡선이 여기에 표시됩니다.",
            font=FONT_BODY, text_color="gray",
        ).grid(row=0, column=0)

    # ── Huff 프리셋 ──────────────────────────────────────────────────────────

    def _on_huff_change(self, choice: str):
        preset = HUFF_PRESETS.get(choice)
        if preset is not None:
            self._pc_values = list(preset)

    # ── Config I/O ───────────────────────────────────────────────────────────

    def _set_entry(self, w, val: str):
        """CTkEntry와 CTkComboBox를 구분하여 값 설정"""
        if isinstance(w, ctk.CTkComboBox):
            w.set(val)
        else:
            w.delete(0, "end")
            w.insert(0, val)

    def _load_config(self):
        try:
            with open(self.config_file, encoding='utf-8') as f:
                cfg = json.load(f)
            s5 = cfg.get('step5', {})
            all_fields = self._BASE_FIELDS + self._CLARK_FIELDS + self._SCS_FIELDS + self._NAKA_FIELDS
            for _, key, _, _ in all_fields:
                if key in s5 and key in self.entries:
                    self._set_entry(self.entries[key], str(s5[key]))
            for k in ('SCS_TC_HR', 'L_KM', 'ALPHA'):
                if k in s5:
                    try: self._scs_naka_cfg[k] = float(s5[k])
                    except (ValueError, TypeError): pass
            if 't03_method' in s5:
                self._scs_naka_cfg['T03_METHOD'] = s5['t03_method']
            if 'huff_interp' in s5:
                self._interp_var.set(s5['huff_interp'])
            if 'huff_rounding' in s5:
                self._round_var.set(s5['huff_rounding'])
            if 'wb_check' in s5:
                v = bool(s5['wb_check'])
                self._wb_check_var.set(v)
                self._scs_naka_cfg['WB_CHECK'] = v
            if 'huff_quartile' in s5:
                q = s5['huff_quartile']
                # 구버전 "사분위" → "분위" 마이그레이션
                q = q.replace("사분위", "분위")
                if q not in HUFF_PRESETS:
                    q = "3분위"
                self._huff_var.set(q)
                self._on_huff_change(q)
            if 'huff_pc' in s5 and isinstance(s5['huff_pc'], list):
                self._pc_values = list(s5['huff_pc'])

            # CN III 자동 로드 (Step 4 결과)
            s4 = cfg.get('step4_effective_rainfall', {})
            cn3 = s4.get('cn3')
            if cn3 is not None and 'CN' in self.entries:
                self._set_entry(self.entries['CN'], f"{cn3:.2f}")

        except Exception:
            pass

        # 재현기간·지속기간은 항상 "전체"로 초기화 (배치계산 기본)
        if 'RETURN_PERIOD' in self.entries:
            self.entries['RETURN_PERIOD'].set("전체")
        if 'TR_MIN' in self.entries:
            self.entries['TR_MIN'].set("전체")
        # 분위·소수점처리·PC값 기본값 강제 유지
        self._huff_var.set("3분위")
        self._pc_values = list(HUFF_PRESETS["3분위"])
        self._round_var.set("반올림 없음")

        # 확률강우량 옵션 로드 (Step 3 Excel)
        self._load_prob_rainfall_options()
        # config에서 SHORT_DT_THRESHOLD 복원 (_load_prob_rainfall_options 이후 적용)
        try:
            with open(self.config_file, encoding='utf-8') as _f2:
                _cfg2 = json.load(_f2)
            _sdt = _cfg2.get('step5', {}).get('SHORT_DT_THRESHOLD')
            if _sdt is not None and 'SHORT_DT_THRESHOLD' in self.entries:
                self.entries['SHORT_DT_THRESHOLD'].set(str(_sdt))
        except Exception:
            pass
        # 재현기간/지속기간 상태에 따라 확률강우량 항목 활성/비활성 동기화
        self._on_rp_tr_change()

    def _save_config(self):
        try:
            cfg = {}
            try:
                with open(self.config_file, encoding='utf-8') as f:
                    cfg = json.load(f)
            except Exception:
                pass

            s5 = {}
            all_fields = self._BASE_FIELDS + self._CLARK_FIELDS + self._SCS_FIELDS + self._NAKA_FIELDS
            for _, key, dtype, _ in all_fields:
                raw = self.entries[key].get().strip() if key in self.entries else ""
                try:
                    s5[key] = int(raw) if dtype == 'int' else (raw if dtype == 'str' else float(raw))
                except ValueError:
                    s5[key] = raw
            s5['t03_method']    = self._scs_naka_cfg['T03_METHOD']
            s5['SCS_TC_HR']     = self._scs_naka_cfg['SCS_TC_HR']
            s5['L_KM']          = self._scs_naka_cfg['L_KM']
            s5['ALPHA']         = self._scs_naka_cfg['ALPHA']
            s5['huff_interp']   = self._interp_var.get()
            s5['huff_rounding'] = self._round_var.get()
            s5['wb_check']      = self._wb_check_var.get()
            s5['huff_quartile'] = self._huff_var.get()
            s5['huff_pc']       = self._parse_pc()
            if 'SHORT_DT_THRESHOLD' in self.entries:
                s5['SHORT_DT_THRESHOLD'] = self.entries['SHORT_DT_THRESHOLD'].get()
            cfg['step5'] = s5

            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # ── 확률강우량 옵션 로드 (Step 2 Excel 연동) ──────────────────────────────

    def _load_prob_rainfall_options(self):
        """Step 3 강우강도식 Excel (Intensity_Table)에서 확률강우량 목록을 읽어 채운다.

        Intensity_Table 구조:
            행 1 (헤더): ["Duration(min)", rp1, rp2, ...]   ← 열 = 재현기간(년)
            행 2~:       [tr_min, I1, I2, ...]              ← 강우강도(mm/hr)

        변환: 총강우량(mm) = 강우강도(mm/hr) × 지속기간(min) / 60
        """
        import glob as _glob
        import openpyxl as _opx

        self._prob_rain_table     = {}
        self._available_periods   = []
        self._available_durations = []
        options = []

        try:
            pattern = os.path.join(self.project_path, '*_C_Rainfall_Intensity.xlsx')
            files   = _glob.glob(pattern)
            if not files:
                return

            wb = _opx.load_workbook(files[0], read_only=True, data_only=True)
            if 'Intensity_Table' not in wb.sheetnames:
                wb.close()
                return

            ws   = wb['Intensity_Table']
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(rows) < 2:
                return

            # ── 헤더 파싱: ["Duration(min)", "2", "5", "10", ...] ──────────
            header  = rows[0]
            rp_list = []
            for v in header[1:]:
                if v is None:
                    continue
                try:
                    rp_list.append(int(float(str(v))))
                except (ValueError, TypeError):
                    pass

            # ── 데이터 행 파싱 ───────────────────────────────────────────────
            tr_list = []
            for row in rows[1:]:
                if not row or row[0] is None:
                    continue
                try:
                    tr_min = int(float(str(row[0])))
                except (ValueError, TypeError):
                    continue

                tr_list.append(tr_min)
                for ci, rp in enumerate(rp_list):
                    if ci + 1 >= len(row):
                        continue
                    intensity = row[ci + 1]   # mm/hr
                    if intensity is None:
                        continue
                    try:
                        mm = float(intensity) * tr_min / 60.0   # mm/hr → mm
                        self._prob_rain_table[(rp, tr_min)] = mm
                        options.append(f"{rp}년/{tr_min}분: {mm:.1f}mm")
                    except (ValueError, TypeError):
                        continue

            self._available_periods   = sorted(set(rp_list))
            self._available_durations = sorted(set(tr_list))

            # ── ComboBox 값 동적 업데이트 ────────────────────────────────────
            rp_values = ["전체"] + [str(r) for r in self._available_periods]
            tr_values = ["전체"] + [str(t) for t in self._available_durations]
            if 'RETURN_PERIOD' in self.entries:
                self.entries['RETURN_PERIOD'].configure(values=rp_values)
            if 'TR_MIN' in self.entries:
                self.entries['TR_MIN'].configure(values=tr_values)
            # ── 단기 최대 지속기간 ComboBox: ≤300분 지속기간만 ──────────────
            short_vals = [str(t) for t in self._available_durations if t <= 300]
            if short_vals and 'SHORT_DT_THRESHOLD' in self.entries:
                self.entries['SHORT_DT_THRESHOLD'].configure(values=short_vals)
                self.entries['SHORT_DT_THRESHOLD'].set(short_vals[-1])

        except Exception:
            pass

        # OptionMenu 제거됨 — options는 _prob_rain_table 로만 보관

    def _on_prob_rain_select(self, choice: str):
        """확률강우량 드롭다운 선택 시 재현기간 / 지속기간 자동 입력 (강우량은 자동 조회)"""
        if not choice.startswith("──"):
            try:
                # "50년/120분: 110.4mm" 파싱
                part_rp, rest = choice.split('년/')
                part_tr, _ = rest.split('분: ')
                rp = int(part_rp.strip())
                tr = int(part_tr.strip())
                self._set_entry(self.entries['RETURN_PERIOD'], str(rp))
                self._set_entry(self.entries['TR_MIN'],        str(tr))
                self._on_rp_tr_change()
            except Exception:
                pass

    def _on_rp_tr_change(self, _=None):
        """재현기간·지속기간 변경 시 확률강우량 자동 조회 및 HEC-1 버튼 상태 갱신"""
        rp_raw = self.entries['RETURN_PERIOD'].get() if 'RETURN_PERIOD' in self.entries else "전체"
        tr_raw = self.entries['TR_MIN'].get()        if 'TR_MIN'        in self.entries else "전체"
        is_batch = (rp_raw == '전체' or tr_raw == '전체')

        # 확률강우량 자동 표시
        if hasattr(self, '_precip_lbl'):
            if not is_batch:
                try:
                    mm = self._prob_rain_table.get((int(rp_raw), int(tr_raw)))
                    if mm is not None:
                        self._precip_lbl.configure(
                            text=f"{mm:.1f} mm", text_color="#f0a500")
                    else:
                        self._precip_lbl.configure(
                            text="자료 없음", text_color="#e74c3c")
                except (ValueError, TypeError):
                    self._precip_lbl.configure(text="─", text_color="#f0a500")
            else:
                self._precip_lbl.configure(text="─", text_color="#f0a500")


    def _open_huff_db_dialog(self):
        """HuffDBDialog 자식창 열기"""
        HuffDBDialog(
            self,
            on_apply=self._apply_huff_settings,
            project_name=self.project_name,
            initial_huff={
                'quartile': self._huff_var.get(),
                'interp':   self._interp_var.get(),
                'rounding': self._round_var.get(),
                'pc':       list(self._pc_values),
            },
        )

    def _open_scs_naka_dialog(self):
        """SCSNakayasuDialog 자식창 열기"""
        self._scs_naka_cfg['WB_CHECK'] = self._wb_check_var.get()
        def _on_apply(c):
            self._scs_naka_cfg.update(c)
            self._wb_check_var.set(c.get('WB_CHECK', False))
        SCSNakayasuDialog(
            self, cfg=self._scs_naka_cfg,
            on_apply=_on_apply,
        )

    def _apply_huff_settings(self, cfg: dict):
        """HuffDBDialog 적용 콜백 — pc, quartile, interp, rounding 갱신"""
        self._pc_values = list(cfg.get('pc', self._pc_values))
        self._huff_var.set(cfg.get('quartile', '직접 입력'))
        if 'interp' in cfg:
            self._interp_var.set(cfg['interp'])
        if 'rounding' in cfg:
            self._round_var.set(cfg['rounding'])

    # ── 파라미터 파싱 ─────────────────────────────────────────────────────────

    def _parse_pc(self) -> list:
        return list(self._pc_values)

    def _get_values(self) -> dict | None:
        try:
            e = self.entries
            values = {
                'STATION_ID':   e['STATION_ID'].get().strip() or "STATION",
                'RETURN_PERIOD': int(e['RETURN_PERIOD'].get()),
                'TR_MIN':        int(e['TR_MIN'].get()),
                'DT_MIN':        int(e['DT_MIN'].get()),
                'NQ':            int(e['NQ'].get()),
                'AREA_KM2':      float(e['AREA_KM2'].get()),
                'CN':            float(e['CN'].get()),
                'TC_HR':         float(e['TC_HR'].get()),
                'R_HR':          float(e['R_HR'].get()),
                'SCS_TC_HR':     self._scs_naka_cfg['SCS_TC_HR'],
                'L_KM':          self._scs_naka_cfg['L_KM'],
                'ALPHA':         self._scs_naka_cfg['ALPHA'],
                'T03_METHOD':    self._scs_naka_cfg['T03_METHOD'],
                'HUFF_INTERP':   self._interp_var.get(),
                'HUFF_ROUNDING': self._round_var.get(),
            }
        except ValueError as exc:
            messagebox.showerror("입력 오류", f"숫자 형식 오류:\n{exc}", parent=self)
            return None

        # 확률강우량 자동 조회 (단건 모드에서만 호출되므로 RP·TR 항상 특정값)
        _rp = values['RETURN_PERIOD']
        _tr = values['TR_MIN']
        _mm = self._prob_rain_table.get((_rp, _tr))
        if _mm is None:
            messagebox.showerror(
                "강우량 없음",
                f"확률강우량 자료가 없습니다: {_rp}년/{_tr}분\n"
                "Step 3 강우강도식 결과 파일을 먼저 생성해주세요.",
                parent=self)
            return None
        values['TOTAL_PRECIP'] = _mm

        pc = self._parse_pc()
        if not pc:
            messagebox.showwarning("입력 오류", "Huff PC 값이 없습니다.", parent=self)
            return None

        if not (0 < values['CN'] <= 100):
            messagebox.showwarning("입력 오류", "CN은 0 초과 100 이하여야 합니다.", parent=self)
            return None

        values['HUFF_PC'] = pc
        return values

    # ── 분석 실행 ─────────────────────────────────────────────────────────────

    def _log(self, msg: str):
        self._txt_log.insert(tk.END, msg + "\n")
        self._txt_log.see(tk.END)
        self.update_idletasks()

    def _run_analysis(self):
        # 배치 모드 분기: 재현기간 또는 지속기간이 "전체"이면 Step 3 전체 조합 즉시 실행
        rp_raw = self.entries['RETURN_PERIOD'].get()
        tr_raw = self.entries['TR_MIN'].get()
        if rp_raw == '전체' or tr_raw == '전체':
            if not self._available_periods or not self._available_durations:
                messagebox.showwarning(
                    "데이터 없음",
                    "Step 3 강우강도식 결과 파일(*_C_Rainfall_Intensity.xlsx)을\n"
                    "찾을 수 없습니다.\n먼저 Step 3을 실행한 후 다시 시도하세요.",
                    parent=self)
                return
            # "전체" = Step 3 파일에 있는 모든 값, 선택창 없이 바로 실행
            sel_periods   = self._available_periods   if rp_raw == '전체' else [int(rp_raw)]
            sel_durations = self._available_durations if tr_raw == '전체' else [int(tr_raw)]
            self._run_batch_analysis(sel_periods, sel_durations)
            return

        values = self._get_values()
        if values is None:
            return

        # SHORT_DT_THRESHOLD 적용 (단건 계산 시에도 배치와 동일하게 DT 결정)
        _sdt_raw = self.entries.get('SHORT_DT_THRESHOLD')
        _short_threshold = int(_sdt_raw.get()) if _sdt_raw else 300
        if values['TR_MIN'] > _short_threshold:
            values['DT_MIN'] = 10

        try:
            self._btn_run.configure(state="disabled", text="계산 중…")
            self.update()

            self._log("\n▶ 분석 시작")
            self.engine = RainfallRunoffEngine()

            # 보간 방법 / 소수점 처리 옵션 변환
            _interp_map = {
                "PCHIP (기본)":   "pchip",
                "선형 (Linear)":  "linear",
                "CubicSpline":    "cubic",
            }
            interp_method = _interp_map.get(values['HUFF_INTERP'], 'pchip')
            pc_rounding   = (values['HUFF_ROUNDING'] == "3째자리 반올림")
            self._log(f"  Huff 보간: {values['HUFF_INTERP']}  /  소수점: {values['HUFF_ROUNDING']}")

            # 유효우량
            self.engine.calculate_effective_rainfall(
                values['TOTAL_PRECIP'], values['TR_MIN'],
                values['DT_MIN'],       values['CN'],
                values['HUFF_PC'],
                interp_method=interp_method,
                pc_rounding=pc_rounding,
            )
            total_excess = float(np.sum(self.engine.rainfall_data['excess_inc']))
            total_vol_m3 = total_excess * values['AREA_KM2'] * 1000.0
            self._log(f"  총 강우량   : {np.sum(self.engine.rainfall_data['rain_inc']):.2f} mm")
            self._log(f"  총 유효우량 : {total_excess:.2f} mm  →  총 유출량 : {total_vol_m3:,.0f} m³")

            common = {'A': values['AREA_KM2']}

            # Clark
            t_clark, q_clark = self.engine.convolve_runoff('Clark', {
                **common, 'Tc': values['TC_HR'], 'R': values['R_HR'],
            })
            peak_clark  = float(np.max(q_clark))
            tpeak_clark = float(t_clark[np.argmax(q_clark)])
            self._log(f"  Clark    첨두: {peak_clark:.3f} cms  @ {tpeak_clark:.0f} min")

            # SCS
            t_scs, q_scs = self.engine.convolve_runoff('SCS', {
                **common, 'Tc': values['SCS_TC_HR'],
            })
            peak_scs  = float(np.max(q_scs))
            tpeak_scs = float(t_scs[np.argmax(q_scs)])
            self._log(f"  SCS      첨두: {peak_scs:.3f} cms  @ {tpeak_scs:.0f} min")

            # Nakayasu
            t_naka, q_naka = self.engine.convolve_runoff('Nakayasu', {
                **common,
                'L':          values['L_KM'],
                't03_method': values['T03_METHOD'],
                'alpha':      values['ALPHA'],
            })
            peak_naka  = float(np.max(q_naka))
            tpeak_naka = float(t_naka[np.argmax(q_naka)])
            self._log(f"  Nakayasu 첨두: {peak_naka:.3f} cms  @ {tpeak_naka:.0f} min")

            # ── 물수지 검토 (체크박스 ON 시에만) ─────────────────────
            if self._wb_check_var.get():
                wb_check = self.engine.check_water_balance(
                    values['AREA_KM2'], values['DT_MIN']
                )
                self._log(f"\n  ┌─ 물수지 검토 (유효우량 vs 유출체적) ──────────────────")
                self._log(f"  │  총 강우량    : {wb_check['rain_mm']:.4f} mm")
                self._log(f"  │  총 유효우량  : {wb_check['Pe_mm']:.4f} mm")
                self._log(f"  │  {'방법':<10}  {'유출체적':>10}  {'차이(mm)':>10}  {'오차(%)':>9}")
                self._log(f"  │  {'-'*46}")
                for mname, m in wb_check['methods'].items():
                    flag = "✔" if abs(m['diff_pct']) < 0.5 else "⚠"
                    self._log(
                        f"  │  {flag} {mname:<9}  {m['V_mm']:>10.4f}  "
                        f"{m['diff_mm']:>+10.4f}  {m['diff_pct']:>+8.3f}%"
                    )
                self._log(f"  └{'─'*50}")
            else:
                wb_check = None

            # 수문곡선 적분 → 총 유출량 (m³)
            _dt_sec = values['DT_MIN'] * 60.0
            vol_clark = float(np.sum(q_clark)) * _dt_sec
            vol_scs   = float(np.sum(q_scs))   * _dt_sec
            vol_naka  = float(np.sum(q_naka))  * _dt_sec

            # 결과 저장
            self.results = {
                't_clark': t_clark, 'q_clark': q_clark,
                't_scs':   t_scs,   'q_scs':   q_scs,
                't_naka':  t_naka,  'q_naka':  q_naka,
                'values':  values,
                'peak_clark':  peak_clark,  'tpeak_clark':  tpeak_clark,
                'peak_scs':    peak_scs,    'tpeak_scs':    tpeak_scs,
                'peak_naka':   peak_naka,   'tpeak_naka':   tpeak_naka,
                'vol_clark':   vol_clark,   'vol_scs':      vol_scs,
                'vol_naka':    vol_naka,
                'total_excess':  total_excess,
                'water_balance': wb_check,
            }

            # 결과 레이블 갱신
            self._result_labels['clark'].configure(
                text=f"{peak_clark:.3f} cms  (@ {tpeak_clark:.0f} min)")
            self._result_labels['scs'].configure(
                text=f"{peak_scs:.3f} cms  (@ {tpeak_scs:.0f} min)")
            self._result_labels['nakayasu'].configure(
                text=f"{peak_naka:.3f} cms  (@ {tpeak_naka:.0f} min)")
            self._result_labels['vol_clark'].configure(
                text=f"{vol_clark:,.0f} m³")
            self._result_labels['vol_scs'].configure(
                text=f"{vol_scs:,.0f} m³")
            self._result_labels['vol_nakayasu'].configure(
                text=f"{vol_naka:,.0f} m³")

            self._draw_graph()
            self._save_config()
            self._auto_save_single()
            self._log("  ✔ 완료")

        except Exception as exc:
            messagebox.showerror("계산 오류", traceback.format_exc(), parent=self)
            self._log(f"  ✘ 오류: {exc}")
        finally:
            self._btn_run.configure(state="normal", text="분석 실행  ▶")

    # ── 그래프 ───────────────────────────────────────────────────────────────

    def _draw_graph(self):
        for w in self._graph_frame.winfo_children():
            w.destroy()
        if self.fig:
            plt.close(self.fig)

        res = self.results
        rd  = self.engine.rainfall_data
        DT  = res['values']['DT_MIN']

        self.fig, ax1 = plt.subplots(figsize=(9, 6), dpi=100)
        self.fig.patch.set_facecolor('#1c1c1e')
        ax1.set_facecolor('#1c1c1e')
        ax2 = ax1.twinx()

        # 강우 바 (위쪽이 0, 아래로 증가)
        ax2.bar(rd['time_min'], rd['rain_inc'], width=DT,
                color='#5dade2', alpha=0.35, label='강우량', align='edge')
        ax2.bar(rd['time_min'], rd['excess_inc'], width=DT,
                color='#2980b9', alpha=0.75, label='유효우량', align='edge')
        max_rain = float(np.max(rd['rain_inc'])) if len(rd['rain_inc']) > 0 else 1.0
        ax2.set_ylim(0, max_rain * 3.5)
        ax2.invert_yaxis()
        ax2.set_ylabel('강우량 / 유효우량 (mm)', color='#85c1e9')
        ax2.tick_params(axis='y', colors='#85c1e9')
        ax2.spines['right'].set_edgecolor('#555')

        # 유출곡선 3개
        ax1.plot(res['t_clark'], res['q_clark'],
                 color=METHOD_COLORS["Clark"],    lw=2.0, label=f"Clark ({res['peak_clark']:.3f} cms)")
        ax1.plot(res['t_scs'],   res['q_scs'],
                 color=METHOD_COLORS["SCS"],      lw=2.0, linestyle='--',
                 label=f"SCS ({res['peak_scs']:.3f} cms)")
        ax1.plot(res['t_naka'],  res['q_naka'],
                 color=METHOD_COLORS["Nakayasu"], lw=2.0, linestyle='-.',
                 label=f"Nakayasu ({res['peak_naka']:.3f} cms)")

        ax1.set_xlabel('시간 (분)', color='#cccccc', fontweight='bold')
        ax1.set_ylabel('유량 (m³/s)', color='#cccccc', fontweight='bold')
        ax1.tick_params(axis='both', colors='#cccccc')
        ax1.set_ylim(bottom=0)
        ax1.grid(True, linestyle='--', alpha=0.25, color='#888')
        for sp in ax1.spines.values():
            sp.set_edgecolor('#444')

        # ── 피크 수선: 피크 시각에서 x축까지 수직 점선 + 시각 텍스트 ──
        _, y_top = ax1.get_ylim()
        for t_arr, q_arr, clr in [
            (res['t_clark'], res['q_clark'], METHOD_COLORS["Clark"]),
            (res['t_scs'],   res['q_scs'],   METHOD_COLORS["SCS"]),
            (res['t_naka'],  res['q_naka'],  METHOD_COLORS["Nakayasu"]),
        ]:
            idx = int(np.argmax(q_arr))
            tp  = float(t_arr[idx])
            qp  = float(q_arr[idx])
            ax1.vlines(tp, 0, qp, colors=clr, linestyles=':', linewidth=1.2, alpha=0.75)
            ax1.text(tp, y_top * 0.03, f"{tp:.0f}분",
                     color=clr, fontsize=8, ha='center', va='bottom', fontweight='bold')

        sid = res['values']['STATION_ID']
        rp  = res['values']['RETURN_PERIOD']
        plt.title(f"수문곡선 비교  ─  {sid}  /  재현기간 {rp}년",
                  color='#eeeeee', fontweight='bold')

        l1, b1 = ax1.get_legend_handles_labels()
        l2, b2 = ax2.get_legend_handles_labels()
        ax1.legend(l1 + l2, b1 + b2,
                   loc='upper right', facecolor='#2c2c2e',
                   edgecolor='#555', labelcolor='#dddddd', fontsize=9)

        plt.tight_layout()

        canvas = FigureCanvasTkAgg(self.fig, master=self._graph_frame)
        canvas.draw()
        canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        self._graph_frame.grid_columnconfigure(0, weight=1)
        self._graph_frame.grid_rowconfigure(0, weight=1)

    def _draw_batch_graph(self, results: dict, common_vals: dict, disp_rp: int):
        """배치 결과에서 각 방법 고유 임계TR 수문곡선을 GUI에 표시.
        강우 바는 공통 TR → 회색, 고유 TR → 방법 색상으로 구분."""
        try:
            rp_items = [(tr, r) for (rp, tr), r in results.items() if rp == disp_rp]
            if not rp_items:
                return

            # 각 방법별 임계TR / 강우량
            tr_c, r_c = max(rp_items, key=lambda x: x[1]['clark'])
            tr_s, r_s = max(rp_items, key=lambda x: x[1]['scs'])
            tr_n, r_n = max(rp_items, key=lambda x: x[1]['naka'])

            short_threshold = common_vals.get('SHORT_DT_THRESHOLD', 300)
            _imap    = {"PCHIP (기본)": "pchip", "선형 (Linear)": "linear",
                        "CubicSpline": "cubic"}
            interp_m = _imap.get(common_vals['HUFF_INTERP'], 'pchip')
            pc_rnd   = (common_vals['HUFF_ROUNDING'] == "3째자리 반올림")
            _A       = {'A': common_vals['AREA_KM2']}

            # 엔진 별도 생성 → rainfall_data 보존
            def _calc(tr, mm):
                dt = common_vals['DT_MIN'] if tr <= short_threshold else 10
                e  = RainfallRunoffEngine()
                e.calculate_effective_rainfall(
                    mm, tr, dt, common_vals['CN'], self._pc_values,
                    interp_method=interp_m, pc_rounding=pc_rnd)
                return e, dt

            eng_c, dt_c = _calc(tr_c, r_c['mm'])
            eng_s, dt_s = _calc(tr_s, r_s['mm'])
            eng_n, dt_n = _calc(tr_n, r_n['mm'])

            tc, qc = eng_c.convolve_runoff('Clark',
                {**_A, 'Tc': common_vals['TC_HR'], 'R': common_vals['R_HR']})
            ts, qs = eng_s.convolve_runoff('SCS',
                {**_A, 'Tc': common_vals['SCS_TC_HR']})
            tn, qn = eng_n.convolve_runoff('Nakayasu',
                {**_A, 'L': common_vals['L_KM'],
                 't03_method': common_vals['T03_METHOD'],
                 'alpha': common_vals['ALPHA']})

            # 기존 그래프 정리
            for w in self._graph_frame.winfo_children():
                w.destroy()
            if self.fig:
                plt.close(self.fig)

            self.fig, ax1 = plt.subplots(figsize=(9, 6), dpi=100)
            self.fig.patch.set_facecolor('#1c1c1e')
            ax1.set_facecolor('#1c1c1e')
            ax2 = ax1.twinx()

            # ── 강우 바: TR별 그룹핑 → 공통=회색, 고유=방법 색상 ──────
            # {tr: {'eng':, 'dt':, 'color':, 'methods': []}}
            rain_groups: dict = {}
            for tr_val, eng, dt_val, mname, clr in [
                (tr_c, eng_c, dt_c, 'Clark',    METHOD_COLORS["Clark"]),
                (tr_s, eng_s, dt_s, 'SCS',      METHOD_COLORS["SCS"]),
                (tr_n, eng_n, dt_n, 'Nakayasu', METHOD_COLORS["Nakayasu"]),
            ]:
                if tr_val not in rain_groups:
                    rain_groups[tr_val] = {
                        'eng': eng, 'dt': dt_val,
                        'color': clr, 'methods': [mname],
                    }
                else:
                    rain_groups[tr_val]['methods'].append(mname)
                    rain_groups[tr_val]['color'] = '#aaaaaa'  # 공통 → 회색

            max_rain_all = max(
                float(np.max(g['eng'].rainfall_data['rain_inc']))
                for g in rain_groups.values()
            ) or 1.0
            ax2.set_ylim(0, max_rain_all * 3.5)
            ax2.invert_yaxis()
            ax2.set_ylabel('강우량 / 유효우량 (mm)', color='#85c1e9')
            ax2.tick_params(axis='y', colors='#85c1e9')
            ax2.spines['right'].set_edgecolor('#555')

            for tr_val, g in rain_groups.items():
                rd  = g['eng'].rainfall_data
                clr = g['color']
                dt_val = g['dt']
                ax2.bar(rd['time_min'], rd['rain_inc'],   width=dt_val,
                        color=clr, alpha=0.18, align='edge')
                ax2.bar(rd['time_min'], rd['excess_inc'], width=dt_val,
                        color=clr, alpha=0.45, align='edge')
                # 지속기간 텍스트 (강우 종료 시각 상단)
                ax2.text(tr_val, max_rain_all * 0.06,
                         f"TR={tr_val}분",
                         color=clr, fontsize=8, fontweight='bold',
                         ha='right', va='bottom')

            # ── 수문곡선 3개 ────────────────────────────────────────────
            lbl_c = f"Clark     {float(np.max(qc)):.3f} cms  ({disp_rp}년/{tr_c}분)"
            lbl_s = f"SCS       {float(np.max(qs)):.3f} cms  ({disp_rp}년/{tr_s}분)"
            lbl_n = f"Nakayasu  {float(np.max(qn)):.3f} cms  ({disp_rp}년/{tr_n}분)"

            ax1.plot(tc, qc, color=METHOD_COLORS["Clark"],    lw=2.0, label=lbl_c)
            ax1.plot(ts, qs, color=METHOD_COLORS["SCS"],      lw=2.0, linestyle='--',  label=lbl_s)
            ax1.plot(tn, qn, color=METHOD_COLORS["Nakayasu"], lw=2.0, linestyle='-.', label=lbl_n)

            ax1.set_xlabel('시간 (분)', color='#cccccc', fontweight='bold')
            ax1.set_ylabel('유량 (m³/s)', color='#cccccc', fontweight='bold')
            ax1.tick_params(axis='both', colors='#cccccc')
            ax1.set_ylim(bottom=0)
            ax1.grid(True, linestyle='--', alpha=0.25, color='#888')
            for sp in ax1.spines.values():
                sp.set_edgecolor('#444')

            # ── 피크 수선: 피크 시각에서 x축까지 수직 점선 + 시각 텍스트 ──
            _, y_top = ax1.get_ylim()
            for t_arr, q_arr, clr in [
                (tc, qc, METHOD_COLORS["Clark"]),
                (ts, qs, METHOD_COLORS["SCS"]),
                (tn, qn, METHOD_COLORS["Nakayasu"]),
            ]:
                idx = int(np.argmax(q_arr))
                tp  = float(t_arr[idx])
                qp  = float(q_arr[idx])
                ax1.vlines(tp, 0, qp, colors=clr, linestyles=':', linewidth=1.2, alpha=0.75)
                ax1.text(tp, y_top * 0.03, f"{tp:.0f}분",
                         color=clr, fontsize=8, ha='center', va='bottom', fontweight='bold')

            sid = common_vals.get('STATION_ID', '')
            plt.title(
                f"수문곡선 비교 [배치]  ─  {sid}  /  {disp_rp}년 빈도",
                color='#eeeeee', fontweight='bold')

            l1, b1 = ax1.get_legend_handles_labels()
            ax1.legend(l1, b1, loc='upper right', facecolor='#2c2c2e',
                       edgecolor='#555', labelcolor='#dddddd', fontsize=9)
            plt.tight_layout()

            canvas = FigureCanvasTkAgg(self.fig, master=self._graph_frame)
            canvas.draw()
            canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
            self._graph_frame.grid_columnconfigure(0, weight=1)
            self._graph_frame.grid_rowconfigure(0, weight=1)
        except Exception:
            pass

    # ── Excel 저장 ────────────────────────────────────────────────────────────

    def _auto_save_single(self):
        """단건 분석 완료 후 자동 저장: Excel(결과요약+PNG시트) + PNG 파일."""
        if not self.results:
            return
        res    = self.results
        values = res['values']
        rp     = values['RETURN_PERIOD']
        tr     = values['TR_MIN']
        stem     = f"{self.project_name}_E_{rp:04d}_{tr:04d}"
        out_path = os.path.join(self.project_path, f"{stem}.xlsx")
        png_path = os.path.join(self.project_path, f"{stem}_수문곡선.png")
        try:
            # 1. 기존 5개 시트 Excel 저장
            self.engine.save_excel(out_path, values)

            # 2. PNG 저장
            self.fig.savefig(png_path, dpi=150, bbox_inches='tight',
                             facecolor='#1c1c1e')

            # 3. openpyxl로 재열어 결과요약 + PNG 시트 추가
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as XLImage
            wb = load_workbook(out_path)
            _fn, _fs = "맑은 고딕", 11
            _hfnt  = Font(bold=True,  name=_fn, size=_fs, color="FFFFFF")
            _hfill = PatternFill("solid", fgColor="4F81BD")
            _bfnt  = Font(bold=True,  name=_fn, size=_fs)
            _dfnt  = Font(bold=False, name=_fn, size=_fs)

            # ── 결과요약 시트 ──────────────────────────────────────────
            ws_sum = wb.create_sheet("결과요약")
            for ci, h in enumerate(["방법","첨두유량(m³/s)","발생시각(min)","총유출량(m³)"], 1):
                c = ws_sum.cell(1, ci, h); c.font = _hfnt; c.fill = _hfill
            for ri, (nm, pk, tp, vl) in enumerate([
                ("Clark",    res['peak_clark'],  res['tpeak_clark'],  res['vol_clark']),
                ("SCS",      res['peak_scs'],    res['tpeak_scs'],    res['vol_scs']),
                ("Nakayasu", res['peak_naka'],   res['tpeak_naka'],   res['vol_naka']),
            ], 2):
                ws_sum.cell(ri, 1, nm).font = _bfnt
                ws_sum.cell(ri, 2, round(pk, 3)).font = _dfnt
                ws_sum.cell(ri, 3, round(tp, 1)).font  = _dfnt
                ws_sum.cell(ri, 4, round(vl, 0)).font  = _dfnt
            for ri, (lbl, val) in enumerate([
                ("재현기간 (년)",         rp),
                ("강우지속기간 (분)",     tr),
                ("총 유효우량 (mm)",      round(res['total_excess'], 2)),
                ("총 유출량-강우기반(m³)", round(res['total_excess'] * values['AREA_KM2'] * 1000, 0)),
            ], 6):
                ws_sum.cell(ri, 1, lbl).font = _bfnt
                ws_sum.cell(ri, 2, val).font  = _dfnt
            for col, w in zip(['A','B','C','D'], [24, 18, 16, 20]):
                ws_sum.column_dimensions[col].width = w

            # ── 수문곡선 시트 ──────────────────────────────────────────
            ws_img = wb.create_sheet("수문곡선")
            ws_img.add_image(XLImage(png_path), 'A1')

            wb.save(out_path)

            # 4. config 업데이트
            cfg = {}
            if os.path.exists(self.config_file):
                with open(self.config_file, encoding='utf-8') as f:
                    cfg = json.load(f)
            cfg['step5_flood_discharge'] = {
                'status':          'completed',
                'output_file':     f"{stem}.xlsx",
                'full_path':       out_path,
                'peak_clark_cms':  round(res['peak_clark'],  4),
                'peak_scs_cms':    round(res['peak_scs'],    4),
                'peak_naka_cms':   round(res['peak_naka'],   4),
                'total_excess_mm': round(res['total_excess'], 4),
                'timestamp':       datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=2, ensure_ascii=False)

            # 5. 프로젝트 로그
            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(
                    f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S'):<20} | "
                    f"{'5. 홍수량 산정':<28} | "
                    f"Clark={res['peak_clark']:.3f}, SCS={res['peak_scs']:.3f}, "
                    f"Nakayasu={res['peak_naka']:.3f} cms  → {stem}.xlsx\n"
                )

            self._log(f"  💾 저장: {stem}.xlsx  /  {stem}_수문곡선.png")

            # 6. HEC-1 입출력파일 자동 생성
            self._generate_hec1_files(auto=True)

        except Exception as exc:
            self._log(f"  ✘ 자동저장 오류: {exc}")

    # ── HEC-1 입출력파일 생성 ─────────────────────────────────────────────────

    def _generate_hec1_files(self, auto: bool = False):
        """단건 분석 결과를 HEC-1 형식 .DAT / .OUT 파일로 저장."""
        if not self.results:
            if not auto:
                messagebox.showwarning("알림", "먼저 단건 분석을 실행해주세요.", parent=self)
            return
        res    = self.results
        values = res['values']
        rp     = values['RETURN_PERIOD']
        tr     = values['TR_MIN']
        stem   = f"{self.project_name}_{rp:03d}_{tr:04d}_E_FOR_HEC1"
        dat_path = os.path.join(self.project_path, f"{stem}.DAT")
        out_path = os.path.join(self.project_path, f"{stem}.OUT")
        try:
            self._write_hec1_dat(dat_path, values)
            self._write_hec1_out(out_path, values)
            self._log(f"  💾 HEC-1: {stem}.DAT / .OUT")
        except Exception as exc:
            self._log(f"  ✘ HEC-1 파일 생성 오류: {exc}")
            if not auto:
                messagebox.showerror("오류", str(exc), parent=self)

    def _write_hec1_dat(self, dat_path: str, values: dict):
        """HEC-1 형식 입력파일(.DAT) 작성."""
        from datetime import datetime as _dt
        rd       = self.engine.rainfall_data
        pc_interp = list(rd['pc_interp'])

        # 1.000 연속 중복 제거 (마지막 1.000 하나만 유지)
        if len(pc_interp) >= 2:
            i = len(pc_interp) - 1
            while i > 0 and abs(pc_interp[i - 1] - 1.0) < 1e-6:
                i -= 1
            pc_interp = pc_interp[:i + 1]

        start_date = _dt(2005, 7, 11, 18, 0)
        d_str = start_date.strftime("%d%b%y").upper()
        t_str = start_date.strftime("%H%M")
        nmin  = values['DT_MIN']
        nq    = values['NQ']
        sid   = values['STATION_ID']

        # 면적 / 강우량 포맷 (1 미만이면 선행 0 제거)
        def _fmt(v, dec):
            s = f"{v:.{dec}f}"
            return s.lstrip('0') if v < 1 else s

        ba_str = _fmt(values['AREA_KM2'],     1)
        pb_str = _fmt(values['TOTAL_PRECIP'], 1)
        cn_str = _fmt(values['CN'],           1)
        tc_str = _fmt(values['TC_HR'],        2)
        r_str  = _fmt(values['R_HR'],         2)

        # PC 카드 (10개씩)
        pc_lines = []
        for i in range(0, len(pc_interp), 10):
            chunk = pc_interp[i:i + 10]
            line  = "PC"
            for idx, val in enumerate(chunk):
                vs = f"{val:.3f}".lstrip('0') if val < 1 else f"{val:.3f}"
                line += f"{vs:>6s}" if idx == 0 else f"{vs:>8s}"
            pc_lines.append(line)

        lines = [
            f"ID Project(Flood)",
            f"*DIAGRAM",
            f"IM",
            f"IO     5       1",
            f"IT    {nmin:2d} {d_str}    {t_str}     {nq}",
            f"VS{sid}",
            f"VV  2.11",
            f"* ",
            f"KK{sid}",
            f"IN    {nmin:2d} {d_str}    {t_str}",
            f"BA{ba_str:>6s}",
            f"PB{pb_str:>6s}",
        ] + pc_lines + [
            f"LS{'':10s}{cn_str}",
            f"UC{'':3s}{tc_str}{'':5s}{r_str}",
            f"*",
            f"ZZ",
        ]

        with open(dat_path, 'w', encoding='utf-8', newline='') as f:
            f.write('\r\n'.join(lines) + '\r\n')

    def _write_hec1_out(self, out_path: str, values: dict):
        """HEC-1 형식 출력파일(.OUT) 작성 — Clark 유출 결과 사용."""
        from datetime import datetime as _dt, timedelta as _td
        clark_data = self.engine.runoff_data.get('Clark', {})
        flows      = clark_data.get('runoff', [])
        nmin       = values['DT_MIN']
        sid        = values['STATION_ID']
        rp         = values['RETURN_PERIOD']
        tr         = values['TR_MIN']
        start_date = _dt(2005, 7, 11, 18, 0)

        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(f"1TABLE 1    STATION      {sid}\n")
            f.write(f"           재현기간: {rp}년  강우지속기간: {tr}분\n")
            f.write("                           FLOW\n\n")
            f.write(" PER  DAY MON  HRMN\n\n")
            for i, flow in enumerate(flows):
                curr = start_date + _td(minutes=i * nmin)
                per  = f"{i + 1:4d}"
                day  = f"{curr.day:2d}"
                mon  = curr.strftime("%b").upper()
                hhmm = curr.strftime("%H%M")
                if abs(flow) < 0.005:
                    fs = "         .00"
                else:
                    fs = f"{flow:12.2f}"
                    if 0 < abs(flow) < 1:
                        fs = fs.replace("0.", " .")
                f.write(f" {per}   {day} {mon}  {hhmm}    {fs}\n")
            f.write("\n")
            peak = float(max(flows)) if len(flows) > 0 else 0.0
            f.write(f"              MAX         {peak:.2f}\n")

    # ── 배치 계산 ────────────────────────────────────────────────────────────

    def _run_batch_analysis(self, sel_periods: list, sel_durations: list):
        """선택된 재현기간×지속기간 전체 조합을 배치 계산하고 Excel로 저장."""
        if not self._prob_rain_table:
            messagebox.showwarning(
                "데이터 없음",
                "Step 3 강우강도식 결과 파일(*_C_Rainfall_Intensity.xlsx)을\n"
                "찾을 수 없습니다.\n먼저 Step 3을 실행한 후 다시 시도하세요.",
                parent=self)
            return

        # 공통 매개변수 수집
        try:
            e = self.entries
            short_dt_raw = self.entries.get('SHORT_DT_THRESHOLD')
            short_dt_threshold = int(short_dt_raw.get()) if short_dt_raw else 300
            common_vals = {
                'STATION_ID':  e['STATION_ID'].get().strip() or "STATION",
                'DT_MIN':      int(e['DT_MIN'].get()),
                'SHORT_DT_THRESHOLD': short_dt_threshold,
                'NQ':          int(e['NQ'].get()),
                'AREA_KM2':    float(e['AREA_KM2'].get()),
                'CN':          float(e['CN'].get()),
                'TC_HR':       float(e['TC_HR'].get()),
                'R_HR':        float(e['R_HR'].get()),
                'SCS_TC_HR':   self._scs_naka_cfg['SCS_TC_HR'],
                'L_KM':        self._scs_naka_cfg['L_KM'],
                'ALPHA':       self._scs_naka_cfg['ALPHA'],
                'T03_METHOD':  self._scs_naka_cfg['T03_METHOD'],
                'HUFF_INTERP': self._interp_var.get(),
                'HUFF_ROUNDING': self._round_var.get(),
            }
        except ValueError as exc:
            messagebox.showerror("입력 오류", f"숫자 형식 오류:\n{exc}", parent=self)
            return

        pc = self._parse_pc()
        if not pc:
            messagebox.showwarning("입력 오류", "Huff PC 값이 없습니다.", parent=self)
            return

        _interp_map = {"PCHIP (기본)": "pchip", "선형 (Linear)": "linear",
                       "CubicSpline": "cubic"}
        interp_method = _interp_map.get(common_vals['HUFF_INTERP'], 'pchip')
        pc_rounding   = (common_vals['HUFF_ROUNDING'] == "3째자리 반올림")

        # 유효 조합 목록
        combinations = [
            (rp, tr, self._prob_rain_table[(rp, tr)])
            for rp in sorted(sel_periods)
            for tr in sorted(sel_durations)
            if (rp, tr) in self._prob_rain_table
        ]
        skipped = (len(sel_periods) * len(sel_durations)) - len(combinations)
        if not combinations:
            messagebox.showwarning("결과 없음",
                "계산 가능한 조합이 없습니다.\n"
                "Step 2 Excel의 지속기간/재현기간이 선택값과 일치하는지 확인하세요.",
                parent=self)
            return

        # ── 컴퓨팅 백엔드 감지 ────────────────────────────────────────────
        xp, use_gpu, backend_label = _detect_compute_backend()
        self._log(f"\n▶ 배치 계산 시작  ·  {len(combinations)}개 조합")
        self._log(f"  백엔드: {backend_label}")

        # ── 진행 팝업 ─────────────────────────────────────────────────────
        prog_win = ctk.CTkToplevel(self)
        prog_win.title("배치 계산 진행 중...")
        prog_win.geometry("420x140")
        prog_win.grab_set()
        ctk.CTkLabel(prog_win, text=f"백엔드: {backend_label}",
                     font=FONT_SMALL, text_color="gray").pack(pady=(10, 2))
        prog_lbl = ctk.CTkLabel(prog_win, text="계산 준비 중...", font=FONT_BODY)
        prog_lbl.pack()
        prog_bar = ctk.CTkProgressBar(prog_win, width=360)
        prog_bar.pack(pady=6)
        prog_bar.set(0)
        prog_win.update()

        # ── 스레드 + Queue 기반 비동기 실행 ──────────────────────────────
        import threading, queue as _queue
        result_q: _queue.Queue = _queue.Queue()

        def _worker():
            try:
                results = _run_batch_compute(
                    combinations, common_vals, pc,
                    interp_method, pc_rounding,
                    xp, use_gpu,
                    progress_cb=lambda d, t, rp, tr:
                        result_q.put(('prog', d, t, rp, tr)),
                )
                result_q.put(('done', results))
            except Exception as exc:
                result_q.put(('err', exc))

        threading.Thread(target=_worker, daemon=True).start()

        # ── 큐 폴링 (50 ms) — 메인 스레드에서 UI 업데이트 ───────────────
        def _poll():
            try:
                while True:
                    msg = result_q.get_nowait()
                    if msg[0] == 'prog':
                        _, done, total, rp, tr = msg
                        prog_bar.set(done / max(total, 1))
                        prog_lbl.configure(
                            text=f"{rp}년/{tr}분  ({done}/{total})")
                    elif msg[0] == 'done':
                        batch_results = msg[1]
                        prog_win.destroy()
                        self._log(
                            f"  완료: {len(batch_results)}개 계산 / {skipped}개 스킵")
                        self._save_config()
                        # ── 결과 레이블 업데이트 ──────────────────────────
                        if batch_results:
                            # 표시 기준 재현기간 결정
                            rp_raw = self.entries['RETURN_PERIOD'].get()
                            if rp_raw == '전체':
                                disp_rp = 50 if 50 in sel_periods else max(sel_periods)
                            else:
                                disp_rp = int(rp_raw)
                            # disp_rp 기준 각 method 최대 피크 추출
                            rp_items = {tr: r for (rp, tr), r in batch_results.items()
                                        if rp == disp_rp}
                            if rp_items:
                                # 각 방법별 임계TR 기준 피크 + 유출량
                                best_c = max(rp_items.items(), key=lambda x: x[1]['clark'])
                                best_s = max(rp_items.items(), key=lambda x: x[1]['scs'])
                                best_n = max(rp_items.items(), key=lambda x: x[1]['naka'])
                                # 각 방법 고유 임계TR 포함 태그
                                tag_c = f"({disp_rp}년/{best_c[0]}분)"
                                tag_s = f"({disp_rp}년/{best_s[0]}분)"
                                tag_n = f"({disp_rp}년/{best_n[0]}분)"
                                self._result_labels['clark'].configure(
                                    text=f"{best_c[1]['clark']:.3f} cms  {tag_c}")
                                self._result_labels['scs'].configure(
                                    text=f"{best_s[1]['scs']:.3f} cms  {tag_s}")
                                self._result_labels['nakayasu'].configure(
                                    text=f"{best_n[1]['naka']:.3f} cms  {tag_n}")
                                self._result_labels['vol_clark'].configure(
                                    text=f"{best_c[1]['vol_clark']:,.0f} m³  {tag_c}")
                                self._result_labels['vol_scs'].configure(
                                    text=f"{best_s[1]['vol_scs']:,.0f} m³  {tag_s}")
                                self._result_labels['vol_nakayasu'].configure(
                                    text=f"{best_n[1]['vol_naka']:,.0f} m³  {tag_n}")
                                self._draw_batch_graph(batch_results, common_vals, disp_rp)
                        if batch_results:
                            self._save_batch_excel(
                                batch_results,
                                sorted(sel_periods), sorted(sel_durations),
                                common_vals)
                        else:
                            messagebox.showwarning("결과 없음", "계산 결과가 없습니다.",
                                                   parent=self)
                        return
                    elif msg[0] == 'err':
                        prog_win.destroy()
                        messagebox.showerror("배치 오류",
                                             traceback.format_exc(), parent=self)
                        self._log(f"  ✘ 오류: {msg[1]}")
                        return
            except _queue.Empty:
                pass
            self.after(50, _poll)

        self.after(50, _poll)

    def _save_batch_excel(self, results: dict, periods: list, durations: list,
                          common_vals: dict):
        """배치 계산 결과를 Excel matrix로 저장."""
        fname    = f"{self.project_name}_E_Flood_Batch.xlsx"
        out_path = os.path.join(self.project_path, fname)

        wb   = Workbook()
        _fn  = "맑은 고딕"
        _fs  = 11
        hfnt      = Font(bold=True,  name=_fn, size=_fs, color="FFFFFF")           # 헤더 흰글씨
        hfill     = PatternFill("solid", fgColor="4F81BD")                          # 헤더 파란 배경
        a1fnt     = Font(bold=True,  name=_fn, size=_fs)                            # A1·A열 라벨
        dfnt      = Font(bold=False, name=_fn, size=_fs)                            # 데이터 셀 기본색
        pk_lbl_fnt = Font(bold=True, name=_fn, size=_fs, color=Color(theme=3))     # 피크·임계 라벨
        pk_val_fnt = Font(bold=True, name=_fn, size=_fs, color="FF0000")           # 홍수량피크 값 빨강
        cr_val_fnt = Font(bold=True, name=_fn, size=_fs, color=Color(theme=1))     # 임계지속기간 값 기본색

        def make_sheet(ws, method_key, method_name):
            ws.title = method_name
            n_dur = len(durations)
            last_data_row = n_dur + 1  # 헤더=1행, 데이터=2~n_dur+1행
            # 헤더 행: 재현기간
            ws.cell(1, 1, "지속기간(분)\\재현기간(년)")
            ws.cell(1, 1).font = a1fnt
            for ci, rp in enumerate(periods, 2):
                c = ws.cell(1, ci, f"{rp}년")
                c.font, c.fill = hfnt, hfill
            # 데이터 행
            for ri, tr in enumerate(durations, 2):
                ws.cell(ri, 1, tr).font = a1fnt
                for ci, rp in enumerate(periods, 2):
                    r = results.get((rp, tr))
                    val = round(r[method_key], 3) if r else None
                    c = ws.cell(ri, ci, val)
                    if val is not None:
                        c.font = dfnt
            # 홍수량피크 행 (재현기간별 최대값)
            peak_row = last_data_row + 1
            ws.cell(peak_row, 1, "홍수량피크").font = pk_lbl_fnt
            for ci, rp in enumerate(periods, 2):
                col_letter = get_column_letter(ci)
                c = ws.cell(peak_row, ci,
                            f"=MAX({col_letter}$2:{col_letter}${last_data_row})")
                c.font = pk_val_fnt
            # 임계지속기간(분) 행 (최대값 발생 지속기간)
            crit_row = peak_row + 1
            ws.cell(crit_row, 1, "임계지속기간(분)").font = pk_lbl_fnt
            for ci, rp in enumerate(periods, 2):
                col_letter = get_column_letter(ci)
                c = ws.cell(crit_row, ci,
                            f"=INDEX($A$2:$A${last_data_row},"
                            f"MATCH(MAX({col_letter}$2:{col_letter}${last_data_row}),"
                            f"{col_letter}$2:{col_letter}${last_data_row},0))")
                c.font = cr_val_fnt
            # 열 너비
            ws.column_dimensions['A'].width = 22
            for ci in range(2, len(periods) + 2):
                ws.column_dimensions[get_column_letter(ci)].width = 10

        # Sheet 1: Clark
        ws1 = wb.active
        make_sheet(ws1, 'clark', 'Clark_첨두(cms)')
        # Sheet 2: SCS
        make_sheet(wb.create_sheet(), 'scs', 'SCS_첨두(cms)')
        # Sheet 3: Nakayasu
        make_sheet(wb.create_sheet(), 'naka', 'Nakayasu_첨두(cms)')

        # Sheet 4: 입력 매개변수
        ws4 = wb.create_sheet("입력매개변수")
        params = [
            ("유역면적 (km²)", common_vals['AREA_KM2']),
            ("유출곡선지수 CN", common_vals['CN']),
            ("단기지속기간 계산간격(분)", common_vals['DT_MIN']),
            ("단기지속기간 임계값(분)", common_vals.get('SHORT_DT_THRESHOLD', 300)),
            ("Clark Tc (hr)", common_vals['TC_HR']),
            ("Clark R (hr)", common_vals['R_HR']),
            ("SCS Tc (hr)", common_vals['SCS_TC_HR']),
            ("유로연장 L (km)", common_vals['L_KM']),
            ("Huff 보간 방법", common_vals['HUFF_INTERP']),
            ("소수점 처리", common_vals['HUFF_ROUNDING']),
        ]
        for ri, (lbl, val) in enumerate(params, 1):
            ws4.cell(ri, 1, lbl).font = Font(bold=True)
            ws4.cell(ri, 2, val)
        ws4.column_dimensions['A'].width = 24
        ws4.column_dimensions['B'].width = 20

        # ── 결과요약 시트 ────────────────────────────────────────────────────────
        ws_sum = wb.create_sheet("결과요약")
        _fn, _fs = "맑은 고딕", 11
        _hfnt2  = Font(bold=True,  name=_fn, size=_fs, color="FFFFFF")
        _hfill2 = PatternFill("solid", fgColor="4F81BD")
        _bfnt2  = Font(bold=True,  name=_fn, size=_fs)
        _dfnt2  = Font(bold=False, name=_fn, size=_fs)
        sum_hdr = ["재현기간(년)",
                   "Clark 피크(m³/s)",    "Clark 임계TR(분)",    "Clark 유출량(m³)",
                   "SCS 피크(m³/s)",      "SCS 임계TR(분)",      "SCS 유출량(m³)",
                   "Nakayasu 피크(m³/s)", "Nakayasu 임계TR(분)", "Nakayasu 유출량(m³)"]
        for ci, h in enumerate(sum_hdr, 1):
            c = ws_sum.cell(1, ci, h); c.font = _hfnt2; c.fill = _hfill2
        for ri, rp in enumerate(periods, 2):
            ws_sum.cell(ri, 1, rp).font = _bfnt2
            rp_items = [(tr, r) for (rp2, tr), r in results.items() if rp2 == rp]
            if rp_items:
                for col_off, key in [(2,'clark'),(5,'scs'),(8,'naka')]:
                    best = max(rp_items, key=lambda x, k=key: x[1][k])
                    ws_sum.cell(ri, col_off,     round(best[1][key], 3)).font             = _dfnt2
                    ws_sum.cell(ri, col_off + 1, best[0]).font                            = _dfnt2
                    ws_sum.cell(ri, col_off + 2, round(best[1].get(f'vol_{key}', 0))).font = _dfnt2
        for col, w in zip(['A','B','C','D','E','F','G','H','I','J'],
                          [14, 18, 16, 18, 18, 16, 18, 20, 18, 20]):
            ws_sum.column_dimensions[col].width = w

        # ── 수문곡선 PNG (각 방법별 고유 임계TR 기준 재계산) ────────────────────
        _png_path = None
        try:
            _disp_rp  = 50 if 50 in periods else max(periods)
            _rp_items = [(tr, r) for (rp2, tr), r in results.items() if rp2 == _disp_rp]
            if _rp_items:
                # 각 방법별 임계TR / 강우량
                _tr_c, _r_c = max(_rp_items, key=lambda x: x[1]['clark'])
                _tr_s, _r_s = max(_rp_items, key=lambda x: x[1]['scs'])
                _tr_n, _r_n = max(_rp_items, key=lambda x: x[1]['naka'])

                _short_thr = common_vals.get('SHORT_DT_THRESHOLD', 300)
                _imap      = {"PCHIP (기본)": "pchip", "선형 (Linear)": "linear",
                              "CubicSpline": "cubic"}
                _interp_m  = _imap.get(common_vals['HUFF_INTERP'], 'pchip')
                _pc_rnd    = (common_vals['HUFF_ROUNDING'] == "3째자리 반올림")
                _A         = {'A': common_vals['AREA_KM2']}

                def _mk_eng(tr, mm):
                    dt = common_vals['DT_MIN'] if tr <= _short_thr else 10
                    e  = RainfallRunoffEngine()
                    e.calculate_effective_rainfall(
                        mm, tr, dt, common_vals['CN'], self._pc_values,
                        interp_method=_interp_m, pc_rounding=_pc_rnd)
                    return e, dt

                _eng_c, _dt_c = _mk_eng(_tr_c, _r_c['mm'])
                _eng_s, _dt_s = _mk_eng(_tr_s, _r_s['mm'])
                _eng_n, _dt_n = _mk_eng(_tr_n, _r_n['mm'])

                _tc, _qc = _eng_c.convolve_runoff('Clark',
                    {**_A, 'Tc': common_vals['TC_HR'], 'R': common_vals['R_HR']})
                _ts, _qs = _eng_s.convolve_runoff('SCS',
                    {**_A, 'Tc': common_vals['SCS_TC_HR']})
                _tn, _qn = _eng_n.convolve_runoff('Nakayasu',
                    {**_A, 'L': common_vals['L_KM'],
                     't03_method': common_vals['T03_METHOD'],
                     'alpha': common_vals['ALPHA']})

                _fig, _ax1 = plt.subplots(figsize=(9, 6), dpi=150)
                _fig.patch.set_facecolor('#1c1c1e')
                _ax1.set_facecolor('#1c1c1e')
                _ax2 = _ax1.twinx()

                # ── 강우 바: TR별 그룹핑 → 공통=회색, 고유=방법 색상 ──
                _rain_groups: dict = {}
                for _tv, _eng, _dv, _mn, _clr in [
                    (_tr_c, _eng_c, _dt_c, 'Clark',    METHOD_COLORS["Clark"]),
                    (_tr_s, _eng_s, _dt_s, 'SCS',      METHOD_COLORS["SCS"]),
                    (_tr_n, _eng_n, _dt_n, 'Nakayasu', METHOD_COLORS["Nakayasu"]),
                ]:
                    if _tv not in _rain_groups:
                        _rain_groups[_tv] = {'eng': _eng, 'dt': _dv,
                                             'color': _clr, 'methods': [_mn]}
                    else:
                        _rain_groups[_tv]['methods'].append(_mn)
                        _rain_groups[_tv]['color'] = '#aaaaaa'

                _max_rain = max(
                    float(np.max(g['eng'].rainfall_data['rain_inc']))
                    for g in _rain_groups.values()
                ) or 1.0
                _ax2.set_ylim(0, _max_rain * 3.5)
                _ax2.invert_yaxis()
                _ax2.set_ylabel('강우량 / 유효우량 (mm)', color='#85c1e9')
                _ax2.tick_params(axis='y', colors='#85c1e9')
                _ax2.spines['right'].set_edgecolor('#555')

                for _tv, _g in _rain_groups.items():
                    _rd  = _g['eng'].rainfall_data
                    _gc  = _g['color']
                    _gdv = _g['dt']
                    _ax2.bar(_rd['time_min'], _rd['rain_inc'],   width=_gdv,
                             color=_gc, alpha=0.18, align='edge')
                    _ax2.bar(_rd['time_min'], _rd['excess_inc'], width=_gdv,
                             color=_gc, alpha=0.45, align='edge')
                    _ax2.text(_tv, _max_rain * 0.06, f"TR={_tv}분",
                              color=_gc, fontsize=8, fontweight='bold',
                              ha='right', va='bottom')

                # ── 수문곡선 3개 ──
                _lbl_c = f"Clark     {float(np.max(_qc)):.3f} cms  ({_disp_rp}년/{_tr_c}분)"
                _lbl_s = f"SCS       {float(np.max(_qs)):.3f} cms  ({_disp_rp}년/{_tr_s}분)"
                _lbl_n = f"Nakayasu  {float(np.max(_qn)):.3f} cms  ({_disp_rp}년/{_tr_n}분)"
                _ax1.plot(_tc, _qc, color=METHOD_COLORS["Clark"],    lw=2.0, label=_lbl_c)
                _ax1.plot(_ts, _qs, color=METHOD_COLORS["SCS"],      lw=2.0,
                          linestyle='--',  label=_lbl_s)
                _ax1.plot(_tn, _qn, color=METHOD_COLORS["Nakayasu"], lw=2.0,
                          linestyle='-.', label=_lbl_n)

                _ax1.set_xlabel('시간 (분)', color='#cccccc', fontweight='bold')
                _ax1.set_ylabel('유량 (m³/s)', color='#cccccc', fontweight='bold')
                _ax1.tick_params(axis='both', colors='#cccccc')
                _ax1.set_ylim(bottom=0)
                _ax1.grid(True, linestyle='--', alpha=0.25, color='#888')
                for _sp in _ax1.spines.values():
                    _sp.set_edgecolor('#444')

                # ── 피크 수선 ──
                _, _y_top = _ax1.get_ylim()
                for _ta, _qa, _clr in [
                    (_tc, _qc, METHOD_COLORS["Clark"]),
                    (_ts, _qs, METHOD_COLORS["SCS"]),
                    (_tn, _qn, METHOD_COLORS["Nakayasu"]),
                ]:
                    _pi  = int(np.argmax(_qa))
                    _tp  = float(_ta[_pi])
                    _qp  = float(_qa[_pi])
                    _ax1.vlines(_tp, 0, _qp, colors=_clr, linestyles=':', linewidth=1.2, alpha=0.75)
                    _ax1.text(_tp, _y_top * 0.03, f"{_tp:.0f}분",
                              color=_clr, fontsize=8, ha='center', va='bottom', fontweight='bold')

                _sid = common_vals.get('STATION_ID', '')
                plt.title(
                    f"수문곡선 비교 [배치]  ─  {_sid}  /  {_disp_rp}년 빈도",
                    color='#eeeeee', fontweight='bold')
                _l1, _b1 = _ax1.get_legend_handles_labels()
                _ax1.legend(_l1, _b1, loc='upper right',
                            facecolor='#2c2c2e', edgecolor='#555',
                            labelcolor='#dddddd', fontsize=9)
                plt.tight_layout()

                _png_stem = f"{self.project_name}_E_수문곡선"
                _png_path = os.path.join(self.project_path, f"{_png_stem}.png")
                _fig.savefig(_png_path, dpi=150, bbox_inches='tight', facecolor='#1c1c1e')
                plt.close(_fig)

                from openpyxl.drawing.image import Image as XLImage
                ws_img = wb.create_sheet("수문곡선")
                ws_img.add_image(XLImage(_png_path), 'A1')
        except Exception as _e:
            self._log(f"  ⚠ 수문곡선 PNG 생성 오류: {_e}")

        wb.save(out_path)
        self._log(f"  💾 배치 결과 저장: {fname}")
        if _png_path:
            self._log(f"  💾 수문곡선 PNG: {os.path.basename(_png_path)}")

        # config 업데이트
        try:
            cfg = {}
            if os.path.exists(self.config_file):
                with open(self.config_file, encoding='utf-8') as f:
                    cfg = json.load(f)
            cfg['step5_flood_batch'] = {
                'status':     'completed',
                'output_file': fname,
                'full_path':   out_path,
                'n_results':  len(results),
                'timestamp':  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(cfg, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

        messagebox.showinfo("배치 계산 완료",
            f"총 {len(results)}개 조합 계산 완료\n저장: {fname}", parent=self)

    # ── 종료 ─────────────────────────────────────────────────────────────────

    def _on_closing(self):
        try:
            if self.fig:
                plt.close(self.fig)
            plt.close('all')
            self.quit()
            self.destroy()
        except Exception:
            pass
        finally:
            sys.exit(0)


# ============================================================================
# 배치 계산 백엔드 — GPU(CuPy) / CPU(NumPy+ThreadPool) 자동 선택
# ============================================================================

def _detect_compute_backend():
    """CuPy(CUDA GPU) 또는 NumPy(CPU) 백엔드를 감지하여 반환.

    Returns
    -------
    (xp, use_gpu, label)
        xp       : cupy 또는 numpy 모듈
        use_gpu  : bool
        label    : 사용자에게 표시할 백엔드 문자열
    """
    try:
        import cupy as cp
        if cp.cuda.is_available():
            cp.array([1.0])  # smoke-test
            dev   = cp.cuda.Device()
            label = (f"GPU CUDA  ·  cupy {cp.__version__}"
                     f"  ·  {dev.pci_bus_id}  "
                     f"({dev.mem_info[1] // 1024**2} MB VRAM)")
            return cp, True, label
    except Exception:
        pass
    import os as _os
    workers = _os.cpu_count() or 4
    return np, False, f"CPU NumPy  ·  {workers} core"


def _batch_fft_convolve(signals: np.ndarray, kernel: np.ndarray, xp=np) -> np.ndarray:
    """FFT 기반 배치 합성곱.

    Parameters
    ----------
    signals : (N, L1)  — N개 유효우량 증분 벡터 (패딩 완료)
    kernel  : (L2,)    — 단위도 벡터
    xp      : numpy 또는 cupy (GPU)

    Returns
    -------
    (N, L1+L2-1) 합성곱 행렬 (항상 numpy ndarray로 반환)
    """
    N, L1 = signals.shape
    L2    = len(kernel)
    n_out = L1 + L2 - 1
    # FFT 크기: n_out 이상의 2의 거듭제곱
    fft_n = 1 << int(np.ceil(np.log2(n_out)))

    if xp.__name__ == 'cupy':
        sig_d  = xp.asarray(signals,       dtype=xp.float64)
        kern_d = xp.asarray(kernel,        dtype=xp.float64)
        F_sig  = xp.fft.rfft(sig_d,  n=fft_n, axis=1)
        F_kern = xp.fft.rfft(kern_d, n=fft_n)
        result = xp.fft.irfft(F_sig * F_kern[None, :], n=fft_n, axis=1)
        return xp.asnumpy(result[:, :n_out])
    else:
        F_sig  = np.fft.rfft(signals, n=fft_n, axis=1)
        F_kern = np.fft.rfft(kernel,  n=fft_n)
        result = np.fft.irfft(F_sig * F_kern[None, :], n=fft_n, axis=1)
        return result[:, :n_out]


def _run_batch_compute(combinations, common_vals, pc,
                       interp_method, pc_rounding,
                       xp, use_gpu,
                       progress_cb=None):
    """TR별 그룹화 → UH 1회 계산 → 배치 FFT 합성곱 → ThreadPool 병렬.

    Parameters
    ----------
    combinations : list of (rp, tr, mm)
    common_vals  : dict  (DT_MIN, AREA_KM2, TC_HR, R_HR, SCS_TC_HR,
                          L_KM, ALPHA, T03_METHOD, CN)
    pc           : list  Huff 누가백분율
    xp           : numpy 또는 cupy
    progress_cb  : callable(done:int, total:int, rp:int, tr:int) | None

    Returns
    -------
    dict  {(rp, tr): {'clark': float, 'scs': float, 'naka': float, 'mm': float}}
    """
    import threading
    from collections import defaultdict
    from concurrent.futures import ThreadPoolExecutor, as_completed

    dt  = common_vals['DT_MIN']
    short_threshold = common_vals.get('SHORT_DT_THRESHOLD', 300)
    A   = common_vals['AREA_KM2']
    cn  = common_vals['CN']

    # TR별 그룹화
    by_tr: dict = defaultdict(list)     # tr: [(rp, mm), ...]
    for rp, tr, mm in combinations:
        by_tr[tr].append((rp, mm))

    batch_results: dict = {}
    lock         = threading.Lock()
    done_counter = [0]
    total        = len(combinations)

    def process_tr_group(tr, rp_mm_list):
        dt_for_tr = dt if tr <= short_threshold else 10
        dt_hr = dt_for_tr / 60.0
        tr_hr = tr / 60.0

        # ── UH 1회 계산 (TR당) ───────────────────────────────────────────
        eng = RainfallRunoffEngine()
        uh_clark = eng.get_clark_uh(
            A, common_vals['TC_HR'], common_vals['R_HR'], dt_hr)
        uh_scs   = eng.get_scs_uh(
            A, common_vals['SCS_TC_HR'], tr_hr, dt_hr)
        uh_naka  = eng.get_nakayasu_uh(
            A, common_vals['L_KM'], tr_hr, dt_hr,
            common_vals['T03_METHOD'], common_vals['ALPHA'])

        # ── 유효우량 배열 계산 ────────────────────────────────────────────
        excess_list = []
        for _, mm in rp_mm_list:
            e = RainfallRunoffEngine()
            e.calculate_effective_rainfall(
                mm, tr, dt_for_tr, cn, pc,
                interp_method=interp_method,
                pc_rounding=pc_rounding)
            excess_list.append(e.rainfall_data['excess_inc'])

        # ── 행렬 패딩 → (N, max_steps) ───────────────────────────────────
        max_len = max(len(v) for v in excess_list)
        N       = len(excess_list)
        mat     = np.zeros((N, max_len))
        for i, v in enumerate(excess_list):
            mat[i, :len(v)] = v

        # ── 배치 FFT 합성곱 (GPU 또는 CPU) ───────────────────────────────
        conv_clark = _batch_fft_convolve(mat, uh_clark, xp) / 10.0
        conv_scs   = _batch_fft_convolve(mat, uh_scs,   xp) / 10.0
        conv_naka  = _batch_fft_convolve(mat, uh_naka,  xp) / 10.0

        # ── 결과 수집 ─────────────────────────────────────────────────────
        dt_sec = dt_for_tr * 60.0
        local = {}
        for i, (rp, mm) in enumerate(rp_mm_list):
            local[(rp, tr)] = {
                'clark':     float(np.max(conv_clark[i])),
                'scs':       float(np.max(conv_scs[i])),
                'naka':      float(np.max(conv_naka[i])),
                'vol_clark': float(np.sum(conv_clark[i])) * dt_sec,
                'vol_scs':   float(np.sum(conv_scs[i]))   * dt_sec,
                'vol_naka':  float(np.sum(conv_naka[i]))  * dt_sec,
                'mm':        mm,
            }
            with lock:
                done_counter[0] += 1
                if progress_cb:
                    progress_cb(done_counter[0], total, rp, tr)

        return local

    # ── ThreadPool: TR 그룹 병렬 처리 ─────────────────────────────────────
    import os as _os
    max_workers = min(len(by_tr), _os.cpu_count() or 4)
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {
            pool.submit(process_tr_group, tr, rp_mm_list): tr
            for tr, rp_mm_list in by_tr.items()
        }
        for future in as_completed(futures):
            try:
                batch_results.update(future.result())
            except Exception:
                pass   # 오류 조합 스킵

    return batch_results


# ============================================================================
# HuffDBDialog — Huff 강우시간분포 설정 자식창 (지점해석 / 지역해석 / 사용자입력)
# ============================================================================

class HuffDBDialog(ctk.CTkToplevel):
    """Huff 강우시간분포 설정 자식창.
    상단 3개 버튼(지점해석 / 지역해석 / 사용자입력)으로 패널을 전환한다."""

    _PC_COLS = ['P0','P10','P20','P30','P40','P50','P60','P70','P80','P90','P100']

    def __init__(self, parent, on_apply, project_name: str = "",
                 initial_huff: dict = None):
        super().__init__(parent)
        self.on_apply      = on_apply
        self._project_name = project_name
        self._initial      = initial_huff or {}
        self.title("Huff 강우시간분포 설정")
        self.geometry("600x500")
        self.resizable(False, False)
        self.grab_set()

        # ── 상단 네비게이션 버튼 3개 ─────────────────────────────────────────
        nav = ctk.CTkFrame(self, fg_color="#2b2b2b", corner_radius=0)
        nav.pack(fill="x")
        self._nav_btns:    dict = {}
        self._panel_frames: dict = {}
        self._active_panel = None

        for key, label in [("local", "지점해석"), ("area", "지역해석"), ("manual", "사용자입력")]:
            b = ctk.CTkButton(
                nav, text=label,
                command=lambda k=key: self._show_panel(k),
                font=FONT_BTN, height=34,
                fg_color="#3a3a3a", hover_color="#555555",
                corner_radius=0,
            )
            b.pack(side="left", fill="x", expand=True, padx=1)
            self._nav_btns[key] = b

        # ── 콘텐츠 영역 ──────────────────────────────────────────────────────
        self._content_area = ctk.CTkFrame(self, fg_color="transparent")
        self._content_area.pack(fill="both", expand=True, padx=8, pady=6)

        # 각 패널 빌드 (아직 pack하지 않음)
        self._panel_frames['local']  = self._build_local()
        self._panel_frames['area']   = self._build_area()
        self._panel_frames['manual'] = self._build_manual()

        # 기본 패널: 지점해석
        self._show_panel("local")

    # ── 패널 전환 ────────────────────────────────────────────────────────────

    def _show_panel(self, key: str):
        if self._active_panel:
            self._panel_frames[self._active_panel].pack_forget()
            self._nav_btns[self._active_panel].configure(fg_color="#3a3a3a")
        self._panel_frames[key].pack(fill="both", expand=True)
        self._nav_btns[key].configure(fg_color="#1a5276")
        self._active_panel = key

    # ── 공통 유틸 ────────────────────────────────────────────────────────────

    def _db_con(self):
        import sqlite3
        con = sqlite3.connect(DB_PATH)
        con.row_factory = sqlite3.Row
        return con

    def _preview_text(self, pc_values: list) -> str:
        return "  ".join(f"{v/100:.3f}" for v in pc_values)

    # ── 지점해석 패널 ────────────────────────────────────────────────────────

    def _build_local(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._content_area, fg_color="transparent")
        self._loc_stn        = ctk.StringVar()
        self._loc_q          = ctk.StringVar(value="3")
        self._loc_prob       = ctk.StringVar(value="50")
        self._loc_prev       = ctk.StringVar(value="")
        self._loc_interp_var = ctk.StringVar(value=self._initial.get('interp',   'PCHIP (기본)'))
        self._loc_round_var  = ctk.StringVar(value=self._initial.get('rounding', '반올림 없음'))

        try:
            con  = self._db_con()
            stns = [f"{r['NAME_STN']} ({r['STATN']})" for r in con.execute(
                "SELECT DISTINCT STATN, NAME_STN FROM Huff_Local ORDER BY NAME_STN"
            ).fetchall()]
            con.close()
        except Exception:
            stns = []

        def row(lbl, var, vals, w=200):
            f = ctk.CTkFrame(frame, fg_color="transparent")
            f.pack(fill="x", padx=8, pady=3)
            ctk.CTkLabel(f, text=lbl, font=FONT_SMALL, width=110, anchor="w").pack(side="left")
            ctk.CTkComboBox(f, values=vals, variable=var, font=FONT_SMALL,
                            width=w, command=lambda _=None: self._refresh_local()
                            ).pack(side="left", padx=4)

        row("관측소", self._loc_stn, stns or ["─"])
        default_stn = stns[0] if stns else ""
        if self._project_name and stns:
            pname = self._project_name.lower()
            for stn in stns:
                if pname in stn.lower() or stn.split("(")[0].strip().lower() in pname:
                    default_stn = stn
                    break
        if default_stn:
            self._loc_stn.set(default_stn)

        row("분위 (Quartile)",    self._loc_q,    ["1","2","3","4"], 130)
        row("확률 % (EXCD_PROB)", self._loc_prob,
            ["10","20","30","40","50","60","70","80","90"], 130)
        row("보간 방법",          self._loc_interp_var,
            ["PCHIP (기본)", "선형 (Linear)", "CubicSpline"], 160)
        row("소수점 처리",        self._loc_round_var,
            ["3째자리 반올림", "반올림 없음"], 160)

        ctk.CTkFrame(frame, height=1, fg_color="#444").pack(fill="x", padx=8, pady=6)
        ctk.CTkLabel(frame, text="미리보기 (P0 → P100):", font=FONT_SMALL,
                     anchor="w").pack(fill="x", padx=10)
        self._loc_lbl = ctk.CTkLabel(frame, text="─", font=FONT_LOG,
                                     text_color="#aaddff", anchor="w", wraplength=560)
        self._loc_lbl.pack(fill="x", padx=10, pady=(2, 6))
        ctk.CTkButton(frame, text="적용", command=self._apply_local,
                      font=FONT_BTN, height=32,
                      fg_color="#1a5276", hover_color="#2471a3").pack(pady=4)
        self._refresh_local()
        return frame

    def _refresh_local(self, _=None):
        try:
            raw   = self._loc_stn.get()
            statn = int(raw.split("(")[-1].rstrip(")").strip())
            q     = int(self._loc_q.get())
            prob  = int(self._loc_prob.get())
            con   = self._db_con()
            row   = con.execute(
                f"SELECT {','.join(self._PC_COLS)} FROM Huff_Local "
                "WHERE STATN=? AND QUARTILE=? AND EXCD_PROB=?",
                (statn, q, prob)).fetchone()
            con.close()
            if row:
                vals = [row[c] for c in self._PC_COLS]
                self._loc_lbl.configure(text=self._preview_text(vals))
                self._loc_prev.set(",".join(str(v) for v in vals))
            else:
                self._loc_lbl.configure(text="(해당 데이터 없음)")
                self._loc_prev.set("")
        except Exception:
            self._loc_lbl.configure(text="(오류)")

    def _apply_local(self):
        raw = self._loc_prev.get()
        if not raw:
            return
        try:
            vals = [float(v)/100 for v in raw.split(",")]
            self.on_apply({
                'pc':       vals,
                'interp':   self._loc_interp_var.get(),
                'rounding': self._loc_round_var.get(),
            })
            self.destroy()
        except Exception:
            pass

    # ── 지역해석 패널 ────────────────────────────────────────────────────────

    def _build_area(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._content_area, fg_color="transparent")
        self._area_gid        = ctk.StringVar()
        self._area_q          = ctk.StringVar(value="3")
        self._area_prob       = ctk.StringVar(value="50")
        self._area_prev       = ctk.StringVar(value="")
        self._area_interp_var = ctk.StringVar(value=self._initial.get('interp',   'PCHIP (기본)'))
        self._area_round_var  = ctk.StringVar(value=self._initial.get('rounding', '반올림 없음'))

        try:
            con  = self._db_con()
            gids = [str(r[0]) for r in con.execute(
                "SELECT DISTINCT GROUP_ID FROM HUFF_Area ORDER BY GROUP_ID"
            ).fetchall()]
            con.close()
        except Exception:
            gids = []

        def row(lbl, var, vals, w=130):
            f = ctk.CTkFrame(frame, fg_color="transparent")
            f.pack(fill="x", padx=8, pady=3)
            ctk.CTkLabel(f, text=lbl, font=FONT_SMALL, width=110, anchor="w").pack(side="left")
            ctk.CTkComboBox(f, values=vals, variable=var, font=FONT_SMALL,
                            width=w, command=lambda _=None: self._refresh_area()
                            ).pack(side="left", padx=4)

        row("지역그룹 (GROUP_ID)", self._area_gid, gids or ["─"])
        if gids:
            self._area_gid.set(gids[0])
        row("분위 (Quartile)",    self._area_q,    ["1","2","3","4"])
        row("확률 % (PROB)",      self._area_prob,
            ["10","20","30","40","50","60","70","80","90"])
        row("보간 방법",          self._area_interp_var,
            ["PCHIP (기본)", "선형 (Linear)", "CubicSpline"], 160)
        row("소수점 처리",        self._area_round_var,
            ["3째자리 반올림", "반올림 없음"], 160)

        ctk.CTkFrame(frame, height=1, fg_color="#444").pack(fill="x", padx=8, pady=6)
        ctk.CTkLabel(frame, text="미리보기 (P0 → P100):", font=FONT_SMALL,
                     anchor="w").pack(fill="x", padx=10)
        self._area_lbl = ctk.CTkLabel(frame, text="─", font=FONT_LOG,
                                      text_color="#aaddff", anchor="w", wraplength=560)
        self._area_lbl.pack(fill="x", padx=10, pady=(2, 6))
        ctk.CTkButton(frame, text="적용", command=self._apply_area,
                      font=FONT_BTN, height=32,
                      fg_color="#1a5276", hover_color="#2471a3").pack(pady=4)
        self._refresh_area()
        return frame

    def _refresh_area(self, _=None):
        try:
            gid  = int(self._area_gid.get())
            q    = int(self._area_q.get())
            prob = int(self._area_prob.get())
            con  = self._db_con()
            row  = con.execute(
                f"SELECT {','.join(self._PC_COLS)} FROM HUFF_Area "
                "WHERE GROUP_ID=? AND QUARTILE=? AND PROB=?",
                (gid, q, prob)).fetchone()
            con.close()
            if row:
                vals = [row[c] for c in self._PC_COLS]
                self._area_lbl.configure(text=self._preview_text(vals))
                self._area_prev.set(",".join(str(v) for v in vals))
            else:
                self._area_lbl.configure(text="(해당 데이터 없음)")
                self._area_prev.set("")
        except Exception:
            self._area_lbl.configure(text="(오류)")

    def _apply_area(self):
        raw = self._area_prev.get()
        if not raw:
            return
        try:
            vals = [float(v)/100 for v in raw.split(",")]
            self.on_apply({
                'pc':       vals,
                'interp':   self._area_interp_var.get(),
                'rounding': self._area_round_var.get(),
            })
            self.destroy()
        except Exception:
            pass

    # ── 사용자입력 패널 ──────────────────────────────────────────────────────

    def _build_manual(self) -> ctk.CTkFrame:
        frame = ctk.CTkFrame(self._content_area, fg_color="transparent")
        init  = self._initial

        def opt_row(lbl, var, vals):
            r = ctk.CTkFrame(frame, fg_color="transparent")
            r.pack(fill="x", padx=8, pady=3)
            ctk.CTkLabel(r, text=lbl, font=FONT_SMALL, width=110, anchor="w").pack(side="left")
            ctk.CTkOptionMenu(r, values=vals, variable=var,
                              font=FONT_SMALL, width=160).pack(side="left", padx=4)

        self._man_q_var = ctk.StringVar(value=init.get('quartile', '3분위'))
        self._man_q_var.trace_add("write", self._on_manual_q_change)
        opt_row("분위 선택", self._man_q_var, list(HUFF_PRESETS.keys()))

        self._man_prob_var = ctk.StringVar(value=init.get('prob', '50'))
        opt_row("확률 % (EXCD_PROB)", self._man_prob_var,
                ["10","20","30","40","50","60","70","80","90"])

        self._man_interp_var = ctk.StringVar(value=init.get('interp', 'PCHIP (기본)'))
        opt_row("보간 방법", self._man_interp_var,
                ["PCHIP (기본)", "선형 (Linear)", "CubicSpline"])

        self._man_round_var = ctk.StringVar(value=init.get('rounding', '반올림 없음'))
        opt_row("소수점 처리", self._man_round_var, ["3째자리 반올림", "반올림 없음"])

        ctk.CTkLabel(frame, text="누가 백분율 (0~1, 공백/쉼표 구분):",
                     font=FONT_SMALL, anchor="w").pack(fill="x", padx=10, pady=(6, 1))
        self._man_txt = ctk.CTkTextbox(frame, height=56, font=FONT_LOG)
        self._man_txt.pack(fill="x", padx=8, pady=(0, 4))
        self._fill_txt(init.get('pc', list(HUFF_PRESETS.get('3분위', []))))

        ctk.CTkButton(frame, text="적용", command=self._apply_manual,
                      font=FONT_BTN, height=32,
                      fg_color="#1a5276", hover_color="#2471a3").pack(pady=4)
        return frame

    def _fill_txt(self, pc: list):
        text = "  ".join(f"{v:.3f}" for v in pc)
        self._man_txt.configure(state="normal")
        self._man_txt.delete("1.0", "end")
        self._man_txt.insert("1.0", text)

    def _on_manual_q_change(self, *_):
        preset = HUFF_PRESETS.get(self._man_q_var.get())
        if preset is not None:
            self._fill_txt(preset)

    def _apply_manual(self):
        raw    = self._man_txt.get("1.0", "end")
        tokens = re.split(r'[,\s\t\n]+', raw.strip())
        pc = []
        for t in tokens:
            t = t.strip()
            if t:
                try:
                    pc.append(float(t))
                except ValueError:
                    pass
        if not pc:
            messagebox.showwarning("입력 오류", "PC 값이 없습니다.", parent=self)
            return
        self.on_apply({
            'pc':       pc,
            'quartile': self._man_q_var.get(),
            'prob':     self._man_prob_var.get(),
            'interp':   self._man_interp_var.get(),
            'rounding': self._man_round_var.get(),
        })
        self.destroy()


# ============================================================================
# BatchSelectionDialog — 다중 재현기간×지속기간 선택창
# ============================================================================
# SCSNakayasuDialog — SCS / Nakayasu 매개변수 설정 자식창
# ============================================================================

class SCSNakayasuDialog(ctk.CTkToplevel):
    """SCS / Nakayasu 합성단위도 매개변수 설정 자식창."""

    def __init__(self, parent, cfg: dict, on_apply):
        super().__init__(parent)
        self.on_apply = on_apply
        self._cfg = dict(cfg)
        self.title("SCS / Nakayasu 합성단위도 매개변수")
        self.geometry("360x340")
        self.resizable(False, False)
        self.grab_set()

        frame = ctk.CTkFrame(self, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=16, pady=12)

        def add_row(parent_f, label, key):
            r = ctk.CTkFrame(parent_f, fg_color="transparent")
            r.pack(fill="x", pady=3)
            ctk.CTkLabel(r, text=label, font=FONT_SMALL,
                         width=180, anchor="w").pack(side="left")
            e = ctk.CTkEntry(r, font=FONT_SMALL, width=110, justify="right")
            e.insert(0, str(self._cfg.get(key, "")))
            e.pack(side="right")
            return e

        def sep():
            ctk.CTkFrame(frame, height=1, fg_color="#444").pack(fill="x", pady=6)

        ctk.CTkLabel(frame, text="[ SCS 합성단위도 ]", font=FONT_HEADER,
                     text_color="#5dade2", anchor="w").pack(fill="x", pady=(0, 4))
        self._e_scs_tc = add_row(frame, "도달시간 Tc (hr)", "SCS_TC_HR")

        sep()

        ctk.CTkLabel(frame, text="[ Nakayasu 합성단위도 ]", font=FONT_HEADER,
                     text_color="#5dade2", anchor="w").pack(fill="x", pady=(0, 4))
        self._e_l     = add_row(frame, "유로연장 L (km)", "L_KM")
        self._e_alpha = add_row(frame, "alpha 값",        "ALPHA")

        t03_row = ctk.CTkFrame(frame, fg_color="transparent")
        t03_row.pack(fill="x", pady=3)
        ctk.CTkLabel(t03_row, text="T0.3 산정 방법", font=FONT_SMALL,
                     width=180, anchor="w").pack(side="left")
        self._t03_var = ctk.StringVar(value=self._cfg.get('T03_METHOD', 'alpha'))
        ctk.CTkOptionMenu(t03_row, values=["alpha", "empirical"],
                          variable=self._t03_var,
                          font=FONT_SMALL, width=110).pack(side="right")

        sep()

        self._wb_var = ctk.BooleanVar(value=bool(self._cfg.get('WB_CHECK', False)))
        ctk.CTkCheckBox(
            frame,
            text="물수지 검토 수행",
            variable=self._wb_var,
            font=FONT_SMALL,
            checkbox_width=16, checkbox_height=16,
        ).pack(anchor="w", pady=(4, 0))

        ctk.CTkButton(frame, text="적용", command=self._apply,
                      font=FONT_BTN, height=34,
                      fg_color="#1a5276", hover_color="#2471a3").pack(pady=(8, 0))

    def _apply(self):
        try:
            self._cfg['SCS_TC_HR']  = float(self._e_scs_tc.get())
            self._cfg['L_KM']       = float(self._e_l.get())
            self._cfg['ALPHA']      = float(self._e_alpha.get())
            self._cfg['T03_METHOD'] = self._t03_var.get()
            self._cfg['WB_CHECK']   = self._wb_var.get()
            self.on_apply(self._cfg)
            self.destroy()
        except ValueError as exc:
            messagebox.showerror("입력 오류", f"숫자 형식 오류:\n{exc}", parent=self)


# ============================================================================

class BatchSelectionDialog(ctk.CTkToplevel):
    """배치 계산을 위한 재현기간·지속기간 복수 선택창."""

    def __init__(self, parent, default_periods, default_durations, on_confirm):
        super().__init__(parent)
        self.on_confirm = on_confirm
        self.title("배치 계산 — 재현기간·지속기간 선택")
        self.geometry("520x480")
        self.grab_set()

        # BooleanVar 딕셔너리 — Step 3에서 동적으로 전달된 목록 사용
        self._rp_vars = {r: ctk.BooleanVar(value=True)
                         for r in sorted(default_periods)}
        self._tr_vars = {d: ctk.BooleanVar(value=True)
                         for d in sorted(default_durations)}

        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=12, pady=12)
        main.grid_columnconfigure((0, 1), weight=1)
        main.grid_rowconfigure(0, weight=1)

        # 재현기간 컬럼
        self._build_col(main, 0, "재현기간 (년)", self._rp_vars,
                        lambda v: f"{v}년")
        # 지속기간 컬럼
        self._build_col(main, 1, "강우지속기간 (분)", self._tr_vars,
                        lambda v: f"{v}분")

        # 확인/취소
        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=12, pady=(0, 12))
        ctk.CTkButton(btn_row, text="확인", command=self._confirm,
                      font=FONT_BTN, fg_color="#1a5276",
                      hover_color="#2471a3").pack(side="right", padx=4)
        ctk.CTkButton(btn_row, text="취소", command=self.destroy,
                      font=FONT_BTN, fg_color="transparent",
                      border_width=1).pack(side="right", padx=4)

    def _build_col(self, parent, col, title, var_dict, fmt):
        frame = ctk.CTkFrame(parent)
        frame.grid(row=0, column=col, sticky="nsew", padx=4)
        frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(frame, text=title, font=FONT_HEADER,
                     text_color="#5dade2").grid(row=0, column=0, columnspan=2,
                                                pady=(8, 4))
        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=4, pady=4)

        for val, var in var_dict.items():
            ctk.CTkCheckBox(scroll, text=fmt(val), variable=var,
                            font=FONT_SMALL,
                            checkbox_width=14, checkbox_height=14,
                            ).pack(anchor="w", pady=1)

        btn_f = ctk.CTkFrame(frame, fg_color="transparent")
        btn_f.grid(row=2, column=0, columnspan=2, pady=4)
        ctk.CTkButton(btn_f, text="전체선택", width=70, height=24,
                      font=FONT_SMALL,
                      command=lambda d=var_dict: [v.set(True) for v in d.values()]
                      ).pack(side="left", padx=2)
        ctk.CTkButton(btn_f, text="전체해제", width=70, height=24,
                      font=FONT_SMALL, fg_color="transparent", border_width=1,
                      command=lambda d=var_dict: [v.set(False) for v in d.values()]
                      ).pack(side="left", padx=2)

    def _confirm(self):
        sel_rp = [r for r, v in self._rp_vars.items() if v.get()]
        sel_tr = [d for d, v in self._tr_vars.items() if v.get()]
        if not sel_rp or not sel_tr:
            messagebox.showwarning("선택 오류", "재현기간과 지속기간을 1개 이상 선택하세요.",
                                   parent=self)
            return
        self.destroy()
        self.on_confirm(sel_rp, sel_tr)


# ============================================================================
# Entry point
# ============================================================================

if __name__ == "__main__":
    plt.rcParams['axes.unicode_minus'] = False

    _project_path = sys.argv[1] if len(sys.argv) > 1 else ""
    _input_file   = sys.argv[2] if len(sys.argv) > 2 else ""

    try:
        app = FloodDischargeApp(_project_path, _input_file)
        app.mainloop()
    except KeyboardInterrupt:
        pass
    except Exception:
        traceback.print_exc()
    finally:
        plt.close('all')
        sys.exit(0)
