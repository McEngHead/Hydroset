import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib.backends.backend_tkagg as tkagg
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class RainfallRunoffEngine:
    """강우-유출 해석 엔진 (SCS, Nakayasu만 포함)"""
    
    def __init__(self):
        self.rainfall_data = {}
        self.uh_data = {}
        self.convolution_data = {}
        
    def calculate_effective_rainfall(self, total_precip, tr_min, dt_min, cn, huff_pc=None):
        """Huff 분포와 SCS CN 방법을 이용한 유효우량 산정"""
        if huff_pc is None:
            huff_pc = [0.000, 0.027, 0.076, 0.142, 0.237, 0.449, 0.686, 0.838, 0.916, 0.962, 1.000]
        
        pc_huff = np.array(huff_pc, dtype=float)
        n_step = int(tr_min / dt_min) + 1
        
        t_sim_norm = np.linspace(0.0, 1.0, n_step)
        t_huff_norm = np.linspace(0.0, 1.0, len(pc_huff))
        pc_interp = np.interp(t_sim_norm, t_huff_norm, pc_huff)
        
        cumulative_rainfall = pc_interp * total_precip
        
        S = (25400.0 / cn) - 254.0
        Ia = 0.2 * S
        
        cumulative_excess = np.zeros_like(cumulative_rainfall)
        for i, P in enumerate(cumulative_rainfall):
            if P > Ia:
                cumulative_excess[i] = ((P - Ia) ** 2) / (P - Ia + S)
            else:
                cumulative_excess[i] = 0.0
                
        rain_inc = np.diff(cumulative_rainfall, prepend=0)
        excess_inc = np.diff(cumulative_excess, prepend=0)
        time_min = np.arange(0, n_step * dt_min, dt_min)
        
        self.rainfall_data = {
            'time_min': time_min,
            'rain_inc': rain_inc,
            'excess_inc': excess_inc,
            'cumulative_rainfall': cumulative_rainfall,
            'cumulative_excess': cumulative_excess,
            'pc_interp': pc_interp,
            'huff_pc': huff_pc,
            'tr_min': tr_min,
            'dt_min': dt_min,
            'S': S,
            'Ia': Ia
        }
        
        return time_min, rain_inc, excess_inc
        
    def get_scs_uh(self, A, Tc, tr_hr, dt_hr):
        """SCS 단위유량도 (수정된 무차원 수문곡선 + 체적 보정)"""
        tp = 0.6 * Tc
        Tp = (tr_hr / 2.0) + tp
        Qp = (2.08 * A) / Tp
        
        # 수정된 SCS 무차원 수문곡선 (이미지2 참조)
        scs_ratios = np.array([
            [0.0, 0.0], [0.1, 0.03], [0.2, 0.10], [0.3, 0.19], [0.4, 0.31],
            [0.5, 0.47], [0.6, 0.66], [0.7, 0.82], [0.8, 0.93], [0.9, 0.99],
            [1.0, 1.00], [1.1, 0.99], [1.2, 0.93], [1.3, 0.86], [1.4, 0.78],
            [1.5, 0.68], [1.6, 0.56], [1.7, 0.46], [1.8, 0.39], [1.9, 0.33],
            [2.0, 0.28], [2.2, 0.207], [2.4, 0.147], [2.6, 0.107], [2.8, 0.077],
            [3.0, 0.055], [3.2, 0.04], [3.4, 0.029], [3.6, 0.021], [3.8, 0.015],
            [4.0, 0.011], [4.5, 0.005], [5.0, 0.0]
        ])
        
        t_real = scs_ratios[:, 0] * Tp
        q_real = scs_ratios[:, 1] * Qp
        
        f = interp1d(t_real, q_real, kind='linear', bounds_error=False, fill_value=0)
        t_base = t_real[-1]
        time_axis = np.arange(0, t_base + dt_hr, dt_hr)
        uh = f(time_axis)
        
        # 체적 보정: 정확히 1cm가 되도록
        volume_m3 = np.sum(uh) * dt_hr * 3600  # cms * s = m³
        depth_m = volume_m3 / (A * 1e6)        # m³ / m² = m
        target_depth_m = 0.01                   # 1cm = 0.01m
        correction_factor = target_depth_m / depth_m
        uh = uh * correction_factor
        
        self.uh_data['scs'] = {
            'time_hr': time_axis, 'uh': uh,
            'tp': tp, 'Tp': Tp, 'Qp': Qp,
            'scs_ratios': scs_ratios,
            'correction_factor': correction_factor
        }
        
        return uh

    def get_nakayasu_uh(self, A, L, tr_hr, dt_hr, t03_method='alpha', alpha=1.5):
        """Nakayasu 단위유량도 (벡터 연산 + 체적 보정)"""
        if L > 15:
            tg = 0.4 + 0.058 * L
        else:
            tg = 0.21 * (L ** 0.7)
            
        Tp = tg + 0.8 * tr_hr
        
        # T0.3 계산 (유연화)
        if t03_method == 'alpha':
            T03 = alpha * tg
        else:  # 'empirical'
            T03 = 0.47 * (A * L) ** 0.25
        
        Qp = (A * 10.0) / (3.6 * (0.3 * Tp + T03))
        
        max_time = Tp + 5.0 * T03
        time_axis = np.arange(0, max_time + dt_hr, dt_hr)
        
        # 벡터 연산으로 단위유량도 계산
        ratio = (time_axis - Tp) / T03
        
        uh = np.where(
            time_axis <= Tp,
            # 상승부: Q = Qp * (t/Tp)^2.4
            Qp * ((time_axis / Tp) ** 2.4),
            # 하강부: 3단계 감쇠
            np.where(
                ratio <= 1.0,
                # 0 < (t-Tp)/T0.3 <= 1.0
                Qp * (0.3 ** ratio),
                np.where(
                    ratio <= 2.5,
                    # 1.0 < (t-Tp)/T0.3 <= 2.5
                    Qp * (0.3 ** ((time_axis - Tp + 0.5 * T03) / (1.5 * T03))),
                    # 2.5 < (t-Tp)/T0.3
                    Qp * (0.3 ** ((time_axis - Tp + 1.5 * T03) / (2.0 * T03)))
                )
            )
        )
        
        # 체적 보정: 정확히 1cm가 되도록
        volume_m3 = np.sum(uh) * dt_hr * 3600  # cms * s = m³
        depth_m = volume_m3 / (A * 1e6)        # m³ / m² = m
        target_depth_m = 0.01                   # 1cm = 0.01m
        correction_factor = target_depth_m / depth_m
        uh = uh * correction_factor
        
        self.uh_data['nakayasu'] = {
            'time_hr': time_axis, 'uh': uh,
            'tg': tg, 'Tp': Tp, 'T03': T03, 'Qp': Qp,
            'correction_factor': correction_factor
        }
        
        return uh

    def convolve_runoff(self, method_params):
        """유효우량과 단위유량도 합성"""
        if not self.rainfall_data:
            raise ValueError("먼저 calculate_effective_rainfall을 실행하여 유효우량을 산정해야 합니다.")
        
        excess_inc = self.rainfall_data['excess_inc']
        dt_min = self.rainfall_data['dt_min']
        tr_min = self.rainfall_data['tr_min']  # 수정: TR (강우지속기간) 사용
        
        tr_hr = tr_min / 60.0  # D (hr) = TR/60
        dt_hr = dt_min / 60.0
        
        method = method_params.get('type')
        A = method_params.get('A')
        
        if method == 'SCS':
            Tc = method_params['Tc']
            uh = self.get_scs_uh(A, Tc, tr_hr, dt_hr)
            
        elif method == 'Nakayasu':
            L = method_params['L']
            t03_method = method_params.get('t03_method', 'alpha')
            alpha = method_params.get('alpha', 1.5)
            uh = self.get_nakayasu_uh(A, L, tr_hr, dt_hr, t03_method, alpha)
            
        else:
            raise ValueError(f"Unknown method: {method}")
        
        # 전체 Convolution 계산 (0.01 이하가 될 때까지)
        full_convolution = np.convolve(excess_inc, uh) / 10.0
        
        # 0.01 이하가 되는 지점 찾기
        threshold = 0.01
        end_idx = len(full_convolution)
        
        # 뒤에서부터 검색하여 마지막으로 threshold를 초과하는 지점 찾기
        for i in range(len(full_convolution) - 1, -1, -1):
            if full_convolution[i] > threshold:
                end_idx = i + 10  # 여유있게 10개 더 추가
                break
        
        # end_idx가 전체 길이를 초과하지 않도록 제한
        end_idx = min(end_idx, len(full_convolution))
        
        # 결과 자르기
        direct_runoff = full_convolution[:end_idx]
        
        # Convolution matrix 생성 (확장된 길이로)
        nq = len(direct_runoff)
        n_excess = len(excess_inc)
        conv_matrix = np.zeros((nq, n_excess))
        
        for i, excess in enumerate(excess_inc):
            if excess > 0:
                for j in range(min(len(uh), nq - i)):
                    conv_matrix[i + j, i] = excess * uh[j] / 10.0
        
        # Convolution 데이터 저장
        self.convolution_data[method] = {
            'conv_matrix': conv_matrix,
            'direct_runoff': direct_runoff
        }
        
        total_steps = len(direct_runoff)
        runoff_time_min = np.arange(0, total_steps * dt_min, dt_min)
        
        return runoff_time_min, direct_runoff, uh


class HydrologyGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("합성단위유량도 분석 프로그램 (SCS, Nakayasu)")
        self.root.geometry("1200x800")
        
        self.engine = RainfallRunoffEngine()
        self.results = {}
        
        self.create_widgets()
        
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        left_frame = ttk.LabelFrame(main_frame, text="입력 매개변수", padding="10")
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5)
        
        right_frame = ttk.LabelFrame(main_frame, text="분석 결과", padding="10")
        right_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5)
        
        self.entries = {}
        
        ttk.Label(left_frame, text="=== 기본 정보 ===", font=('', 10, 'bold')).grid(row=0, column=0, columnspan=2, pady=5)
        self.add_entry(left_frame, "관측소 ID", "STATION_ID", "080000", 1)
        self.add_entry(left_frame, "유역면적 (km²)", "AREA_KM2", "0.48", 2)
        
        ttk.Label(left_frame, text="=== 강우 정보 ===", font=('', 10, 'bold')).grid(row=3, column=0, columnspan=2, pady=5)
        self.add_entry(left_frame, "총 강우량 (mm)", "TOTAL_PRECIP", "65.59", 4)
        self.add_entry(left_frame, "강우지속기간 (분)", "TR_MIN", "40", 5)
        self.add_entry(left_frame, "계산시간간격 (분)", "DT_MIN", "1", 6)
        self.add_entry(left_frame, "유출곡선지수 (CN)", "CN", "91.61", 7)
        
        ttk.Label(left_frame, text="=== Huff 분포 (쉼표 구분) ===", font=('', 10, 'bold')).grid(row=8, column=0, columnspan=2, pady=5)
        self.add_entry(left_frame, "Huff 분포", "HUFF_PC", "0.000,0.018,0.056,0.108,0.191,0.328,0.492,0.736,0.901,0.969,1.000", 9, width=50)
        
        ttk.Label(left_frame, text="=== SCS 매개변수 ===", font=('', 10, 'bold')).grid(row=10, column=0, columnspan=2, pady=5)
        self.add_entry(left_frame, "도달시간 Tc (hr)", "TC_HR", "0.2525", 11)
        
        ttk.Label(left_frame, text="=== Nakayasu 매개변수 ===", font=('', 10, 'bold')).grid(row=12, column=0, columnspan=2, pady=5)
        self.add_entry(left_frame, "유로연장 L (km)", "L_KM", "1.42", 13)
        self.add_entry(left_frame, "T0.3 계산방법 (alpha/empirical)", "T03_METHOD", "alpha", 14)
        self.add_entry(left_frame, "alpha 값 (1.5~3.0)", "ALPHA", "1.5", 15)
        
        button_frame = ttk.Frame(left_frame)
        button_frame.grid(row=16, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="분석 실행", command=self.run_analysis).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Excel 저장", command=self.save_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="초기화", command=self.reset_fields).pack(side=tk.LEFT, padx=5)
        
        self.result_text = tk.Text(right_frame, width=60, height=15, wrap=tk.WORD)
        self.result_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        scrollbar = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=self.result_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.result_text['yscrollcommand'] = scrollbar.set
        
        ttk.Button(right_frame, text="그래프 표시", command=self.show_plot).grid(row=1, column=0, pady=10)
        
    def add_entry(self, parent, label_text, key, default_value, row, width=20):
        ttk.Label(parent, text=label_text + ":").grid(row=row, column=0, sticky=tk.W, pady=2)
        entry = ttk.Entry(parent, width=width)
        entry.insert(0, default_value)
        entry.grid(row=row, column=1, sticky=tk.W, pady=2)
        self.entries[key] = entry
        
    def get_values(self):
        try:
            values = {}
            values['STATION_ID'] = self.entries['STATION_ID'].get()
            values['AREA_KM2'] = float(self.entries['AREA_KM2'].get())
            values['TOTAL_PRECIP'] = float(self.entries['TOTAL_PRECIP'].get())
            values['TR_MIN'] = float(self.entries['TR_MIN'].get())
            values['DT_MIN'] = float(self.entries['DT_MIN'].get())
            values['CN'] = float(self.entries['CN'].get())
            
            huff_str = self.entries['HUFF_PC'].get()
            values['HUFF_PC'] = [float(x.strip()) for x in huff_str.split(',')]
            
            values['TC_HR'] = float(self.entries['TC_HR'].get())
            values['L_KM'] = float(self.entries['L_KM'].get())
            values['T03_METHOD'] = self.entries['T03_METHOD'].get().strip().lower()
            values['ALPHA'] = float(self.entries['ALPHA'].get())
            
            return values
        except ValueError as e:
            messagebox.showerror("입력 오류", f"잘못된 입력값이 있습니다: {str(e)}")
            return None
            
    def run_analysis(self):
        values = self.get_values()
        if values is None:
            return
            
        try:
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, f"=== 분석 시작: {values['STATION_ID']} ===\n\n")
            
            time_rain, rain_inc, excess_inc = self.engine.calculate_effective_rainfall(
                total_precip=values['TOTAL_PRECIP'],
                tr_min=values['TR_MIN'],
                dt_min=values['DT_MIN'],
                cn=values['CN'],
                huff_pc=values['HUFF_PC']
            )
            
            total_rain = np.sum(rain_inc)
            total_excess = np.sum(excess_inc)
            
            self.result_text.insert(tk.END, f"총 강우량: {total_rain:.2f} mm\n")
            self.result_text.insert(tk.END, f"총 유효우량: {total_excess:.2f} mm\n")
            self.result_text.insert(tk.END, f"유효우량율: {total_excess/total_rain*100:.2f}%\n")
            self.result_text.insert(tk.END, "-" * 40 + "\n\n")
            
            # SCS
            t_scs, q_scs, uh_scs = self.engine.convolve_runoff({
                'type': 'SCS', 'A': values['AREA_KM2'], 'Tc': values['TC_HR']
            })
            
            peak_scs = np.max(q_scs)
            time_scs = t_scs[np.argmax(q_scs)]
            
            self.result_text.insert(tk.END, f"[SCS 방법]\n")
            self.result_text.insert(tk.END, f"  첨두홍수량: {peak_scs:.2f} cms\n")
            self.result_text.insert(tk.END, f"  발생시간: {time_scs:.0f} 분\n\n")
            
            # Nakayasu
            t_nakayasu, q_nakayasu, uh_nakayasu = self.engine.convolve_runoff({
                'type': 'Nakayasu', 'A': values['AREA_KM2'], 'L': values['L_KM'],
                't03_method': values['T03_METHOD'], 'alpha': values['ALPHA']
            })
            
            peak_nakayasu = np.max(q_nakayasu)
            time_nakayasu = t_nakayasu[np.argmax(q_nakayasu)]
            
            self.result_text.insert(tk.END, f"[Nakayasu 방법]\n")
            self.result_text.insert(tk.END, f"  첨두홍수량: {peak_nakayasu:.2f} cms\n")
            self.result_text.insert(tk.END, f"  발생시간: {time_nakayasu:.0f} 분\n\n")
            
            self.result_text.insert(tk.END, "=" * 40 + "\n")
            self.result_text.insert(tk.END, "분석 완료! 'Excel 저장' 버튼을 눌러 상세 계산과정을 확인하세요.")
            
            self.results = {
                'time_rain': time_rain, 'rain_inc': rain_inc, 'excess_inc': excess_inc,
                't_scs': t_scs, 'q_scs': q_scs,
                't_nakayasu': t_nakayasu, 'q_nakayasu': q_nakayasu,
                'values': values
            }
            
            messagebox.showinfo("완료", "분석이 성공적으로 완료되었습니다!")
            
        except Exception as e:
            import traceback
            messagebox.showerror("오류", f"분석 중 오류가 발생했습니다:\n{str(e)}\n\n{traceback.format_exc()}")
    
    def save_excel(self):
        """Excel 파일로 저장"""
        if not self.results:
            messagebox.showwarning("경고", "먼저 분석을 실행해주세요.")
            return
        
        folder = filedialog.askdirectory(title="Excel 파일 저장 폴더 선택")
        if not folder:
            return
        
        try:
            values = self.results['values']
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            station_str = f"{values['STATION_ID']:<6s}"[:6].replace(' ', '0')
            filename = f"UnitHydrograph_{station_str}_{timestamp}.xlsx"
            filepath = os.path.join(folder, filename)
            
            wb = Workbook()
            wb.remove(wb.active)
            
            # Raw Data 시트 생성
            self.create_raw_data_sheet(wb, values)
            
            # 2가지 방법별 시트 생성
            self.create_scs_sheet(wb, values)
            self.create_nakayasu_sheet(wb, values)
            
            # Convolution 시트 생성
            self.create_convolution_sheets(wb, values)
            
            wb.save(filepath)
            
            messagebox.showinfo("저장 완료", 
                f"Excel 파일이 저장되었습니다!\n\n"
                f"📁 위치: {folder}\n"
                f"📄 파일: {filename}\n\n"
                f"생성된 시트:\n"
                f"• Raw Data (공통 강우/유효우량)\n"
                f"• SCS 방법\n"
                f"• Nakayasu 방법\n"
                f"• Convolution (각 방법별)")
            
        except Exception as e:
            import traceback
            messagebox.showerror("저장 오류", f"Excel 저장 중 오류 발생:\n{str(e)}\n\n{traceback.format_exc()}")
    
    def create_raw_data_sheet(self, wb, values):
        """공통 Raw Data 시트 생성"""
        ws = wb.create_sheet("Raw Data")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        red_font = Font(color="FF0000", bold=True)
        black_font = Font(color="000000")
        
        for col_idx in range(1, 13):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
        
        headers = ["Time (min)", "Inc Rain", "Inc Excess", "Cum Rain", "Cum Excess"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        ws.cell(row=1, column=8, value="PARAMETER").font = header_font
        ws.cell(row=1, column=8).fill = header_fill
        ws.cell(row=1, column=9, value="VALUES").font = header_font
        ws.cell(row=1, column=9).fill = header_fill
        
        param_list = [
            ("Area (km2)", values['AREA_KM2'], True),
            ("Total Precip (mm)", values['TOTAL_PRECIP'], True),
            ("CN", values['CN'], True),
            ("dt (min)", values['DT_MIN'], True),
            ("TR (min)", values['TR_MIN'], True),
            ("S (mm)", "=(25400/I4)-254", False),
            ("Ia (mm)", "=0.2*I7", False),
        ]
        
        for i, (label, value, is_user_input) in enumerate(param_list, 2):
            ws.cell(row=i, column=8, value=label)
            value_cell = ws.cell(row=i, column=9, value=value)
            value_cell.font = red_font if is_user_input else black_font
        
        ws.cell(row=1, column=11, value="Huff Input").font = header_font
        ws.cell(row=1, column=11).fill = header_fill
        ws.cell(row=1, column=12, value="Interpolated PC").font = header_font
        ws.cell(row=1, column=12).fill = header_fill
        
        huff_pc = self.engine.rainfall_data.get('huff_pc', [])
        for i, pc_val in enumerate(huff_pc):
            cell = ws.cell(row=i+3, column=11, value=pc_val)
            cell.font = red_font
        
        pc_interp = self.engine.rainfall_data.get('pc_interp', [])
        for i, val in enumerate(pc_interp):
            row = i + 3
            ws.cell(row=row, column=1, value=i * values['DT_MIN'])
            cell = ws.cell(row=row, column=12, value=float(val))
            cell.font = black_font
        
        nq = len(pc_interp)
        for i in range(nq):
            r = i + 3
            
            ws.cell(row=r, column=4, value=f"=L{r}*$I$3").font = black_font
            
            if i > 0:
                ws.cell(row=r, column=2, value=f"=D{r}-D{r-1}").font = black_font
            else:
                ws.cell(row=r, column=2, value=f"=D{r}").font = black_font
            
            ws.cell(row=r, column=5, 
                   value=f"=IF(D{r}>$I$8, POWER(D{r}-$I$8,2)/(D{r}-$I$8+$I$7), 0)").font = black_font
            
            if i > 0:
                ws.cell(row=r, column=3, value=f"=E{r}-E{r-1}").font = black_font
            else:
                ws.cell(row=r, column=3, value="=E3").font = black_font
    
    def create_scs_sheet(self, wb, values):
        """SCS 방법 상세 계산 시트 (이미지2 형식)"""
        ws = wb.create_sheet("SCS 방법")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        red_font = Font(color="FF0000", bold=True)
        black_font = Font(color="000000")
        
        for col_idx in range(1, 10):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        
        ws.merge_cells('A1:H1')
        ws['A1'] = "SCS 합성단위유량도 계산"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 파라미터 (F열=PARAMETER, G열=VALUES)
        ws.cell(row=2, column=6, value="PARAMETER").font = header_font
        ws.cell(row=2, column=6).fill = header_fill
        ws.cell(row=2, column=7, value="VALUES").font = header_font
        ws.cell(row=2, column=7).fill = header_fill
        
        scs_params = [
            ("Area A (km²)", values['AREA_KM2'], True),   # F4, G4
            ("Tc (hr)", values['TC_HR'], True),           # F5, G5
            ("TR (min)", values['TR_MIN'], True),         # F6, G6
            ("D (hr)", "=G6/60", False),                  # F7, G7 = TR/60
            ("tp (hr)", "=0.6*G5", False),                # F8, G8 = 0.6*Tc
            ("Tp (hr)", "=(G7/2)+G8", False),             # F9, G9 = (D/2)+tp
            ("Qp (cms)", "=(2.08*G4)/G9", False),         # F10, G10 = (2.08*Area)/Tp
        ]
        
        for i, (label, value, is_user_input) in enumerate(scs_params, 4):
            ws.cell(row=i, column=6, value=label)
            value_cell = ws.cell(row=i, column=7, value=value)
            value_cell.font = red_font if is_user_input else black_font
        
        # 무차원 수문곡선 + 단위유량도 (A~D열)
        headers = ["t/Tp", "q/Qp", "시간 (hr)", "유량 (cms)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h).font = header_font
            ws.cell(row=3, column=col).fill = header_fill
        
        scs_data = self.engine.uh_data.get('scs', {})
        
        # 보간된 전체 단위유량도에서 비율 계산
        time_hr = scs_data.get('time_hr', [])
        uh = scs_data.get('uh', [])
        Tp = scs_data.get('Tp', 1)
        Qp = scs_data.get('Qp', 1)
        
        # 전체 보간된 데이터 출력 (4행부터)
        for i, (t, q) in enumerate(zip(time_hr, uh)):
            r = i + 4
            # A열: t/Tp 비율 (값)
            ratio_t = t / Tp if Tp > 0 else 0
            ws.cell(row=r, column=1, value=ratio_t).number_format = '0.00'
            
            # B열: q/Qp 비율 (값)
            ratio_q = q / Qp if Qp > 0 else 0
            ws.cell(row=r, column=2, value=ratio_q).number_format = '0.0000'
            
            # C열: 시간 (수식)
            ws.cell(row=r, column=3, value=f"=$G$9*A{r}").font = black_font
            ws.cell(row=r, column=3).number_format = '0.00'
            
            # D열: 유량 (수식)
            ws.cell(row=r, column=4, value=f"=$G$10*B{r}").font = black_font
            ws.cell(row=r, column=4).number_format = '0.0000'
    
    def create_nakayasu_sheet(self, wb, values):
        """Nakayasu 방법 상세 계산 시트"""
        ws = wb.create_sheet("Nakayasu 방법")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        red_font = Font(color="FF0000", bold=True)
        black_font = Font(color="000000")
        
        for col_idx in range(1, 10):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        ws.merge_cells('A1:H1')
        ws['A1'] = "Nakayasu 합성단위유량도 계산"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 파라미터 (F-G열)
        ws.cell(row=2, column=6, value="PARAMETER").font = header_font
        ws.cell(row=2, column=6).fill = header_fill
        ws.cell(row=2, column=7, value="VALUES").font = header_font
        ws.cell(row=2, column=7).fill = header_fill
        
        nakayasu_data = self.engine.uh_data.get('nakayasu', {})
        
        # T0.3 수식 생성 (조건부)
        t03_method = values.get('T03_METHOD', 'alpha').lower()
        if t03_method == 'alpha':
            t03_formula = "=G12*G7"  # alpha * tg
        else:  # 'empirical'
            t03_formula = "=0.47*(G4*G5)^0.25"
        
        nakayasu_params = [
            ("Area A (km²)", values['AREA_KM2'], True),                # F4, G4 (빨강)
            ("L (km)", values['L_KM'], True),                          # F5, G5 (빨강)
            ("TR (min)", values['TR_MIN'], True),                      # F6, G6 (빨강)
            ("tg (hr)", "=IF(G5>15,0.4+0.058*G5,0.21*G5^0.7)", False), # F7, G7 (수식)
            ("Tp (hr)", "=G7+0.8*G6/60", False),                       # F8, G8 (수식)
            ("T0.3 (hr)", t03_formula, False),                         # F9, G9 (수식)
            ("Qp (cms)", "=(G4*10)/(3.6*(0.3*G8+G9))", False),         # F10, G10 (수식)
            ("", "", None),                                             # F11, G11 (빈 행)
            ("T0.3 Method", values.get('T03_METHOD', 'alpha'), True),  # F12, G12 (빨강)
            ("alpha", values.get('ALPHA', 1.5), True),                 # F13, G13 (빨강)
        ]
        
        for i, (label, value, is_user_input) in enumerate(nakayasu_params, 4):
            if is_user_input is None:  # 빈 행
                continue
            ws.cell(row=i, column=6, value=label)
            value_cell = ws.cell(row=i, column=7, value=value)
            if is_user_input:
                value_cell.font = red_font
            else:
                value_cell.font = black_font
                value_cell.number_format = '0.0000'
        
        # 무차원 수문곡선 + 단위유량도 (A~D열)
        headers = ["t/Tp", "q/Qp", "시간 (hr)", "유량 (cms)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h).font = header_font
            ws.cell(row=3, column=col).fill = header_fill
        
        # 보간된 전체 단위유량도에서 비율 계산
        time_hr = nakayasu_data.get('time_hr', [])
        uh = nakayasu_data.get('uh', [])
        Tp = nakayasu_data.get('Tp', 1)
        Qp = nakayasu_data.get('Qp', 1)
        
        # 전체 보간된 데이터 출력 (4행부터)
        for i, (t, q) in enumerate(zip(time_hr, uh)):
            r = i + 4
            # A열: t/Tp 비율 (값)
            ratio_t = t / Tp if Tp > 0 else 0
            ws.cell(row=r, column=1, value=ratio_t).number_format = '0.00'
            
            # B열: q/Qp 비율 (값)
            ratio_q = q / Qp if Qp > 0 else 0
            ws.cell(row=r, column=2, value=ratio_q).number_format = '0.0000'
            
            # C열: 시간 (수식)
            ws.cell(row=r, column=3, value=f"=$G$8*A{r}").font = black_font
            ws.cell(row=r, column=3).number_format = '0.00'
            
            # D열: 유량 (수식)
            ws.cell(row=r, column=4, value=f"=$G$10*B{r}").font = black_font
            ws.cell(row=r, column=4).number_format = '0.0000'
    
    def create_convolution_sheets(self, wb, values):
        """Convolution 계산 시트 생성 (2개 방법) - 수식 기반"""
        methods = ['SCS', 'Nakayasu']
        
        for method in methods:
            ws = wb.create_sheet(f"Convolution_{method}")
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            black_font = Font(color="000000")
            
            ws.cell(row=1, column=1, value="Time (min)").font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=2, value="Total Runoff (cms)").font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            
            conv_data = self.engine.convolution_data.get(method, {})
            conv_matrix = conv_data.get('conv_matrix', np.array([]))
            
            if len(conv_matrix) == 0:
                continue
            
            nq = conv_matrix.shape[0]
            n_excess = conv_matrix.shape[1]
            
            # 헤더 - 각 강우 기간
            for idx in range(n_excess):
                col = 3 + idx
                ws.cell(row=1, column=col, value=f"P_{idx+1}").font = header_font
                ws.cell(row=1, column=col).fill = header_fill
            
            # 데이터 - 수식으로 작성
            for r_idx in range(nq):
                r_num = r_idx + 2
                
                # 시간 (분)
                ws.cell(row=r_num, column=1, value=r_idx * values['DT_MIN'])
                
                # Total Runoff = SUM(P_1:P_n)
                last_col_letter = get_column_letter(2 + n_excess)
                ws.cell(row=r_num, column=2, value=f"=SUM(C{r_num}:{last_col_letter}{r_num})").font = black_font
                ws.cell(row=r_num, column=2).number_format = '0.0000'
                
                # 각 P_i 열의 수식
                # Convolution: P_i의 row에서의 기여 = 유효우량[i] * 단위유량도[row - i] / 10
                # 조건: row >= i (시간 인덱스 >= 유효우량 인덱스)
                for c_idx in range(n_excess):
                    col = 3 + c_idx
                    
                    # Raw Data의 유효우량: C열의 (3+c_idx)행
                    # 즉, C3, C4, C5, ..., C42
                    excess_row = 3 + c_idx
                    
                    # 단위유량도 행 계산:
                    #   - 시간 인덱스 = r_idx = r_num - 2
                    #   - 유효우량 인덱스 = c_idx
                    #   - UH 인덱스 = r_idx - c_idx
                    #   - UH 행 = 4 + UH_index = 4 + (r_num-2) - c_idx = r_num + 2 - c_idx
                    uh_row = r_num + 2 - c_idx
                    
                    # 조건: r_idx >= c_idx
                    # 즉, (r_num-2) >= c_idx
                    # r_num >= c_idx + 2
                    # ROW() >= (COLUMN()-3) + 2
                    # ROW() >= COLUMN() - 1
                    # 또는 ROW() > COLUMN() - 2
                    formula = f"=IF(ROW()>COLUMN()-2, IFERROR('Raw Data'!$C${excess_row}*'{method} 방법'!D{uh_row}/10, \"\"), \"\")"
                    
                    ws.cell(row=r_num, column=col, value=formula).font = black_font
                    ws.cell(row=r_num, column=col).number_format = '0.0000'

    
    def show_plot(self):
        """그래프 표시"""
        if not self.results:
            messagebox.showwarning("경고", "먼저 분석을 실행해주세요.")
            return
            
        plot_window = tk.Toplevel(self.root)
        plot_window.title("수문곡선 분석 결과")
        plot_window.geometry("1000x700")
        
        fig = Figure(figsize=(10, 7))
        ax1 = fig.add_subplot(111)
        ax2 = ax1.twinx()
        
        DT_MIN = self.results['values']['DT_MIN']
        time_rain = self.results['time_rain']
        rain_inc = self.results['rain_inc']
        excess_inc = self.results['excess_inc']
        
        ax2.bar(time_rain, rain_inc, width=DT_MIN, color='gray', alpha=0.3, label='Total Rainfall', align='edge')
        ax2.bar(time_rain, excess_inc, width=DT_MIN, color='blue', alpha=0.5, label='Effective Rainfall', align='edge')
        ax2.set_ylabel('Rainfall (mm)')
        ax2.set_ylim(0, max(rain_inc) * 3)
        ax2.invert_yaxis()
        ax2.legend(loc='upper center')
        
        ax1.plot(self.results['t_scs'], self.results['q_scs'], 'g-.', label='SCS', linewidth=2)
        ax1.plot(self.results['t_nakayasu'], self.results['q_nakayasu'], 'k-', label='Nakayasu', linewidth=2)
        
        ax1.set_xlabel('Time (min)', fontsize=12)
        ax1.set_ylabel('Discharge (cms)', fontsize=12)
        ax1.set_title(f"Flood Hydrograph Synthesis (Area={self.results['values']['AREA_KM2']}km², CN={self.results['values']['CN']})", 
                     fontsize=14, fontweight='bold')
        ax1.legend(loc='upper right')
        ax1.grid(True, alpha=0.3)
        
        fig.tight_layout()
        
        canvas = tkagg.FigureCanvasTkAgg(fig, master=plot_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        toolbar = tkagg.NavigationToolbar2Tk(canvas, plot_window)
        toolbar.update()
        
    def reset_fields(self):
        """입력 필드 초기화"""
        defaults = {
            'STATION_ID': '080000', 'AREA_KM2': '250', 'TOTAL_PRECIP': '221.4',
            'TR_MIN': '120', 'DT_MIN': '10', 'CN': '84.3',
            'HUFF_PC': '0.000,0.009,0.053,0.101,0.184,0.331,0.538,0.756,0.916,0.975,1.000',
            'TC_HR': '9.5', 'L_KM': '3.5'
        }
        
        for key, value in defaults.items():
            self.entries[key].delete(0, tk.END)
            self.entries[key].insert(0, value)
            
        self.result_text.delete(1.0, tk.END)
        self.results = {}


if __name__ == "__main__":
    root = tk.Tk()
    app = HydrologyGUI(root)
    root.mainloop()